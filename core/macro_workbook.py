"""Load macro workbook settings (Name_Exclusions, task colors).

Ports Module7 sheet reads from the autopodbor `.xlsm`. The web UI and CLI
use these rules automatically; no extra upload is required when the workbook
is on disk (see `data/config/macro.json`).
"""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.exclusions import NameExclusionRule, TaskColorEntry

NAME_EXCLUSIONS_SHEET = "Name_Exclusions"
_CONFIG_RELATIVE_PATH = ("data", "config", "macro.json")
_DEFAULT_WORKBOOK_RELATIVE = ("data", "config", "autopodbor.xlsm")
_DEFAULT_SEARCH_DIRS = ("data", "config"), ("data", "real")


@dataclass(frozen=True)
class MacroWorkbookSettings:
    """Name-exclusion rules and task-color metadata from one macro workbook."""

    workbook_path: Path | None
    name_exclusion_rules: list[NameExclusionRule]
    task_color_entries: list[TaskColorEntry]


def repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def default_macro_config_path() -> Path:
    return repo_root().joinpath(*_CONFIG_RELATIVE_PATH)


def resolve_macro_workbook_path(config_path: str | Path | None = None) -> Path | None:
    """Pick the macro workbook path without requiring a web upload."""
    env_path = os.environ.get("ESTIMATE_AI_MACRO_WORKBOOK", "").strip()
    if env_path:
        candidate = Path(env_path)
        if candidate.is_file():
            return candidate.resolve()

    config = _load_macro_config(config_path)
    workbook_value = str(config.get("workbook") or "").strip()
    if workbook_value:
        candidate = _resolve_path(workbook_value)
        if candidate.is_file():
            return candidate.resolve()

    default_path = repo_root().joinpath(*_DEFAULT_WORKBOOK_RELATIVE)
    if default_path.is_file():
        return default_path.resolve()

    search_dirs = config.get("search_dirs")
    if isinstance(search_dirs, list) and search_dirs:
        directories = [_resolve_path(str(item)) for item in search_dirs]
    else:
        directories = [repo_root().joinpath(*parts) for parts in _DEFAULT_SEARCH_DIRS]

    return _find_newest_macro_workbook(directories)


def load_default_macro_settings(resolved_path: str | Path | None = None) -> MacroWorkbookSettings:
    """Load rules from the resolved macro workbook, or return empty settings."""
    path = Path(resolved_path) if resolved_path else resolve_macro_workbook_path()
    if path is None or not path.is_file():
        return MacroWorkbookSettings(None, [], [])

    rules, colors = load_name_exclusions_workbook(path)
    return MacroWorkbookSettings(path.resolve(), rules, colors)


def load_name_exclusions_workbook(
    workbook_path: str | Path,
) -> tuple[list[NameExclusionRule], list[TaskColorEntry]]:
    """Read enabled rules (A:F) and task colors (H:K) from Name_Exclusions."""
    path = Path(workbook_path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if NAME_EXCLUSIONS_SHEET not in workbook.sheetnames:
            return [], []
        worksheet = workbook[NAME_EXCLUSIONS_SHEET]
        return _read_name_rules(worksheet), _read_task_colors(worksheet)
    finally:
        workbook.close()


def _load_macro_config(config_path: str | Path | None) -> dict[str, Any]:
    path = Path(config_path) if config_path is not None else default_macro_config_path()
    if not path.is_file():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _resolve_path(value: str) -> Path:
    path = Path(value)
    if path.is_file():
        return path
    return repo_root() / path


def _find_newest_macro_workbook(directories: list[Path]) -> Path | None:
    candidates: list[Path] = []
    for directory in directories:
        if not directory.is_dir():
            continue
        for path in directory.glob("*.xlsm"):
            if _workbook_has_exclusions_sheet(path):
                candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime).resolve()


def _workbook_has_exclusions_sheet(path: Path) -> bool:
    workbook = load_workbook(path, read_only=True)
    try:
        return NAME_EXCLUSIONS_SHEET in workbook.sheetnames
    finally:
        workbook.close()


def _read_name_rules(worksheet) -> list[NameExclusionRule]:
    rules: list[NameExclusionRule] = []
    for row in range(2, worksheet.max_row + 1):
        enabled_value = worksheet.cell(row=row, column=1).value
        scope = worksheet.cell(row=row, column=2).value
        match_mode = worksheet.cell(row=row, column=3).value
        pattern = worksheet.cell(row=row, column=4).value
        group = worksheet.cell(row=row, column=5).value
        comment = worksheet.cell(row=row, column=6).value
        if not _is_enabled(enabled_value):
            continue
        pattern_text = _cell_text(pattern)
        if pattern_text == "":
            continue
        rules.append(
            NameExclusionRule(
                enabled=True,
                scope=_cell_text(scope) or "BOTH",
                match_mode=_cell_text(match_mode) or "ALL_WORDS",
                pattern=pattern_text,
                group=_cell_text(group),
                comment=_cell_text(comment),
            )
        )
    return rules


def _read_task_colors(worksheet) -> list[TaskColorEntry]:
    entries: list[TaskColorEntry] = []
    for row in range(2, worksheet.max_row + 1):
        enabled_value = worksheet.cell(row=row, column=8).value
        task_number = worksheet.cell(row=row, column=9).value
        reason = worksheet.cell(row=row, column=10).value
        comment = worksheet.cell(row=row, column=11).value
        if not _is_enabled(enabled_value):
            continue
        task_text = _cell_text(task_number)
        if task_text == "":
            continue
        entries.append(
            TaskColorEntry(
                enabled=True,
                task_number=task_text,
                reason=_cell_text(reason),
                comment=_cell_text(comment),
            )
        )
    return entries


def _is_enabled(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    return str(value).strip().upper() in {"1", "TRUE"}


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()

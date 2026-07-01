"""Folder-walk RNMC catalog ingestion."""

# Design notes for future persistence layers:
# - force_reimport=True only marks parser output. This module never deletes or
#   replaces stored rows; the DB/storage layer that persists IngestFileResult
#   must perform the explicit delete-old-rows-then-insert-new operation.
# - Both fully failed files and successful files are added to imported_keys.
#   A later normal run can only report skipped=True for either case; the
#   success-vs-failure history must come from a future import log table.

from dataclasses import dataclass, field
from numbers import Real
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.catalog import CatalogRow
from core.normalize import NormCode, NormUnit


NBSP = "\u00a0"
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}

HEADER_NAME_WORKS = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
HEADER_NAME_SHORT = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
HEADER_UNIT = "\u0415\u0434.\u0438\u0437\u043c."
HEADER_QTY = "\u041a\u043e\u043b-\u0432\u043e"
HEADER_QTY_LONG = "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e"
TASK_LABEL = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438 1\u0424"
TASK_LABEL_SHORT = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438"
NUMBER_SIGN = "\u2116"


@dataclass(frozen=True)
class IngestSettings:
    """Header names used to map source tables into CatalogRow fields."""

    work_name_header: str = HEADER_NAME_WORKS
    unit_header: str = HEADER_UNIT
    price_header: str = "\u0426\u0435\u043d\u0430"
    code_header: str = "\u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c"
    added_date_header: str = "CatalogAddedDate"


@dataclass(frozen=True)
class FileIdentity:
    """Import identity, intentionally scoped by region folder plus filename."""

    region_folder: str
    filename: str


@dataclass(frozen=True)
class ValidationIssue:
    """Per-row validation issue for the import report."""

    row_number: int
    reason: str


@dataclass(frozen=True)
class IngestFileResult:
    """Result of processing one file during folder ingestion."""

    file_identity: FileIdentity
    path: Path
    region: str
    task_number: str = ""
    rows: list[CatalogRow] = field(default_factory=list)
    validation_issues: list[ValidationIssue] = field(default_factory=list)
    failed: bool = False
    failure_reason: str = ""
    skipped: bool = False
    force_reimport: bool = False


@dataclass(frozen=True)
class HeaderMatch:
    row_number: int
    name_col: int
    unit_col: int
    qty_col: int


def ingest_folder(
    root_path: str | Path,
    already_imported: set[FileIdentity] | None = None,
    force_reimport_keys: set[FileIdentity] | None = None,
    settings: IngestSettings | None = None,
) -> list[IngestFileResult]:
    """Walk a root folder recursively and ingest unprocessed Excel files."""
    root = Path(root_path)
    imported_keys = set() if already_imported is None else already_imported
    force_keys = set() if force_reimport_keys is None else set(force_reimport_keys)
    active_settings = IngestSettings() if settings is None else settings
    results: list[IngestFileResult] = []

    for file_path in sorted(_iter_excel_files(root), key=lambda path: str(path).casefold()):
        identity = _file_identity(file_path)

        if file_path.parent == root:
            results.append(
                _record_processed(
                    imported_keys,
                    _failed_result(
                        file_path,
                        identity,
                        "file is not inside a region subfolder",
                        force_reimport=identity in force_keys,
                    ),
                )
            )
            continue

        if identity in imported_keys and identity not in force_keys:
            results.append(
                IngestFileResult(
                    file_identity=identity,
                    path=file_path,
                    region=identity.region_folder,
                    skipped=True,
                )
            )
            continue

        results.append(
            _record_processed(
                imported_keys,
                _process_file(
                    file_path,
                    identity,
                    active_settings,
                    force_reimport=identity in force_keys,
                ),
            )
        )

    return results


def force_reimport_file(
    file_path: str | Path,
    settings: IngestSettings | None = None,
) -> IngestFileResult:
    """Explicit single-file re-import path; callers handle clean replacement."""
    path = Path(file_path)
    return _process_file(
        path,
        _file_identity(path),
        IngestSettings() if settings is None else settings,
        force_reimport=True,
    )


def _iter_excel_files(root: Path) -> list[Path]:
    return [
        path
        for path in root.rglob("*")
        if path.is_file() and path.suffix.casefold() in EXCEL_EXTENSIONS
    ]


def _file_identity(file_path: Path) -> FileIdentity:
    return FileIdentity(
        region_folder=file_path.parent.name,
        filename=file_path.name,
    )


def _process_file(
    file_path: Path,
    identity: FileIdentity,
    settings: IngestSettings,
    *,
    force_reimport: bool = False,
) -> IngestFileResult:
    workbook: Workbook | None = None

    try:
        workbook = load_workbook(file_path, data_only=False, read_only=False)
        task_number = _extract_task_number(workbook)

        for worksheet in workbook.worksheets:
            header_match = _find_header_row(worksheet)
            if header_match is None:
                continue

            rows, issues = _extract_rows(
                worksheet,
                header_match,
                task_number,
                identity,
                settings,
            )
            if rows or issues:
                return IngestFileResult(
                    file_identity=identity,
                    path=file_path,
                    region=identity.region_folder,
                    task_number=task_number,
                    rows=rows,
                    validation_issues=issues,
                    force_reimport=force_reimport,
                )

        return _failed_result(
            file_path,
            identity,
            "no matching header row found",
            task_number=task_number,
            force_reimport=force_reimport,
        )
    except Exception as exc:
        return _failed_result(
            file_path,
            identity,
            str(exc),
            force_reimport=force_reimport,
        )
    finally:
        if workbook is not None:
            workbook.close()


def _record_processed(
    imported_keys: set[FileIdentity],
    result: IngestFileResult,
) -> IngestFileResult:
    imported_keys.add(result.file_identity)
    return result


def _failed_result(
    file_path: Path,
    identity: FileIdentity,
    reason: str,
    *,
    task_number: str = "",
    force_reimport: bool = False,
) -> IngestFileResult:
    return IngestFileResult(
        file_identity=identity,
        path=file_path,
        region=identity.region_folder,
        task_number=task_number,
        failed=True,
        failure_reason=reason,
        force_reimport=force_reimport,
    )


def _find_header_row(worksheet: Worksheet) -> HeaderMatch | None:
    max_row = min(worksheet.max_row, 400)
    max_col = min(worksheet.max_column, 150)

    for row_number in range(1, max_row + 1):
        name_col = 0
        unit_col = 0
        qty_col = 0

        for col_number in range(1, max_col + 1):
            text = _normalize_header_text(_cell_value(worksheet, row_number, col_number))
            if _is_name_header(text):
                name_col = col_number
            if _is_unit_header(text):
                unit_col = col_number
            if _is_qty_header(text):
                qty_col = col_number

        if name_col > 0 and unit_col > 0 and qty_col > 0:
            return HeaderMatch(
                row_number=row_number,
                name_col=name_col,
                unit_col=unit_col,
                qty_col=qty_col,
            )

    return None


def _extract_rows(
    worksheet: Worksheet,
    header_match: HeaderMatch,
    task_number: str,
    identity: FileIdentity,
    settings: IngestSettings,
) -> tuple[list[CatalogRow], list[ValidationIssue]]:
    source_map = _build_source_header_map(worksheet, header_match.row_number)
    numbering_col = _find_numbering_col(source_map) or 1
    field_columns = _field_columns(source_map, settings)
    rows: list[CatalogRow] = []
    issues: list[ValidationIssue] = []
    started = False
    blank_streak = 0

    for row_number in range(header_match.row_number + 1, worksheet.max_row + 1):
        if not started:
            if (
                _is_empty_or_blank(_cell_value(worksheet, row_number, numbering_col))
                and _is_empty_or_blank(_cell_value(worksheet, row_number, header_match.unit_col))
                and _is_empty_or_blank(_cell_value(worksheet, row_number, header_match.qty_col))
            ):
                continue
            started = True

        is_blank_end_row = (
            _is_empty_or_blank(_cell_value(worksheet, row_number, numbering_col))
            and _is_empty_or_blank(_cell_value(worksheet, row_number, header_match.name_col))
            and _is_empty_or_blank(_cell_value(worksheet, row_number, header_match.unit_col))
            and _is_empty_or_blank(_cell_value(worksheet, row_number, header_match.qty_col))
        )

        if is_blank_end_row:
            blank_streak += 1
        else:
            blank_streak = 0

        if blank_streak >= 3:
            break

        unit_value = _cell_value(worksheet, row_number, header_match.unit_col)
        qty_value = _cell_value(worksheet, row_number, header_match.qty_col)
        if _is_empty_or_blank(unit_value) and _is_empty_or_blank(qty_value):
            continue

        catalog_row = CatalogRow(
            task_id=task_number,
            price=_field_value(worksheet, row_number, field_columns, "price"),
            code=_field_value(worksheet, row_number, field_columns, "code"),
            unit=_field_value(worksheet, row_number, field_columns, "unit"),
            work_name=_field_value(worksheet, row_number, field_columns, "work_name"),
            region=identity.region_folder,
            added_date=_field_value(worksheet, row_number, field_columns, "added_date"),
        )

        row_issues = _validate_ingested_row(row_number, catalog_row)
        if row_issues:
            issues.extend(row_issues)
        else:
            rows.append(catalog_row)

    return rows, issues


def _build_source_header_map(worksheet: Worksheet, header_row: int) -> dict[str, int]:
    result: dict[str, int] = {}
    for col_number in range(1, worksheet.max_column + 1):
        value = _cell_value(worksheet, header_row, col_number)
        if not _is_empty_or_blank(value):
            result[_normalize_header_text(value)] = col_number
    return result


def _field_columns(source_map: dict[str, int], settings: IngestSettings) -> dict[str, int]:
    return {
        "work_name": source_map.get(_normalize_header_text(settings.work_name_header), 0),
        "unit": source_map.get(_normalize_header_text(settings.unit_header), 0),
        "price": source_map.get(_normalize_header_text(settings.price_header), 0),
        "code": source_map.get(_normalize_header_text(settings.code_header), 0),
        "added_date": source_map.get(_normalize_header_text(settings.added_date_header), 0),
    }


def _field_value(
    worksheet: Worksheet,
    row_number: int,
    field_columns: dict[str, int],
    field_name: str,
) -> Any:
    column = field_columns[field_name]
    if column <= 0:
        return None
    return _cell_value(worksheet, row_number, column)


def _validate_ingested_row(row_number: int, catalog_row: CatalogRow) -> list[ValidationIssue]:
    issues: list[ValidationIssue] = []

    if _parse_positive_number(catalog_row.price) is None:
        issues.append(ValidationIssue(row_number=row_number, reason="missing_or_invalid_price"))

    if NormCode(catalog_row.code) == "":
        issues.append(ValidationIssue(row_number=row_number, reason="missing_or_invalid_code"))

    if NormUnit(catalog_row.unit) == "":
        issues.append(ValidationIssue(row_number=row_number, reason="missing_or_invalid_unit"))

    return issues


def _extract_task_number(workbook: Workbook) -> str:
    for worksheet in workbook.worksheets:
        for row_number in range(1, 51):
            for col_number in range(1, 21):
                value = _cell_value(worksheet, row_number, col_number)
                text = "" if value is None else str(value)
                if text == "":
                    continue

                for label in (TASK_LABEL, TASK_LABEL_SHORT):
                    position = text.casefold().find(label.casefold())
                    if position == -1:
                        continue

                    tail = _cleanup_task_tail(text[position + len(label) :])
                    if tail != "":
                        return tail

                    neighbor = _neighbor_task_value(worksheet, row_number, col_number)
                    if neighbor != "":
                        return neighbor

    return ""


def _cleanup_task_tail(value: str) -> str:
    text = value.strip()
    text = text.replace(":", "")
    text = text.replace("#", "")
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def _neighbor_task_value(worksheet: Worksheet, row_number: int, col_number: int) -> str:
    for offset in range(1, 4):
        value = _cell_value(worksheet, row_number, col_number + offset)
        text = "" if value is None else str(value).strip()
        if text != "":
            return text
    return ""


def _find_numbering_col(source_map: dict[str, int]) -> int:
    for key, col_number in source_map.items():
        text = key.casefold()
        if (
            NUMBER_SIGN.casefold() in text
            or text.startswith("no")
            or "pp" in text
            or "p/p" in text
            or "p-p" in text
        ):
            return col_number
    return 0


def _normalize_header_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().casefold()
    text = text.replace(NBSP, " ")
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = text.replace(" ", "")
    text = text.replace("\t", "")
    return text


def _is_name_header(value: str) -> bool:
    name_works = _normalize_header_text(HEADER_NAME_WORKS)
    name_short = _normalize_header_text(HEADER_NAME_SHORT)
    return value == name_works or value == name_short or value.startswith(name_short)


def _is_unit_header(value: str) -> bool:
    unit = _normalize_header_text(HEADER_UNIT)
    return value.startswith(unit) or unit in value


def _is_qty_header(value: str) -> bool:
    qty = _normalize_header_text(HEADER_QTY)
    qty_long = _normalize_header_text(HEADER_QTY_LONG)
    return qty in value or qty_long in value


def _is_empty_or_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def _parse_positive_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Real):
        number = float(value)
    else:
        text = str(value).strip()
        if text == "":
            return None
        try:
            number = float(text)
        except ValueError:
            return None

    if number <= 0:
        return None
    return number


def _cell_value(worksheet: Worksheet, row: int, column: int) -> Any:
    return worksheet.cell(row=row, column=column).value

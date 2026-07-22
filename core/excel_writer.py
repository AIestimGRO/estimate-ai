"""Write a matching run result back into a copy of the estimate workbook.

Pure Excel output primitive (symmetric to core/excel_io.py): given a source
estimate file, the structured run result, the physical worksheet row for each
result row, and an explicit column plan, it writes analog columns, the
average-price formula, the section code, the `/KR` code suffix, and cell
colouring into a *copy* of the workbook. The source file is never modified.

Ports the output side-effects of ProcessSmeta (Module4), DOMAIN_RULES.md
section 6, and uses the average-column placement rule from core/layout.py
(R12). It makes no matching or pricing decisions of its own. The column plan
is supplied by the caller so both the fixed template layout and a detected
layout (Step 4c) can be written correctly.
"""

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.run_matching import EstimateRowResult, MatchingRunResult
from core.excel_io import find_estimate_sheet
from core.exclusions import (
    LEGACY_REASON_COLOR,
    TaskColorEntry,
    TaskHighlightReason,
    is_task_marked,
    resolve_task_highlight,
)
from core.layout import resolve_average_placement
from core.risk import REASON_RATIO_EXCEEDED

# Module4_updated.bas RGB fills
HEADER_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
TASK_FILL = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
REASON_LABEL_FONT = Font(bold=True, size=8, italic=True)
DUP_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
PROBLEM_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

PRICE_NUMBER_FORMAT = "#,##0.00"
HEADER_FONT = Font(bold=True, size=9)
REGION_FONT = Font(italic=True, size=9)
AVG_FONT = Font(bold=True)

# Header labels for columns inserted when absent (synonyms live in layout.json).
KR_HEADER_LABEL = "/\u041a\u0420"
SECTION_HEADER_LABEL = "\u041a\u043e\u0434 \u0440\u0430\u0437\u0434\u0435\u043b\u0430"
TKP_PRICE_HEADER_LABEL = "\u0410\u043d\u0430\u043b\u043e\u0433 \u0438\u0437 \u0422\u041a\u041f"
TKP_NAME_HEADER_LABEL = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0438\u0437 \u0422\u041a\u041f"
TKP_TASK_HEADER_LABEL = "\u041d\u043e\u043c\u0435\u0440 \u0437\u0430\u0434\u0430\u0447\u0438 \u0422\u041a\u041f"
TKP_COLUMN_COUNT = 3


@dataclass(frozen=True)
class WriterColumns:
    """Where the writer should place each output on the worksheet."""

    base_price: int
    code: int
    code_kr: int
    section: int | None
    analog_start: int | None
    header_row: int = 0


@dataclass(frozen=True)
class WriteReport:
    """Summary of what the writer produced."""

    output_path: Path
    written_rows: int
    inserted_average_column: bool
    average_column: int
    analog_start_column: int
    analog_column_count: int
    tkp_start_column: int | None = None


@dataclass(frozen=True)
class AnalogColumnDef:
    """One global analog output column (task header + region sub-header)."""

    column: int
    task_id: str
    price_position: int
    region: str


@dataclass(frozen=True)
class GlobalAnalogPlan:
    """Stable (task_id, price_position) -> column map for the whole run."""

    by_key: dict[tuple[str, int], int]
    columns: tuple[AnalogColumnDef, ...]
    last_column: int


def write_run_result(
    source_path: str | Path,
    output_path: str | Path,
    result: MatchingRunResult,
    row_numbers: list[int],
    *,
    columns: WriterColumns,
    regional_coefficient: float = 1.0,
    sheet_title: str | None = None,
    task_color_entries: list[TaskColorEntry] | None = None,
    task_highlight_reasons: list[TaskHighlightReason] | None = None,
    target_region: str | None = None,
    use_tkp_analogs: bool = False,
) -> WriteReport:
    """Write `result` into a copy of the estimate workbook at `output_path`.

    `target_region` is the estimate file's own declared region (e.g. from
    "Регион:"); when given, analog columns whose region matches it are placed
    first (R.. rule, 2026-07). See `_regions_match` for how matching works.

    `task_highlight_reasons` is the admin-configured reason->colour registry
    (the "Синие задачи" page). A task_color_entries row whose `reason` does
    not match any entry here falls back to the legacy blue fill and gets no
    label written above its column (backward compatibility).
    """
    if len(row_numbers) != len(result.rows):
        raise ValueError("row_numbers length must match result.rows length")

    task_colors = [] if task_color_entries is None else task_color_entries
    reasons = [] if task_highlight_reasons is None else task_highlight_reasons
    workbook = load_workbook(source_path, data_only=False)

    try:
        if sheet_title is not None:
            worksheet = workbook[sheet_title]
        else:
            worksheet = find_estimate_sheet(workbook)

        placement = _plan_average_column(worksheet, columns.base_price, row_numbers)
        if placement.needs_insert:
            worksheet.insert_cols(placement.column)
            _shift_formulas_after_insert(worksheet, placement.column)

        layout_columns = _ensure_kr_and_section_columns(
            worksheet,
            columns,
            placement,
            row_numbers,
        )
        analog_start_base = (
            max(worksheet.max_column + 1, placement.column + 1)
            if layout_columns.analog_start is None
            else layout_columns.analog_start
        )
        output_columns = _plan_output_columns(layout_columns, placement, analog_start_base)

        header_row = _effective_header_row(layout_columns, row_numbers)
        analog_plan = _build_global_analog_plan(
            result, output_columns.analog_start, target_region
        )
        tkp_start_column = (
            max(analog_plan.last_column + 1, output_columns.analog_start)
            if use_tkp_analogs
            else None
        )
        last_output_column = (
            tkp_start_column + TKP_COLUMN_COUNT - 1
            if tkp_start_column is not None
            else analog_plan.last_column
        )
        last_data_row = max(row_numbers) if row_numbers else output_columns.analog_start
        if analog_plan.columns or tkp_start_column is not None:
            _clear_analog_block(
                worksheet,
                header_row or output_columns.analog_start,
                last_data_row,
                output_columns.analog_start,
                last_output_column,
            )
        if analog_plan.columns:
            _write_analog_headers(
                worksheet,
                header_row,
                analog_plan,
                last_data_row,
                task_colors,
                reasons,
            )
        if tkp_start_column is not None:
            _write_tkp_headers(worksheet, header_row, tkp_start_column)

        written_rows = 0
        for row_number, row in zip(row_numbers, result.rows):
            if _write_row(
                worksheet,
                row_number,
                row,
                output_columns,
                analog_plan,
                regional_coefficient,
                task_colors,
                reasons,
                tkp_start_column,
            ):
                written_rows += 1

        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    finally:
        workbook.close()

    return WriteReport(
        output_path=destination,
        written_rows=written_rows,
        inserted_average_column=placement.needs_insert,
        average_column=output_columns.average,
        analog_start_column=output_columns.analog_start,
        analog_column_count=len(analog_plan.columns),
        tkp_start_column=tkp_start_column,
    )


@dataclass(frozen=True)
class _OutputColumns:
    base_price: int
    average: int
    code_kr: int
    section: int | None
    analog_start: int


# Matches a column reference inside a formula, e.g. "H29", "$H$29", "AVERAGE(F32,Q32:CK32)".
# Deliberately does not match things preceded by a letter/digit/underscore, so it
# skips sheet-name prefixes (e.g. "Дефлятор!$P$9" is still matched correctly on P9,
# which is what we want -- the fix must also re-point same-sheet formulas that
# happen to reference another sheet's fixed cells is out of scope here, since those
# don't move).
_CELL_REF_RE = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)")


def _shift_formula_columns(formula: str, insert_at_column: int) -> str:
    """Re-point column references in `formula` after a column was inserted.

    ``Worksheet.insert_cols()`` moves cell values/styles to the right but does
    NOT rewrite formulas living in *other* cells of the sheet (this is a known
    openpyxl limitation -- real Excel does this automatically). Any formula
    written before the insert that references a column >= insert_at_column
    therefore silently ends up pointing at the wrong data once the insert has
    shifted that data one column to the right (typical symptoms: an "ИТОГО"
    SUM formula summing the neighbouring column instead of its own, or a
    leftover average-price formula whose AVERAGE() range is off by one).

    Call this immediately after every ``worksheet.insert_cols(insert_at_column)``.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _replace(match: re.Match[str]) -> str:
        dollar_col, col_letters, dollar_row, row_digits = match.groups()
        col_index = column_index_from_string(col_letters)
        if col_index >= insert_at_column:
            col_index += 1
        return f"{dollar_col}{get_column_letter(col_index)}{dollar_row}{row_digits}"

    return _CELL_REF_RE.sub(_replace, formula)


def _shift_formulas_after_insert(worksheet: Worksheet, insert_at_column: int) -> None:
    """Fix up every existing formula on the sheet after a column insert.

    The newly inserted column itself (``insert_at_column``) is left alone --
    it is blank until the caller writes the new formula/header into it.
    """
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.column == insert_at_column:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                shifted = _shift_formula_columns(cell.value, insert_at_column)
                if shifted != cell.value:
                    cell.value = shifted


def _plan_average_column(
    worksheet: Worksheet,
    base_price_column: int,
    row_numbers: list[int],
):
    neighbour = base_price_column + 1
    occupied: set[int] = set()
    for row_number in row_numbers:
        value = worksheet.cell(row=row_number, column=neighbour).value
        if value in (None, ""):
            continue
        if _looks_like_average_formula(value):
            # A previous run already placed the average-price formula here
            # (re-processing an already-processed file). Reuse this column --
            # `_write_average_formula` overwrites it with a fresh formula
            # below -- instead of inserting a duplicate one next to it.
            continue
        occupied.add(neighbour)
        break
    return resolve_average_placement(base_price_column, occupied)


def _looks_like_average_formula(value: object) -> bool:
    if not isinstance(value, str) or not value.startswith("="):
        return False
    upper = value.upper().replace(" ", "")
    return upper.startswith("=MAX(") and "AVERAGE(" in upper


def _plan_output_columns(columns: WriterColumns, placement, analog_start_base: int) -> _OutputColumns:
    return _OutputColumns(
        base_price=columns.base_price,
        average=placement.column,
        code_kr=columns.code_kr,
        section=columns.section,
        analog_start=analog_start_base,
    )


def _effective_header_row(columns: WriterColumns, row_numbers: list[int]) -> int:
    if columns.header_row > 0:
        return columns.header_row
    if not row_numbers:
        return 0
    return min(row_numbers) - 2


def _shifted_column(column: int, placement) -> int:
    if placement.needs_insert and column >= placement.column:
        return column + 1
    return column


def _ensure_kr_and_section_columns(
    worksheet: Worksheet,
    columns: WriterColumns,
    placement,
    row_numbers: list[int],
) -> WriterColumns:
    """Ensure dedicated `/KR` and section columns exist (R11, DOMAIN_RULES.md section 6).

    The GESN code column keeps the plain code; rows with analogs receive
    ``code + /KR`` in the adjacent `/KR` column. Section codes go into their
    own column. Missing columns are inserted only when the target slot is
    already occupied; a blank column (as in many eV-grup exports) is reused.
    """
    header_row = _effective_header_row(columns, row_numbers)
    code_col = _shifted_column(columns.code, placement)

    kr_col = _find_kr_column(worksheet, header_row, code_col)
    section_col = (
        _shifted_column(columns.section, placement)
        if columns.section is not None
        else _find_section_column(worksheet, header_row, code_col)
    )

    pending_inserts: list[tuple[int, str]] = []
    pending_headers: list[tuple[int, str]] = []

    if kr_col is None:
        kr_col = code_col + 1
        if _column_needs_insert_for_kr(worksheet, header_row, kr_col, section_col):
            pending_inserts.append((kr_col, KR_HEADER_LABEL))
            if section_col is not None and section_col >= kr_col:
                section_col += 1
        else:
            pending_headers.append((kr_col, KR_HEADER_LABEL))

    if section_col is None or section_col <= kr_col:
        section_col = kr_col + 1
        if not any(insert_at == section_col for insert_at, _ in pending_inserts):
            if _column_needs_insert_for_section(worksheet, header_row, section_col):
                pending_inserts.append((section_col, SECTION_HEADER_LABEL))
            else:
                pending_headers.append((section_col, SECTION_HEADER_LABEL))

    for insert_at, label in sorted(pending_inserts, key=lambda item: item[0], reverse=True):
        worksheet.insert_cols(insert_at)
        _shift_formulas_after_insert(worksheet, insert_at)
        _write_output_header(worksheet, header_row, insert_at, label)

    for column, label in pending_headers:
        _write_output_header(worksheet, header_row, column, label)

    analog_start = section_col + 1
    return WriterColumns(
        base_price=columns.base_price,
        code=code_col,
        code_kr=kr_col,
        section=section_col,
        analog_start=analog_start,
        header_row=columns.header_row,
    )


def _write_output_header(
    worksheet: Worksheet,
    header_row: int,
    column: int,
    label: str,
) -> None:
    if header_row <= 0:
        return

    header_cell = worksheet.cell(row=header_row, column=column)
    if header_cell.value not in (None, ""):
        return

    header_cell.value = label
    header_cell.font = HEADER_FONT
    header_cell.fill = HEADER_FILL


def _column_needs_insert_for_kr(
    worksheet: Worksheet,
    header_row: int,
    kr_col: int,
    section_col: int | None,
) -> bool:
    if section_col == kr_col:
        return True

    if header_row <= 0:
        return False

    value = worksheet.cell(row=header_row, column=kr_col).value
    if value in (None, ""):
        return False

    if _is_kr_header(value) or _is_section_header(value):
        return False

    return True


def _column_needs_insert_for_section(
    worksheet: Worksheet,
    header_row: int,
    section_col: int,
) -> bool:
    if header_row <= 0:
        return False

    value = worksheet.cell(row=header_row, column=section_col).value
    if value in (None, ""):
        return False

    if _is_section_header(value):
        return False

    return True


def _find_kr_column(
    worksheet: Worksheet,
    header_row: int,
    code_column: int,
) -> int | None:
    if header_row <= 0:
        return None

    for column in range(code_column + 1, code_column + 4):
        if _is_kr_header(worksheet.cell(row=header_row, column=column).value):
            return column
    return None


def _find_section_column(
    worksheet: Worksheet,
    header_row: int,
    code_column: int,
) -> int | None:
    if header_row <= 0:
        return None

    for column in range(code_column + 1, code_column + 6):
        if _is_section_header(worksheet.cell(row=header_row, column=column).value):
            return column
    return None


def _is_kr_header(value: object) -> bool:
    text = _normalize_header_text(value)
    return text == "\u043a\u0440" or text.startswith("/")


def _is_section_header(value: object) -> bool:
    text = _normalize_header_text(value)
    return "\u043a\u043e\u0434 \u0440\u0430\u0437\u0434\u0435\u043b\u0430" in text


def _build_global_analog_plan(
    result: MatchingRunResult,
    analog_start: int,
    target_region: str | None = None,
) -> GlobalAnalogPlan:
    """Assign one worksheet column per (task_id, price_position) pair (Module4 step 2).

    Column order (2026-07 rule): all analog columns whose region matches the
    region of the file being processed come first (original first-seen order
    kept as a stable tiebreaker within that group); then every remaining
    region, grouped together, ordered alphabetically А-Я. A task's columns
    (its price positions) are always kept adjacent.
    """
    task_order: list[str] = []
    seen_tasks: set[str] = set()
    task_max_pi: dict[str, int] = {}
    region_by_key: dict[tuple[str, int], str] = {}
    task_region: dict[str, str] = {}

    for row in result.rows:
        for analog in row.analogs:
            task_id = analog.task_id
            if task_id not in seen_tasks:
                seen_tasks.add(task_id)
                task_order.append(task_id)
            task_max_pi[task_id] = max(task_max_pi.get(task_id, 0), analog.price_position)
            region_by_key[(task_id, analog.price_position)] = analog.entry.region
            task_region.setdefault(task_id, analog.entry.region)

    ordered_tasks = _order_tasks_by_region(task_order, task_region, target_region)

    by_key: dict[tuple[str, int], int] = {}
    column_defs: list[AnalogColumnDef] = []
    next_column = analog_start

    for task_id in ordered_tasks:
        for price_position in range(1, task_max_pi[task_id] + 1):
            key = (task_id, price_position)
            by_key[key] = next_column
            column_defs.append(
                AnalogColumnDef(
                    column=next_column,
                    task_id=task_id,
                    price_position=price_position,
                    region=region_by_key.get(key, ""),
                )
            )
            next_column += 1

    last_column = next_column - 1 if column_defs else analog_start - 1
    return GlobalAnalogPlan(by_key=by_key, columns=tuple(column_defs), last_column=last_column)


def _order_tasks_by_region(
    task_order: list[str],
    task_region: dict[str, str],
    target_region: str | None,
) -> list[str]:
    appearance_index = {task_id: index for index, task_id in enumerate(task_order)}

    def sort_key(task_id: str) -> tuple[int, str, int]:
        region = task_region.get(task_id, "")
        is_target = bool(target_region) and _regions_match(region, target_region)
        rank = 0 if is_target else 1
        region_sort = "" if is_target else _normalize_region_text(region)
        return (rank, region_sort, appearance_index[task_id])

    return sorted(task_order, key=sort_key)


# --- Region-name matching -------------------------------------------------
#
# The catalog's analog region (a free-text folder name picked by whoever
# built the catalog, e.g. "Якутия", "Тула" -- see DOMAIN_RULES.md 9.5) and
# the estimate file's own declared region (typed by hand next to "Регион:",
# e.g. "71. Тульская область", "14. Республика Саха (Якутия)") are two
# unrelated, differently-worded naming systems; there is no shared code or
# lookup table between them elsewhere in this project. The matching below is
# therefore a best-effort heuristic, not an exact lookup:
#   - numeric prefixes ("71. ") and common administrative suffix words
#     ("область", "край", "округ", "республика", ...) are stripped;
#   - text inside parentheses is also tried on its own, since official names
#     often carry the informal/short name in parentheses, e.g.
#     "Республика Саха (Якутия)" -> "Якутия";
#   - a small curated table of synonym roots below handles the common
#     Russian noun/adjective mismatch a plain prefix check misses (e.g.
#     "Тула" vs "Тульская область" -- the adjective softens "л" to "ль",
#     so they only share a 3-letter root, "тул", not a straightforward
#     prefix or substring match);
#   - anything not in the table falls back to: normalized forms equal, one
#     contains the other, or they share the same 4-letter prefix.
# If real-world files turn up a region pair this still gets wrong, add its
# root to _REGION_ALIAS_GROUPS rather than special-casing whole names.

_REGION_PREFIX_RE = re.compile(r"^\s*\d+\s*[.)]\s*")
_REGION_PAREN_RE = re.compile(r"\(([^)]+)\)")
_REGION_NON_LETTER_RE = re.compile(r"[^\u0430-\u044f\s]")
_REGION_WHITESPACE_RE = re.compile(r"\s+")
_REGION_SUFFIX_WORDS = (
    "область", "обл", "край", "округ", "республика", "автономный",
    "авт", "ао", "город", "г", "район", "р-н",
)
_REGION_STEM_LENGTH = 4

# Curated synonym roots for regions seen in real files, where the file's
# "Регион:" wording and the catalog's short folder name diverge more than a
# simple prefix/substring check can bridge. A region matches a group if its
# normalized text *contains* any one of the group's roots.
_REGION_ALIAS_GROUPS: tuple[frozenset[str], ...] = (
    frozenset({"тула", "тульск"}),
    frozenset({"кострома", "костромск"}),
    frozenset({"якутия", "саха"}),
    frozenset({"башкортостан", "башкирия", "башкирск"}),
    frozenset({"челябинск"}),
    frozenset({"забайкал"}),
    frozenset({"ямал", "янао", "ямао"}),
    frozenset({"урал"}),
    frozenset({"москва", "московск"}),
)


def _normalize_region_text(text: str) -> str:
    normalized = text.strip().lower().replace("\u0451", "\u0435")  # ё -> е
    for suffix in _REGION_SUFFIX_WORDS:
        normalized = re.sub(rf"\b{suffix}\.?\b", " ", normalized)
    normalized = _REGION_NON_LETTER_RE.sub(" ", normalized)
    return _REGION_WHITESPACE_RE.sub(" ", normalized).strip()


def _region_aliases(raw_text: object) -> set[str]:
    if not isinstance(raw_text, str) or not raw_text.strip():
        return set()

    text = _REGION_PREFIX_RE.sub("", raw_text)
    aliases = {_normalize_region_text(text)}
    for parenthetical in _REGION_PAREN_RE.findall(text):
        aliases.add(_normalize_region_text(parenthetical))
    aliases.add(_normalize_region_text(_REGION_PAREN_RE.sub(" ", text)))
    return {alias for alias in aliases if alias}


def _regions_match(region_a: object, region_b: object) -> bool:
    aliases_a = _region_aliases(region_a)
    aliases_b = _region_aliases(region_b)

    for group in _REGION_ALIAS_GROUPS:
        a_hits = any(root in alias for alias in aliases_a for root in group)
        b_hits = any(root in alias for alias in aliases_b for root in group)
        if a_hits and b_hits:
            return True

    for alias_a in aliases_a:
        for alias_b in aliases_b:
            if alias_a == alias_b:
                return True
            if len(alias_a) >= _REGION_STEM_LENGTH and len(alias_b) >= _REGION_STEM_LENGTH:
                if alias_a[:_REGION_STEM_LENGTH] == alias_b[:_REGION_STEM_LENGTH]:
                    return True
                if alias_a in alias_b or alias_b in alias_a:
                    return True
    return False


def _clear_analog_block(
    worksheet: Worksheet,
    header_row: int,
    last_row: int,
    first_col: int,
    last_col: int,
) -> None:
    start_row = header_row if header_row > 0 else 1
    for row in range(start_row, last_row + 1):
        for column in range(first_col, last_col + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.value = None
            cell.fill = PatternFill()


def _fill_for_marked_task(
    task_colors: list[TaskColorEntry],
    reasons: list[TaskHighlightReason],
    task_id: str,
) -> tuple[PatternFill, str | None] | None:
    """Fill + label for a marked task, or None if the task is not marked.

    Colour comes from the matched TaskHighlightReason; if the task is marked
    but its stored reason does not match any registered reason (legacy
    data), falls back to the historical TASK_FILL colour with no label.
    """
    if not is_task_marked(task_colors, task_id):
        return None
    highlight = resolve_task_highlight(task_colors, reasons, task_id)
    if highlight is not None:
        color_hex = highlight.color_hex
        label = highlight.label
    else:
        color_hex = LEGACY_REASON_COLOR
        label = None
    fill = PatternFill(start_color=f"FF{color_hex}", end_color=f"FF{color_hex}", fill_type="solid")
    return fill, label


def _write_analog_headers(
    worksheet: Worksheet,
    header_row: int,
    plan: GlobalAnalogPlan,
    last_data_row: int,
    task_colors: list[TaskColorEntry],
    reasons: list[TaskHighlightReason],
) -> None:
    if header_row <= 0 or not plan.columns:
        return

    region_row = header_row + 1
    # Analog columns are always newly placed to the right of the source
    # table, so the row directly above the header is free to use as a
    # dedicated reason-label row for marked tasks.
    reason_row = header_row - 1

    for column_def in plan.columns:
        column = column_def.column
        task_cell = worksheet.cell(row=header_row, column=column)
        task_cell.value = column_def.task_id
        task_cell.font = HEADER_FONT
        task_cell.fill = HEADER_FILL

        region_cell = worksheet.cell(row=region_row, column=column)
        region_cell.value = column_def.region
        region_cell.font = REGION_FONT
        region_cell.fill = HEADER_FILL

        highlighted = _fill_for_marked_task(task_colors, reasons, column_def.task_id)
        if highlighted is not None:
            fill, label = highlighted
            for row in range(header_row, last_data_row + 1):
                worksheet.cell(row=row, column=column).fill = fill
            task_cell.fill = HEADER_FILL
            region_cell.fill = HEADER_FILL

            if reason_row > 0 and label:
                reason_cell = worksheet.cell(row=reason_row, column=column)
                reason_cell.value = label
                reason_cell.font = REASON_LABEL_FONT
                reason_cell.fill = fill


def _write_tkp_headers(
    worksheet: Worksheet,
    header_row: int,
    start_column: int,
) -> None:
    if header_row <= 0:
        return

    labels = (
        TKP_PRICE_HEADER_LABEL,
        TKP_NAME_HEADER_LABEL,
        TKP_TASK_HEADER_LABEL,
    )
    for offset, label in enumerate(labels):
        cell = worksheet.cell(row=header_row, column=start_column + offset)
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _write_row(
    worksheet: Worksheet,
    row_number: int,
    row: EstimateRowResult,
    columns: _OutputColumns,
    plan: GlobalAnalogPlan,
    coefficient: float,
    task_colors: list[TaskColorEntry],
    reasons: list[TaskHighlightReason],
    tkp_start_column: int | None,
) -> bool:
    if columns.section is not None and row.section_code:
        section_cell = worksheet.cell(row=row_number, column=columns.section)
        section_cell.number_format = "@"
        section_cell.value = row.section_code

    _write_average_formula(
        worksheet,
        row_number,
        columns,
        plan,
        tkp_start_column,
    )

    if row.kr_code is not None:
        kr_cell = worksheet.cell(row=row_number, column=columns.code_kr)
        kr_cell.number_format = "@"
        kr_cell.value = row.kr_code

    for analog in row.analogs:
        column = plan.by_key[(analog.task_id, analog.price_position)]
        price_cell = worksheet.cell(row=row_number, column=column)
        price_cell.value = round(analog.entry.price * coefficient, 2)
        price_cell.number_format = PRICE_NUMBER_FORMAT
        _apply_cell_colour(price_cell, row, analog, task_colors, reasons)

    if tkp_start_column is not None and row.tkp_match is not None:
        entry = row.tkp_match.entry
        price_cell = worksheet.cell(row=row_number, column=tkp_start_column)
        price_cell.value = round(float(entry.winner_unit_price_no_vat), 2)
        price_cell.number_format = PRICE_NUMBER_FORMAT
        worksheet.cell(row=row_number, column=tkp_start_column + 1).value = entry.item_name
        worksheet.cell(row=row_number, column=tkp_start_column + 2).value = entry.task_no

    return row.has_analogs or row.has_tkp_analog


def _write_average_formula(
    worksheet: Worksheet,
    row_number: int,
    columns: _OutputColumns,
    plan: GlobalAnalogPlan,
    tkp_start_column: int | None,
) -> None:
    avg_cell = worksheet.cell(row=row_number, column=columns.average)
    avg_cell.value = _average_formula(
        row_number,
        columns.base_price,
        columns.analog_start,
        (
            tkp_start_column
            if tkp_start_column is not None
            else plan.last_column
        ),
    )
    avg_cell.number_format = PRICE_NUMBER_FORMAT
    avg_cell.font = AVG_FONT


def _apply_cell_colour(
    cell,
    row: EstimateRowResult,
    analog,
    task_colors: list[TaskColorEntry],
    reasons: list[TaskHighlightReason],
) -> None:
    highlighted = _fill_for_marked_task(task_colors, reasons, analog.task_id)
    if highlighted is not None:
        cell.fill = highlighted[0]
        return

    flagged_ids = {id(entry) for entry in row.risk_result.flagged_entries}
    colour_all = (
        row.risk_result.is_flagged
        and row.risk_result.reason == REASON_RATIO_EXCEEDED
    )

    if colour_all or id(analog.entry) in flagged_ids:
        cell.fill = PROBLEM_FILL
    elif analog.price_position > 1:
        cell.fill = DUP_FILL


def _average_formula(
    row_number: int,
    base_column: int,
    analog_start: int,
    last_analog_column: int,
) -> str:
    base = f"{get_column_letter(base_column)}{row_number}"
    if last_analog_column < analog_start:
        return f"={base}"

    first = f"{get_column_letter(analog_start)}{row_number}"
    last = f"{get_column_letter(last_analog_column)}{row_number}"
    return f"=MAX({base}, IFERROR(AVERAGE({base}, {first}:{last}), {base}))"


def resolve_kr_column(
    worksheet: Worksheet,
    header_row: int,
    code_column: int,
    settings_code_kr: int,
    settings_code_search: int,
) -> int:
    """Pick the `/KR` destination column; never reuse the plain code column."""
    found = _find_kr_column(worksheet, header_row, code_column)
    if found is not None:
        return found

    if settings_code_kr != settings_code_search:
        return settings_code_kr
    return code_column + 1


def _normalize_header_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).lower().strip()
    for char in ("\r", "\n", "\t", ".", ",", ";", ":"):
        text = text.replace(char, " ")
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()

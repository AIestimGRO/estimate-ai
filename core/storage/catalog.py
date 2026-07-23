"""Catalog source and item persistence."""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date, datetime
from numbers import Real
from pathlib import Path
from typing import BinaryIO, Iterator

from openpyxl import Workbook, load_workbook

from core.catalog import CatalogRow
from core.excel_io import Settings, read_catalog_rows_with_positions
from core.normalize import NormCode, NormUnit


BATCH_SIZE = 2000
DEFAULT_SOURCE_NAME = "main"
LEGACY_FILE_LOG_SOURCE_NAME = "legacy_file_log"
LEGACY_FILE_LOG_SOURCE_KIND = "legacy_file_log"
STATUS_LEGACY_IMPORTED = "legacy_imported"
STATUS_SUCCESS = "success"
STATUS_SKIPPED = "skipped"
STATUS_NO_DATA = "no_data"
STATUS_FAILED = "failed"
STATUS_DUPLICATE_NAME = "duplicate_name"
STATUS_MANUAL_CHECKED = "manual_checked"
STATUS_PENDING = "pending"
ROW_LOG_STATUS_REJECTED = "rejected"
RNMC_ZIP_SOURCE_NAME = "rnmc_zip_upload"
RNMC_ZIP_SOURCE_KIND = "rnmc_zip"
PROCESSED_FILE_STATUSES = frozenset(
    {
        STATUS_LEGACY_IMPORTED,
        STATUS_SUCCESS,
        STATUS_SKIPPED,
        STATUS_NO_DATA,
        STATUS_DUPLICATE_NAME,
        STATUS_MANUAL_CHECKED,
        STATUS_PENDING,
    }
)
FINAL_PREVIEW_SKIP_STATUSES = frozenset(
    {
        STATUS_LEGACY_IMPORTED,
        STATUS_SUCCESS,
        STATUS_SKIPPED,
        STATUS_NO_DATA,
        STATUS_DUPLICATE_NAME,
        STATUS_MANUAL_CHECKED,
    }
)

CATALOG_EDITOR_TEXT_FIELDS = frozenset(
    {
        "task_id",
        "region",
        "code",
        "unit",
        "work_name",
        "source_region_folder",
        "source_filename",
        "lsr_quarter",
        "planned_start",
        "planned_finish",
    }
)
CATALOG_EDITOR_NUMERIC_FIELDS = frozenset(
    {
        "quantity",
        "price",
        "price_original",
        "price_zlvl",
        "total_price",
        "labor_unit",
        "labor_total",
        "machine_labor_unit",
        "machine_labor_total",
        "regional_coefficient",
        "source_row_number",
    }
)
REQUIRED_CATALOG_TEXT_FIELDS = frozenset({"task_id", "code", "unit"})
REQUIRED_CATALOG_NUMERIC_FIELDS = frozenset({"price"})


@dataclass(frozen=True)
class CatalogSource:
    id: int
    name: str
    kind: str
    created_at: str
    item_count: int


@dataclass(frozen=True)
class ImportedFileRecord:
    id: int
    source_name: str
    source_kind: str
    region_folder: str
    filename: str
    status: str
    imported_at: str
    task_number: str
    rows_ok: int
    rows_rejected: int
    failure_reason: str
    filename_key: str
    legacy_note: str
    lsr_quarter: str
    planned_start: str
    planned_finish: str
    regional_coefficient: float | None


@dataclass(frozen=True)
class CatalogItemRecord:
    id: int
    task_id: str
    region: str
    code: str
    unit: str
    quantity: float | None
    work_name: str
    price: float
    price_original: float | None
    price_zlvl: float | None
    total_price: float | None
    labor_unit: float | None
    labor_total: float | None
    machine_labor_unit: float | None
    machine_labor_total: float | None
    regional_coefficient: float | None
    lsr_quarter: str
    planned_start: str
    planned_finish: str
    source_row_number: int


@dataclass(frozen=True)
class CatalogEditorRow:
    id: int
    source_name: str
    source_kind: str
    task_id: str
    region: str
    code: str
    unit: str
    quantity: float | None
    work_name: str
    price: float
    price_original: float | None
    price_zlvl: float | None
    total_price: float | None
    labor_unit: float | None
    labor_total: float | None
    machine_labor_unit: float | None
    machine_labor_total: float | None
    regional_coefficient: float | None
    lsr_quarter: str
    planned_start: str
    planned_finish: str
    source_region_folder: str
    source_filename: str
    source_row_number: int
    added_date: str


@dataclass(frozen=True)
class CatalogEditorPage:
    rows: list[CatalogEditorRow]
    total_rows: int
    page: int
    page_size: int
    total_pages: int
    filters: dict[str, str]


@dataclass(frozen=True)
class ImportRowLogRecord:
    id: int
    file_id: int
    row_number: int
    status: str
    reason: str


@dataclass(frozen=True)
class LegacyFileLogImportResult:
    rows_seen: int
    rows_imported: int
    duplicates: int
    empty_rows: int


@dataclass(frozen=True)
class CatalogImportResult:
    source_name: str
    source_id: int
    rows_imported: int
    rows_skipped: int
    source_filename: str


@dataclass(frozen=True)
class CatalogRowStorageItem:
    catalog_row: CatalogRow
    source_region_folder: str
    source_filename: str
    source_row_number: int


@dataclass(frozen=True)
class CatalogFileRowsImportResult:
    source_name: str
    source_id: int
    rows_imported: int
    rows_skipped: int
    source_filename: str
    region_folder: str


def list_catalog_sources(connection: sqlite3.Connection) -> list[CatalogSource]:
    rows = connection.execute(
        """
        SELECT
            catalog_sources.id,
            catalog_sources.name,
            catalog_sources.kind,
            catalog_sources.created_at,
            COUNT(catalog_items.id) AS item_count
        FROM catalog_sources
        LEFT JOIN catalog_items ON catalog_items.source_id = catalog_sources.id
        GROUP BY
            catalog_sources.id,
            catalog_sources.name,
            catalog_sources.kind,
            catalog_sources.created_at
        ORDER BY catalog_sources.name
        """
    ).fetchall()
    return [
        CatalogSource(
            id=int(row["id"]),
            name=str(row["name"]),
            kind=str(row["kind"]),
            created_at=str(row["created_at"]),
            item_count=int(row["item_count"]),
        )
        for row in rows
    ]


def list_imported_files(
    connection: sqlite3.Connection,
    *,
    status: str = "",
) -> list[ImportedFileRecord]:
    status_filter = _text(status)
    where = ""
    params: tuple[object, ...] = ()
    if status_filter:
        where = "WHERE imported_files.status = ?"
        params = (status_filter,)
    return _imported_file_records_from_query(connection, where, params)


def _imported_file_records_from_query(
    connection: sqlite3.Connection,
    where: str = "",
    params: tuple[object, ...] = (),
) -> list[ImportedFileRecord]:
    rows = connection.execute(
        """
        SELECT
            imported_files.id,
            COALESCE(catalog_sources.name, '') AS source_name,
            COALESCE(catalog_sources.kind, '') AS source_kind,
            imported_files.region_folder,
            imported_files.filename,
            imported_files.status,
            imported_files.imported_at,
            imported_files.task_number,
            imported_files.rows_ok,
            imported_files.rows_rejected,
            imported_files.failure_reason,
            imported_files.filename_key,
            imported_files.legacy_note,
            imported_files.lsr_quarter,
            imported_files.planned_start,
            imported_files.planned_finish,
            imported_files.regional_coefficient
        FROM imported_files
        LEFT JOIN catalog_sources ON catalog_sources.id = imported_files.source_id
        {where}
        ORDER BY imported_files.imported_at DESC, imported_files.id DESC
        """.format(where=where),
        params,
    ).fetchall()
    return [
        ImportedFileRecord(
            id=int(row["id"]),
            source_name=str(row["source_name"]),
            source_kind=str(row["source_kind"]),
            region_folder=str(row["region_folder"]),
            filename=str(row["filename"]),
            status=str(row["status"]),
            imported_at=str(row["imported_at"]),
            task_number=str(row["task_number"]),
            rows_ok=int(row["rows_ok"]),
            rows_rejected=int(row["rows_rejected"]),
            failure_reason=str(row["failure_reason"]),
            filename_key=str(row["filename_key"]),
            legacy_note=str(row["legacy_note"]),
            lsr_quarter=str(row["lsr_quarter"]),
            planned_start=str(row["planned_start"]),
            planned_finish=str(row["planned_finish"]),
            regional_coefficient=_optional_float(row["regional_coefficient"]),
        )
        for row in rows
    ]


def list_catalog_rows(
    connection: sqlite3.Connection,
    *,
    source_name: str = DEFAULT_SOURCE_NAME,
) -> list[CatalogRow]:
    source_id = _require_source_id(connection, source_name)
    rows = connection.execute(
        """
        SELECT
            task_id, region, code, unit, quantity, work_name, price,
            price_original, price_zlvl, added_date,
            total_price, labor_unit, labor_total,
            machine_labor_unit, machine_labor_total, regional_coefficient,
            lsr_quarter, planned_start, planned_finish, source_filename
        FROM catalog_items
        WHERE source_id = ?
        ORDER BY id
        """,
        (source_id,),
    ).fetchall()
    return [_row_to_catalog_row(row) for row in rows]


def count_catalog_rows(
    connection: sqlite3.Connection,
    *,
    source_name: str = DEFAULT_SOURCE_NAME,
) -> int:
    row = connection.execute(
        """
        SELECT COUNT(*) AS row_count
        FROM catalog_items
        INNER JOIN catalog_sources ON catalog_sources.id = catalog_items.source_id
        WHERE catalog_sources.name = ?
        """,
        (source_name,),
    ).fetchone()
    return 0 if row is None else int(row["row_count"])




def list_catalog_editor_page(
    connection: sqlite3.Connection,
    *,
    filters: dict[str, str] | None = None,
    page: int = 1,
    page_size: int = 100,
) -> CatalogEditorPage:
    normalized_filters = _catalog_editor_filters(filters or {})
    safe_page = max(1, int(page))
    safe_page_size = min(500, max(25, int(page_size)))
    where_sql, params = _catalog_editor_where(normalized_filters)
    total_row = connection.execute(
        f"""
        SELECT COUNT(*) AS row_count
        FROM catalog_items
        INNER JOIN catalog_sources ON catalog_sources.id = catalog_items.source_id
        {where_sql}
        """,
        params,
    ).fetchone()
    total_rows = 0 if total_row is None else int(total_row["row_count"])
    total_pages = max(1, (total_rows + safe_page_size - 1) // safe_page_size)
    safe_page = min(safe_page, total_pages)
    offset = (safe_page - 1) * safe_page_size
    rows = connection.execute(
        f"""
        SELECT
            catalog_items.id,
            catalog_sources.name AS source_name,
            catalog_sources.kind AS source_kind,
            catalog_items.task_id,
            catalog_items.region,
            catalog_items.code,
            catalog_items.unit,
            catalog_items.quantity,
            catalog_items.work_name,
            catalog_items.price,
            catalog_items.price_original,
            catalog_items.price_zlvl,
            catalog_items.total_price,
            catalog_items.labor_unit,
            catalog_items.labor_total,
            catalog_items.machine_labor_unit,
            catalog_items.machine_labor_total,
            catalog_items.regional_coefficient,
            catalog_items.lsr_quarter,
            catalog_items.planned_start,
            catalog_items.planned_finish,
            catalog_items.source_region_folder,
            catalog_items.source_filename,
            catalog_items.source_row_number,
            catalog_items.added_date
        FROM catalog_items
        INNER JOIN catalog_sources ON catalog_sources.id = catalog_items.source_id
        {where_sql}
        ORDER BY catalog_items.id DESC
        LIMIT ? OFFSET ?
        """,
        (*params, safe_page_size, offset),
    ).fetchall()
    return CatalogEditorPage(
        rows=[_catalog_editor_row(row) for row in rows],
        total_rows=total_rows,
        page=safe_page,
        page_size=safe_page_size,
        total_pages=total_pages,
        filters=normalized_filters,
    )


# Fixed, human-facing export layout (2026-07). Kept separate from
# CatalogEditorRow's internal field names/order on purpose: the export is a
# stable external contract the user asked for, while the DB/editor columns
# may keep evolving independently.
CATALOG_EXPORT_HEADERS = (
    "№_пп",
    "Номер задачи",
    "Наименование работ",
    "Ед.изм.",
    "Кол-во",
    "Цена единицы работ (с учетом вспомогательных материалов), руб. без НДС",
    "Цена единицы работ (с учетом вспомогательных материалов), руб. без НДС ZLVL",
    "Итого стоимость, руб. с НДС",
    "Итого стоимость, руб. без НДС",
    "ТЗ на ед., чел-час",
    "ТЗ всего, чел-час",
    "ТЗм на ед., чел-час",
    "ТЗм всего, чел-час",
    "Перечень ГЭСН/ФЕР/ТЕР/КР",
    "source_file",
    "Регион",
    "Год Квартал ЛСР",
    "Планируемый срок начала работ",
    "Планируемый срок окончания работ",
    "Региональный коэффициент",
    "Дата добавления в каталог",
)

# catalog_items.total_price (and every price_* column) is normalized to
# "без НДС" at import time regardless of what the source file said -- see
# _vat_divisor() in app/services/rnmc_excel.py. There is no separate
# "с НДС" total stored anywhere, so the export derives it from scratch.
CATALOG_EXPORT_VAT_MULTIPLIER = 1.2


def iter_catalog_export_rows(
    connection: sqlite3.Connection,
    *,
    filters: dict[str, str] | None = None,
) -> Iterator[CatalogEditorRow]:
    """Stream every catalog row matching `filters`, unpaginated."""
    normalized_filters = _catalog_editor_filters(filters or {})
    where_sql, params = _catalog_editor_where(normalized_filters)
    cursor = connection.execute(
        f"""
        SELECT
            catalog_items.id,
            catalog_sources.name AS source_name,
            catalog_sources.kind AS source_kind,
            catalog_items.task_id,
            catalog_items.region,
            catalog_items.code,
            catalog_items.unit,
            catalog_items.quantity,
            catalog_items.work_name,
            catalog_items.price,
            catalog_items.price_original,
            catalog_items.price_zlvl,
            catalog_items.total_price,
            catalog_items.labor_unit,
            catalog_items.labor_total,
            catalog_items.machine_labor_unit,
            catalog_items.machine_labor_total,
            catalog_items.regional_coefficient,
            catalog_items.lsr_quarter,
            catalog_items.planned_start,
            catalog_items.planned_finish,
            catalog_items.source_region_folder,
            catalog_items.source_filename,
            catalog_items.source_row_number,
            catalog_items.added_date
        FROM catalog_items
        INNER JOIN catalog_sources ON catalog_sources.id = catalog_items.source_id
        {where_sql}
        ORDER BY catalog_items.id ASC
        """,
        params,
    )
    for row in cursor:
        yield _catalog_editor_row(row)


def write_catalog_export_xlsx(
    connection: sqlite3.Connection,
    output_path: str | Path,
    *,
    filters: dict[str, str] | None = None,
) -> int:
    """Export catalog_items into an .xlsx with the fixed layout above."""
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Каталог")
    sheet.append(CATALOG_EXPORT_HEADERS)

    row_count = 0
    for index, row in enumerate(iter_catalog_export_rows(connection, filters=filters), start=1):
        total_without_vat = row.total_price
        total_with_vat = (
            None
            if total_without_vat is None
            else total_without_vat * CATALOG_EXPORT_VAT_MULTIPLIER
        )
        sheet.append(
            [
                index,
                row.task_id,
                row.work_name,
                row.unit,
                row.quantity,
                row.price_original,
                row.price_zlvl,
                total_with_vat,
                total_without_vat,
                row.labor_unit,
                row.labor_total,
                row.machine_labor_unit,
                row.machine_labor_total,
                row.code,
                row.source_filename,
                row.region,
                row.lsr_quarter,
                row.planned_start,
                row.planned_finish,
                row.regional_coefficient,
                row.added_date,
            ]
        )
        row_count += 1

    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return row_count


def update_catalog_item(
    connection: sqlite3.Connection,
    item_id: int,
    *,
    values: dict[str, object],
) -> bool:
    cleaned = _clean_catalog_editor_values(values)
    if not cleaned:
        return False
    assignments = ", ".join(f"{field} = ?" for field in cleaned)
    cursor = connection.execute(
        f"UPDATE catalog_items SET {assignments} WHERE id = ?",
        (*cleaned.values(), int(item_id)),
    )
    connection.commit()
    return cursor.rowcount > 0


def delete_catalog_item(connection: sqlite3.Connection, item_id: int) -> bool:
    cursor = connection.execute(
        "DELETE FROM catalog_items WHERE id = ?",
        (int(item_id),),
    )
    connection.commit()
    return cursor.rowcount > 0


def bulk_delete_catalog_items(connection: sqlite3.Connection, item_ids: list[int]) -> int:
    ids = _valid_catalog_item_ids(item_ids)
    if not ids:
        return 0
    placeholders = ", ".join("?" for _ in ids)
    cursor = connection.execute(
        f"DELETE FROM catalog_items WHERE id IN ({placeholders})",
        tuple(ids),
    )
    connection.commit()
    return int(cursor.rowcount)


def clear_catalog_for_rebuild(connection: sqlite3.Connection) -> tuple[int, int]:
    """Clear catalog rows while preserving processed-file history.

    imported_files is the durable replacement for legacy File_Log.xlsx and must
    survive catalog rebuilds so already processed RNMC files remain skippable.
    """
    catalog_count = int(
        connection.execute("SELECT COUNT(*) AS count FROM catalog_items").fetchone()["count"]
    )
    import_count = int(
        connection.execute("SELECT COUNT(*) AS count FROM imported_files").fetchone()["count"]
    )
    with connection:
        connection.execute("DELETE FROM import_row_log")
        connection.execute("DELETE FROM catalog_items")
    return catalog_count, import_count


def bulk_update_catalog_items(
    connection: sqlite3.Connection,
    item_ids: list[int],
    *,
    field: str,
    operation: str,
    value: object,
) -> int:
    ids = _valid_catalog_item_ids(item_ids)
    if not ids:
        return 0
    field_name = _text(field)
    op = _text(operation)
    placeholders = ", ".join("?" for _ in ids)

    if field_name in CATALOG_EDITOR_TEXT_FIELDS:
        if op != "set":
            raise ValueError("Text fields support only set operation")
        cleaned = _clean_catalog_editor_values({field_name: value})
        if not cleaned:
            return 0
        cursor = connection.execute(
            f"UPDATE catalog_items SET {field_name} = ? WHERE id IN ({placeholders})",
            (cleaned[field_name], *ids),
        )
    elif field_name in CATALOG_EDITOR_NUMERIC_FIELDS:
        number = _parse_catalog_editor_number(field_name, value)
        if op == "set":
            cursor = connection.execute(
                f"UPDATE catalog_items SET {field_name} = ? WHERE id IN ({placeholders})",
                (number, *ids),
            )
        elif op == "add":
            if field_name in REQUIRED_CATALOG_NUMERIC_FIELDS:
                cursor = connection.execute(
                    f"""
                    UPDATE catalog_items
                    SET {field_name} = {field_name} + ?
                    WHERE id IN ({placeholders})
                      AND {field_name} + ? > 0
                    """,
                    (number, *ids, number),
                )
            else:
                cursor = connection.execute(
                    f"""
                    UPDATE catalog_items
                    SET {field_name} = COALESCE({field_name}, 0) + ?
                    WHERE id IN ({placeholders})
                    """,
                    (number, *ids),
                )
        elif op == "multiply":
            if number == 0 and field_name in REQUIRED_CATALOG_NUMERIC_FIELDS:
                raise ValueError("Required numeric fields cannot be multiplied by zero")
            cursor = connection.execute(
                f"""
                UPDATE catalog_items
                SET {field_name} = CASE
                    WHEN {field_name} IS NULL THEN NULL
                    ELSE {field_name} * ?
                END
                WHERE id IN ({placeholders})
                """,
                (number, *ids),
            )
        else:
            raise ValueError("Unsupported bulk operation")
    else:
        raise ValueError("Unsupported catalog field")

    connection.commit()
    return int(cursor.rowcount)


def _catalog_editor_filters(filters: dict[str, str]) -> dict[str, str]:
    allowed = {"q", "source", "region", "task_id", "code", "unit", "filename"}
    return {key: _text(filters.get(key)) for key in allowed}


def _catalog_editor_where(filters: dict[str, str]) -> tuple[str, tuple[object, ...]]:
    clauses: list[str] = []
    params: list[object] = []
    if filters.get("source"):
        clauses.append("catalog_sources.name LIKE ?")
        params.append(f'%{filters["source"]}%')
    if filters.get("region"):
        clauses.append("catalog_items.region LIKE ?")
        params.append(f'%{filters["region"]}%')
    if filters.get("task_id"):
        clauses.append("catalog_items.task_id LIKE ?")
        params.append(f'%{filters["task_id"]}%')
    if filters.get("code"):
        clauses.append("catalog_items.code LIKE ?")
        params.append(f'%{filters["code"]}%')
    if filters.get("unit"):
        clauses.append("catalog_items.unit LIKE ?")
        params.append(f'%{filters["unit"]}%')
    if filters.get("filename"):
        clauses.append("catalog_items.source_filename LIKE ?")
        params.append(f'%{filters["filename"]}%')
    if filters.get("q"):
        search = f'%{filters["q"]}%'
        clauses.append(
            "(catalog_items.work_name LIKE ? OR catalog_items.code LIKE ? "
            "OR catalog_items.task_id LIKE ? OR catalog_items.region LIKE ? "
            "OR catalog_items.source_filename LIKE ? OR catalog_items.lsr_quarter LIKE ?)"
        )
        params.extend([search, search, search, search, search, search])
    where = "" if not clauses else "WHERE " + " AND ".join(clauses)
    return where, tuple(params)


def _catalog_editor_row(row: sqlite3.Row) -> CatalogEditorRow:
    return CatalogEditorRow(
        id=int(row["id"]),
        source_name=str(row["source_name"]),
        source_kind=str(row["source_kind"]),
        task_id=str(row["task_id"]),
        region=str(row["region"]),
        code=str(row["code"]),
        unit=str(row["unit"]),
        quantity=_optional_float(row["quantity"]),
        work_name=str(row["work_name"]),
        price=float(row["price"]),
        price_original=_optional_float(row["price_original"]),
        price_zlvl=_optional_float(row["price_zlvl"]),
        total_price=_optional_float(row["total_price"]),
        labor_unit=_optional_float(row["labor_unit"]),
        labor_total=_optional_float(row["labor_total"]),
        machine_labor_unit=_optional_float(row["machine_labor_unit"]),
        machine_labor_total=_optional_float(row["machine_labor_total"]),
        regional_coefficient=_optional_float(row["regional_coefficient"]),
        lsr_quarter=str(row["lsr_quarter"]),
        planned_start=str(row["planned_start"]),
        planned_finish=str(row["planned_finish"]),
        source_region_folder=str(row["source_region_folder"]),
        source_filename=str(row["source_filename"]),
        source_row_number=int(row["source_row_number"]),
        added_date="" if row["added_date"] is None else str(row["added_date"]),
    )


def _clean_catalog_editor_values(values: dict[str, object]) -> dict[str, object]:
    cleaned: dict[str, object] = {}
    for field, value in values.items():
        field_name = _text(field)
        if field_name in CATALOG_EDITOR_TEXT_FIELDS:
            text = _text(value)
            if field_name in REQUIRED_CATALOG_TEXT_FIELDS and text == "":
                raise ValueError(f"{field_name} cannot be empty")
            cleaned[field_name] = text
        elif field_name in CATALOG_EDITOR_NUMERIC_FIELDS:
            cleaned[field_name] = _parse_catalog_editor_number(field_name, value)
        else:
            raise ValueError(f"unsupported catalog field: {field_name}")
    return cleaned


def _parse_catalog_editor_number(field_name: str, value: object) -> float | int | None:
    number = _parse_optional_number(value)
    if number is None:
        if field_name in REQUIRED_CATALOG_NUMERIC_FIELDS:
            raise ValueError(f"{field_name} cannot be empty")
        return None
    if field_name in REQUIRED_CATALOG_NUMERIC_FIELDS and number <= 0:
        raise ValueError(f"{field_name} must be positive")
    if field_name == "source_row_number":
        return max(0, int(number))
    return number


def _valid_catalog_item_ids(item_ids: list[int]) -> list[int]:
    result = []
    for item_id in item_ids:
        try:
            value = int(item_id)
        except (TypeError, ValueError):
            continue
        if value > 0:
            result.append(value)
    return result



def import_legacy_file_log(
    connection: sqlite3.Connection,
    workbook_path: str | Path | BinaryIO,
    *,
    source_name: str = LEGACY_FILE_LOG_SOURCE_NAME,
) -> LegacyFileLogImportResult:
    source_id = _get_or_create_source(
        connection,
        source_name,
        kind=LEGACY_FILE_LOG_SOURCE_KIND,
    )
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["FileLog"] if "FileLog" in workbook.sheetnames else workbook.active
        headers = _legacy_file_log_headers(sheet)
        seen: dict[str, str] = {}
        rows_seen = 0
        rows_imported = 0
        duplicates = 0
        empty_rows = 0

        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = _legacy_file_log_row(values, headers)
            if _row_is_empty(row):
                empty_rows += 1
                continue
            filename = _text(row.get("filename"))
            if filename == "":
                empty_rows += 1
                continue
            rows_seen += 1

            region = _region_from_folder_text(row.get("region_folder"))
            key = normalize_import_filename(filename)
            legacy_note = _text(row.get("legacy_note"))
            rows_ok = _legacy_rows_ok(legacy_note)
            failure_reason = ""
            status = STATUS_LEGACY_IMPORTED

            previous_region = seen.get(key)
            if previous_region is not None:
                duplicates += 1
                status = STATUS_DUPLICATE_NAME
                failure_reason = f"Duplicate filename in FileLog; first region: {previous_region}"
            else:
                seen[key] = region

            _record_imported_file(
                connection,
                source_id=source_id,
                region_folder=region,
                filename=filename,
                status=status,
                rows_ok=rows_ok,
                rows_rejected=0,
                failure_reason=failure_reason,
                legacy_note=legacy_note,
                lsr_quarter=_text(row.get("lsr_quarter")),
                planned_start=_text(row.get("planned_start")),
                planned_finish=_text(row.get("planned_finish")),
            )
            rows_imported += 1
        connection.commit()
        return LegacyFileLogImportResult(
            rows_seen=rows_seen,
            rows_imported=rows_imported,
            duplicates=duplicates,
            empty_rows=empty_rows,
        )
    finally:
        workbook.close()


def sync_legacy_file_log_history(
    connection: sqlite3.Connection,
    workbook_path: str | Path | BinaryIO,
    *,
    source_name: str = LEGACY_FILE_LOG_SOURCE_NAME,
) -> LegacyFileLogImportResult:
    """Add missing legacy File_Log names without overwriting current import records."""
    source_id = _get_or_create_source(
        connection,
        source_name,
        kind=LEGACY_FILE_LOG_SOURCE_KIND,
    )
    existing_keys = {
        _text(row["filename_key"])
        for row in connection.execute(
            "SELECT DISTINCT filename_key FROM imported_files WHERE filename_key <> ''"
        ).fetchall()
        if _text(row["filename_key"])
    }
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["FileLog"] if "FileLog" in workbook.sheetnames else workbook.active
        headers = _legacy_file_log_headers(sheet)
        seen: set[str] = set()
        rows_seen = 0
        rows_imported = 0
        duplicates = 0
        empty_rows = 0
        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = _legacy_file_log_row(values, headers)
            if _row_is_empty(row):
                empty_rows += 1
                continue
            filename = _text(row.get("filename"))
            key = normalize_import_filename(filename)
            if not filename or not key:
                empty_rows += 1
                continue
            rows_seen += 1
            if key in seen:
                duplicates += 1
                continue
            seen.add(key)
            if key in existing_keys:
                continue
            region = _region_from_folder_text(row.get("region_folder"))
            legacy_note = _text(row.get("legacy_note"))
            _record_imported_file(
                connection,
                source_id=source_id,
                region_folder=region,
                filename=filename,
                status=STATUS_LEGACY_IMPORTED,
                rows_ok=_legacy_rows_ok(legacy_note),
                rows_rejected=0,
                legacy_note=legacy_note,
                lsr_quarter=_text(row.get("lsr_quarter")),
                planned_start=_text(row.get("planned_start")),
                planned_finish=_text(row.get("planned_finish")),
            )
            existing_keys.add(key)
            rows_imported += 1
        connection.commit()
        return LegacyFileLogImportResult(
            rows_seen=rows_seen,
            rows_imported=rows_imported,
            duplicates=duplicates,
            empty_rows=empty_rows,
        )
    finally:
        workbook.close()


def filename_is_processed(connection: sqlite3.Connection, filename: str) -> bool:
    return _filename_has_status(connection, filename, PROCESSED_FILE_STATUSES)


def filename_is_final_for_preview(connection: sqlite3.Connection, filename: str) -> bool:
    """Return True when a file name should not be parsed in row preview.

    Pending and failed files are intentionally not final: a later preview/retry
    can parse them again from a newly uploaded zip.
    """
    return _filename_has_status(connection, filename, FINAL_PREVIEW_SKIP_STATUSES)


def final_filename_keys_for_preview(connection: sqlite3.Connection) -> set[str]:
    """Return normalized file names that should be skipped before workbook parsing.

    imported_files is authoritative, while catalog_items.source_filename is a
    recovery source for databases created by older legacy-catalog imports that
    did not fully reconstruct File_Log history.
    """
    placeholders = ", ".join("?" for _ in FINAL_PREVIEW_SKIP_STATUSES)
    rows = connection.execute(
        f"""
        SELECT DISTINCT filename_key
        FROM imported_files
        WHERE filename_key <> ''
          AND status IN ({placeholders})
        """,
        tuple(sorted(FINAL_PREVIEW_SKIP_STATUSES)),
    ).fetchall()
    keys = {_text(row["filename_key"]) for row in rows if _text(row["filename_key"])}
    catalog_rows = connection.execute(
        """
        SELECT DISTINCT source_filename
        FROM catalog_items
        WHERE TRIM(COALESCE(source_filename, '')) <> ''
        """
    ).fetchall()
    keys.update(
        normalize_import_filename(row["source_filename"])
        for row in catalog_rows
        if normalize_import_filename(row["source_filename"])
    )
    return keys


def _filename_has_status(
    connection: sqlite3.Connection,
    filename: str,
    statuses: frozenset[str],
) -> bool:
    key = normalize_import_filename(filename)
    if key == "":
        return False
    placeholders = ", ".join("?" for _ in statuses)
    row = connection.execute(
        f"""
        SELECT 1
        FROM imported_files
        WHERE filename_key = ?
          AND status IN ({placeholders})
        LIMIT 1
        """,
        (key, *sorted(statuses)),
    ).fetchone()
    return row is not None


def imported_file_exists_for_region(
    connection: sqlite3.Connection,
    *,
    region_folder: str,
    filename: str,
) -> bool:
    row = connection.execute(
        """
        SELECT 1
        FROM imported_files
        WHERE region_folder = ? AND filename = ?
        LIMIT 1
        """,
        (_text(region_folder), _text(filename)),
    ).fetchone()
    return row is not None


def record_imported_file(
    connection: sqlite3.Connection,
    *,
    region_folder: str,
    filename: str,
    status: str,
    rows_ok: int = 0,
    rows_rejected: int = 0,
    failure_reason: str = "",
    task_number: str = "",
    legacy_note: str = "",
    lsr_quarter: str = "",
    planned_start: str = "",
    planned_finish: str = "",
    regional_coefficient: object = None,
    source_name: str = RNMC_ZIP_SOURCE_NAME,
    source_kind: str = RNMC_ZIP_SOURCE_KIND,
) -> int:
    source_id = _get_or_create_source(connection, source_name, kind=source_kind)
    return _record_imported_file(
        connection,
        source_id=source_id,
        region_folder=region_folder,
        filename=filename,
        status=status,
        rows_ok=rows_ok,
        rows_rejected=rows_rejected,
        failure_reason=failure_reason,
        task_number=task_number,
        legacy_note=legacy_note,
        lsr_quarter=lsr_quarter,
        planned_start=planned_start,
        planned_finish=planned_finish,
        regional_coefficient=regional_coefficient,
    )


def get_imported_file(connection: sqlite3.Connection, import_id: int) -> ImportedFileRecord | None:
    records = _imported_file_records_from_query(
        connection,
        "WHERE imported_files.id = ?",
        (int(import_id),),
    )
    return records[0] if records else None


def update_imported_file_metadata(
    connection: sqlite3.Connection,
    import_id: int,
    *,
    region_folder: str,
    task_number: str,
    lsr_quarter: str,
    planned_start: str,
    planned_finish: str,
    regional_coefficient: object = None,
) -> bool:
    existing = get_imported_file(connection, import_id)
    coefficient = _parse_optional_number(regional_coefficient)
    cursor = connection.execute(
        """
        UPDATE imported_files
        SET region_folder = ?,
            task_number = ?,
            lsr_quarter = ?,
            planned_start = ?,
            planned_finish = ?,
            regional_coefficient = ?
        WHERE id = ?
        """,
        (
            _text(region_folder),
            _text(task_number),
            _text(lsr_quarter),
            _text(planned_start),
            _text(planned_finish),
            coefficient,
            int(import_id),
        ),
    )
    if cursor.rowcount > 0 and existing is not None:
        _apply_import_metadata_to_catalog_items(
            connection,
            existing,
            region_folder=_text(region_folder),
            task_number=_text(task_number),
            lsr_quarter=_text(lsr_quarter),
            planned_start=_text(planned_start),
            planned_finish=_text(planned_finish),
            regional_coefficient=coefficient,
        )
    connection.commit()
    return cursor.rowcount > 0


def _apply_import_metadata_to_catalog_items(
    connection: sqlite3.Connection,
    record: ImportedFileRecord,
    *,
    region_folder: str,
    task_number: str,
    lsr_quarter: str,
    planned_start: str,
    planned_finish: str,
    regional_coefficient: float | None,
) -> None:
    source_name = record.source_name or RNMC_ZIP_SOURCE_NAME
    source_row = connection.execute(
        "SELECT id FROM catalog_sources WHERE name = ? LIMIT 1",
        (source_name,),
    ).fetchone()
    if source_row is None:
        return
    source_id = int(source_row["id"])
    if regional_coefficient is not None and regional_coefficient > 0:
        connection.execute(
            """
            UPDATE catalog_items
            SET region = ?,
                task_id = CASE WHEN ? <> '' THEN ? ELSE task_id END,
                regional_coefficient = ?,
                lsr_quarter = ?,
                planned_start = ?,
                planned_finish = ?,
                source_region_folder = ?,
                price_zlvl = CASE
                    WHEN price_original IS NOT NULL AND price_original > 0
                    THEN price_original / ?
                    ELSE price_zlvl
                END,
                price = CASE
                    WHEN price_original IS NOT NULL AND price_original > 0
                    THEN price_original / ?
                    ELSE price
                END
            WHERE source_id = ?
              AND source_region_folder = ?
              AND source_filename = ?
            """,
            (
                region_folder,
                task_number,
                task_number,
                regional_coefficient,
                lsr_quarter,
                planned_start,
                planned_finish,
                region_folder,
                regional_coefficient,
                regional_coefficient,
                source_id,
                record.region_folder,
                record.filename,
            ),
        )
    else:
        connection.execute(
            """
            UPDATE catalog_items
            SET region = ?,
                task_id = CASE WHEN ? <> '' THEN ? ELSE task_id END,
                regional_coefficient = NULL,
                lsr_quarter = ?,
                planned_start = ?,
                planned_finish = ?,
                source_region_folder = ?
            WHERE source_id = ?
              AND source_region_folder = ?
              AND source_filename = ?
            """,
            (
                region_folder,
                task_number,
                task_number,
                lsr_quarter,
                planned_start,
                planned_finish,
                region_folder,
                source_id,
                record.region_folder,
                record.filename,
            ),
        )

def allow_import_retry(connection: sqlite3.Connection, import_id: int) -> bool:
    cursor = connection.execute(
        """
        UPDATE imported_files
        SET status = ?,
            failure_reason = 'Retry allowed; upload the ZIP again to reprocess this file'
        WHERE id = ?
          AND status IN (?, ?)
        """,
        (STATUS_PENDING, int(import_id), STATUS_FAILED, STATUS_NO_DATA),
    )
    connection.commit()
    return cursor.rowcount > 0


def list_catalog_items_for_imported_file(
    connection: sqlite3.Connection,
    import_id: int,
) -> list[CatalogItemRecord]:
    record = get_imported_file(connection, import_id)
    if record is None:
        return []
    rows = connection.execute(
        """
        SELECT
            catalog_items.id,
            catalog_items.task_id,
            catalog_items.region,
            catalog_items.code,
            catalog_items.unit,
            catalog_items.quantity,
            catalog_items.work_name,
            catalog_items.price,
            catalog_items.price_original,
            catalog_items.price_zlvl,
            catalog_items.total_price,
            catalog_items.labor_unit,
            catalog_items.labor_total,
            catalog_items.machine_labor_unit,
            catalog_items.machine_labor_total,
            catalog_items.regional_coefficient,
            catalog_items.lsr_quarter,
            catalog_items.planned_start,
            catalog_items.planned_finish,
            catalog_items.source_row_number
        FROM catalog_items
        WHERE catalog_items.source_region_folder = ?
          AND catalog_items.source_filename = ?
          AND catalog_items.source_id = (
              SELECT id FROM catalog_sources WHERE name = ? LIMIT 1
          )
        ORDER BY catalog_items.source_row_number, catalog_items.id
        """,
        (record.region_folder, record.filename, RNMC_ZIP_SOURCE_NAME),
    ).fetchall()
    return [
        CatalogItemRecord(
            id=int(row["id"]),
            task_id=str(row["task_id"]),
            region=str(row["region"]),
            code=str(row["code"]),
            unit=str(row["unit"]),
            quantity=_optional_float(row["quantity"]),
            work_name=str(row["work_name"]),
            price=float(row["price"]),
            price_original=_optional_float(row["price_original"]),
            price_zlvl=_optional_float(row["price_zlvl"]),
            total_price=_optional_float(row["total_price"]),
            labor_unit=_optional_float(row["labor_unit"]),
            labor_total=_optional_float(row["labor_total"]),
            machine_labor_unit=_optional_float(row["machine_labor_unit"]),
            machine_labor_total=_optional_float(row["machine_labor_total"]),
            regional_coefficient=_optional_float(row["regional_coefficient"]),
            lsr_quarter=str(row["lsr_quarter"]),
            planned_start=str(row["planned_start"]),
            planned_finish=str(row["planned_finish"]),
            source_row_number=int(row["source_row_number"]),
        )
        for row in rows
    ]


def list_import_row_logs(
    connection: sqlite3.Connection,
    import_id: int,
) -> list[ImportRowLogRecord]:
    rows = connection.execute(
        """
        SELECT id, file_id, row_number, status, reason
        FROM import_row_log
        WHERE file_id = ?
        ORDER BY row_number, id
        """,
        (int(import_id),),
    ).fetchall()
    return [
        ImportRowLogRecord(
            id=int(row["id"]),
            file_id=int(row["file_id"]),
            row_number=int(row["row_number"]),
            status=str(row["status"]),
            reason=str(row["reason"]),
        )
        for row in rows
    ]


def replace_import_row_logs(
    connection: sqlite3.Connection,
    import_id: int,
    row_logs: list[tuple[int, str, str]],
) -> None:
    connection.execute("DELETE FROM import_row_log WHERE file_id = ?", (int(import_id),))
    if row_logs:
        connection.executemany(
            """
            INSERT INTO import_row_log(file_id, row_number, status, reason)
            VALUES (?, ?, ?, ?)
            """,
            [
                (int(import_id), int(row_number), _text(status), _text(reason))
                for row_number, status, reason in row_logs
            ],
        )


def normalize_import_filename(filename: str) -> str:
    return Path(str(filename).replace("\\", "/")).name.strip().casefold()


def replace_catalog_rows_for_file(
    connection: sqlite3.Connection,
    items: list[CatalogRowStorageItem],
    *,
    region_folder: str,
    filename: str,
    source_name: str = RNMC_ZIP_SOURCE_NAME,
    source_kind: str = RNMC_ZIP_SOURCE_KIND,
) -> CatalogFileRowsImportResult:
    source_id = _get_or_create_source(connection, source_name, kind=source_kind)
    region = _text(region_folder)
    source_filename = _text(filename)

    connection.execute(
        """
        DELETE FROM catalog_items
        WHERE source_id = ?
          AND source_region_folder = ?
          AND source_filename = ?
        """,
        (source_id, region, source_filename),
    )

    payload: list[tuple] = []
    skipped = 0
    for item in items:
        row = item.catalog_row
        if not _is_storable_row(row):
            skipped += 1
            continue
        price = _parse_positive_price(row.price)
        if price is None:
            skipped += 1
            continue
        price_original = _parse_optional_number(row.price_original)
        price_zlvl = _parse_optional_number(row.price_zlvl)
        payload.append(
            (
                source_id,
                _text(row.task_id),
                _text(row.region),
                _text(row.code),
                _text(row.unit),
                _parse_optional_number(row.quantity),
                _text(row.work_name),
                price,
                price_original if price_original is not None else price,
                price_zlvl if price_zlvl is not None else price,
                _parse_optional_number(row.total_price),
                _parse_optional_number(row.labor_unit),
                _parse_optional_number(row.labor_total),
                _parse_optional_number(row.machine_labor_unit),
                _parse_optional_number(row.machine_labor_total),
                _parse_optional_number(row.regional_coefficient),
                _metadata_text(row.lsr_quarter),
                _metadata_text(row.planned_start),
                _metadata_text(row.planned_finish),
                _serialize_date(row.added_date),
                _text(item.source_region_folder),
                _text(item.source_filename),
                int(item.source_row_number),
            )
        )

    for offset in range(0, len(payload), BATCH_SIZE):
        connection.executemany(
            """
            INSERT INTO catalog_items (
                source_id, task_id, region, code, unit, quantity, work_name, price,
                price_original, price_zlvl, total_price, labor_unit, labor_total, machine_labor_unit,
                machine_labor_total, regional_coefficient, lsr_quarter, planned_start,
                planned_finish, added_date, source_region_folder, source_filename, source_row_number
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload[offset : offset + BATCH_SIZE],
        )

    from core.storage.corrections import synchronize_catalog_corrections

    synchronize_catalog_corrections(connection)
    return CatalogFileRowsImportResult(
        source_name=source_name,
        source_id=source_id,
        rows_imported=len(payload),
        rows_skipped=skipped,
        source_filename=source_filename,
        region_folder=region,
    )


def import_catalog_from_excel(
    connection: sqlite3.Connection,
    workbook_path: str | Path,
    *,
    source_name: str = DEFAULT_SOURCE_NAME,
    replace: bool = True,
    settings: Settings | None = None,
) -> CatalogImportResult:
    path = Path(workbook_path).resolve()
    source_id = _get_or_create_source(
        connection,
        source_name,
        kind="excel_bulk",
    )

    if replace:
        connection.execute("DELETE FROM catalog_items WHERE source_id = ?", (source_id,))
        connection.execute("DELETE FROM imported_files WHERE source_id = ?", (source_id,))

    positioned_rows = read_catalog_rows_with_positions(path, settings)
    payload: list[tuple] = []
    skipped = 0

    for row_number, catalog_row in positioned_rows:
        if not _is_storable_row(catalog_row):
            skipped += 1
            continue
        price = _parse_positive_price(catalog_row.price)
        if price is None:
            skipped += 1
            continue
        price_original = _parse_optional_number(catalog_row.price_original)
        price_zlvl = _parse_optional_number(catalog_row.price_zlvl)
        payload.append(
            (
                source_id,
                _text(catalog_row.task_id),
                _text(catalog_row.region),
                _text(catalog_row.code),
                _text(catalog_row.unit),
                _parse_optional_number(catalog_row.quantity),
                _text(catalog_row.work_name),
                price,
                price_original if price_original is not None else price,
                price_zlvl if price_zlvl is not None else price,
                _parse_optional_number(catalog_row.total_price),
                _parse_optional_number(catalog_row.labor_unit),
                _parse_optional_number(catalog_row.labor_total),
                _parse_optional_number(catalog_row.machine_labor_unit),
                _parse_optional_number(catalog_row.machine_labor_total),
                _parse_optional_number(catalog_row.regional_coefficient),
                _metadata_text(catalog_row.lsr_quarter),
                _metadata_text(catalog_row.planned_start),
                _metadata_text(catalog_row.planned_finish),
                _serialize_date(catalog_row.added_date),
                "",
                        _text(catalog_row.source_filename) or path.name,
                row_number,
            )
        )

    for offset in range(0, len(payload), BATCH_SIZE):
        connection.executemany(
            """
            INSERT INTO catalog_items (
                source_id, task_id, region, code, unit, quantity, work_name, price,
                price_original, price_zlvl, total_price, labor_unit, labor_total,
                machine_labor_unit, machine_labor_total, regional_coefficient,
                lsr_quarter, planned_start, planned_finish, added_date,
                source_region_folder, source_filename, source_row_number
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload[offset : offset + BATCH_SIZE],
        )

    _record_catalog_source_files(
        connection,
        source_id=source_id,
        positioned_rows=positioned_rows,
    )

    _record_imported_file(
        connection,
        source_id=source_id,
        region_folder="",
        filename=path.name,
        status=STATUS_SUCCESS,
        rows_ok=len(payload),
        rows_rejected=skipped,
    )
    from core.storage.corrections import synchronize_catalog_corrections

    synchronize_catalog_corrections(connection)
    connection.commit()

    return CatalogImportResult(
        source_name=source_name,
        source_id=source_id,
        rows_imported=len(payload),
        rows_skipped=skipped,
        source_filename=path.name,
    )



def _record_catalog_source_files(
    connection: sqlite3.Connection,
    *,
    source_id: int,
    positioned_rows: list[tuple[int, CatalogRow]],
) -> None:
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for _row_number, catalog_row in positioned_rows:
        filename = _text(catalog_row.source_filename)
        if filename == "":
            continue
        region = _text(catalog_row.region)
        key = (region, filename)
        item = grouped.setdefault(
            key,
            {
                "rows_ok": 0,
                "rows_rejected": 0,
                "task_number": "",
                "lsr_quarter": "",
                "planned_start": "",
                "planned_finish": "",
                "regional_coefficient": None,
            },
        )
        if _is_storable_row(catalog_row):
            item["rows_ok"] = int(item["rows_ok"]) + 1
        else:
            item["rows_rejected"] = int(item["rows_rejected"]) + 1
        if item["task_number"] == "":
            item["task_number"] = _text(catalog_row.task_id)
        if item["lsr_quarter"] == "":
            item["lsr_quarter"] = _metadata_text(catalog_row.lsr_quarter)
        if item["planned_start"] == "":
            item["planned_start"] = _metadata_text(catalog_row.planned_start)
        if item["planned_finish"] == "":
            item["planned_finish"] = _metadata_text(catalog_row.planned_finish)
        if item["regional_coefficient"] is None:
            item["regional_coefficient"] = _parse_optional_number(catalog_row.regional_coefficient)

    for (region, filename), item in grouped.items():
        _record_imported_file(
            connection,
            source_id=source_id,
            region_folder=region,
            filename=filename,
            status=STATUS_LEGACY_IMPORTED,
            rows_ok=int(item["rows_ok"]),
            rows_rejected=int(item["rows_rejected"]),
            failure_reason="",
            task_number=_text(item["task_number"]),
            legacy_note="legacy catalog source_file",
            lsr_quarter=_text(item["lsr_quarter"]),
            planned_start=_text(item["planned_start"]),
            planned_finish=_text(item["planned_finish"]),
            regional_coefficient=item["regional_coefficient"],
        )


def _legacy_file_log_headers(sheet) -> dict[str, int]:
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    normalized = {_normalize_header(value): index for index, value in enumerate(raw_headers)}
    return {
        "region_folder": _find_header(normalized, ["folder"]),
        "filename": _find_header(normalized, ["file"]),
        "legacy_note": _find_header(normalized, ["status"]),
        "lsr_quarter": _find_header(normalized, ["год квартал лср"]),
        "planned_start": _find_header(
            normalized,
            [
                "планирумый срок начала работ",
                "планируемый срок начала работ",
            ],
        ),
        "planned_finish": _find_header(
            normalized,
            ["планируемый срок окончания работ"],
        ),
    }


def _legacy_file_log_row(values: tuple[object, ...], headers: dict[str, int]) -> dict[str, object]:
    result: dict[str, object] = {}
    for name, index in headers.items():
        result[name] = values[index] if 0 <= index < len(values) else None
    return result


def _find_header(normalized_headers: dict[str, int], aliases: list[str]) -> int:
    for alias in aliases:
        if alias in normalized_headers:
            return normalized_headers[alias]
    return -1


def _normalize_header(value: object) -> str:
    return " ".join(_text(value).casefold().split())


def _row_is_empty(row: dict[str, object]) -> bool:
    return all(_text(value) == "" for value in row.values())


def _region_from_folder_text(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    parts = [part.strip() for part in text.replace("/", "\\").split("\\") if part.strip()]
    return parts[-1] if parts else text


def _legacy_rows_ok(value: object) -> int:
    if value is None or isinstance(value, bool):
        return 0
    if isinstance(value, Real):
        return max(0, int(value))
    text = _text(value).replace(",", ".")
    try:
        return max(0, int(float(text)))
    except ValueError:
        return 0



def _get_or_create_source(
    connection: sqlite3.Connection,
    name: str,
    *,
    kind: str,
) -> int:
    existing = connection.execute(
        "SELECT id FROM catalog_sources WHERE name = ?",
        (name,),
    ).fetchone()
    if existing is not None:
        return int(existing["id"])

    cursor = connection.execute(
        "INSERT INTO catalog_sources(name, kind) VALUES (?, ?)",
        (name, kind),
    )
    return int(cursor.lastrowid)


def _require_source_id(connection: sqlite3.Connection, source_name: str) -> int:
    row = connection.execute(
        "SELECT id FROM catalog_sources WHERE name = ?",
        (source_name,),
    ).fetchone()
    if row is None:
        raise LookupError(f"catalog source not found: {source_name}")
    return int(row["id"])


def _record_imported_file(
    connection: sqlite3.Connection,
    *,
    source_id: int,
    region_folder: str,
    filename: str,
    status: str,
    rows_ok: int,
    rows_rejected: int,
    failure_reason: str = "",
    task_number: str = "",
    legacy_note: str = "",
    lsr_quarter: str = "",
    planned_start: str = "",
    planned_finish: str = "",
    regional_coefficient: object = None,
) -> int:
    connection.execute(
        """
        INSERT INTO imported_files (
            source_id, region_folder, filename, status, task_number,
            rows_ok, rows_rejected, failure_reason, filename_key, legacy_note,
            lsr_quarter, planned_start, planned_finish, regional_coefficient
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(region_folder, filename) DO UPDATE SET
            source_id = excluded.source_id,
            status = excluded.status,
            imported_at = datetime('now'),
            task_number = excluded.task_number,
            rows_ok = excluded.rows_ok,
            rows_rejected = excluded.rows_rejected,
            failure_reason = excluded.failure_reason,
            filename_key = excluded.filename_key,
            legacy_note = excluded.legacy_note,
            lsr_quarter = excluded.lsr_quarter,
            planned_start = excluded.planned_start,
            planned_finish = excluded.planned_finish,
            regional_coefficient = excluded.regional_coefficient
        """,
        (
            source_id,
            region_folder,
            filename,
            status,
            task_number,
            rows_ok,
            rows_rejected,
            failure_reason,
            normalize_import_filename(filename),
            legacy_note,
            lsr_quarter,
            planned_start,
            planned_finish,
            _parse_optional_number(regional_coefficient),
        ),
    )
    row = connection.execute(
        """
        SELECT id FROM imported_files
        WHERE region_folder = ? AND filename = ?
        LIMIT 1
        """,
        (region_folder, filename),
    ).fetchone()
    if row is None:  # pragma: no cover - sqlite upsert defensive boundary
        raise RuntimeError("imported file row was not recorded")
    return int(row["id"])


def _is_storable_row(catalog_row: CatalogRow) -> bool:
    if _text(catalog_row.task_id) == "":
        return False
    if NormCode(catalog_row.code) == "":
        return False
    if NormUnit(catalog_row.unit) == "":
        return False
    return _parse_positive_price(catalog_row.price) is not None


def _row_to_catalog_row(row: sqlite3.Row) -> CatalogRow:
    added_date = row["added_date"]
    return CatalogRow(
        task_id=row["task_id"],
        price=row["price"],
        code=row["code"],
        unit=row["unit"],
        work_name=row["work_name"],
        quantity=row["quantity"],
        region=row["region"],
        added_date=None if added_date is None else str(added_date),
        price_original=row["price_original"],
        price_zlvl=row["price_zlvl"],
        total_price=row["total_price"],
        labor_unit=row["labor_unit"],
        labor_total=row["labor_total"],
        machine_labor_unit=row["machine_labor_unit"],
        machine_labor_total=row["machine_labor_total"],
        regional_coefficient=row["regional_coefficient"],
        lsr_quarter=row["lsr_quarter"],
        planned_start=row["planned_start"],
        planned_finish=row["planned_finish"],
        source_filename=row["source_filename"],
    )


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _metadata_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    return "" if text in {"-", "—"} else text


def _parse_positive_price(value: object) -> float | None:
    number = _parse_optional_number(value)
    if number is None or number <= 0:
        return None
    return number


def _parse_optional_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Real):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    text = text.replace(" ", " ").replace(" ", "")
    text = "".join(char for char in text if char.isdigit() or char in {".", ",", "-"})
    if text in {"", ".", ",", "-"}:
        return None
    if "," in text and "." in text:
        comma_pos = text.rfind(",")
        dot_pos = text.rfind(".")
        decimal_sep = "," if comma_pos > dot_pos else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "")
        text = text.replace(decimal_sep, ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _optional_float(value: object) -> float | None:
    if value is None:
        return None
    return float(value)


def _serialize_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None

"""Parse the KL20 CatalogBuilder aggregate workbook (TKP winner catalog).

Ports the *output format* of modKL20_CatalogBuilder_ASCII_V3.bas
(`BuildKL20WinnerCatalog_ASCII_V3`), not the macro's own file-walk/parse
logic: the macro already does the hard work of locating each tender's
comparison sheet, picking the winning bidder's price block, and assembling
two flat sheets in a workbook:

- `KL20_FileCatalog` - one row per source file scanned, with parse status
  (`OK` / `Skipped` / `Needs review`), the winner's identity/total, and the
  tender's own metadata (task number, customer, procedure name).
- `KL20_WOR_Catalog` - one row per priced work item, carrying the same
  per-tender metadata plus the item's own name/unit/quantity/prices.

This module only reads that already-aggregated workbook into plain
dataclasses (`TkpSourceFile`, `TkpItem`); core/storage/tkp.py is
responsible for the delta-import/dedup and persistence, the same split as
core/ingest.py vs core/storage/catalog.py for the RNMC pipeline.
"""

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

FILE_CATALOG_SHEET = "KL20_FileCatalog"
WOR_CATALOG_SHEET = "KL20_WOR_Catalog"

STATUS_OK = "OK"

# Column headers exactly as written by WriteHeaders(...) in the macro
# (see BuildKL20WinnerCatalog_ASCII_V3), kept as the single source of truth
# for column position instead of hardcoded indices.
FILE_CATALOG_HEADERS = (
    "RunId", "FilePath", "FileName", "ModifiedDate", "SheetName",
    "ParseStatus", "ParseMessage", "UsedRows", "UsedCols", "WorStartRow",
    "WorEndRow", "WorEndMethod", "WorPositionCount", "WorSectionCount",
    "ParticipantCount", "WinnerGroupIndex", "WinnerStartCol",
    "WinnerStartColLetter", "WinnerMethod", "WinnerName", "WinnerINN",
    "WinnerUIN", "WinnerTaskNoBK", "WinnerHeader1", "WinnerHeader2",
    "WinnerHeader3", "WinnerHeader4", "WinnerBlockName", "WinnerBlockUIN",
    "WinnerBlockTotalVat", "WinnerBlockReason", "WinnerBlockSource",
    "TaskNo", "RequestDate", "Version", "Customer", "GeneralContractor",
    "ProcedureName", "WinnerTotalNoVat", "WinnerTotalVat", "RnmcTotalNoVat",
)

FILE_CATALOG_OPTIONAL_HEADERS = (
    "ReserveMethod", "ReserveName", "ReserveINN", "ReserveUIN",
    "ReserveTotalNoVat", "ReserveTotalVat",
)

WOR_CATALOG_HEADERS = (
    "RunId", "FilePath", "FileName", "SheetName", "SourceRow",
    "SectionCode", "SectionName", "SubsectionName", "ItemCode", "ItemName",
    "Unit", "Qty", "QtySourceText", "RnmcUnitPriceNoVat",
    "RnmcLineTotalNoVat", "WinnerUnitPriceNoVat", "WinnerLineTotalNoVat",
    "WinnerName", "WinnerINN", "WinnerUIN", "WinnerGroupIndex",
    "WinnerStartCol", "WinnerStartColLetter", "WinnerUnitHeader",
    "WinnerTotalHeader", "TaskNo", "RequestDate", "Version", "Customer",
    "GeneralContractor", "ProcedureName", "WinnerMethod", "WinnerBlockName",
    "WinnerBlockUIN", "WinnerBlockTotalVat", "WinnerBlockReason",
    "WinnerBlockSource",
)

WOR_CATALOG_OPTIONAL_HEADERS = (
    "ReserveUnitPriceNoVat", "ReserveLineTotalNoVat", "ReserveName",
    "ReserveINN", "ReserveUIN", "ReserveGroupIndex", "ReserveStartCol",
    "ReserveStartColLetter", "ReserveUnitHeader", "ReserveTotalHeader",
    "ReserveMethod", "WorSchema", "QualityFlags",
)


@dataclass(frozen=True)
class TkpSourceFile:
    """One row from KL20_FileCatalog: one scanned tender-comparison file."""

    run_id: str
    file_path: str
    file_name: str
    modified_date: datetime | None
    sheet_name: str
    parse_status: str
    parse_message: str
    task_no: str
    request_date: datetime | None
    customer: str
    general_contractor: str
    procedure_name: str
    winner_name: str
    winner_inn: str
    winner_uin: str
    winner_total_no_vat: float | None
    winner_total_vat: float | None
    rnmc_total_no_vat: float | None
    reserve_name: str = ""
    reserve_inn: str = ""
    reserve_uin: str = ""
    reserve_total_no_vat: float | None = None
    reserve_total_vat: float | None = None
    reserve_method: str = ""

    @property
    def is_parsed(self) -> bool:
        return self.parse_status == STATUS_OK


@dataclass(frozen=True)
class TkpItem:
    """One row from KL20_WOR_Catalog: one priced work item from a winner."""

    run_id: str
    file_path: str
    file_name: str
    sheet_name: str
    source_row: int
    section_code: str
    section_name: str
    subsection_name: str
    item_code: str
    item_name: str
    unit: str
    qty: float | None
    qty_source_text: str
    rnmc_unit_price_no_vat: float | None
    rnmc_line_total_no_vat: float | None
    winner_unit_price_no_vat: float | None
    winner_line_total_no_vat: float | None
    winner_name: str
    winner_inn: str
    winner_uin: str
    winner_group_index: int
    winner_start_col: int
    winner_start_col_letter: str
    winner_unit_header: str
    winner_total_header: str
    task_no: str
    request_date: datetime | None
    version: str
    customer: str
    general_contractor: str
    procedure_name: str
    winner_method: str
    winner_block_name: str
    winner_block_uin: str
    winner_block_total_vat: float | None
    winner_block_reason: str
    reserve_unit_price_no_vat: float | None = None
    reserve_line_total_no_vat: float | None = None
    reserve_name: str = ""
    reserve_inn: str = ""
    reserve_uin: str = ""
    reserve_group_index: int = 0
    reserve_start_col: int = 0
    reserve_start_col_letter: str = ""
    reserve_unit_header: str = ""
    reserve_total_header: str = ""
    reserve_method: str = ""
    wor_schema: str = ""
    quality_flags: str = ""


def is_usable_tkp_unit_price(value: float | None) -> bool:
    """Return True only for an explicit positive participant unit price."""
    return value is not None and value > 0


def tkp_item_has_usable_unit_price(item: TkpItem) -> bool:
    """A TKP row is importable when winner or reserve has an explicit unit price."""
    return (
        is_usable_tkp_unit_price(item.winner_unit_price_no_vat)
        or is_usable_tkp_unit_price(item.reserve_unit_price_no_vat)
    )


def tkp_item_price_status(item: TkpItem) -> str:
    """Return a stable preview status for participant unit-price completeness."""
    winner_ok = is_usable_tkp_unit_price(item.winner_unit_price_no_vat)
    reserve_ok = is_usable_tkp_unit_price(item.reserve_unit_price_no_vat)
    if winner_ok and reserve_ok:
        return "ready_both_prices"
    if winner_ok:
        return "ready_winner_price"
    if reserve_ok:
        return "ready_reserve_price"
    return "missing_both_unit_prices"


@dataclass(frozen=True)
class TkpBlockDiagnostic:
    """One macro-level source block check used by the staged TKP preview."""

    code: str
    label: str
    found: bool
    value: str = ""
    row: int = 0


@dataclass(frozen=True)
class TkpSourceDiagnostics:
    """Macro-level diagnostics for one original KL workbook."""

    file_path: str
    file_name: str
    sheet_name: str
    wor_start_row: int = 0
    wor_end_row: int = 0
    wor_end_method: str = ""
    wor_schema: str = ""
    participant_count: int = 0
    rows_found: int = 0
    rows_with_usable_price: int = 0
    rows_rejected_missing_prices: int = 0
    blocks: tuple[TkpBlockDiagnostic, ...] = ()


@dataclass(frozen=True)
class TkpCatalogParseResult:
    """Everything read from one KL20 CatalogBuilder aggregate workbook."""

    run_ids: tuple[str, ...]
    files: list[TkpSourceFile] = field(default_factory=list)
    items: list[TkpItem] = field(default_factory=list)
    diagnostics: list[TkpSourceDiagnostics] = field(default_factory=list)


class TkpCatalogFormatError(ValueError):
    """The workbook does not look like a KL20 CatalogBuilder export."""


def parse_tkp_catalog_workbook(path: str | Path) -> TkpCatalogParseResult:
    """Read a KL20 aggregate workbook into source files and priced items.

    `KL20_WOR_Catalog` is mandatory. `KL20_FileCatalog` is optional because
    business users may keep only the selected WOR columns; source rows are
    derived from WOR metadata in that case. Header validation remains strict
    so a shifted or renamed column is never read silently as another field.
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if WOR_CATALOG_SHEET not in workbook.sheetnames:
            raise TkpCatalogFormatError(f"sheet '{WOR_CATALOG_SHEET}' not found")

        items = _read_wor_catalog(workbook[WOR_CATALOG_SHEET])
        files = (
            _read_file_catalog(workbook[FILE_CATALOG_SHEET])
            if FILE_CATALOG_SHEET in workbook.sheetnames
            else []
        )
        files = _include_sources_derived_from_items(files, items)
        run_ids = tuple(dict.fromkeys(
            [*(source.run_id for source in files), *(item.run_id for item in items)]
        ))
        return TkpCatalogParseResult(run_ids=run_ids, files=files, items=items)
    finally:
        workbook.close()


def _header_index(
    worksheet: Worksheet,
    expected: tuple[str, ...],
    sheet_label: str,
    optional: tuple[str, ...] = (),
) -> dict[str, int]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    index = {str(value): position for position, value in enumerate(header_row) if value is not None}
    missing = [name for name in expected if name not in index]
    if missing:
        raise TkpCatalogFormatError(
            f"{sheet_label}: missing expected column(s) {missing}"
        )
    for name in optional:
        index.setdefault(name, -1)
    return index


def _optional(row: tuple, index: dict[str, int], name: str) -> object:
    position = index.get(name, -1)
    if position < 0 or position >= len(row):
        return None
    return row[position]


def _read_file_catalog(worksheet: Worksheet) -> list[TkpSourceFile]:
    index = _header_index(
        worksheet,
        FILE_CATALOG_HEADERS,
        FILE_CATALOG_SHEET,
        FILE_CATALOG_OPTIONAL_HEADERS,
    )
    files: list[TkpSourceFile] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[index["FilePath"]] in (None, ""):
            continue
        files.append(
            TkpSourceFile(
                run_id=_text(row[index["RunId"]]),
                file_path=_text(row[index["FilePath"]]),
                file_name=_text(row[index["FileName"]]),
                modified_date=_as_datetime(row[index["ModifiedDate"]]),
                sheet_name=_text(row[index["SheetName"]]),
                parse_status=_text(row[index["ParseStatus"]]),
                parse_message=_text(row[index["ParseMessage"]]),
                task_no=_text(row[index["TaskNo"]]),
                request_date=_as_datetime(row[index["RequestDate"]]),
                customer=_text(row[index["Customer"]]),
                general_contractor=_text(row[index["GeneralContractor"]]),
                procedure_name=_text(row[index["ProcedureName"]]),
                winner_name=_text(row[index["WinnerName"]]),
                winner_inn=_text(row[index["WinnerINN"]]),
                winner_uin=_text(row[index["WinnerUIN"]]),
                winner_total_no_vat=_as_float(row[index["WinnerTotalNoVat"]]),
                winner_total_vat=_as_float(row[index["WinnerTotalVat"]]),
                rnmc_total_no_vat=_as_float(row[index["RnmcTotalNoVat"]]),
                reserve_name=_text(_optional(row, index, "ReserveName")),
                reserve_inn=_text(_optional(row, index, "ReserveINN")),
                reserve_uin=_text(_optional(row, index, "ReserveUIN")),
                reserve_total_no_vat=_as_float(_optional(row, index, "ReserveTotalNoVat")),
                reserve_total_vat=_as_float(_optional(row, index, "ReserveTotalVat")),
                reserve_method=_text(_optional(row, index, "ReserveMethod")),
            )
        )
    return files


def _read_wor_catalog(worksheet: Worksheet) -> list[TkpItem]:
    index = _header_index(
        worksheet,
        WOR_CATALOG_HEADERS,
        WOR_CATALOG_SHEET,
        WOR_CATALOG_OPTIONAL_HEADERS,
    )
    items: list[TkpItem] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item_name = _text(row[index["ItemName"]])
        if item_name == "":
            continue
        items.append(
            TkpItem(
                run_id=_text(row[index["RunId"]]),
                file_path=_text(row[index["FilePath"]]),
                file_name=_text(row[index["FileName"]]),
                sheet_name=_text(row[index["SheetName"]]),
                source_row=_as_int(row[index["SourceRow"]]),
                section_code=_text(row[index["SectionCode"]]),
                section_name=_text(row[index["SectionName"]]),
                subsection_name=_text(row[index["SubsectionName"]]),
                item_code=_text(row[index["ItemCode"]]),
                item_name=item_name,
                unit=_text(row[index["Unit"]]),
                qty=_as_float(row[index["Qty"]]),
                qty_source_text=_text(row[index["QtySourceText"]]),
                rnmc_unit_price_no_vat=_as_float(row[index["RnmcUnitPriceNoVat"]]),
                rnmc_line_total_no_vat=_as_float(row[index["RnmcLineTotalNoVat"]]),
                winner_unit_price_no_vat=_as_float(row[index["WinnerUnitPriceNoVat"]]),
                winner_line_total_no_vat=_as_float(row[index["WinnerLineTotalNoVat"]]),
                winner_name=_text(row[index["WinnerName"]]),
                winner_inn=_text(row[index["WinnerINN"]]),
                winner_uin=_text(row[index["WinnerUIN"]]),
                winner_group_index=_as_int(row[index["WinnerGroupIndex"]]),
                winner_start_col=_as_int(row[index["WinnerStartCol"]]),
                winner_start_col_letter=_text(row[index["WinnerStartColLetter"]]),
                winner_unit_header=_text(row[index["WinnerUnitHeader"]]),
                winner_total_header=_text(row[index["WinnerTotalHeader"]]),
                task_no=_text(row[index["TaskNo"]]),
                request_date=_as_datetime(row[index["RequestDate"]]),
                version=_text(row[index["Version"]]),
                customer=_text(row[index["Customer"]]),
                general_contractor=_text(row[index["GeneralContractor"]]),
                procedure_name=_text(row[index["ProcedureName"]]),
                winner_method=_text(row[index["WinnerMethod"]]),
                winner_block_name=_text(row[index["WinnerBlockName"]]),
                winner_block_uin=_text(row[index["WinnerBlockUIN"]]),
                winner_block_total_vat=_as_float(row[index["WinnerBlockTotalVat"]]),
                winner_block_reason=_text(row[index["WinnerBlockReason"]]),
                reserve_unit_price_no_vat=_as_float(_optional(row, index, "ReserveUnitPriceNoVat")),
                reserve_line_total_no_vat=_as_float(_optional(row, index, "ReserveLineTotalNoVat")),
                reserve_name=_text(_optional(row, index, "ReserveName")),
                reserve_inn=_text(_optional(row, index, "ReserveINN")),
                reserve_uin=_text(_optional(row, index, "ReserveUIN")),
                reserve_group_index=_as_int(_optional(row, index, "ReserveGroupIndex")),
                reserve_start_col=_as_int(_optional(row, index, "ReserveStartCol")),
                reserve_start_col_letter=_text(_optional(row, index, "ReserveStartColLetter")),
                reserve_unit_header=_text(_optional(row, index, "ReserveUnitHeader")),
                reserve_total_header=_text(_optional(row, index, "ReserveTotalHeader")),
                reserve_method=_text(_optional(row, index, "ReserveMethod")),
                wor_schema=_text(_optional(row, index, "WorSchema")),
                quality_flags=_text(_optional(row, index, "QualityFlags")),
            )
        )
    return items


def _include_sources_derived_from_items(
    files: list[TkpSourceFile],
    items: list[TkpItem],
) -> list[TkpSourceFile]:
    """Add source records for WOR rows absent from `KL20_FileCatalog`."""
    result = list(files)
    known = {(source.file_path, source.file_name) for source in files}
    first_items: dict[tuple[str, str], TkpItem] = {}
    for item in items:
        first_items.setdefault((item.file_path, item.file_name), item)

    for key, item in first_items.items():
        if key in known:
            continue
        result.append(
            TkpSourceFile(
                run_id=item.run_id,
                file_path=item.file_path,
                file_name=item.file_name,
                modified_date=None,
                sheet_name=item.sheet_name,
                parse_status=STATUS_OK,
                parse_message="Derived from KL20_WOR_Catalog.",
                task_no=item.task_no,
                request_date=item.request_date,
                customer=item.customer,
                general_contractor=item.general_contractor,
                procedure_name=item.procedure_name,
                winner_name=item.winner_name,
                winner_inn=item.winner_inn,
                winner_uin=item.winner_uin,
                winner_total_no_vat=None,
                winner_total_vat=None,
                rnmc_total_no_vat=None,
                reserve_name=item.reserve_name,
                reserve_inn=item.reserve_inn,
                reserve_uin=item.reserve_uin,
                reserve_total_no_vat=None,
                reserve_total_vat=None,
                reserve_method=item.reserve_method,
            )
        )
    return result


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_float(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _as_int(value: object) -> int:
    parsed = _as_float(value)
    return int(parsed) if parsed is not None else 0


def _as_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    return None

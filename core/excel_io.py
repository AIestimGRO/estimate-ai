"""Excel readers for catalog and estimate workbooks."""

from dataclasses import dataclass
from numbers import Real
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.catalog import CatalogRow
from core.layout import (
    CATALOG_FIELD_ADDED_DATE,
    CATALOG_FIELD_CODE,
    CATALOG_FIELD_PRICE,
    CATALOG_FIELD_REGION,
    CATALOG_FIELD_TASK_ID,
    CATALOG_FIELD_UNIT,
    CATALOG_FIELD_WORK_NAME,
    DEFAULT_MAX_BLANK_RUN,
    LayoutResult,
    data_row_numbers,
    load_catalog_layout_config,
    resolve_catalog_layout,
)
from core.matching import EstimateRow
from core.normalize import NormCode, NormUnit


CATALOG_SHEET_PART = "\u041a\u0430\u0442"
ESTIMATE_SHEET_PART = "\u041e\u0421"
GESN_MARKER = "\u0413\u042d\u0421\u041d"
FER_MARKER = "\u0424\u0415\u0420"
PERECHEN_MARKER = "\u041f\u0435\u0440"


@dataclass(frozen=True)
class Settings:
    """Column settings mirroring Module1.TSettings defaults."""

    col_search: int = 14
    col_avg: int = 7
    col_kr: int = 14
    col_f: int = 6
    col_section: int = 15
    col_analog_start: int = 16
    col_smeta_work_name: int = 3
    col_smeta_unit: int = 4
    cat_task_col: int = 2
    cat_price_col: int = 7
    cat_code_col: int = 14
    cat_region_col: int = 16
    cat_work_name_col: int = 3
    cat_unit_col: int = 4
    cat_added_date_col: int = 17
    price_spread_limit: float = 3.0


def catalog_default_columns(settings: Settings) -> dict[str, int]:
    """Template column map used as the catalog resolver's fallback defaults."""
    return {
        CATALOG_FIELD_TASK_ID: settings.cat_task_col,
        CATALOG_FIELD_WORK_NAME: settings.cat_work_name_col,
        CATALOG_FIELD_UNIT: settings.cat_unit_col,
        CATALOG_FIELD_PRICE: settings.cat_price_col,
        CATALOG_FIELD_CODE: settings.cat_code_col,
        CATALOG_FIELD_REGION: settings.cat_region_col,
        CATALOG_FIELD_ADDED_DATE: settings.cat_added_date_col,
    }


def _catalog_column(
    layout: LayoutResult,
    field: str,
    defaults: dict[str, int],
) -> int:
    column = layout.column(field)
    if column is not None:
        return column
    return defaults[field]


def read_catalog_rows(
    workbook_path: str | Path,
    settings: Settings | None = None,
) -> list[CatalogRow]:
    """Read raw catalog rows from the catalog sheet.

    Uses ``data_only=True`` so formula-driven price/date cells resolve to the
    values Excel cached on save (real files store prices as formulas), instead
    of returning the formula text (OPEN_ITEMS #2).
    """
    return [row for _, row in read_catalog_rows_with_positions(workbook_path, settings)]


def read_catalog_rows_with_positions(
    workbook_path: str | Path,
    settings: Settings | None = None,
) -> list[tuple[int, CatalogRow]]:
    """Read catalog rows with their 1-based Excel row numbers."""
    active_settings = Settings() if settings is None else settings
    workbook = load_workbook(workbook_path, data_only=True)

    try:
        worksheet = _find_sheet_by_part(workbook.worksheets, CATALOG_SHEET_PART)
        data_limit = int(getattr(worksheet, "max_row", 0) or 0)
        named_rows = _read_named_catalog_rows_with_positions(worksheet)
        if named_rows is not None:
            return named_rows
        defaults = catalog_default_columns(active_settings)
        catalog_config = load_catalog_layout_config()
        layout = (
            resolve_catalog_layout(worksheet, catalog_config, default_columns=defaults)
            if catalog_config is not None
            else None
        )

        if layout is not None and layout.ok and layout.header_row > 0:
            data_start = layout.header_row + catalog_config.data_start_offset
            task_col = _catalog_column(layout, CATALOG_FIELD_TASK_ID, defaults)
            price_col = _catalog_column(layout, CATALOG_FIELD_PRICE, defaults)
            code_col = _catalog_column(layout, CATALOG_FIELD_CODE, defaults)
            unit_col = _catalog_column(layout, CATALOG_FIELD_UNIT, defaults)
            work_name_col = _catalog_column(layout, CATALOG_FIELD_WORK_NAME, defaults)
            region_col = _catalog_column(layout, CATALOG_FIELD_REGION, defaults)
            added_date_col = _catalog_column(layout, CATALOG_FIELD_ADDED_DATE, defaults)
        else:
            data_start = 4
            task_col = active_settings.cat_task_col
            price_col = active_settings.cat_price_col
            code_col = active_settings.cat_code_col
            unit_col = active_settings.cat_unit_col
            work_name_col = active_settings.cat_work_name_col
            region_col = active_settings.cat_region_col
            added_date_col = active_settings.cat_added_date_col

        rows: list[tuple[int, CatalogRow]] = []

        for row_number in range(data_start, data_limit + 1):
            rows.append(
                (
                    row_number,
                    CatalogRow(
                        task_id=_cell_value(worksheet, row_number, task_col),
                        price=_cell_value(worksheet, row_number, price_col),
                        code=_cell_value(worksheet, row_number, code_col),
                        unit=_cell_value(worksheet, row_number, unit_col),
                        work_name=_cell_value(
                            worksheet,
                            row_number,
                            work_name_col,
                        ),
                        region=_cell_value(worksheet, row_number, region_col),
                        added_date=_cell_value(
                            worksheet,
                            row_number,
                            added_date_col,
                        ),
                    ),
                )
            )

        return rows
    finally:
        workbook.close()


def read_estimate_rows(
    workbook_path: str | Path,
    settings: Settings | None = None,
) -> list[EstimateRow]:
    """Read raw working estimate rows from the estimate sheet."""
    return [row for _, row in read_estimate_rows_with_positions(workbook_path, settings)]


def read_estimate_rows_with_positions(
    workbook_path: str | Path,
    settings: Settings | None = None,
) -> list[tuple[int, EstimateRow]]:
    """Read working estimate rows tagged with their physical worksheet row.

    The physical row number is needed by the Excel writer to write results
    back onto the exact source row; matching itself stays position-agnostic.
    Values are read with ``data_only=True`` so formula prices resolve to their
    cached numbers (OPEN_ITEMS #2).
    """
    active_settings = Settings() if settings is None else settings
    workbook = load_workbook(workbook_path, data_only=True)

    try:
        worksheet = find_estimate_sheet(workbook, active_settings)
        return _collect_estimate_rows(worksheet, active_settings)
    finally:
        workbook.close()


def _collect_estimate_rows(
    worksheet: Worksheet,
    settings: Settings,
) -> list[tuple[int, EstimateRow]]:
    header_row = _detect_estimate_header_row(worksheet, settings)
    if header_row == 0:
        return []

    rows: list[tuple[int, EstimateRow]] = []
    for row_number in range(header_row + 2, worksheet.max_row + 1):
        code = _cell_value(worksheet, row_number, settings.col_search)
        unit = _cell_value(worksheet, row_number, settings.col_smeta_unit)
        base_price = _cell_value(worksheet, row_number, settings.col_f)

        if not _is_working_estimate_row(code, unit, base_price):
            continue

        rows.append(
            (
                row_number,
                EstimateRow(
                    code=code,
                    unit=unit,
                    work_name=_cell_value(
                        worksheet,
                        row_number,
                        settings.col_smeta_work_name,
                    ),
                    base_price=base_price,
                ),
            )
        )

    return rows


def find_estimate_sheet(
    workbook: Any,
    settings: Settings | None = None,
) -> Worksheet:
    """Locate the estimate worksheet (name contains the estimate marker)."""
    return _find_sheet_by_part(workbook.worksheets, ESTIMATE_SHEET_PART)


def read_estimate_rows_from_worksheet(
    worksheet: Worksheet,
    settings: Settings | None = None,
) -> list[tuple[int, EstimateRow]]:
    """Template read of an already-open worksheet (fixed Settings columns)."""
    active_settings = Settings() if settings is None else settings
    return _collect_estimate_rows(worksheet, active_settings)


def read_estimate_rows_by_columns(
    worksheet: Worksheet,
    *,
    header_row: int,
    code_column: int,
    unit_column: int,
    work_name_column: int,
    base_price_column: int,
    max_blank_run: int = DEFAULT_MAX_BLANK_RUN,
) -> list[tuple[int, EstimateRow]]:
    """Read estimate rows using explicitly resolved columns (flexible path).

    Body rows are scanned tolerating isolated blank rows (data_row_numbers,
    R5); each row is still validated as a working estimate row.
    """
    key_columns = [code_column, unit_column, base_price_column]
    rows: list[tuple[int, EstimateRow]] = []

    for row_number in data_row_numbers(
        worksheet,
        header_row + 1,
        key_columns,
        max_blank_run=max_blank_run,
    ):
        code = _cell_value(worksheet, row_number, code_column)
        unit = _cell_value(worksheet, row_number, unit_column)
        base_price = _cell_value(worksheet, row_number, base_price_column)

        if not _is_working_estimate_row(code, unit, base_price):
            continue

        # Skip the column-enumeration row under the header (e.g. "1 2 3 ...").
        # Real codes always contain letters/dashes, so a bare integer in the
        # code cell is a header artefact, not a work item (R4-lite).
        if str(code).strip().isdigit():
            continue

        rows.append(
            (
                row_number,
                EstimateRow(
                    code=code,
                    unit=unit,
                    work_name=_cell_value(worksheet, row_number, work_name_column),
                    base_price=base_price,
                ),
            )
        )

    return rows


NAMED_CATALOG_HEADER_ALIASES = {
    "task_id": ("номерзадачи",),
    "work_name": ("наименованиеработ", "наименование"),
    "unit": ("ед.изм.", "ед.изм", "единицаизмерения"),
    "quantity": ("кол-во", "количество"),
    "total_price": ("итогостоимостьруб.безндс", "итогостоимостьбезндс"),
    "labor_unit": ("тзнаед.чел-час", "тзнаедчелчас"),
    "labor_total": ("тзвсегочел-час", "тзвсегочелчас"),
    "machine_labor_unit": ("тзмнаед.чел-час", "тзмнаедчелчас"),
    "machine_labor_total": ("тзмвсегочел-час", "тзмвсегочелчас"),
    "code": ("переченьгэсн/фер/тер/кр", "переченьгэсн/фер/гэсн/кр"),
    "source_filename": ("source_file", "исходныйфайл"),
    "region": ("регион",),
    "lsr_quarter": ("годкварталлср", "год/кварталлср"),
    "planned_start": ("планирумыйсрокначаларабот", "планируемыйсрокначаларабот"),
    "planned_finish": ("планируемыйсрококончанияработ",),
    "regional_coefficient": ("региональныйкоэффициент",),
    "added_date": ("датадобавлениявкаталог", "датадобавления", "датадобав"),
}


def _read_named_catalog_rows_with_positions(worksheet: Worksheet) -> list[tuple[int, CatalogRow]] | None:
    header_row, headers = _detect_named_catalog_header(worksheet)
    if header_row == 0:
        return None

    required = ("task_id", "work_name", "unit", "quantity", "price_original", "price_zlvl", "code")
    if any(headers.get(name, 0) <= 0 for name in required):
        return None

    rows: list[tuple[int, CatalogRow]] = []
    max_row = int(getattr(worksheet, "max_row", 0) or 0)
    max_col = max(headers.values())
    for row_offset, values in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=1,
    ):
        row_number = header_row + row_offset
        price_zlvl = _value_by_header(values, headers, "price_zlvl")
        rows.append(
            (
                row_number,
                CatalogRow(
                    task_id=_value_by_header(values, headers, "task_id"),
                    price=price_zlvl,
                    price_original=_value_by_header(values, headers, "price_original"),
                    price_zlvl=price_zlvl,
                    code=_value_by_header(values, headers, "code"),
                    unit=_value_by_header(values, headers, "unit"),
                    quantity=_value_by_header(values, headers, "quantity"),
                    work_name=_value_by_header(values, headers, "work_name"),
                    region=_value_by_header(values, headers, "region"),
                    added_date=_value_by_header(values, headers, "added_date"),
                    total_price=_value_by_header(values, headers, "total_price"),
                    labor_unit=_value_by_header(values, headers, "labor_unit"),
                    labor_total=_value_by_header(values, headers, "labor_total"),
                    machine_labor_unit=_value_by_header(values, headers, "machine_labor_unit"),
                    machine_labor_total=_value_by_header(values, headers, "machine_labor_total"),
                    regional_coefficient=_value_by_header(values, headers, "regional_coefficient"),
                    lsr_quarter=_clean_metadata_value(_value_by_header(values, headers, "lsr_quarter")),
                    planned_start=_clean_metadata_value(_value_by_header(values, headers, "planned_start")),
                    planned_finish=_clean_metadata_value(_value_by_header(values, headers, "planned_finish")),
                    source_filename=_value_by_header(values, headers, "source_filename"),
                ),
            )
        )
    return rows


def _detect_named_catalog_header(worksheet: Worksheet) -> tuple[int, dict[str, int]]:
    max_row = min(50, int(getattr(worksheet, "max_row", 0) or 0))
    max_col = min(80, int(getattr(worksheet, "max_column", 0) or 0))
    for row_number in range(1, max_row + 1):
        values = [worksheet.cell(row=row_number, column=column).value for column in range(1, max_col + 1)]
        normalized = {_compact_header(value): index + 1 for index, value in enumerate(values) if _compact_header(value)}
        headers = _named_catalog_header_map(normalized)
        if headers.get("price_original", 0) > 0 and headers.get("price_zlvl", 0) > 0 and headers.get("code", 0) > 0:
            return row_number, headers
    return 0, {}


def _named_catalog_header_map(normalized: dict[str, int]) -> dict[str, int]:
    result: dict[str, int] = {}
    for key, column in normalized.items():
        if "ценаединицыработ" in key and "zlvl" in key and "итого" not in key:
            result["price_zlvl"] = column
        elif "ценаединицыработ" in key and "zlvl" not in key and "итого" not in key:
            result.setdefault("price_original", column)
        elif "итогостоимость" in key and "безндс" in key:
            result.setdefault("total_price", column)

    for field, aliases in NAMED_CATALOG_HEADER_ALIASES.items():
        if field in result:
            continue
        for alias in aliases:
            column = normalized.get(_compact_header(alias))
            if column is not None:
                result[field] = column
                break
    return result


def _value_by_header(values: tuple[object, ...], headers: dict[str, int], field: str) -> object:
    column = headers.get(field, 0)
    if column <= 0 or column > len(values):
        return None
    return values[column - 1]


def _compact_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).casefold().replace("ё", "е")
    text = text.replace("\xa0", " ")
    return "".join(ch for ch in text if ch.isalnum() or ch in {"/", ".", "-", "_"})


def _clean_metadata_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, str) and value.strip() in {"", "-", "—"}:
        return ""
    return value


def _find_sheet_by_part(worksheets: list[Worksheet], part: str) -> Worksheet:
    part_key = part.casefold()
    for worksheet in worksheets:
        if part_key in worksheet.title.casefold():
            return worksheet
    return worksheets[0]


def _detect_estimate_header_row(worksheet: Worksheet, settings: Settings) -> int:
    markers = tuple(marker.casefold() for marker in (GESN_MARKER, FER_MARKER, PERECHEN_MARKER))

    for row_number in range(1, 51):
        value = _cell_value(worksheet, row_number, settings.col_search)
        value_key = "" if value is None else str(value).casefold()
        if any(marker in value_key for marker in markers):
            return row_number

    return 0


def _is_working_estimate_row(code: object, unit: object, base_price: object) -> bool:
    if NormCode(code) == "":
        return False
    if NormUnit(unit) == "":
        return False

    parsed_base_price = _parse_number(base_price)
    return parsed_base_price is not None and parsed_base_price > 0


def _parse_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Real):
        return float(value)

    text = str(value).strip()
    if text == "":
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _cell_value(worksheet: Worksheet, row: int, column: int) -> Any:
    return worksheet.cell(row=row, column=column).value

"""Parse original KL 2.0 workbooks selected directly in the admin UI.

The legacy VBA CatalogBuilder remains the reference for the business rules:
participant price groups start at column K and have a width of four columns,
only the selected winner is exported, and the winner priority is the final
recommendation, preliminary recommendation, single participant, minimum
offer total, then minimum WOR total.

This implementation intentionally improves two brittle VBA details:

* a KL sheet may be found by its structure when its title is not exactly
  ``KL 2.0`` (real files use titles such as ``KL 4``);
* note columns are not treated as participant groups unless they have a
  participant name or a valid unit-price/line-total header pair.
"""

from __future__ import annotations

import ast
import hashlib
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

from core.tkp_ingest import (
    STATUS_OK,
    TkpBlockDiagnostic,
    TkpCatalogParseResult,
    TkpItem,
    TkpSourceDiagnostics,
    TkpSourceFile,
    tkp_item_has_usable_unit_price,
)

SUPPORTED_SOURCE_SUFFIXES = frozenset({".xlsx", ".xlsm"})
PARSER_VERSION = "tkp-folder-v2"

_WOR_BLOCK = "\u0411\u041b\u041e\u041a \u0412\u041e\u0420 \u0418 \u0426\u0415\u041d\u0410"
_ADVANCE_LIMIT = (
    "\u041f\u0420\u0415\u0414\u0415\u041b\u042c\u041d\u042b\u0419 "
    "\u0420\u0410\u0417\u041c\u0415\u0420 \u0410\u0412\u0410\u041d\u0421\u0410"
)
_TOTAL_OFFER = (
    "\u0418\u0422\u041e\u0413\u041e\u0412\u0410\u042f \u0421\u0423\u041c\u041c\u0410 "
    "\u041f\u0420\u0415\u0414\u041b\u041e\u0416\u0415\u041d\u0418\u042f"
)
_NO_VAT = "\u0411\u0415\u0417 \u041d\u0414\u0421"
_WITH_VAT = "\u0421 \u041d\u0414\u0421"
_NAME = "\u041d\u0410\u0418\u041c\u0415\u041d\u041e\u0412\u0410\u041d\u0418\u0415"
_WINNER = "\u041f\u041e\u0411\u0415\u0414\u0418\u0422\u0415\u041b\u042c"
_PRICE = "\u0426\u0415\u041d"
_UNIT = "\u0415\u0414"
_TOTAL = "\u0412\u0421\u0415\u0413\u041e"
_COST = "\u0421\u0422"

_MAX_ROWS = 20_000
_MAX_COLS = 500


@dataclass(frozen=True)
class TkpSourceInput:
    """One uploaded source workbook and its user-facing path."""

    path: Path
    display_path: str
    fingerprint: str | None = None


@dataclass(frozen=True)
class _Participant:
    index: int
    start_col: int
    name: str
    name_col: int = 0


@dataclass(frozen=True)
class _WorLayout:
    schema_name: str
    code_col: int
    item_name_col: int
    unit_col: int
    qty_col: int
    rnmc_unit_price_col: int
    rnmc_line_total_col: int
    first_participant_col: int


class _WorkbookPair:
    """Formula workbook plus cached-value workbook for the same file."""

    def __init__(self, path: Path) -> None:
        keep_vba = path.suffix.casefold() == ".xlsm"
        self.formulas = load_workbook(
            path,
            data_only=False,
            read_only=False,
            keep_vba=keep_vba,
            keep_links=False,
        )
        self.values = load_workbook(
            path,
            data_only=True,
            read_only=False,
            keep_vba=keep_vba,
            keep_links=False,
        )

    def close(self) -> None:
        self.formulas.close()
        self.values.close()


def source_file_fingerprint(path: str | Path) -> str:
    """Return a deterministic content revision key for folder imports."""
    digest = hashlib.sha256()
    digest.update(PARSER_VERSION.encode("ascii"))
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_tkp_source_workbooks(
    sources: Iterable[TkpSourceInput],
) -> TkpCatalogParseResult:
    """Parse multiple original KL workbooks into the aggregate data model."""
    files: list[TkpSourceFile] = []
    items: list[TkpItem] = []
    diagnostics: list[TkpSourceDiagnostics] = []
    run_ids: list[str] = []
    for source in sources:
        try:
            result = parse_tkp_source_workbook(
                source.path,
                display_path=source.display_path,
                fingerprint=source.fingerprint,
            )
        except Exception as exc:
            fingerprint = source.fingerprint or source_file_fingerprint(source.path)
            run_id = _run_id(fingerprint)
            file_name = Path(source.display_path).name or source.path.name
            files.append(
                _empty_source(
                    run_id=run_id,
                    file_path=source.display_path,
                    file_name=file_name,
                    status="Needs review",
                    message=f"Workbook read error: {exc}",
                )
            )
            diagnostics.append(
                TkpSourceDiagnostics(
                    file_path=source.display_path,
                    file_name=file_name,
                    sheet_name="",
                    blocks=(
                        TkpBlockDiagnostic(
                            code="workbook",
                            label="Workbook",
                            found=False,
                            value=f"Workbook read error: {exc}",
                        ),
                    ),
                )
            )
            run_ids.append(run_id)
            continue
        files.extend(result.files)
        items.extend(result.items)
        diagnostics.extend(result.diagnostics)
        run_ids.extend(result.run_ids)
    return TkpCatalogParseResult(
        run_ids=tuple(dict.fromkeys(run_ids)),
        files=files,
        items=items,
        diagnostics=diagnostics,
    )


def parse_tkp_source_workbook(
    path: str | Path,
    *,
    display_path: str | None = None,
    fingerprint: str | None = None,
) -> TkpCatalogParseResult:
    """Parse one original KL workbook and export only the winner's WOR rows."""
    workbook_path = Path(path)
    if workbook_path.suffix.casefold() not in SUPPORTED_SOURCE_SUFFIXES:
        raise ValueError(f"unsupported source workbook type: {workbook_path.suffix}")

    source_path = display_path or workbook_path.name
    file_name = Path(source_path).name or workbook_path.name
    revision = fingerprint or source_file_fingerprint(workbook_path)
    run_id = _run_id(revision)
    pair = _WorkbookPair(workbook_path)
    try:
        sheet_name = _find_kl_sheet_name(pair.formulas)
        if sheet_name is None:
            source = _empty_source(
                run_id=run_id,
                file_path=source_path,
                file_name=file_name,
                status="Skipped",
                message="KL worksheet not found by name or strict KL structure.",
            )
            diagnostics = TkpSourceDiagnostics(
                file_path=source_path,
                file_name=file_name,
                sheet_name="",
                blocks=(
                    TkpBlockDiagnostic(
                        code="KL",
                        label="KL worksheet",
                        found=False,
                        value="KL worksheet not found by name or strict KL structure.",
                    ),
                ),
            )
            return TkpCatalogParseResult(
                run_ids=(run_id,),
                files=[source],
                items=[],
                diagnostics=[diagnostics],
            )

        formula_sheet = pair.formulas[sheet_name]
        value_sheet = pair.values[sheet_name]
        reader = _SheetReader(formula_sheet, value_sheet)
        start_row = _find_wor_start(reader)
        end_row, end_method = _find_wor_end(reader, start_row)
        layout = _detect_wor_layout(reader, start_row)
        participants = _find_participants(reader, start_row, end_row, layout)
        winner, winner_method = _determine_winner(
            reader,
            participants,
            start_row,
            end_row,
            layout,
        )
        reserve, reserve_method = _determine_reserve(
            reader,
            participants,
            winner,
            start_row,
            end_row,
            layout,
        )

        problems: list[str] = []
        if start_row == 0:
            problems.append("WOR start not found.")
        if start_row > 0 and end_row == 0:
            problems.append("WOR end not found.")
        if not participants:
            problems.append("No participant price groups found.")
        if winner is None:
            problems.append("Winner not detected.")

        metadata_col = (
            winner.start_col
            if winner is not None
            else participants[0].start_col
            if participants
            else 11
        )
        task_no = reader.text_at_code("1.3", metadata_col)
        if not task_no:
            problems.append("Task number not found.")
        request_date = _as_datetime(reader.value_at_code("1.1", metadata_col))
        version = reader.text_at_code("1.2", metadata_col)
        customer = reader.text_at_code("1.5", metadata_col)
        general_contractor = reader.text_at_code("1.7", metadata_col)
        procedure_name = _first_nonblank(reader.text(1, 2), reader.text(2, 2))

        winner_block_10 = _recommended_fields(reader, (("10.1", "block10"),))
        winner_block_8 = _recommended_fields(reader, (("8.1", "block8"),))
        reserve_block_10 = _recommended_fields(reader, (("10.2", "block10_reserve"),))
        reserve_block_8 = _recommended_fields(reader, (("8.2", "block8_reserve"),))
        recommended = winner_block_10 if winner_block_10["source"] else winner_block_8
        reserve_recommended = (
            reserve_block_10 if reserve_block_10["source"] else reserve_block_8
        )
        winner_name = winner.name if winner is not None else ""
        winner_inn = (
            reader.text_at_code("2.3", winner.start_col)
            if winner is not None
            else ""
        )
        winner_uin = _first_nonblank(
            recommended["uin"],
            reader.text_at_code("1.4", winner.start_col) if winner is not None else "",
        )
        total_no_vat_row = _find_total_row(reader, start_row + 1, no_vat=True)
        total_vat_row = _find_total_row(reader, start_row + 1, no_vat=False)
        winner_total_no_vat = _participant_total_no_vat(
            reader, winner, total_no_vat_row, start_row, end_row, layout
        )
        winner_total_vat = (
            reader.number(total_vat_row, winner.start_col + 1)
            if winner is not None and total_vat_row
            else None
        )
        reserve_total_no_vat = _participant_total_no_vat(
            reader, reserve, total_no_vat_row, start_row, end_row, layout
        )
        reserve_total_vat = (
            reader.number(total_vat_row, reserve.start_col + 1)
            if reserve is not None and total_vat_row
            else None
        )
        rnmc_total_no_vat = (
            reader.number(total_no_vat_row, layout.rnmc_line_total_col)
            if total_no_vat_row
            else None
        )
        reserve_name = reserve.name if reserve is not None else ""
        reserve_inn = (
            _participant_text_at_code(reader, "2.3", reserve)
            if reserve is not None
            else ""
        )
        reserve_uin = _first_nonblank(
            _text(reserve_recommended["uin"]),
            _participant_text_at_code(reader, "1.4", reserve) if reserve is not None else "",
        )

        parsed_items: list[TkpItem] = []
        if (
            winner is not None
            and start_row > 0
            and end_row > start_row
        ):
            parsed_items = _parse_wor_items(
                reader=reader,
                run_id=run_id,
                file_path=source_path,
                file_name=file_name,
                start_row=start_row,
                end_row=end_row,
                winner=winner,
                winner_method=winner_method,
                winner_inn=winner_inn,
                winner_uin=winner_uin,
                task_no=task_no,
                request_date=request_date,
                version=version,
                customer=customer,
                general_contractor=general_contractor,
                procedure_name=procedure_name,
                recommended=recommended,
                reserve=reserve,
                reserve_method=reserve_method,
                reserve_inn=reserve_inn,
                reserve_uin=reserve_uin,
                layout=layout,
            )
            if not parsed_items:
                problems.append("No WOR positions found.")

        status = STATUS_OK if not problems else "Needs review"
        message = "Parsed." if not problems else "Parsed. " + " ".join(problems)
        source = TkpSourceFile(
            run_id=run_id,
            file_path=source_path,
            file_name=file_name,
            modified_date=None,
            sheet_name=sheet_name,
            parse_status=status,
            parse_message=message,
            task_no=task_no,
            request_date=request_date,
            customer=customer,
            general_contractor=general_contractor,
            procedure_name=procedure_name,
            winner_name=winner_name,
            winner_inn=winner_inn,
            winner_uin=winner_uin,
            winner_total_no_vat=winner_total_no_vat,
            winner_total_vat=winner_total_vat,
            rnmc_total_no_vat=rnmc_total_no_vat,
            reserve_name=reserve_name,
            reserve_inn=reserve_inn,
            reserve_uin=reserve_uin,
            reserve_total_no_vat=reserve_total_no_vat,
            reserve_total_vat=reserve_total_vat,
            reserve_method=reserve_method,
        )
        usable_rows = sum(tkp_item_has_usable_unit_price(item) for item in parsed_items)
        diagnostics = TkpSourceDiagnostics(
            file_path=source_path,
            file_name=file_name,
            sheet_name=sheet_name,
            wor_start_row=start_row,
            wor_end_row=end_row,
            wor_end_method=end_method,
            wor_schema=layout.schema_name,
            participant_count=len(participants),
            rows_found=len(parsed_items),
            rows_with_usable_price=usable_rows,
            rows_rejected_missing_prices=len(parsed_items) - usable_rows,
            blocks=(
                _block_diagnostic(reader, "1.1", "Request date", reader.text_at_code("1.1", metadata_col)),
                _block_diagnostic(reader, "1.2", "Version", version),
                _block_diagnostic(reader, "1.3", "Task number", task_no),
                _block_diagnostic(reader, "1.5", "Customer", customer),
                _block_diagnostic(reader, "1.7", "General contractor", general_contractor),
                _block_diagnostic(
                    reader,
                    "2.2",
                    "Participants",
                    ", ".join(participant.name for participant in participants),
                ),
                TkpBlockDiagnostic(
                    code="4",
                    label="WOR and Price",
                    found=start_row > 0,
                    value=(
                        f"row {start_row}; end {end_row}; schema {layout.schema_name}"
                        if start_row > 0
                        else ""
                    ),
                    row=start_row,
                ),
                _block_diagnostic(reader, "8.1", "Preliminary winner", _text(winner_block_8["name"])),
                _block_diagnostic(reader, "8.2", "Preliminary reserve", _text(reserve_block_8["name"])),
                _block_diagnostic(reader, "10.1", "Recommended winner", _text(winner_block_10["name"])),
                _block_diagnostic(reader, "10.2", "Reserve winner", _text(reserve_block_10["name"])),
            ),
        )
        return TkpCatalogParseResult(
            run_ids=(run_id,),
            files=[source],
            items=parsed_items,
            diagnostics=[diagnostics],
        )
    finally:
        pair.close()


def _block_diagnostic(
    reader: "_SheetReader",
    code: str,
    label: str,
    value: str,
) -> TkpBlockDiagnostic:
    row = reader.find_code_row(code)
    return TkpBlockDiagnostic(
        code=code,
        label=label,
        found=row > 0,
        value=value,
        row=row,
    )


class _SheetReader:
    def __init__(self, formulas: Worksheet, values: Worksheet) -> None:
        self.formulas = formulas
        self.values = values
        self.max_row = min(
            max(formulas.max_row or 1, values.max_row or 1),
            _MAX_ROWS,
        )
        self.max_col = min(
            max(formulas.max_column or 1, values.max_column or 1),
            _MAX_COLS,
        )
        self._evaluating: set[tuple[int, int]] = set()

    def raw(self, row: int, col: int) -> object:
        if row <= 0 or col <= 0:
            return None
        return self.formulas.cell(row, col).value

    def value(self, row: int, col: int) -> object:
        if row <= 0 or col <= 0:
            return None
        cached = self.values.cell(row, col).value
        if cached is not None:
            return cached
        raw = self.raw(row, col)
        if not (isinstance(raw, str) and raw.startswith("=")):
            return raw
        key = (row, col)
        if key in self._evaluating:
            return None
        self._evaluating.add(key)
        try:
            return self._evaluate_formula(raw)
        finally:
            self._evaluating.remove(key)

    def text(self, row: int, col: int) -> str:
        return _text(self.value(row, col))

    def number(self, row: int, col: int) -> float | None:
        return _as_float(self.value(row, col))

    def row_text(self, row: int, start_col: int = 1, end_col: int | None = None) -> str:
        last_col = self.max_col if end_col is None else min(end_col, self.max_col)
        return " ".join(
            value
            for value in (
                self.text(row, col)
                for col in range(max(1, start_col), last_col + 1)
            )
            if value
        )

    def find_code_row(self, code: str) -> int:
        target = _code_norm(code)
        for row in range(1, self.max_row + 1):
            if _code_norm(self.text(row, 1)) == target:
                return row
        return 0

    def value_at_code(self, code: str, preferred_col: int, fallback_col: int = 11) -> object:
        row = self.find_code_row(code)
        if row == 0:
            return None
        preferred = self.value(row, preferred_col)
        if _text(preferred):
            return preferred
        for col in range(max(1, fallback_col), self.max_col + 1):
            value = self.value(row, col)
            if _text(value):
                return value
        return None

    def text_at_code(self, code: str, preferred_col: int, fallback_col: int = 11) -> str:
        return _text(self.value_at_code(code, preferred_col, fallback_col))

    def _evaluate_formula(self, formula: str) -> object:
        expression = formula[1:].strip()
        direct = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", expression, re.IGNORECASE)
        if direct:
            return self.value(int(direct.group(2)), _column_number(direct.group(1)))

        concatenate = re.fullmatch(r"CONCATENATE\((.*)\)", expression, re.IGNORECASE)
        if concatenate:
            values = []
            for part in _split_formula_arguments(concatenate.group(1)):
                part = part.strip()
                if len(part) >= 2 and part[0] == part[-1] == '"':
                    values.append(part[1:-1])
                    continue
                reference = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", part, re.IGNORECASE)
                if reference:
                    values.append(
                        _text(
                            self.value(
                                int(reference.group(2)),
                                _column_number(reference.group(1)),
                            )
                        )
                    )
                    continue
                return None
            return "".join(values)

        expression = re.sub(
            r"SUM\(\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)\)",
            lambda match: str(self._sum_range(match.group(0))),
            expression,
            flags=re.IGNORECASE,
        )

        unresolved = False

        def replace_reference(match: re.Match[str]) -> str:
            nonlocal unresolved
            value = self.number(
                int(match.group(2)),
                _column_number(match.group(1)),
            )
            if value is None:
                unresolved = True
                return "0"
            return repr(value)

        expression = re.sub(
            r"\$?([A-Z]{1,3})\$?(\d+)",
            replace_reference,
            expression,
            flags=re.IGNORECASE,
        )
        if unresolved or re.search(r"[A-Za-z_]", expression):
            return None
        return _safe_arithmetic(expression)

    def _sum_range(self, range_formula: str) -> float:
        address = range_formula[4:-1].replace("$", "")
        min_col, min_row, max_col, max_row = range_boundaries(address)
        total = 0.0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = self.number(row, col)
                if value is not None:
                    total += value
        return total


def _find_kl_sheet_name(workbook) -> str | None:
    """Return a KL worksheet without mistaking RNMC/estimate sheets for KL.

    A sheet whose title explicitly looks like a KL sheet remains eligible so
    incomplete KL workbooks can still reach the Needs review stage. A sheet
    with an unrelated title is accepted only when several independent KL
    structures agree: participant metadata, WOR block, participant price
    headers, offer total, and a recommendation block.
    """
    named: list[str] = []
    structural: list[tuple[int, str]] = []
    for worksheet in workbook.worksheets:
        normalized = _norm_sheet_name(worksheet.title)
        if re.fullmatch(r"\u041a\u041b\d+", normalized):
            named.append(worksheet.title)
            continue

        evidence = _kl_structural_evidence(worksheet)
        if evidence["strict_match"]:
            structural.append((int(evidence["score"]), worksheet.title))

    if named:
        return named[0]
    if not structural:
        return None
    structural.sort(key=lambda item: (-item[0], workbook.sheetnames.index(item[1])))
    return structural[0][1]


def _kl_structural_evidence(worksheet: Worksheet) -> dict[str, object]:
    """Collect independent KL markers for a nonstandard worksheet title."""
    max_row = min(worksheet.max_row or 1, 500)
    max_col = min(worksheet.max_column or 1, 120)
    participant_row = 0
    wor_row = 0
    total_offer_row = 0
    recommendation_row = 0
    participant_identity_row = 0

    for row in range(1, max_row + 1):
        code = _code_norm(worksheet.cell(row, 1).value)
        label = _norm(worksheet.cell(row, 2).value)
        row_text = " ".join(
            _text(worksheet.cell(row, col).value)
            for col in range(1, max_col + 1)
            if _text(worksheet.cell(row, col).value)
        )
        normalized_row = _norm(row_text)

        if code == "2.2" and _NAME in label:
            participant_row = participant_row or row
        if code == "2.3":
            participant_identity_row = participant_identity_row or row
        if _WOR_BLOCK in normalized_row:
            wor_row = wor_row or row
        if _TOTAL_OFFER in normalized_row:
            total_offer_row = total_offer_row or row
        if code in {"8.1", "10.1"}:
            recommendation_row = recommendation_row or row

    price_pair = False
    if wor_row:
        for col in range(1, max_col):
            if (
                _looks_like_price_header(worksheet.cell(wor_row, col).value)
                and _looks_like_total_header(worksheet.cell(wor_row, col + 1).value)
            ):
                price_pair = True
                break

    score = (
        (4 if participant_row else 0)
        + (2 if participant_identity_row else 0)
        + (8 if wor_row else 0)
        + (4 if price_pair else 0)
        + (4 if total_offer_row else 0)
        + (6 if recommendation_row else 0)
    )
    strict_match = bool(
        participant_row
        and wor_row
        and price_pair
        and total_offer_row
        and recommendation_row
    )
    return {
        "strict_match": strict_match,
        "score": score,
        "participant_row": participant_row,
        "participant_identity_row": participant_identity_row,
        "wor_row": wor_row,
        "price_pair": price_pair,
        "total_offer_row": total_offer_row,
        "recommendation_row": recommendation_row,
    }


def _find_wor_start(reader: _SheetReader) -> int:
    for row in range(1, reader.max_row + 1):
        if _WOR_BLOCK in _norm(reader.row_text(row)):
            return row
    return 0


def _find_wor_end(reader: _SheetReader, start_row: int) -> tuple[int, str]:
    if start_row <= 0:
        return 0, ""
    for row in range(start_row + 1, reader.max_row + 1):
        if _ADVANCE_LIMIT in _norm(reader.row_text(row)):
            return row, "advance_limit"
    row = _find_total_row(reader, start_row + 1, no_vat=True)
    if row:
        return row, "total_no_vat_fallback"
    row = _find_total_row(reader, start_row + 1, no_vat=False)
    return (row, "total_vat_fallback") if row else (0, "")


def _find_total_row(reader: _SheetReader, from_row: int, *, no_vat: bool) -> int:
    for row in range(max(1, from_row), reader.max_row + 1):
        text = _norm(reader.row_text(row))
        if _TOTAL_OFFER not in text:
            continue
        if no_vat and _NO_VAT in text:
            return row
        if not no_vat and _WITH_VAT in text and _NO_VAT not in text:
            return row
    return 0


def _detect_wor_layout(reader: _SheetReader, start_row: int) -> _WorLayout:
    code_col = 1
    item_name_col = 2
    unit_col = 3
    qty_col = 4
    if (
        start_row > 0
        and _looks_like_price_header(reader.raw(start_row, 9))
        and _looks_like_unit_header(reader.raw(start_row, 10))
        and _looks_like_qty_header(reader.raw(start_row, 11))
        and _looks_like_total_header(reader.raw(start_row, 12))
    ):
        return _WorLayout(
            schema_name="rnmc_lot_i_l",
            code_col=code_col,
            item_name_col=item_name_col,
            unit_col=unit_col,
            qty_col=qty_col,
            rnmc_unit_price_col=9,
            rnmc_line_total_col=12,
            first_participant_col=13,
        )
    if (
        start_row > 0
        and _looks_like_price_header(reader.raw(start_row, 9))
        and _looks_like_total_header(reader.raw(start_row, 10))
    ):
        return _WorLayout(
            schema_name="rnmc_i_j",
            code_col=code_col,
            item_name_col=item_name_col,
            unit_col=unit_col,
            qty_col=qty_col,
            rnmc_unit_price_col=9,
            rnmc_line_total_col=10,
            first_participant_col=11,
        )
    if (
        start_row > 0
        and _looks_like_price_header(reader.raw(start_row, 11))
        and _looks_like_total_header(reader.raw(start_row, 12))
        and _has_numeric_pair(reader, start_row, 9, 10)
    ):
        return _WorLayout(
            schema_name="rnmc_i_j_inferred",
            code_col=code_col,
            item_name_col=item_name_col,
            unit_col=unit_col,
            qty_col=qty_col,
            rnmc_unit_price_col=9,
            rnmc_line_total_col=10,
            first_participant_col=11,
        )
    for col in range(5, min(reader.max_col, 80)):
        if (
            _looks_like_price_header(reader.raw(start_row, col))
            and _looks_like_total_header(reader.raw(start_row, col + 1))
        ):
            return _WorLayout(
                schema_name="detected_adjacent",
                code_col=code_col,
                item_name_col=item_name_col,
                unit_col=unit_col,
                qty_col=qty_col,
                rnmc_unit_price_col=col,
                rnmc_line_total_col=col + 1,
                first_participant_col=col + 2,
            )
    return _WorLayout(
        schema_name="legacy_default",
        code_col=code_col,
        item_name_col=item_name_col,
        unit_col=unit_col,
        qty_col=qty_col,
        rnmc_unit_price_col=9,
        rnmc_line_total_col=10,
        first_participant_col=11,
    )


def _has_numeric_pair(
    reader: _SheetReader,
    start_row: int,
    unit_col: int,
    total_col: int,
) -> bool:
    for row in range(start_row + 1, min(reader.max_row, start_row + 80) + 1):
        if reader.number(row, unit_col) is not None and reader.number(row, total_col) is not None:
            return True
    return False


def _find_participants(
    reader: _SheetReader,
    start_row: int,
    end_row: int,
    layout: _WorLayout,
) -> list[_Participant]:
    if start_row <= 0:
        return []
    name_row = reader.find_code_row("2.2")
    participants: list[_Participant] = []
    last_name = ""
    last_name_col = 0
    for start_col in range(layout.first_participant_col, reader.max_col):
        if not (
            _looks_like_price_header(reader.raw(start_row, start_col))
            and _looks_like_total_header(reader.raw(start_row, start_col + 1))
        ):
            continue
        if start_col in {layout.rnmc_unit_price_col, layout.rnmc_line_total_col}:
            continue
        has_price = (
            end_row > start_row
            and any(
                reader.number(row, start_col) is not None
                or reader.number(row, start_col + 1) is not None
                for row in range(start_row + 1, end_row)
                if _row_type(reader, row, layout) == "Position"
            )
        )
        direct_name = reader.text(name_row, start_col) if name_row else ""
        if direct_name:
            last_name = direct_name
            last_name_col = start_col
        name = direct_name or last_name
        if name or has_price:
            participants.append(
                _Participant(
                    index=len(participants) + 1,
                    start_col=start_col,
                    name=name or f"ParticipantGroup_{len(participants) + 1}",
                    name_col=start_col if direct_name else last_name_col,
                )
            )
    return participants


def _determine_winner(
    reader: _SheetReader,
    participants: list[_Participant],
    start_row: int,
    end_row: int,
    layout: _WorLayout,
) -> tuple[_Participant | None, str]:
    for code, method in (("10.1", "block10_recommended"), ("8.1", "block8_recommended")):
        row = reader.find_code_row(code)
        winner = _winner_from_recommended_row(reader, row, participants)
        if winner is not None:
            return winner, method
    if len(participants) == 1:
        return participants[0], "single_participant"

    total_row = _find_total_row(reader, start_row + 1, no_vat=True)
    winner = _minimum_participant(
        participants,
        lambda participant: reader.number(total_row, participant.start_col + 1)
        if total_row
        else None,
    )
    if winner is not None:
        return winner, "min_total_row"

    winner = _minimum_participant(
        participants,
        lambda participant: _sum_wor_totals(
            reader,
            start_row,
            end_row,
            participant.start_col,
            layout,
        ),
    )
    return (winner, "min_wor_sum") if winner is not None else (None, "")


def _determine_reserve(
    reader: _SheetReader,
    participants: list[_Participant],
    winner: _Participant | None,
    start_row: int,
    end_row: int,
    layout: _WorLayout,
) -> tuple[_Participant | None, str]:
    for code, method in (("10.2", "block10_reserve"), ("8.2", "block8_reserve")):
        row = reader.find_code_row(code)
        reserve = _winner_from_recommended_row(reader, row, participants)
        if reserve is not None and reserve != winner:
            return reserve, method
    candidates = [participant for participant in participants if participant != winner]
    if not candidates:
        return None, ""

    total_row = _find_total_row(reader, start_row + 1, no_vat=True)
    reserve = _minimum_participant(
        candidates,
        lambda participant: reader.number(total_row, participant.start_col + 1)
        if total_row
        else None,
    )
    if reserve is not None:
        return reserve, "inferred_second_min_total"

    reserve = _minimum_participant(
        candidates,
        lambda participant: _sum_wor_totals(
            reader,
            start_row,
            end_row,
            participant.start_col,
            layout,
        ),
    )
    return (reserve, "inferred_second_min_wor_sum") if reserve is not None else (None, "")


def _winner_from_recommended_row(
    reader: _SheetReader,
    row: int,
    participants: list[_Participant],
) -> _Participant | None:
    if row <= 0:
        return None
    recommended_name = reader.text(row, 5)
    row_text = _norm(reader.row_text(row))
    for participant in participants:
        group_text = _norm(
            reader.row_text(
                row,
                participant.start_col,
                participant.start_col + 3,
            )
        )
        if _is_winner_marker(group_text):
            return participant

    for participant in participants:
        name = _norm(participant.name)
        if name and (
            _contains_either(_norm(recommended_name), name)
            or _contains_either(row_text, name)
        ):
            return participant

    populated = [
        participant
        for participant in participants
        if reader.row_text(
            row,
            participant.start_col,
            participant.start_col + 3,
        ).strip()
    ]
    return populated[0] if len(populated) == 1 else None


def _minimum_participant(participants, value_getter):
    best: _Participant | None = None
    best_value: float | None = None
    for participant in participants:
        value = value_getter(participant)
        if value is None or value <= 0:
            continue
        if best_value is None or value < best_value:
            best = participant
            best_value = value
    return best


def _sum_wor_totals(
    reader: _SheetReader,
    start_row: int,
    end_row: int,
    start_col: int,
    layout: _WorLayout,
) -> float | None:
    if start_row <= 0 or end_row <= start_row:
        return None
    total = 0.0
    found = False
    for row in range(start_row + 1, end_row):
        if _row_type(reader, row, layout) != "Position":
            continue
        value = reader.number(row, start_col + 1)
        if value is not None:
            total += value
            found = True
    return total if found else None


def _recommended_fields(
    reader: _SheetReader,
    code_sources: tuple[tuple[str, str], ...],
) -> dict[str, object]:
    for code, source in code_sources:
        row = reader.find_code_row(code)
        if row <= 0:
            continue
        values = [reader.value(row, col) for col in range(5, 9)]
        if any(_text(value) for value in values):
            return {
                "name": _text(values[0]),
                "uin": _text(values[1]),
                "total_vat": _as_float(values[2]),
                "reason": _text(values[3]),
                "source": source,
            }
    return {
        "name": "",
        "uin": "",
        "total_vat": None,
        "reason": "",
        "source": "",
    }


def _parse_wor_items(
    *,
    reader: _SheetReader,
    run_id: str,
    file_path: str,
    file_name: str,
    start_row: int,
    end_row: int,
    winner: _Participant,
    winner_method: str,
    winner_inn: str,
    winner_uin: str,
    task_no: str,
    request_date: datetime | None,
    version: str,
    customer: str,
    general_contractor: str,
    procedure_name: str,
    recommended: dict[str, object],
    reserve: _Participant | None,
    reserve_method: str,
    reserve_inn: str,
    reserve_uin: str,
    layout: _WorLayout,
) -> list[TkpItem]:
    section_code = ""
    section_name = ""
    subsection_name = ""
    items: list[TkpItem] = []
    for row in range(start_row + 1, end_row):
        row_type = _row_type(reader, row, layout)
        if row_type == "Section":
            code = reader.text(row, 1)
            if code:
                section_code = code
                section_name = reader.text(row, 2)
                subsection_name = ""
            else:
                subsection_name = reader.text(row, 2)
            continue
        if row_type != "Position":
            continue
        qty_raw = reader.value(row, layout.qty_col)
        quality_flags = _quality_flags(reader, row, layout, winner, reserve)
        items.append(
            TkpItem(
                run_id=run_id,
                file_path=file_path,
                file_name=file_name,
                sheet_name=reader.formulas.title,
                source_row=row,
                section_code=section_code,
                section_name=section_name,
                subsection_name=subsection_name,
                item_code=reader.text(row, layout.code_col),
                item_name=reader.text(row, layout.item_name_col),
                unit=reader.text(row, layout.unit_col),
                qty=_as_float(qty_raw),
                qty_source_text=_text(qty_raw),
                rnmc_unit_price_no_vat=reader.number(row, layout.rnmc_unit_price_col),
                rnmc_line_total_no_vat=reader.number(row, layout.rnmc_line_total_col),
                winner_unit_price_no_vat=reader.number(row, winner.start_col),
                winner_line_total_no_vat=reader.number(row, winner.start_col + 1),
                winner_name=winner.name,
                winner_inn=winner_inn,
                winner_uin=winner_uin,
                winner_group_index=winner.index,
                winner_start_col=winner.start_col,
                winner_start_col_letter=get_column_letter(winner.start_col),
                winner_unit_header=reader.text(start_row, winner.start_col),
                winner_total_header=reader.text(start_row, winner.start_col + 1),
                task_no=task_no,
                request_date=request_date,
                version=version,
                customer=customer,
                general_contractor=general_contractor,
                procedure_name=procedure_name,
                winner_method=winner_method,
                winner_block_name=_text(recommended["name"]),
                winner_block_uin=_text(recommended["uin"]),
                winner_block_total_vat=_as_float(recommended["total_vat"]),
                winner_block_reason=_text(recommended["reason"]),
                reserve_unit_price_no_vat=(
                    reader.number(row, reserve.start_col) if reserve is not None else None
                ),
                reserve_line_total_no_vat=(
                    reader.number(row, reserve.start_col + 1) if reserve is not None else None
                ),
                reserve_name=reserve.name if reserve is not None else "",
                reserve_inn=reserve_inn,
                reserve_uin=reserve_uin,
                reserve_group_index=reserve.index if reserve is not None else 0,
                reserve_start_col=reserve.start_col if reserve is not None else 0,
                reserve_start_col_letter=(
                    get_column_letter(reserve.start_col) if reserve is not None else ""
                ),
                reserve_unit_header=(
                    reader.text(start_row, reserve.start_col) if reserve is not None else ""
                ),
                reserve_total_header=(
                    reader.text(start_row, reserve.start_col + 1) if reserve is not None else ""
                ),
                reserve_method=reserve_method,
                wor_schema=layout.schema_name,
                quality_flags=quality_flags,
            )
        )
    return items


def _row_type(reader: _SheetReader, row: int, layout: _WorLayout) -> str:
    item_code = _code_norm(reader.text(row, layout.code_col))
    item_name = reader.text(row, layout.item_name_col)
    unit = reader.text(row, layout.unit_col)
    qty = reader.text(row, layout.qty_col)
    rnmc_price = reader.number(row, layout.rnmc_unit_price_col)
    rnmc_total = reader.number(row, layout.rnmc_line_total_col)
    if not item_name and not unit and not qty and rnmc_price is None and rnmc_total is None:
        return "Blank"
    if item_code.endswith("."):
        return "Section"
    if item_name and (unit or qty or rnmc_price is not None or rnmc_total is not None):
        return "Position"
    return "Section"


def _quality_flags(
    reader: _SheetReader,
    row: int,
    layout: _WorLayout,
    winner: _Participant,
    reserve: _Participant | None,
) -> str:
    flags: list[str] = []
    if not reader.text(row, layout.unit_col):
        flags.append("missing_unit")
    if reader.number(row, layout.qty_col) is None:
        flags.append("missing_qty")
    if reader.number(row, layout.rnmc_unit_price_col) is None:
        flags.append("missing_rnmc_unit_price")
    if reader.number(row, winner.start_col) is None:
        flags.append("missing_winner_unit_price")
    if reserve is not None and reader.number(row, reserve.start_col) is None:
        flags.append("missing_reserve_unit_price")
    return ";".join(flags)


def _participant_total_no_vat(
    reader: _SheetReader,
    participant: _Participant | None,
    total_row: int,
    start_row: int,
    end_row: int,
    layout: _WorLayout,
) -> float | None:
    if participant is None:
        return None
    value = reader.number(total_row, participant.start_col + 1) if total_row else None
    if value is not None:
        return value
    return _sum_wor_totals(reader, start_row, end_row, participant.start_col, layout)


def _participant_text_at_code(
    reader: _SheetReader,
    code: str,
    participant: _Participant | None,
) -> str:
    if participant is None:
        return ""
    row = reader.find_code_row(code)
    if row <= 0:
        return ""
    for col in (participant.start_col, participant.name_col):
        value = reader.text(row, col) if col else ""
        if value:
            return value
    return reader.text_at_code(code, participant.start_col)


def _empty_source(
    *,
    run_id: str,
    file_path: str,
    file_name: str,
    status: str,
    message: str,
) -> TkpSourceFile:
    return TkpSourceFile(
        run_id=run_id,
        file_path=file_path,
        file_name=file_name,
        modified_date=None,
        sheet_name="",
        parse_status=status,
        parse_message=message,
        task_no="",
        request_date=None,
        customer="",
        general_contractor="",
        procedure_name="",
        winner_name="",
        winner_inn="",
        winner_uin="",
        winner_total_no_vat=None,
        winner_total_vat=None,
        rnmc_total_no_vat=None,
    )


def _run_id(fingerprint: str) -> str:
    return f"{PARSER_VERSION}:{fingerprint}"


def _norm(value: object) -> str:
    text = _text(value).upper().replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _norm_sheet_name(value: object) -> str:
    return re.sub(r"[\s.,\-_]+", "", _norm(value))


def _code_norm(value: object) -> str:
    return _text(value).replace(",", ".").replace(" ", "").rstrip(".")


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_float(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace("\xa0", "").replace(" ", "")
    if not text:
        return None
    filtered = "".join(
        char for char in text if char.isdigit() or char in ",.-"
    )
    if not any(char.isdigit() for char in filtered):
        return None
    comma = filtered.rfind(",")
    dot = filtered.rfind(".")
    if comma >= 0 and dot >= 0:
        if comma > dot:
            filtered = filtered.replace(".", "")
            filtered = filtered[: filtered.rfind(",")] + "." + filtered[filtered.rfind(",") + 1 :]
        else:
            filtered = filtered.replace(",", "")
    elif comma >= 0:
        filtered = filtered[:comma] + "." + filtered[comma + 1 :]
    try:
        return float(filtered)
    except ValueError:
        return None


def _as_datetime(value: object) -> datetime | None:
    return value if isinstance(value, datetime) else None


def _looks_like_price_header(value: object) -> bool:
    text = _norm(value)
    return _PRICE in text and (_UNIT in text or "\u0417\u0410 1" in text)


def _looks_like_total_header(value: object) -> bool:
    text = _norm(value)
    return _TOTAL in text or _COST in text


def _looks_like_unit_header(value: object) -> bool:
    text = _norm(value)
    return _UNIT in text and _PRICE not in text


def _looks_like_qty_header(value: object) -> bool:
    text = _norm(value)
    return "КОЛ" in text and "ВО" in text


def _contains_either(left: str, right: str) -> bool:
    return bool(left and right and (left in right or right in left))


def _is_winner_marker(value: str) -> bool:
    text = _norm(value)
    return text in {"1", "X", "YES", "DA", "\u0414\u0410"} or _WINNER in text


def _first_nonblank(first: str, second: str) -> str:
    return first if first.strip() else second


def _column_number(letters: str) -> int:
    value = 0
    for char in letters.upper():
        value = value * 26 + ord(char) - ord("A") + 1
    return value


def _split_formula_arguments(value: str) -> list[str]:
    result: list[str] = []
    current = []
    quoted = False
    for char in value:
        if char == '"':
            quoted = not quoted
        if char in ",;" and not quoted:
            result.append("".join(current))
            current = []
        else:
            current.append(char)
    result.append("".join(current))
    return result


def _safe_arithmetic(expression: str) -> float | None:
    try:
        tree = ast.parse(expression, mode="eval")
    except SyntaxError:
        return None

    def evaluate(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return evaluate(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
            value = evaluate(node.operand)
            return value if isinstance(node.op, ast.UAdd) else -value
        if isinstance(node, ast.BinOp) and isinstance(
            node.op,
            (ast.Add, ast.Sub, ast.Mult, ast.Div),
        ):
            left = evaluate(node.left)
            right = evaluate(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            return left / right
        raise ValueError("unsupported formula expression")

    try:
        return evaluate(tree)
    except (ValueError, ZeroDivisionError):
        return None

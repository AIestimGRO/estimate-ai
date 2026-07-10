"""RNMC workbook preview helpers.

This module mirrors the legacy VBA import detection rules, but it does not
write catalog rows. It is used to preview what an RNMC zip import would find.
"""

from __future__ import annotations

import re
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from numbers import Real
from pathlib import PurePosixPath
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.rnmc_zip import EXCEL_SUFFIXES
from core.catalog import CatalogRow
from core.normalize import NormCode, NormUnit
from core.storage.catalog import (
    ROW_LOG_STATUS_REJECTED,
    STATUS_DUPLICATE_NAME as DB_STATUS_DUPLICATE_NAME,
    STATUS_FAILED,
    STATUS_NO_DATA,
    STATUS_SKIPPED,
    STATUS_SUCCESS,
    CatalogRowStorageItem,
    final_filename_keys_for_preview,
    imported_file_exists_for_region,
    normalize_import_filename,
    record_imported_file,
    replace_catalog_rows_for_file,
    replace_import_row_logs,
)

HEADER_NAME_WORKS = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
HEADER_NAME_SHORT = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
HEADER_UNIT = "\u0415\u0434.\u0438\u0437\u043c."
HEADER_UNIT_NO_DOT = "\u0415\u0434.\u0438\u0437\u043c"
HEADER_UNIT_LONG = "\u0415\u0434\u0438\u043d\u0438\u0446\u0430 \u0438\u0437\u043c\u0435\u0440\u0435\u043d\u0438\u044f"
HEADER_QTY_SHORT = "\u041a\u043e\u043b-\u0432\u043e"
HEADER_QTY_LONG = "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e"
TASK_LABEL_FULL = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438 1\u0424"
TASK_LABEL_SHORT = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438"

CODE_HEADER_PATTERNS = (
    "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0422\u0415\u0420/\u041a\u0420",
    "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0413\u042d\u0421\u041d/\u041a\u0420",
    "\u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c",
    "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d",
    "\u041e\u0431\u043e\u0441\u043d\u043e\u0432\u0430\u043d\u0438\u0435",
    "\u0428\u0438\u0444\u0440",
    "\u041a\u043e\u0434",
)
PRICE_HEADER_PATTERNS = (
    "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442",
    "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b",
    "\u0426\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434",
    "\u0421\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c \u0435\u0434\u0438\u043d\u0438\u0446\u044b",
    "\u0421\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c \u0437\u0430 \u0435\u0434",
    "\u0411\u0430\u0437\u043e\u0432\u0430\u044f \u0446\u0435\u043d\u0430",
    "\u0426\u0435\u043d\u0430 \u0437\u0430 1",
    "\u0426\u0435\u043d\u0430",
)
ADDED_DATE_HEADER_PATTERNS = (
    "CatalogAddedDate",
    "\u0414\u0430\u0442\u0430 \u0434\u043e\u0431\u0430\u0432\u043b\u0435\u043d\u0438\u044f",
    "\u0414\u0430\u0442\u0430 \u0434\u043e\u0431",
)

LSR_QUARTER_LABEL_PATTERNS = (
    "\u0413\u043e\u0434 \u041a\u0432\u0430\u0440\u0442\u0430\u043b \u041b\u0421\u0420",
    "\u0413\u043e\u0434/\u043a\u0432\u0430\u0440\u0442\u0430\u043b \u041b\u0421\u0420",
    "\u0413\u043e\u0434 \u0438 \u043a\u0432\u0430\u0440\u0442\u0430\u043b \u041b\u0421\u0420",
    "\u041a\u0432\u0430\u0440\u0442\u0430\u043b \u041b\u0421\u0420",
    "\u041f\u0440\u0435\u0434\u043e\u0441\u0442\u0430\u0432\u043b\u0435\u043d\u043d\u044b\u0445 \u0441\u043c\u0435\u0442 \u0432 \u0446\u0435\u043d\u0430\u0445",
)
PLANNED_START_LABEL_PATTERNS = (
    "\u041f\u043b\u0430\u043d\u0438\u0440\u0443\u043c\u044b\u0439 \u0441\u0440\u043e\u043a \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442",
    "\u041f\u043b\u0430\u043d\u0438\u0440\u0443\u0435\u043c\u044b\u0439 \u0441\u0440\u043e\u043a \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442",
    "\u041f\u043b\u0430\u043d\u043e\u0432\u044b\u0439 \u0441\u0440\u043e\u043a \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442",
    "\u0414\u0430\u0442\u0430 \u043d\u0430\u0447\u0430\u043b\u0430 \u0440\u0430\u0431\u043e\u0442",
    "\u043d\u0430\u0447\u0430\u043b\u043e \u0440\u0430\u0431\u043e\u0442",
)
PLANNED_FINISH_LABEL_PATTERNS = (
    "\u041f\u043b\u0430\u043d\u0438\u0440\u0443\u0435\u043c\u044b\u0439 \u0441\u0440\u043e\u043a \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442",
    "\u041f\u043b\u0430\u043d\u0438\u0440\u0443\u043c\u044b\u0439 \u0441\u0440\u043e\u043a \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442",
    "\u041f\u043b\u0430\u043d\u043e\u0432\u044b\u0439 \u0441\u0440\u043e\u043a \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442",
    "\u0414\u0430\u0442\u0430 \u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u044f \u0440\u0430\u0431\u043e\u0442",
    "\u043e\u043a\u043e\u043d\u0447\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442",
)
REGION_LABEL_PATTERNS = (
    "\u0420\u0435\u0433\u0438\u043e\u043d \u0440\u0430\u0441\u043f\u043e\u043b\u043e\u0436\u0435\u043d\u0438\u044f \u043e\u0431\u044a\u0435\u043a\u0442\u0430",
    "\u0420\u0435\u0433\u0438\u043e\u043d \u043e\u0431\u044a\u0435\u043a\u0442\u0430",
    "\u0420\u0435\u0433\u0438\u043e\u043d",
)
REGIONAL_COEFFICIENT_LABEL_PATTERNS = (
    "\u0420\u0435\u0433\u0438\u043e\u043d\u0430\u043b\u044c\u043d\u044b\u0439 \u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442",
    "\u041a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442",
)

STATUS_PREVIEW_OK = "preview_ok"
STATUS_NO_TABLE = "no_table"
STATUS_NO_ROWS = "no_rows"
STATUS_SKIPPED_PROCESSED = "skipped_processed"
STATUS_DUPLICATE_NAME = "duplicate_name"
STATUS_UNSUPPORTED_FORMAT = "unsupported_format"
STATUS_PARSE_ERROR = "parse_error"

SUPPORTED_PREVIEW_SUFFIXES = frozenset({".xlsx", ".xlsm"})
DEFAULT_ROW_PREVIEW_LIMIT = 30
EXCEL_DATE_BASE = date(1899, 12, 30)
VALUE_WITH_VAT_DIVISOR = 1.2


@dataclass(frozen=True)
class RnmcWorkbookMetadata:
    lsr_quarter: str = ""
    planned_start: str = ""
    planned_finish: str = ""
    region_folder: str = ""
    regional_coefficient: float | None = None


@dataclass(frozen=True)
class RnmcHeaderPreview:
    code: str = ""
    work_name: str = ""
    unit: str = ""
    quantity: str = ""
    unit_price: str = ""
    total_price: str = ""
    labor_unit: str = ""
    labor_total: str = ""
    machine_labor_unit: str = ""
    machine_labor_total: str = ""


@dataclass(frozen=True)
class RnmcRowSample:
    row_number: int
    code: str
    work_name: str
    unit: str
    quantity: str
    unit_price: str
    total_price: str
    labor_unit: str
    labor_total: str
    machine_labor_unit: str
    machine_labor_total: str
    issue: str = ""


@dataclass(frozen=True)
class RnmcWorkbookPreview:
    archive_path: str
    filename: str
    region_folder: str
    status: str
    reason: str
    sheet_name: str
    header_row: int
    task_number: str
    lsr_quarter: str
    planned_start: str
    planned_finish: str
    regional_coefficient: float | None
    rows_ok: int
    rows_rejected: int
    header_preview: RnmcHeaderPreview
    sample_rows: list[RnmcRowSample]
    is_limited: bool = False


@dataclass(frozen=True)
class RnmcZipRowPreviewResult:
    entries: list[RnmcWorkbookPreview]
    ignored_files: int

    @property
    def total_excel_files(self) -> int:
        return len(self.entries)

    @property
    def preview_ok_count(self) -> int:
        return _count_status(self.entries, STATUS_PREVIEW_OK)

    @property
    def skipped_processed_count(self) -> int:
        return _count_status(self.entries, STATUS_SKIPPED_PROCESSED)

    @property
    def duplicate_name_count(self) -> int:
        return _count_status(self.entries, STATUS_DUPLICATE_NAME)

    @property
    def no_table_count(self) -> int:
        return _count_status(self.entries, STATUS_NO_TABLE)

    @property
    def no_rows_count(self) -> int:
        return _count_status(self.entries, STATUS_NO_ROWS)

    @property
    def unsupported_format_count(self) -> int:
        return _count_status(self.entries, STATUS_UNSUPPORTED_FORMAT)

    @property
    def parse_error_count(self) -> int:
        return _count_status(self.entries, STATUS_PARSE_ERROR)

    @property
    def rows_ok_total(self) -> int:
        return sum(entry.rows_ok for entry in self.entries)

    @property
    def rows_rejected_total(self) -> int:
        return sum(entry.rows_rejected for entry in self.entries)

    @property
    def limited_count(self) -> int:
        return sum(1 for entry in self.entries if entry.is_limited)


@dataclass(frozen=True)
class RnmcRejectedRow:
    row_number: int
    reason: str


@dataclass(frozen=True)
class RnmcCatalogRowCandidate:
    row_number: int
    catalog_row: CatalogRow


@dataclass(frozen=True)
class RnmcZipCatalogImportEntry:
    archive_path: str
    filename: str
    region_folder: str
    status: str
    reason: str
    sheet_name: str
    header_row: int
    task_number: str
    lsr_quarter: str
    planned_start: str
    planned_finish: str
    regional_coefficient: float | None
    rows_imported: int
    rows_rejected: int


@dataclass(frozen=True)
class RnmcZipCatalogImportResult:
    entries: list[RnmcZipCatalogImportEntry]
    ignored_files: int

    @property
    def total_excel_files(self) -> int:
        return len(self.entries)

    @property
    def success_count(self) -> int:
        return _count_import_status(self.entries, STATUS_SUCCESS)

    @property
    def no_data_count(self) -> int:
        return _count_import_status(self.entries, STATUS_NO_DATA)

    @property
    def failed_count(self) -> int:
        return _count_import_status(self.entries, STATUS_FAILED)

    @property
    def skipped_count(self) -> int:
        return _count_import_status(self.entries, STATUS_SKIPPED)

    @property
    def duplicate_name_count(self) -> int:
        return _count_import_status(self.entries, DB_STATUS_DUPLICATE_NAME)

    @property
    def rows_imported_total(self) -> int:
        return sum(entry.rows_imported for entry in self.entries)

    @property
    def rows_rejected_total(self) -> int:
        return sum(entry.rows_rejected for entry in self.entries)


def analyze_rnmc_zip_row_preview(
    connection: sqlite3.Connection,
    zip_path: str,
    *,
    region_override: str = "",
    max_preview_rows: int = DEFAULT_ROW_PREVIEW_LIMIT,
) -> RnmcZipRowPreviewResult:
    """Preview RNMC workbook rows inside a zip archive without database writes."""
    manual_region = _text(region_override)
    row_limit = max(1, int(max_preview_rows))
    final_filename_keys = final_filename_keys_for_preview(connection)
    entries: list[RnmcWorkbookPreview] = []
    ignored_files = 0
    seen_keys: set[str] = set()

    try:
        with ZipFile(zip_path) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                path = _normalize_archive_path(info.filename)
                if not _is_supported_excel_path(path):
                    ignored_files += 1
                    continue
                filename = PurePosixPath(path).name
                key = normalize_import_filename(filename)
                region = manual_region or _region_from_archive_path(path)

                if key in seen_keys:
                    entries.append(
                        _empty_preview(
                            path,
                            filename,
                            region,
                            STATUS_DUPLICATE_NAME,
                            "Duplicate filename inside uploaded zip",
                        )
                    )
                elif key in final_filename_keys:
                    entries.append(
                        _empty_preview(
                            path,
                            filename,
                            region,
                            STATUS_SKIPPED_PROCESSED,
                            "Filename already has a final imported_files status; workbook was not opened",
                        )
                    )
                elif PurePosixPath(path).suffix.casefold() not in SUPPORTED_PREVIEW_SUFFIXES:
                    entries.append(
                        _empty_preview(
                            path,
                            filename,
                            region,
                            STATUS_UNSUPPORTED_FORMAT,
                            "Preview supports .xlsx and .xlsm; legacy .xls will need a separate reader",
                        )
                    )
                else:
                    try:
                        data = archive.read(info)
                        entries.append(_preview_workbook_bytes(
                            data,
                            path,
                            filename,
                            region,
                            allow_metadata_region=manual_region == "",
                            row_limit=row_limit,
                        ))
                    except Exception as exc:  # pragma: no cover - defensive UI boundary
                        entries.append(
                            _empty_preview(
                                path,
                                filename,
                                region,
                                STATUS_PARSE_ERROR,
                                str(exc),
                            )
                        )
                seen_keys.add(key)
    except BadZipFile as exc:
        raise ValueError("uploaded file is not a valid zip archive") from exc

    return RnmcZipRowPreviewResult(entries=entries, ignored_files=ignored_files)


def import_rnmc_zip_catalog_rows(
    connection: sqlite3.Connection,
    zip_path: str,
    *,
    region_override: str = "",
) -> RnmcZipCatalogImportResult:
    """Import valid RNMC rows from a zip archive into catalog_items."""
    manual_region = _text(region_override)
    final_filename_keys = final_filename_keys_for_preview(connection)
    entries: list[RnmcZipCatalogImportEntry] = []
    ignored_files = 0
    seen_keys: set[str] = set()

    try:
        with ZipFile(zip_path) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                path = _normalize_archive_path(info.filename)
                if not _is_supported_excel_path(path):
                    ignored_files += 1
                    continue
                filename = PurePosixPath(path).name
                key = normalize_import_filename(filename)
                region = manual_region or _region_from_archive_path(path)

                if key in seen_keys:
                    entry = _record_non_imported_file(
                        connection,
                        path,
                        filename,
                        region,
                        DB_STATUS_DUPLICATE_NAME,
                        "Duplicate filename inside uploaded zip",
                    )
                elif key in final_filename_keys:
                    if imported_file_exists_for_region(
                        connection,
                        region_folder=region,
                        filename=filename,
                    ):
                        entry = _import_entry(
                            path,
                            filename,
                            region,
                            STATUS_SKIPPED,
                            "Filename already has a final imported_files status",
                        )
                    else:
                        entry = _record_non_imported_file(
                            connection,
                            path,
                            filename,
                            region,
                            STATUS_SKIPPED,
                            "Skipped because filename already exists in imported_files",
                        )
                elif PurePosixPath(path).suffix.casefold() not in SUPPORTED_PREVIEW_SUFFIXES:
                    entry = _record_non_imported_file(
                        connection,
                        path,
                        filename,
                        region,
                        STATUS_FAILED,
                        "Import supports .xlsx and .xlsm; legacy .xls will need a separate reader",
                    )
                else:
                    try:
                        data = archive.read(info)
                        entry = _import_workbook_bytes(
                            connection,
                            data,
                            path,
                            filename,
                            region,
                            allow_metadata_region=manual_region == "",
                        )
                    except Exception as exc:  # pragma: no cover - defensive UI boundary
                        entry = _record_non_imported_file(
                            connection,
                            path,
                            filename,
                            region,
                            STATUS_FAILED,
                            str(exc),
                        )
                seen_keys.add(key)
                entries.append(entry)
    except BadZipFile as exc:
        raise ValueError("uploaded file is not a valid zip archive") from exc

    connection.commit()
    return RnmcZipCatalogImportResult(entries=entries, ignored_files=ignored_files)


def _import_workbook_bytes(
    connection: sqlite3.Connection,
    data: bytes,
    archive_path: str,
    filename: str,
    region_folder: str,
    *,
    allow_metadata_region: bool = True,
) -> RnmcZipCatalogImportEntry:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        metadata = _extract_workbook_metadata(workbook)
        resolved_region = _resolve_workbook_region(region_folder, metadata, allow_metadata_region)
        task_number = _extract_task_number(workbook)
        for sheet in workbook.worksheets:
            header = _find_header_row(sheet)
            if header is None:
                continue
            rows, rejected_rows = _extract_catalog_row_candidates(
                sheet,
                header,
                task_number,
                resolved_region,
                filename,
                metadata.regional_coefficient,
            )
            rejected = len(rejected_rows)
            if not rows:
                file_id = record_imported_file(
                    connection,
                    region_folder=resolved_region,
                    filename=filename,
                    status=STATUS_NO_DATA,
                    rows_ok=0,
                    rows_rejected=rejected,
                    failure_reason="Header found, but no valid catalog rows",
                    task_number=task_number,
                    lsr_quarter=metadata.lsr_quarter,
                    planned_start=metadata.planned_start,
                    planned_finish=metadata.planned_finish,
                    regional_coefficient=metadata.regional_coefficient,
                )
                replace_import_row_logs(
                    connection,
                    file_id,
                    [
                        (row.row_number, ROW_LOG_STATUS_REJECTED, row.reason)
                        for row in rejected_rows
                    ],
                )
                return _import_entry(
                    archive_path,
                    filename,
                    resolved_region,
                    STATUS_NO_DATA,
                    "Header found, but no valid catalog rows",
                    sheet_name=str(sheet.title),
                    header_row=header.row_number,
                    task_number=task_number,
                    lsr_quarter=metadata.lsr_quarter,
                    planned_start=metadata.planned_start,
                    planned_finish=metadata.planned_finish,
                    regional_coefficient=metadata.regional_coefficient,
                    rows_rejected=rejected,
                )

            result = replace_catalog_rows_for_file(
                connection,
                [
                    CatalogRowStorageItem(
                        catalog_row=row.catalog_row,
                        source_region_folder=resolved_region,
                        source_filename=filename,
                        source_row_number=row.row_number,
                    )
                    for row in rows
                ],
                region_folder=resolved_region,
                filename=filename,
            )
            file_id = record_imported_file(
                connection,
                region_folder=resolved_region,
                filename=filename,
                status=STATUS_SUCCESS,
                rows_ok=result.rows_imported,
                rows_rejected=rejected + result.rows_skipped,
                failure_reason="",
                task_number=task_number,
                lsr_quarter=metadata.lsr_quarter,
                planned_start=metadata.planned_start,
                planned_finish=metadata.planned_finish,
                regional_coefficient=metadata.regional_coefficient,
            )
            replace_import_row_logs(
                connection,
                file_id,
                [
                    (row.row_number, ROW_LOG_STATUS_REJECTED, row.reason)
                    for row in rejected_rows
                ],
            )
            return _import_entry(
                archive_path,
                filename,
                resolved_region,
                STATUS_SUCCESS,
                "Catalog rows imported",
                sheet_name=str(sheet.title),
                header_row=header.row_number,
                task_number=task_number,
                lsr_quarter=metadata.lsr_quarter,
                planned_start=metadata.planned_start,
                planned_finish=metadata.planned_finish,
                regional_coefficient=metadata.regional_coefficient,
                rows_imported=result.rows_imported,
                rows_rejected=rejected + result.rows_skipped,
            )

        file_id = record_imported_file(
            connection,
            region_folder=resolved_region,
            filename=filename,
            status=STATUS_NO_DATA,
            rows_ok=0,
            rows_rejected=0,
            failure_reason="Required headers were not found",
            task_number=task_number,
            lsr_quarter=metadata.lsr_quarter,
            planned_start=metadata.planned_start,
            planned_finish=metadata.planned_finish,
            regional_coefficient=metadata.regional_coefficient,
        )
        replace_import_row_logs(connection, file_id, [])
        return _import_entry(
            archive_path,
            filename,
            resolved_region,
            STATUS_NO_DATA,
            "Required headers were not found",
            task_number=task_number,
            lsr_quarter=metadata.lsr_quarter,
            planned_start=metadata.planned_start,
            planned_finish=metadata.planned_finish,
            regional_coefficient=metadata.regional_coefficient,
        )
    finally:
        workbook.close()


def _extract_catalog_row_candidates(
    sheet: Worksheet,
    header: _HeaderMatch,
    task_number: str,
    region_folder: str,
    filename: str,
    regional_coefficient: float | None = None,
) -> tuple[list[RnmcCatalogRowCandidate], list[RnmcRejectedRow]]:
    num_col = _find_numbering_col(header.header_map) or 1
    code_col = _find_code_col(header)
    value_columns = _detect_value_columns(header.header_map)
    date_col = _find_pattern_col(header.header_map, ADDED_DATE_HEADER_PATTERNS)
    max_row = int(sheet.max_row or 0)
    max_col = max(
        num_col,
        header.name_col,
        header.unit_col,
        header.qty_col,
        code_col,
        date_col,
        value_columns.unit_price.column,
        value_columns.total_price.column,
        value_columns.labor_unit_col,
        value_columns.labor_total_col,
        value_columns.machine_labor_unit_col,
        value_columns.machine_labor_total_col,
    )
    rows: list[RnmcCatalogRowCandidate] = []
    rejected: list[RnmcRejectedRow] = []
    started = False
    blank_streak = 0

    if max_row <= header.row_number:
        return rows, rejected

    sheet_rows = sheet.iter_rows(
        min_row=header.row_number + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )
    for row_offset, row_values in enumerate(sheet_rows, start=1):
        row_number = header.row_number + row_offset
        num_value = _row_value(row_values, num_col)
        name_value = _row_value(row_values, header.name_col)
        unit_value = _row_value(row_values, header.unit_col)
        qty_value = _row_value(row_values, header.qty_col)
        code_value = _row_value(row_values, code_col)
        price_source_value = _row_value(row_values, value_columns.unit_price.column)
        total_source_value = _row_value(row_values, value_columns.total_price.column)

        if _is_technical_numbering_row(row_values):
            continue

        if not started:
            if _is_blank(num_value) and _is_blank(unit_value) and _is_blank(qty_value):
                continue
            started = True

        is_end_blank = (
            _is_blank(num_value)
            and _is_blank(name_value)
            and _is_blank(unit_value)
            and _is_blank(qty_value)
        )
        if is_end_blank:
            blank_streak += 1
        else:
            blank_streak = 0
        if blank_streak >= 3:
            break
        if is_end_blank:
            continue

        if _is_non_catalog_section_row(
            name_value,
            unit_value,
            qty_value,
            code_value,
            price_source_value,
            total_source_value,
        ):
            continue

        if _is_blank(unit_value) and _is_blank(qty_value):
            rejected.append(RnmcRejectedRow(row_number, "missing_unit_and_quantity"))
            continue

        price_value = price_source_value
        parsed_price = _parse_positive_number(price_value)
        if parsed_price is not None:
            parsed_price = parsed_price / value_columns.unit_price.divisor

        total_price = _parse_optional_value(
            total_source_value,
            divisor=value_columns.total_price.divisor,
        )
        catalog_row = CatalogRow(
            task_id=task_number,
            price=parsed_price if parsed_price is not None else price_value,
            code=code_value,
            unit=unit_value,
            work_name=name_value,
            region=region_folder,
            added_date=_row_value(row_values, date_col),
            total_price=total_price,
            labor_unit=_parse_optional_value(
                _row_value(row_values, value_columns.labor_unit_col)
            ),
            labor_total=_parse_optional_value(
                _row_value(row_values, value_columns.labor_total_col)
            ),
            machine_labor_unit=_parse_optional_value(
                _row_value(row_values, value_columns.machine_labor_unit_col)
            ),
            machine_labor_total=_parse_optional_value(
                _row_value(row_values, value_columns.machine_labor_total_col)
            ),
            regional_coefficient=regional_coefficient,
        )
        issue = _catalog_row_issue(catalog_row)
        if issue != "":
            rejected.append(RnmcRejectedRow(row_number, issue))
            continue
        rows.append(RnmcCatalogRowCandidate(row_number=row_number, catalog_row=catalog_row))

    return rows, rejected

def _record_non_imported_file(
    connection: sqlite3.Connection,
    archive_path: str,
    filename: str,
    region_folder: str,
    status: str,
    reason: str,
) -> RnmcZipCatalogImportEntry:
    file_id = record_imported_file(
        connection,
        region_folder=region_folder,
        filename=filename,
        status=status,
        rows_ok=0,
        rows_rejected=0,
        failure_reason=reason,
    )
    replace_import_row_logs(connection, file_id, [])
    return _import_entry(archive_path, filename, region_folder, status, reason)


def _import_entry(
    archive_path: str,
    filename: str,
    region_folder: str,
    status: str,
    reason: str,
    *,
    sheet_name: str = "",
    header_row: int = 0,
    task_number: str = "",
    lsr_quarter: str = "",
    planned_start: str = "",
    planned_finish: str = "",
    regional_coefficient: float | None = None,
    rows_imported: int = 0,
    rows_rejected: int = 0,
) -> RnmcZipCatalogImportEntry:
    return RnmcZipCatalogImportEntry(
        archive_path=archive_path,
        filename=filename,
        region_folder=region_folder,
        status=status,
        reason=reason,
        sheet_name=sheet_name,
        header_row=header_row,
        task_number=task_number,
        lsr_quarter=lsr_quarter,
        planned_start=planned_start,
        planned_finish=planned_finish,
        regional_coefficient=regional_coefficient,
        rows_imported=rows_imported,
        rows_rejected=rows_rejected,
    )


def _preview_workbook_bytes(
    data: bytes,
    archive_path: str,
    filename: str,
    region_folder: str,
    *,
    allow_metadata_region: bool = True,
    row_limit: int = DEFAULT_ROW_PREVIEW_LIMIT,
) -> RnmcWorkbookPreview:
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        metadata = _extract_workbook_metadata(workbook)
        resolved_region = _resolve_workbook_region(region_folder, metadata, allow_metadata_region)
        task_number = _extract_task_number(workbook)
        for sheet in workbook.worksheets:
            header = _find_header_row(sheet)
            if header is None:
                continue
            rows_ok, rows_rejected, samples, is_limited = _preview_table_rows(
                sheet,
                header,
                row_limit=row_limit,
            )
            status = STATUS_PREVIEW_OK if rows_ok > 0 else STATUS_NO_ROWS
            reason = "Rows found" if rows_ok > 0 else "Header found, but no accepted rows"
            if is_limited:
                reason = f"{reason}; preview stopped at {row_limit} table rows"
            return RnmcWorkbookPreview(
                archive_path=archive_path,
                filename=filename,
                region_folder=resolved_region,
                status=status,
                reason=reason,
                sheet_name=str(sheet.title),
                header_row=header.row_number,
                task_number=task_number,
                lsr_quarter=metadata.lsr_quarter,
                planned_start=metadata.planned_start,
                planned_finish=metadata.planned_finish,
                regional_coefficient=metadata.regional_coefficient,
                rows_ok=rows_ok,
                rows_rejected=rows_rejected,
                header_preview=_build_header_preview(header),
                sample_rows=samples,
                is_limited=is_limited,
            )

        return _empty_preview(
            archive_path,
            filename,
            resolved_region,
            STATUS_NO_TABLE,
            "Required headers were not found",
            task_number=task_number,
            lsr_quarter=metadata.lsr_quarter,
            planned_start=metadata.planned_start,
            planned_finish=metadata.planned_finish,
            regional_coefficient=metadata.regional_coefficient,
        )
    finally:
        workbook.close()


@dataclass(frozen=True)
class _HeaderMatch:
    row_number: int
    name_col: int
    unit_col: int
    qty_col: int
    header_map: dict[str, int]
    header_labels: dict[int, str]


@dataclass(frozen=True)
class _ValueColumn:
    column: int
    divisor: float = 1.0


@dataclass(frozen=True)
class _HeaderValueColumns:
    unit_price: _ValueColumn
    total_price: _ValueColumn
    labor_unit_col: int
    labor_total_col: int
    machine_labor_unit_col: int
    machine_labor_total_col: int


def _find_header_row(sheet: Worksheet) -> _HeaderMatch | None:
    max_row = min(int(sheet.max_row or 0), 400)
    max_col = min(int(sheet.max_column or 0), 150)
    if max_row <= 0 or max_col <= 0:
        return None

    rows = sheet.iter_rows(
        min_row=1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )
    for row_number, row in enumerate(rows, start=1):
        name_col = 0
        unit_col = 0
        qty_col = 0
        header_map: dict[str, int] = {}
        header_labels: dict[int, str] = {}
        for index, value in enumerate(row, start=1):
            header_text = _text(value)
            if header_text != "":
                header_labels.setdefault(index, header_text)
            for normalized in _header_keys(value):
                header_map.setdefault(normalized, index)
                if _is_name_header(normalized):
                    name_col = index
                if _is_unit_header(normalized):
                    unit_col = index
                if _is_qty_header(normalized):
                    qty_col = index
        if name_col and unit_col and qty_col:
            return _HeaderMatch(
                row_number=row_number,
                name_col=name_col,
                unit_col=unit_col,
                qty_col=qty_col,
                header_map=header_map,
                header_labels=header_labels,
            )
    return None


def _preview_table_rows(
    sheet: Worksheet,
    header: _HeaderMatch,
    *,
    row_limit: int = DEFAULT_ROW_PREVIEW_LIMIT,
) -> tuple[int, int, list[RnmcRowSample], bool]:
    num_col = _find_numbering_col(header.header_map) or 1
    code_col = _find_code_col(header)
    value_columns = _detect_value_columns(header.header_map)
    max_row = int(sheet.max_row or 0)
    max_col = max(
        num_col,
        header.name_col,
        header.unit_col,
        header.qty_col,
        code_col,
        value_columns.unit_price.column,
        value_columns.total_price.column,
        value_columns.labor_unit_col,
        value_columns.labor_total_col,
        value_columns.machine_labor_unit_col,
        value_columns.machine_labor_total_col,
    )
    rows_ok = 0
    rows_rejected = 0
    samples: list[RnmcRowSample] = []
    started = False
    blank_streak = 0
    scanned_rows = 0
    is_limited = False

    if max_row <= header.row_number:
        return rows_ok, rows_rejected, samples, is_limited

    rows = sheet.iter_rows(
        min_row=header.row_number + 1,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    )
    for row_offset, row_values in enumerate(rows, start=1):
        row_number = header.row_number + row_offset
        num_value = _row_value(row_values, num_col)
        name_value = _row_value(row_values, header.name_col)
        unit_value = _row_value(row_values, header.unit_col)
        qty_value = _row_value(row_values, header.qty_col)
        code_value = _row_value(row_values, code_col)
        price_source_value = _row_value(row_values, value_columns.unit_price.column)
        total_source_value = _row_value(row_values, value_columns.total_price.column)

        if _is_technical_numbering_row(row_values):
            continue

        if not started:
            if _is_blank(num_value) and _is_blank(unit_value) and _is_blank(qty_value):
                continue
            started = True

        is_end_blank = (
            _is_blank(num_value)
            and _is_blank(name_value)
            and _is_blank(unit_value)
            and _is_blank(qty_value)
        )
        if is_end_blank:
            blank_streak += 1
        else:
            blank_streak = 0
        if blank_streak >= 3:
            break
        if is_end_blank:
            continue

        if _is_non_catalog_section_row(
            name_value,
            unit_value,
            qty_value,
            code_value,
            price_source_value,
            total_source_value,
        ):
            continue

        parsed_price = _parse_positive_number(price_source_value)
        if parsed_price is not None:
            parsed_price = parsed_price / value_columns.unit_price.divisor
        preview_row = CatalogRow(
            task_id="preview-task",
            price=parsed_price if parsed_price is not None else price_source_value,
            code=code_value,
            unit=unit_value,
            work_name=name_value,
            region="preview-region",
        )
        if _is_blank(unit_value) and _is_blank(qty_value):
            issue = "missing_unit_and_quantity"
        else:
            issue = _catalog_row_issue(preview_row)
        if issue:
            rows_rejected += 1
        else:
            rows_ok += 1

        scanned_rows += 1
        samples.append(
            RnmcRowSample(
                row_number=row_number,
                code=_text(code_value),
                work_name=_text(name_value),
                unit=_text(unit_value),
                quantity=_text(qty_value),
                unit_price=_format_preview_number(
                    _parse_optional_value(
                        price_source_value,
                        divisor=value_columns.unit_price.divisor,
                    )
                ),
                total_price=_format_preview_number(
                    _parse_optional_value(
                        total_source_value,
                        divisor=value_columns.total_price.divisor,
                    )
                ),
                labor_unit=_format_preview_number(
                    _parse_optional_value(_row_value(row_values, value_columns.labor_unit_col))
                ),
                labor_total=_format_preview_number(
                    _parse_optional_value(_row_value(row_values, value_columns.labor_total_col))
                ),
                machine_labor_unit=_format_preview_number(
                    _parse_optional_value(_row_value(row_values, value_columns.machine_labor_unit_col))
                ),
                machine_labor_total=_format_preview_number(
                    _parse_optional_value(_row_value(row_values, value_columns.machine_labor_total_col))
                ),
                issue=issue,
            )
        )
        if scanned_rows >= row_limit:
            is_limited = True
            break
    return rows_ok, rows_rejected, samples, is_limited

def _build_header_preview(header: _HeaderMatch) -> RnmcHeaderPreview:
    value_columns = _detect_value_columns(header.header_map)
    return RnmcHeaderPreview(
        code=_header_label(header, _find_code_col(header)),
        work_name=_header_label(header, header.name_col),
        unit=_header_label(header, header.unit_col),
        quantity=_header_label(header, header.qty_col),
        unit_price=_header_label(header, value_columns.unit_price.column),
        total_price=_header_label(header, value_columns.total_price.column),
        labor_unit=_header_label(header, value_columns.labor_unit_col),
        labor_total=_header_label(header, value_columns.labor_total_col),
        machine_labor_unit=_header_label(header, value_columns.machine_labor_unit_col),
        machine_labor_total=_header_label(header, value_columns.machine_labor_total_col),
    )


def _header_label(header: _HeaderMatch, column: int) -> str:
    if column <= 0:
        return ""
    label = header.header_labels.get(column, "")
    if label:
        return label
    if _is_unlabeled_code_col_before_section_code(header, column):
        return "[без заголовка перед Код раздела]"
    return ""


def _format_preview_number(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:g}"


def _row_value(row_values: tuple[object, ...], column: int) -> object:
    index = column - 1
    if index < 0 or index >= len(row_values):
        return None
    return row_values[index]


def _extract_workbook_metadata(workbook) -> RnmcWorkbookMetadata:
    lsr_quarter = ""
    planned_start = ""
    planned_finish = ""
    region_folder = ""
    regional_coefficient: float | None = None
    for sheet in workbook.worksheets:
        if lsr_quarter == "":
            lsr_quarter = _find_metadata_value(
                sheet,
                LSR_QUARTER_LABEL_PATTERNS,
                value_kind="quarter",
            )
        if planned_start == "":
            planned_start = _find_metadata_value(
                sheet,
                PLANNED_START_LABEL_PATTERNS,
                value_kind="date",
            )
        if planned_finish == "":
            planned_finish = _find_metadata_value(
                sheet,
                PLANNED_FINISH_LABEL_PATTERNS,
                value_kind="date",
            )
        if region_folder == "":
            region_folder = _find_region_metadata_value(sheet)
        if regional_coefficient is None:
            regional_coefficient = _find_regional_coefficient_metadata_value(sheet)

        context = _find_metadata_context(sheet)
        if lsr_quarter == "":
            lsr_quarter = context.lsr_quarter
        if planned_start == "":
            planned_start = context.planned_start
        if planned_finish == "":
            planned_finish = context.planned_finish
        if region_folder == "":
            region_folder = context.region_folder
        if regional_coefficient is None:
            regional_coefficient = context.regional_coefficient
        if (
            lsr_quarter
            and planned_start
            and planned_finish
            and region_folder
            and regional_coefficient is not None
        ):
            break
    return RnmcWorkbookMetadata(
        lsr_quarter=lsr_quarter,
        planned_start=planned_start,
        planned_finish=planned_finish,
        region_folder=region_folder,
        regional_coefficient=regional_coefficient,
    )


def _find_metadata_value(
    sheet: Worksheet,
    label_patterns: tuple[str, ...],
    *,
    value_kind: str,
) -> str:
    normalized_patterns = tuple(_normalize_header(pattern) for pattern in label_patterns)
    max_row = min(int(sheet.max_row or 0), 120)
    max_col = min(int(sheet.max_column or 0), 60)
    rows = _metadata_rows(sheet, max_row=max_row, max_col=max_col)
    for row_number, row_values in enumerate(rows, start=1):
        for col_number, value in enumerate(row_values, start=1):
            normalized = _normalize_header(value)
            if normalized == "" or not _metadata_label_matches(normalized, normalized_patterns):
                continue

            inline = _metadata_inline_value(value)
            if inline:
                formatted = _format_metadata_value(inline, value_kind=value_kind)
                if formatted:
                    return formatted

            for candidate in _metadata_neighbor_values(rows, row_number, col_number, max_row, max_col):
                formatted = _format_metadata_value(candidate, value_kind=value_kind)
                if formatted:
                    return formatted
    return ""


def _find_region_metadata_value(sheet: Worksheet) -> str:
    normalized_patterns = tuple(_normalize_header(pattern) for pattern in REGION_LABEL_PATTERNS)
    max_row = min(int(sheet.max_row or 0), 180)
    max_col = min(int(sheet.max_column or 0), 80)
    rows = _metadata_rows(sheet, max_row=max_row, max_col=max_col)
    for row_number, row_values in enumerate(rows, start=1):
        for col_number, value in enumerate(row_values, start=1):
            normalized = _normalize_header(value)
            if normalized == "" or not _region_label_matches(normalized, normalized_patterns):
                continue
            inline = _metadata_inline_value(value)
            formatted = _format_region_metadata(inline)
            if formatted:
                return formatted
            for candidate in _metadata_neighbor_values(rows, row_number, col_number, max_row, max_col):
                formatted = _format_region_metadata(candidate)
                if formatted:
                    return formatted
    return ""


def _find_regional_coefficient_metadata_value(sheet: Worksheet) -> float | None:
    normalized_patterns = tuple(
        _normalize_header(pattern) for pattern in REGIONAL_COEFFICIENT_LABEL_PATTERNS
    )
    max_row = min(int(sheet.max_row or 0), 180)
    max_col = min(int(sheet.max_column or 0), 80)
    rows = _metadata_rows(sheet, max_row=max_row, max_col=max_col)
    for row_number, row_values in enumerate(rows, start=1):
        for col_number, value in enumerate(row_values, start=1):
            normalized = _normalize_header(value)
            if normalized == "" or not _coefficient_label_matches(normalized, normalized_patterns):
                continue
            inline = _metadata_inline_value(value)
            formatted = _format_coefficient_metadata(inline)
            if formatted is not None:
                return formatted
            for candidate in _metadata_neighbor_values(rows, row_number, col_number, max_row, max_col):
                formatted = _format_coefficient_metadata(candidate)
                if formatted is not None:
                    return formatted
    return None



def _region_label_matches(normalized: str, normalized_patterns: tuple[str, ...]) -> bool:
    if "\u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442" in normalized:
        return False
    for pattern in normalized_patterns:
        if pattern == "":
            continue
        if normalized == pattern:
            return True
        if pattern != "\u0440\u0435\u0433\u0438\u043e\u043d" and pattern in normalized:
            return True
    return False


def _coefficient_label_matches(normalized: str, normalized_patterns: tuple[str, ...]) -> bool:
    for pattern in normalized_patterns:
        if pattern and pattern in normalized:
            return True
    return False


def _format_region_metadata(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    if _parse_number(text) is not None:
        return ""
    normalized = " ".join(text.replace("\u00a0", " ").split())
    folded = normalized.casefold()
    if len(folded) < 3:
        return ""
    if (
        "\u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442" in folded
        or folded in {"\u0440\u0435\u0433\u0438\u043e\u043d", "\u0440\u0435\u0433\u0438\u043e\u043d \u043e\u0431\u044a\u0435\u043a\u0442\u0430", "\u0440\u0435\u0433\u0438\u043e\u043d \u0440\u0430\u0441\u043f\u043e\u043b\u043e\u0436\u0435\u043d\u0438\u044f \u043e\u0431\u044a\u0435\u043a\u0442\u0430"}
        or "\u0434\u0430\u043d\u043d\u044b\u0435 \u0434\u043b\u044f \u043a\u043e\u043d\u0441\u043e\u043b\u0438\u0434\u0430\u0446\u0438\u0438" in folded
    ):
        return ""
    return normalized


def _format_coefficient_metadata(value: object) -> float | None:
    number = _parse_number(value)
    if number is None or number <= 0:
        return None
    if number < 0.1 or number > 10:
        return None
    return number


def _resolve_workbook_region(
    default_region: str,
    metadata: RnmcWorkbookMetadata,
    allow_metadata_region: bool,
) -> str:
    if allow_metadata_region and metadata.region_folder:
        return metadata.region_folder
    return default_region


def _metadata_label_matches(normalized: str, normalized_patterns: tuple[str, ...]) -> bool:
    for pattern in normalized_patterns:
        if pattern == "":
            continue
        if pattern in normalized:
            return True
        if len(normalized) >= 8 and normalized in pattern:
            return True
    return False


def _metadata_inline_value(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    for separator in (":", "—", "-", "="):
        if separator in text:
            tail = text.split(separator, 1)[1].strip()
            if tail:
                return tail
    return ""


def _metadata_rows(
    sheet: Worksheet,
    *,
    max_row: int,
    max_col: int,
) -> list[tuple[object, ...]]:
    if max_row <= 0 or max_col <= 0:
        return []
    return list(
        sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
    )


def _metadata_neighbor_values(
    rows: list[tuple[object, ...]],
    row_number: int,
    col_number: int,
    max_row: int,
    max_col: int,
) -> list[object]:
    values: list[object] = []
    for offset in range(1, 5):
        column = col_number + offset
        if column <= max_col:
            values.append(_grid_value(rows, row_number, column))
    if row_number + 1 <= max_row:
        values.append(_grid_value(rows, row_number + 1, col_number))
        for offset in range(1, 3):
            column = col_number + offset
            if column <= max_col:
                values.append(_grid_value(rows, row_number + 1, column))
    return values


def _grid_value(rows: list[tuple[object, ...]], row_number: int, column: int) -> object:
    row_index = row_number - 1
    column_index = column - 1
    if row_index < 0 or row_index >= len(rows):
        return None
    row_values = rows[row_index]
    if column_index < 0 or column_index >= len(row_values):
        return None
    return row_values[column_index]


def _format_metadata_value(value: object, *, value_kind: str) -> str:
    if value_kind == "date":
        return _format_date_metadata(value)
    if value_kind == "quarter":
        return _format_quarter_metadata(value)
    return _text(value)


def _format_date_metadata(value: object) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    serial_date = _excel_serial_date(value)
    if serial_date:
        return serial_date

    text = _text(value)
    if text == "":
        return ""
    normalized = text.replace("\u00a0", " ").strip()
    for pattern in (
        r"(?P<day>\d{1,2})[./-](?P<month>\d{1,2})[./-](?P<year>\d{4})",
        r"(?P<year>\d{4})[./-](?P<month>\d{1,2})[./-](?P<day>\d{1,2})",
    ):
        match = re.search(pattern, normalized)
        if not match:
            continue
        try:
            parsed = date(
                int(match.group("year")),
                int(match.group("month")),
                int(match.group("day")),
            )
        except ValueError:
            continue
        return parsed.isoformat()
    start, finish = _parse_month_period_context(normalized)
    return start or finish


def _format_quarter_metadata(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    return _parse_quarter_context(text)


def _find_metadata_context(sheet: Worksheet) -> RnmcWorkbookMetadata:
    lsr_quarter = ""
    planned_start = ""
    planned_finish = ""
    region_folder = ""
    regional_coefficient: float | None = None
    max_row = min(int(sheet.max_row or 0), 180)
    max_col = min(int(sheet.max_column or 0), 80)
    rows = _metadata_rows(sheet, max_row=max_row, max_col=max_col)
    for row_values in rows:
        text = " ".join(_text(value) for value in row_values if _text(value))
        if text == "":
            continue
        if lsr_quarter == "":
            lsr_quarter = _parse_quarter_context(text)
        if planned_start == "" or planned_finish == "":
            start, finish = _parse_month_period_context(text)
            planned_start = planned_start or start
            planned_finish = planned_finish or finish
        if region_folder == "":
            region_folder = _parse_region_context(text)
        if regional_coefficient is None:
            regional_coefficient = _parse_coefficient_context(text)
        if (
            lsr_quarter
            and planned_start
            and planned_finish
            and region_folder
            and regional_coefficient is not None
        ):
            break
    return RnmcWorkbookMetadata(
        lsr_quarter=lsr_quarter,
        planned_start=planned_start,
        planned_finish=planned_finish,
        region_folder=region_folder,
        regional_coefficient=regional_coefficient,
    )


def _parse_region_context(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    match = re.search(
        r"(?:\u0440\u0435\u0433\u0438\u043e\u043d(?:\s+\u0440\u0430\u0441\u043f\u043e\u043b\u043e\u0436\u0435\u043d\u0438\u044f\s+\u043e\u0431\u044a\u0435\u043a\u0442\u0430|\s+\u043e\u0431\u044a\u0435\u043a\u0442\u0430)?)\s*[:=\-\u2014]\s*(?P<region>[^;,.]+)",
        text.replace("\u00a0", " "),
        flags=re.IGNORECASE,
    )
    if match is None:
        return ""
    return _format_region_metadata(match.group("region"))


def _parse_coefficient_context(value: object) -> float | None:
    text = _text(value)
    if text == "":
        return None
    match = re.search(
        r"(?:\u0440\u0435\u0433\u0438\u043e\u043d\u0430\u043b\u044c\u043d\u044b\u0439\s+\u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442|\u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442)\s*[:=\-\u2014]\s*(?P<value>[0-9]+(?:[,.][0-9]+)?)",
        text.replace("\u00a0", " "),
        flags=re.IGNORECASE,
    )
    if match is None:
        return None
    return _format_coefficient_metadata(match.group("value"))


def _parse_quarter_context(value: object) -> str:
    text = _text(value)
    if text == "":
        return ""
    folded = text.replace(" ", " ").casefold()
    year_pattern = r"(?P<year>(?:20|19)\d{2}|\d{2}\s*г\.?)"
    roman_pattern = "(?P<roman>iv|iii|ii|i)"
    number_pattern = "(?P<number>[1-4])"
    quarter_word = r"(?:\u043a\u0432\.?|\u043a\u0432\u0430\u0440\u0442\u0430\u043b\w*)"
    patterns = (
        roman_pattern + r"\s*" + quarter_word + "[^0-9]{0,20}" + year_pattern,
        number_pattern + r"\s*" + quarter_word + "[^0-9]{0,20}" + year_pattern,
        quarter_word + r"\s*" + number_pattern + "[^0-9]{0,20}" + year_pattern,
    )
    for pattern in patterns:
        match = re.search(pattern, folded)
        if match is None:
            continue
        year = _normalize_year(match.group("year"))
        if not _is_plausible_metadata_year(year):
            continue
        quarter = match.groupdict().get("number") or _roman_quarter(match.groupdict().get("roman", ""))
        if quarter:
            return f"{year} Q{quarter}"
    return ""

def _roman_quarter(value: str) -> str:
    return {"i": "1", "ii": "2", "iii": "3", "iv": "4"}.get(value.casefold(), "")


def _normalize_year(value: str) -> str:
    text = _text(value)
    if text == "":
        return ""
    digits = "".join(char for char in text if char.isdigit())
    if len(digits) == 2:
        return f"20{digits}"
    if len(digits) == 4:
        return digits
    return ""


def _parse_month_period_context(value: object) -> tuple[str, str]:
    text = _text(value)
    if text == "":
        return "", ""
    folded = text.replace(" ", " ").casefold()
    tokens = _month_tokens(folded)
    if not tokens:
        return "", ""
    years = [_normalize_year(item) for item in re.findall(r"(?<!\d)((?:20|19)\d{2})(?!\d)", folded)]
    years = [year for year in years if _is_plausible_metadata_year(year)]
    if not _looks_like_period_text(folded) and not _looks_like_month_date_only_text(folded):
        return "", ""
    if len(tokens) == 1:
        month, year = tokens[0]
        year = year or (years[0] if years else "")
        if year and _is_plausible_metadata_year(year):
            value = _month_date(year, month)
            return value, value
        return "", ""
    start_month, start_year = tokens[0]
    finish_month, finish_year = tokens[1]
    if start_year == "" and finish_year:
        start_year = finish_year
    if finish_year == "" and start_year:
        finish_year = start_year
    if start_year == "" and finish_year == "" and years:
        start_year = years[0]
        finish_year = years[-1]
    if (
        start_year
        and finish_year
        and _is_plausible_metadata_year(start_year)
        and _is_plausible_metadata_year(finish_year)
    ):
        return _month_date(start_year, start_month), _month_date(finish_year, finish_month)
    return "", ""

def _looks_like_period_text(value: str) -> bool:
    if any(
        token in value
        for token in (
            "период",
            "срок",
            "начал",
            "оконч",
        )
    ):
        return True
    return bool(
        re.search(
            r"(?:^|\s)\u0441\s+.+?(?:\s\u043f\u043e\s|[-\u2013\u2014]).+",
            value,
        )
    )

def _month_tokens(value: str) -> list[tuple[int, str]]:
    month_pattern = (
        r"(?<![а-яёa-z])"
        r"(?P<month>"
        r"январ[а-яё]*|"
        r"феврал[а-яё]*|"
        r"март[а-яё]*|"
        r"апрел[а-яё]*|"
        r"ма(?:й|я|е|ю|ем)|"
        r"июн[а-яё]*|"
        r"июл[а-яё]*|"
        r"август[а-яё]*|"
        r"сентябр[а-яё]*|"
        r"октябр[а-яё]*|"
        r"ноябр[а-яё]*|"
        r"декабр[а-яё]*)"
        r"(?![а-яёa-z])"
        r"\s*(?P<year>(?:20|19)\d{2}|\d{2}\s*г\.?)?"
    )
    tokens: list[tuple[int, str]] = []
    for match in re.finditer(month_pattern, value):
        month = _month_number(match.group("month"))
        year = _normalize_year(match.group("year") or "")
        if month:
            tokens.append((month, year if _is_plausible_metadata_year(year) else ""))
    return tokens

def _month_number(value: str) -> int:
    folded = value.casefold()
    prefixes = (
        ("\u044f\u043d\u0432\u0430\u0440", 1),
        ("\u0444\u0435\u0432\u0440\u0430\u043b", 2),
        ("\u043c\u0430\u0440\u0442", 3),
        ("\u0430\u043f\u0440\u0435\u043b", 4),
        ("\u043c\u0430", 5),
        ("\u0438\u044e\u043d", 6),
        ("\u0438\u044e\u043b", 7),
        ("\u0430\u0432\u0433\u0443\u0441\u0442", 8),
        ("\u0441\u0435\u043d\u0442\u044f\u0431\u0440", 9),
        ("\u043e\u043a\u0442\u044f\u0431\u0440", 10),
        ("\u043d\u043e\u044f\u0431\u0440", 11),
        ("\u0434\u0435\u043a\u0430\u0431\u0440", 12),
    )
    for prefix, month in prefixes:
        if folded.startswith(prefix):
            return month
    return 0


def _month_date(year: str, month: int) -> str:
    return date(int(year), int(month), 1).isoformat()


def _excel_serial_date(value: object) -> str:
    if not isinstance(value, Real) or isinstance(value, bool):
        return ""
    serial = float(value)
    if serial < 20000 or serial > 80000:
        return ""
    parsed = EXCEL_DATE_BASE + timedelta(days=int(serial))
    if parsed.year < 2020 or parsed.year > 2035:
        return ""
    return parsed.isoformat()

def _extract_task_number(workbook) -> str:
    full_label = TASK_LABEL_FULL.casefold()
    short_label = TASK_LABEL_SHORT.casefold()
    for sheet in workbook.worksheets:
        max_row = min(int(sheet.max_row or 0), 50)
        max_col = min(int(sheet.max_column or 0), 20)
        rows = sheet.iter_rows(
            min_row=1,
            max_row=max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        )
        for row_values in rows:
            for index, value in enumerate(row_values, start=1):
                text = _text(value)
                if text == "":
                    continue
                folded = text.casefold()
                for label in (full_label, short_label):
                    position = folded.find(label)
                    if position < 0:
                        continue
                    tail = _cleanup_task_tail(text[position + len(label) :])
                    if tail:
                        return tail
                    neighbor = _neighbor_task_value_from_row(row_values, index)
                    if neighbor:
                        return neighbor
    return ""

def _neighbor_task_value_from_row(row_values: tuple[object, ...], col_number: int) -> str:
    for offset in range(1, 4):
        value = _text(_row_value(row_values, col_number + offset))
        if value:
            return value
    return ""


def _cleanup_task_tail(value: str) -> str:
    text = value.replace(":", "").replace("#", "")
    return " ".join(text.split())


def _find_code_col(header: _HeaderMatch) -> int:
    normalized_patterns = tuple(_normalize_header(pattern) for pattern in CODE_HEADER_PATTERNS)
    generic_code = _normalize_header("Код")
    for pattern in normalized_patterns:
        if pattern == generic_code:
            continue
        for key, column in header.header_map.items():
            if _is_section_code_header(key):
                continue
            if key == pattern or pattern in key:
                return int(column)

    fallback = _find_unlabeled_code_col_before_section_code(header)
    if fallback:
        return fallback

    for key, column in header.header_map.items():
        if _is_section_code_header(key):
            continue
        if key == generic_code:
            return int(column)
    return 0


def _is_section_code_header(normalized_header: str) -> bool:
    return "код" in normalized_header and "раздел" in normalized_header


def _find_unlabeled_code_col_before_section_code(header: _HeaderMatch) -> int:
    for key, column in header.header_map.items():
        if not _is_section_code_header(key):
            continue
        candidate = int(column) - 1
        if candidate > 0 and candidate not in header.header_labels:
            return candidate
    return 0


def _is_unlabeled_code_col_before_section_code(header: _HeaderMatch, column: int) -> bool:
    return column > 0 and column == _find_unlabeled_code_col_before_section_code(header)


def _is_technical_numbering_row(row_values: tuple[object, ...]) -> bool:
    non_empty = [_text(value) for value in row_values if _text(value) != ""]
    if len(non_empty) < 4:
        return False
    numeric: list[int] = []
    for value in non_empty:
        if not re.fullmatch(r"\d+", value):
            return False
        numeric.append(int(value))
    if any(value <= 0 or value > 40 for value in numeric):
        return False
    return True


def _is_non_catalog_section_row(
    work_name: object,
    unit_value: object,
    qty_value: object,
    code_value: object,
    price_value: object,
    total_value: object,
) -> bool:
    if not _is_blank(unit_value) or not _is_blank(qty_value):
        return False
    if _parse_number(price_value) is not None or _parse_number(total_value) is not None:
        return False
    if NormCode(code_value) != "":
        return False
    return _text(work_name) != ""


def _is_plausible_metadata_year(value: str) -> bool:
    text = _text(value)
    if text == "":
        return False
    try:
        year = int(text)
    except ValueError:
        return False
    return 2010 <= year <= 2035


def _looks_like_month_date_only_text(value: str) -> bool:
    compact = re.sub(r"[\s.,()]+", " ", value).strip()
    return bool(
        re.fullmatch(
            r"(?:с |по |в )?[а-яё]+\s+(?:(?:20|19)\d{2}|\d{2}\s*г\.?)\s*(?:г|г\.|года)?",
            compact,
        )
    )


def _find_numbering_col(header_map: dict[str, int]) -> int:
    number_sign = chr(8470).casefold()
    for key, column in header_map.items():
        lowered = key.casefold()
        if (
            number_sign in lowered
            or lowered.startswith("no")
            or "pp" in lowered
            or "p/p" in lowered
            or "p-p" in lowered
        ):
            return int(column)
    return 0




def _detect_value_columns(header_map: dict[str, int]) -> _HeaderValueColumns:
    unit_price_candidates: list[_ValueColumn] = []
    total_price_candidates: list[_ValueColumn] = []
    labor_unit_col = 0
    labor_total_col = 0
    machine_labor_unit_col = 0
    machine_labor_total_col = 0

    for key, column in header_map.items():
        if _is_average_value_header(key):
            continue
        if machine_labor_unit_col == 0 and _is_machine_labor_unit_header(key):
            machine_labor_unit_col = int(column)
            continue
        if machine_labor_total_col == 0 and _is_machine_labor_total_header(key):
            machine_labor_total_col = int(column)
            continue
        if labor_unit_col == 0 and _is_labor_unit_header(key):
            labor_unit_col = int(column)
            continue
        if labor_total_col == 0 and _is_labor_total_header(key):
            labor_total_col = int(column)
            continue
        if _is_unit_price_header(key):
            unit_price_candidates.append(_ValueColumn(int(column), _vat_divisor(key)))
            continue
        if _is_total_price_header(key):
            total_price_candidates.append(_ValueColumn(int(column), _vat_divisor(key)))

    return _HeaderValueColumns(
        unit_price=_prefer_without_vat(unit_price_candidates),
        total_price=_prefer_without_vat(total_price_candidates),
        labor_unit_col=labor_unit_col,
        labor_total_col=labor_total_col,
        machine_labor_unit_col=machine_labor_unit_col,
        machine_labor_total_col=machine_labor_total_col,
    )


def _prefer_without_vat(candidates: list[_ValueColumn]) -> _ValueColumn:
    if not candidates:
        return _ValueColumn(0, 1.0)
    for candidate in candidates:
        if candidate.divisor == 1.0:
            return candidate
    return candidates[0]


def _vat_divisor(normalized_header: str) -> float:
    without_vat = "\u0431\u0435\u0437\u043d\u0434\u0441"
    with_vat = "\u0441\u043d\u0434\u0441"
    if with_vat in normalized_header and without_vat not in normalized_header:
        return VALUE_WITH_VAT_DIVISOR
    return 1.0


def _is_average_value_header(normalized_header: str) -> bool:
    return any(
        token in normalized_header
        for token in (
            "\u0441\u0440\u0435\u0434",
            "\u0441\u0440\u0437\u043d\u0430\u0447",
            "\u0441\u0440.\u0437\u043d\u0430\u0447",
        )
    )


def _is_unit_price_header(normalized_header: str) -> bool:
    if _is_total_price_header(normalized_header):
        return False
    price = "\u0446\u0435\u043d\u0430"
    cost = "\u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442"
    if normalized_header == price:
        return True
    if "\u0431\u0430\u0437\u043e\u0432\u0430\u044f\u0446\u0435\u043d\u0430" in normalized_header:
        return True
    unit_markers = (
        "\u0435\u0434\u0438\u043d\u0438\u0446",
        "\u0437\u0430\u0435\u0434",
        "\u0437\u04301",
        "\u0435\u0434.",
    )
    return (price in normalized_header or cost in normalized_header) and any(
        marker in normalized_header for marker in unit_markers
    )


def _is_total_price_header(normalized_header: str) -> bool:
    return (
        "\u0438\u0442\u043e\u0433\u043e" in normalized_header
        and "\u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442" in normalized_header
    )


def _is_labor_unit_header(normalized_header: str) -> bool:
    return (
        "\u0442\u0437\u043d\u0430\u0435\u0434" in normalized_header
        or "\u0442\u0437\u0440\u043d\u0430\u0435\u0434" in normalized_header
        or "\u0437\u0442\u0440\u043d\u0430\u0435\u0434" in normalized_header
    )


def _is_labor_total_header(normalized_header: str) -> bool:
    return (
        "\u0442\u0437\u0432\u0441\u0435\u0433\u043e" in normalized_header
        or "\u0442\u0437\u0440\u0432\u0441\u0435\u0433\u043e" in normalized_header
        or "\u0437\u0442\u0440\u0432\u0441\u0435\u0433\u043e" in normalized_header
    )


def _is_machine_labor_unit_header(normalized_header: str) -> bool:
    return "\u0442\u0437\u043c\u043d\u0430\u0435\u0434" in normalized_header


def _is_machine_labor_total_header(normalized_header: str) -> bool:
    return "\u0442\u0437\u043c\u0432\u0441\u0435\u0433\u043e" in normalized_header


def _find_pattern_col(header_map: dict[str, int], patterns: tuple[str, ...]) -> int:
    normalized_patterns = tuple(_normalize_header(pattern) for pattern in patterns)
    for pattern in normalized_patterns:
        for key, column in header_map.items():
            if key == pattern or pattern in key:
                return int(column)
    return 0


def _cell_value(sheet: Worksheet, row_number: int, col_number: int) -> object:
    if col_number <= 0:
        return None
    return sheet.cell(row_number, col_number).value


def _catalog_row_issue(row: CatalogRow) -> str:
    if _text(row.task_id) == "":
        return "missing_task_number"
    if NormCode(row.code) == "":
        return "missing_or_invalid_code"
    if NormUnit(row.unit) == "":
        return "missing_or_invalid_unit"
    if _parse_positive_number(row.price) is None:
        return "missing_or_invalid_price"
    return ""


def _parse_positive_number(value: object) -> float | None:
    number = _parse_number(value)
    if number is None or number <= 0:
        return None
    return number


def _parse_optional_value(value: object, *, divisor: float = 1.0) -> float | None:
    number = _parse_number(value)
    if number is None:
        return None
    return number / divisor


def _parse_number(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Real):
        return float(value)
    if isinstance(value, (date, datetime)):
        return None
    text = str(value).strip()
    if text == "":
        return None
    text = text.replace(" ", " ").replace(" ", "")
    text = "".join(char for char in text if char.isdigit() or char in {".", ",", "-"})
    if text in {"", ".", ",", "-"}:
        return None
    if "," in text and "." in text:
        comma_pos = text.rfind(",")
        dot_pos = text.rfind(".")
        decimal_sep = "," if comma_pos > dot_pos else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "")
        text = text.replace(decimal_sep, ".")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _count_import_status(entries: list[RnmcZipCatalogImportEntry], status: str) -> int:
    return sum(1 for entry in entries if entry.status == status)


def _is_name_header(value: str) -> bool:
    name_works = _normalize_header(HEADER_NAME_WORKS)
    name_short = _normalize_header(HEADER_NAME_SHORT)
    return value in {name_works, name_short} or value.startswith(name_short)


def _is_unit_header(value: str) -> bool:
    units = {
        _normalize_header(HEADER_UNIT),
        _normalize_header(HEADER_UNIT_NO_DOT),
        _normalize_header(HEADER_UNIT_LONG),
    }
    return any(value.startswith(unit) or unit in value for unit in units if unit)


def _is_qty_header(value: str) -> bool:
    return _normalize_header(HEADER_QTY_SHORT) in value or _normalize_header(HEADER_QTY_LONG) in value


def _header_keys(value: object) -> list[str]:
    text = _text(value)
    if text == "":
        return []
    keys: list[str] = []
    for part in (text, text.splitlines()[0] if text.splitlines() else ""):
        normalized = _normalize_header(part)
        if normalized and normalized not in keys:
            keys.append(normalized)
    return keys


def _normalize_header(value: object) -> str:
    text = _text(value).casefold()
    text = text.replace("\u00a0", " ").replace("\r", " ").replace("\n", " ")
    return "".join(text.split())


def _normalize_archive_path(value: str) -> str:
    path = str(value).replace("\\", "/").strip("/")
    return "/".join(part for part in path.split("/") if part not in {"", "."})


def _is_supported_excel_path(path: str) -> bool:
    pure = PurePosixPath(path)
    name = pure.name
    if name.startswith("~$"):
        return False
    return pure.suffix.casefold() in EXCEL_SUFFIXES


def _region_from_archive_path(path: str) -> str:
    parent = PurePosixPath(path).parent
    if str(parent) in {"", "."}:
        return ""
    return parent.name.strip()


def _empty_preview(
    archive_path: str,
    filename: str,
    region_folder: str,
    status: str,
    reason: str,
    *,
    task_number: str = "",
    lsr_quarter: str = "",
    planned_start: str = "",
    planned_finish: str = "",
    regional_coefficient: float | None = None,
) -> RnmcWorkbookPreview:
    return RnmcWorkbookPreview(
        archive_path=archive_path,
        filename=filename,
        region_folder=region_folder,
        status=status,
        reason=reason,
        sheet_name="",
        header_row=0,
        task_number=task_number,
        lsr_quarter=lsr_quarter,
        planned_start=planned_start,
        planned_finish=planned_finish,
        regional_coefficient=regional_coefficient,
        rows_ok=0,
        rows_rejected=0,
        header_preview=RnmcHeaderPreview(),
        sample_rows=[],
        is_limited=False,
    )


def _count_status(entries: list[RnmcWorkbookPreview], status: str) -> int:
    return sum(1 for entry in entries if entry.status == status)


def _is_blank(value: object) -> bool:
    return _text(value) == ""


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()

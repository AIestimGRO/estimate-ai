"""Durable browser post-processing over a generated estimate workbook."""

from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.write_result import RunAndWriteResult
from core.ooxml_preservation import preserve_workbook_package_features
from core.exclusions import LEGACY_REASON_COLOR, is_task_marked, resolve_task_highlight
from core.storage.rules import list_task_color_entries, list_task_highlight_reasons
from core.storage.workspace import (
    STATUS_READY,
    get_processing_job,
    list_cell_overrides,
    list_processing_rows,
    replace_processing_rows,
    set_job_final_output,
    upsert_processing_job,
)


def persist_processing_job(
    connection,
    *,
    job_id: str,
    owner_user_id: int,
    estimate_filename: str,
    source_path: str | Path,
    outcome: RunAndWriteResult,
    region: str,
    use_tkp_analogs: bool,
) -> None:
    """Snapshot the generated workbook so browser review survives restarts."""
    workbook = load_workbook(outcome.output_path, data_only=False, read_only=True)
    try:
        worksheet = workbook[outcome.sheet_title]
        header_row = int(outcome.write_report.header_row)
        max_column = _max_output_column(outcome, worksheet.max_column)
        schema = _build_column_schema(
            worksheet,
            header_row,
            max_column,
            outcome,
        )
        rows: list[tuple[int, int, list[object], dict[str, object]]] = []
        for row_index, (excel_row, result_row) in enumerate(
            zip(outcome.row_numbers, outcome.result.rows),
            start=1,
        ):
            values = [
                _json_cell_value(
                    worksheet.cell(row=int(excel_row), column=column).value
                )
                for column in range(1, max_column + 1)
            ]
            average_column = int(outcome.write_report.average_column)
            if 0 < average_column <= len(values):
                values[average_column - 1] = result_row.recommended_price
            rows.append(
                (
                    row_index,
                    int(excel_row),
                    values,
                    {
                        "has_analogs": bool(result_row.has_analogs),
                        "risk": bool(result_row.risk_result.is_flagged),
                        "has_tkp": bool(result_row.has_tkp_analog),
                        "status": str(result_row.status),
                        "reason": str(result_row.match_result.reason),
                    },
                )
            )
    finally:
        workbook.close()

    upsert_processing_job(
        connection,
        job_id=str(job_id),
        owner_user_id=int(owner_user_id),
        estimate_filename=str(estimate_filename),
        source_path=str(Path(source_path)),
        output_path=str(outcome.output_path),
        sheet_title=str(outcome.sheet_title),
        header_row=int(outcome.write_report.header_row),
        coefficient=float(outcome.regional_coefficient),
        region=str(region or ""),
        use_tkp_analogs=bool(use_tkp_analogs),
        status=STATUS_READY,
        total_rows=len(rows),
        matched_rows=int(outcome.result.matched_row_count),
        flagged_rows=int(outcome.result.flagged_row_count),
        tkp_matched_rows=int(outcome.result.tkp_matched_row_count),
        column_schema=schema,
    )
    replace_processing_rows(connection, str(job_id), rows)


def build_postprocessed_workbook(connection, job_id: str) -> Path:
    """Build a fresh final workbook from the immutable run output + overrides."""
    job = get_processing_job(connection, str(job_id))
    if job is None:
        raise ValueError("Processing job was not found")

    base_path = Path(job.output_path)
    if not base_path.is_file():
        raise FileNotFoundError(base_path)

    suffix = base_path.suffix or ".xlsx"
    final_path = base_path.with_name(f"{base_path.stem} reviewed{suffix}")
    shutil.copy2(base_path, final_path)

    keep_vba = suffix.lower() == ".xlsm"
    workbook = load_workbook(final_path, data_only=False, keep_vba=keep_vba)
    try:
        worksheet = workbook[job.sheet_title]
        rows_by_id = {
            row.id: row
            for row in list_processing_rows(connection, job.id)
        }
        for override in list_cell_overrides(connection, job.id):
            row = rows_by_id.get(override.row_id)
            if row is None:
                continue
            worksheet.cell(
                row=int(row.excel_row_number),
                column=int(override.column_index),
            ).value = override.current_value

        _apply_current_task_highlights(
            connection,
            worksheet,
            job,
            list(rows_by_id.values()),
        )
        _request_full_recalculation(workbook)
        workbook.save(final_path)
        preserve_workbook_package_features(
            base_path,
            final_path,
            modified_sheet_title=worksheet.title,
        )
    finally:
        workbook.close()

    set_job_final_output(connection, job.id, str(final_path))
    return final_path


def _apply_current_task_highlights(
    connection,
    worksheet,
    job,
    rows,
) -> None:
    """Apply task highlighting that may have been approved after the run."""
    color_entries = list_task_color_entries(connection)
    reasons = list_task_highlight_reasons(connection)
    if not color_entries:
        return

    for column in job.column_schema:
        if str(column.get("kind") or "") != "analog":
            continue
        task_number = str(column.get("task_number") or "").strip()
        if not task_number or not is_task_marked(color_entries, task_number):
            continue

        highlight = resolve_task_highlight(color_entries, reasons, task_number)
        if highlight is None:
            color_hex = LEGACY_REASON_COLOR
            label = ""
        else:
            color_hex = highlight.color_hex
            label = highlight.label

        fill = PatternFill(
            start_color=f"FF{color_hex}",
            end_color=f"FF{color_hex}",
            fill_type="solid",
        )
        column_index = int(column["index"])
        for row in rows:
            worksheet.cell(
                row=int(row.excel_row_number),
                column=column_index,
            ).fill = fill

        reason_row = int(job.header_row) - 1
        if reason_row > 0 and label:
            cell = worksheet.cell(row=reason_row, column=column_index)
            cell.value = label
            cell.font = Font(bold=True, size=8, italic=True)
            cell.fill = fill


def _max_output_column(outcome: RunAndWriteResult, worksheet_max: int) -> int:
    candidates = [
        int(worksheet_max),
        int(outcome.write_report.average_column),
        int(outcome.write_report.analog_start_column)
        + max(0, int(outcome.write_report.analog_column_count) - 1),
    ]
    if outcome.write_report.tkp_start_column is not None:
        candidates.append(int(outcome.write_report.tkp_start_column) + 2)
    return max(candidates)


def _build_column_schema(
    worksheet,
    header_row: int,
    max_column: int,
    outcome: RunAndWriteResult,
) -> list[dict[str, object]]:
    analog_by_column = {
        int(item.column): item
        for item in outcome.write_report.analog_columns
    }
    tkp_start = outcome.write_report.tkp_start_column
    average_column = int(outcome.write_report.average_column)
    schema: list[dict[str, object]] = []

    for column in range(1, max_column + 1):
        raw_header = (
            worksheet.cell(row=header_row, column=column).value
            if header_row > 0
            else None
        )
        label = _display_text(raw_header) or get_column_letter(column)
        sublabel = ""
        kind = "source"
        task_number = ""

        analog = analog_by_column.get(column)
        if analog is not None:
            kind = "analog"
            task_number = str(analog.task_id)
            sublabel = str(analog.region or "")
            label = task_number or label
        elif tkp_start is not None and column == int(tkp_start):
            kind = "tkp_price"
        elif tkp_start is not None and column == int(tkp_start) + 1:
            kind = "tkp_name"
        elif tkp_start is not None and column == int(tkp_start) + 2:
            kind = "tkp_task"
        elif column == average_column:
            kind = "average"

        schema.append(
            {
                "index": column,
                "letter": get_column_letter(column),
                "label": label,
                "sublabel": sublabel,
                "kind": kind,
                "task_number": task_number,
                "editable": True,
            }
        )
    return schema


def _display_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _json_cell_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _request_full_recalculation(workbook) -> None:
    calculation = workbook.calculation
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True
    calculation.calcOnSave = True
    calculation.calcCompleted = False

"""Flexible estimate reading (Step 4c): pick the sheet, resolve columns.

Wraps the deterministic layout resolver (core/layout.py) around the proven
template reader so arbitrary uploaded files degrade gracefully instead of
returning a silent empty result:

- sheet selection: a single "estimate" sheet by name is used directly; when
  several sheets could qualify, `MultipleSheetsError` lists the candidates
  (for the UI to offer a choice); an explicit sheet title can be forced;
- reading: the template layout (fixed Settings columns) is tried first
  because it is what the real client files use and is already verified; if it
  yields no rows, detected columns are used (blank-row tolerant);
- failure: if neither yields rows, `KeyDataNotFoundError` carries the layout
  report so the user learns exactly which key field was not found.
"""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from core.excel_io import (
    ESTIMATE_SHEET_PART,
    Settings,
    find_estimate_sheet,
    read_estimate_rows_by_columns,
    read_estimate_rows_from_worksheet,
)
from core.layout import (
    FIELD_BASE_PRICE,
    FIELD_CODE,
    FIELD_UNIT,
    FIELD_WORK_NAME,
    LayoutConfig,
    LayoutResult,
    format_layout_report,
    load_layout_config,
    resolve_layout,
    select_sheets,
)
from core.matching import EstimateRow

METHOD_TEMPLATE = "template"
METHOD_DETECTED = "detected"


class EstimateReadError(Exception):
    """Base error for estimate reading."""


class KeyDataNotFoundError(EstimateReadError):
    """Required data (code/unit/base price) could not be located."""

    def __init__(self, sheet_title: str, report: str) -> None:
        self.sheet_title = sheet_title
        self.report = report
        super().__init__(f"key data not found on sheet '{sheet_title}'")


class MultipleSheetsError(EstimateReadError):
    """Several sheets could hold the estimate; the user must choose one."""

    def __init__(self, candidates: list[str]) -> None:
        self.candidates = candidates
        super().__init__("multiple candidate sheets; a choice is required")


@dataclass(frozen=True)
class EstimateData:
    """Result of a flexible estimate read."""

    sheet_title: str
    positioned_rows: list[tuple[int, EstimateRow]]
    header_row: int
    code_column: int
    unit_column: int
    work_name_column: int
    base_price_column: int
    method: str
    layout: LayoutResult


def default_columns(settings: Settings) -> dict[str, int]:
    """Template column map used as the resolver's fallback defaults."""
    return {
        FIELD_WORK_NAME: settings.col_smeta_work_name,
        FIELD_UNIT: settings.col_smeta_unit,
        FIELD_CODE: settings.col_search,
        FIELD_BASE_PRICE: settings.col_f,
    }


def load_estimate(
    estimate_path: str | Path,
    *,
    settings: Settings | None = None,
    config: LayoutConfig | None = None,
    selected_sheet_title: str | None = None,
) -> EstimateData:
    """Read the estimate flexibly, raising a clear error when data is missing."""
    active_settings = Settings() if settings is None else settings
    layout_config = load_layout_config() if config is None else config

    workbook = load_workbook(estimate_path, data_only=False)
    try:
        worksheet, title = _choose_sheet(
            workbook,
            active_settings,
            layout_config,
            selected_sheet_title,
        )
        layout = resolve_layout(
            worksheet,
            layout_config,
            default_columns=default_columns(active_settings),
        )

        template_rows = read_estimate_rows_from_worksheet(worksheet, active_settings)
        if template_rows:
            return EstimateData(
                sheet_title=title,
                positioned_rows=template_rows,
                header_row=0,
                code_column=active_settings.col_search,
                unit_column=active_settings.col_smeta_unit,
                work_name_column=active_settings.col_smeta_work_name,
                base_price_column=active_settings.col_f,
                method=METHOD_TEMPLATE,
                layout=layout,
            )

        if layout.ok and layout.header_row > 0:
            detected_rows = read_estimate_rows_by_columns(
                worksheet,
                header_row=layout.header_row,
                code_column=layout.column(FIELD_CODE),
                unit_column=layout.column(FIELD_UNIT),
                work_name_column=_resolved_or(
                    layout, FIELD_WORK_NAME, active_settings.col_smeta_work_name
                ),
                base_price_column=layout.column(FIELD_BASE_PRICE),
                max_blank_run=layout_config.max_blank_run,
            )
            if detected_rows:
                return EstimateData(
                    sheet_title=title,
                    positioned_rows=detected_rows,
                    header_row=layout.header_row,
                    code_column=layout.column(FIELD_CODE),
                    unit_column=layout.column(FIELD_UNIT),
                    work_name_column=_resolved_or(
                        layout, FIELD_WORK_NAME, active_settings.col_smeta_work_name
                    ),
                    base_price_column=layout.column(FIELD_BASE_PRICE),
                    method=METHOD_DETECTED,
                    layout=layout,
                )

        raise KeyDataNotFoundError(title, format_layout_report(layout))
    finally:
        workbook.close()


def _choose_sheet(workbook, settings, config, selected_title):
    if selected_title is not None:
        if selected_title not in workbook.sheetnames:
            raise EstimateReadError(f"sheet '{selected_title}' not found")
        return workbook[selected_title], selected_title

    part = ESTIMATE_SHEET_PART.casefold()
    name_matches = [ws for ws in workbook.worksheets if part in ws.title.casefold()]
    if len(name_matches) == 1:
        return name_matches[0], name_matches[0].title

    selection = select_sheets(
        workbook,
        config,
        default_columns=default_columns(settings),
    )
    if selection.needs_user_choice:
        raise MultipleSheetsError([candidate.title for candidate in selection.candidates])
    if selection.chosen:
        title = selection.chosen[0].title
        return workbook[title], title

    worksheet = find_estimate_sheet(workbook, settings)
    return worksheet, worksheet.title


def _resolved_or(layout: LayoutResult, field: str, fallback: int) -> int:
    column = layout.column(field)
    return fallback if column is None else column

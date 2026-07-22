"""Application service: run matching and write the result into a copy.

Ties the tested pieces together end to end: read catalog + estimate flexibly
(Step 4c: sheet selection, template-or-detected columns, clear errors),
resolve the regional coefficient by label (R16), run matching, then write the
structured result into a `WA` copy of the estimate workbook. One function
controls the row positions, the column plan, and the run result, so they stay
aligned for the writer.
"""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.services.catalog_source import CatalogNotAvailableError, load_catalog_for_run
from app.services.read_estimate import METHOD_TEMPLATE, EstimateData, load_estimate
from app.services.run_matching import EstimateRowResult, MatchingRunResult, run_matching
from core.excel_io import Settings
from core.excel_writer import WriteReport, WriterColumns, resolve_kr_column, write_run_result
from core.exclusions import NameExclusionRule, TaskColorEntry, TaskHighlightReason
from core.layout import FIELD_SECTION, LayoutConfig, load_layout_config, resolve_regional_coefficient
from core.macro_workbook import load_default_macro_settings
from core.risk import GesnException
from core.storage import (
    FlaggedRiskSnapshot,
    connect,
    database_is_available,
    init_database,
    load_gesn_exceptions,
    persist_flagged_risks,
)

WA_SUFFIX = " WA"


@dataclass(frozen=True)
class RunAndWriteResult:
    """End-to-end outcome of a matching run plus its Excel export."""

    result: MatchingRunResult
    write_report: WriteReport
    regional_coefficient: float
    coefficient_method: str
    output_path: Path
    sheet_title: str
    read_method: str
    macro_workbook: Path | None = None
    name_exclusion_rule_count: int = 0
    catalog_source: str = ""
    catalog_row_count: int = 0


@dataclass(frozen=True)
class EstimatePreview:
    """Region + coefficient as detected, shown to the user before the real run.

    Read-only: resolves the sheet and the regional coefficient/region exactly
    the way `run_and_write` will, but does not touch the catalog or run any
    matching. Used by the web UI's confirmation screen (2026-07 rule:
    "защита от дурака" -- the detected region/coefficient must be shown and
    editable before processing, not silently applied).
    """

    sheet_title: str
    region_text: str | None
    coefficient_value: float
    coefficient_method: str


def preview_estimate_context(
    estimate_path: str | Path,
    *,
    settings: Settings | None = None,
    selected_sheet_title: str | None = None,
    layout_config: LayoutConfig | None = None,
) -> EstimatePreview:
    """Resolve sheet + region + coefficient without running the full match.

    Raises the same `MultipleSheetsError` / `KeyDataNotFoundError` /
    `EstimateReadError` as `run_and_write` for the caller to handle (e.g. the
    web UI's existing sheet-choice screen).
    """
    active_settings = Settings() if settings is None else settings
    config = load_layout_config() if layout_config is None else layout_config

    estimate = load_estimate(
        estimate_path,
        settings=active_settings,
        config=config,
        selected_sheet_title=selected_sheet_title,
    )
    coefficient, method, region = _resolve_coefficient(
        estimate_path, estimate.sheet_title, None, config
    )
    return EstimatePreview(
        sheet_title=estimate.sheet_title,
        region_text=region,
        coefficient_value=coefficient,
        coefficient_method=method,
    )


def run_and_write(
    catalog_path: str | Path | None,
    estimate_path: str | Path,
    output_path: str | Path | None = None,
    *,
    settings: Settings | None = None,
    catalog_source_name: str = "main",
    database_path: str | Path | None = None,
    selected_sheet_title: str | None = None,
    name_exclusion_rules: list[NameExclusionRule] | None = None,
    task_color_entries: list[TaskColorEntry] | None = None,
    task_highlight_reasons: list[TaskHighlightReason] | None = None,
    gesn_exceptions: dict[str, GesnException] | None = None,
    demontazh_filter_enabled: bool = True,
    price_spread_limit: float | None = None,
    regional_coefficient: float | None = None,
    target_region: str | None = None,
    layout_config: LayoutConfig | None = None,
) -> RunAndWriteResult:
    """Run matching over the files and write the result into a `WA` copy."""
    active_settings = Settings() if settings is None else settings
    config = load_layout_config() if layout_config is None else layout_config
    spread_limit = (
        active_settings.price_spread_limit
        if price_spread_limit is None
        else price_spread_limit
    )

    exclusion_rules, task_colors, macro_workbook = _resolve_macro_settings(
        name_exclusion_rules,
        task_color_entries,
    )

    resolved_exceptions = _resolve_gesn_exceptions(gesn_exceptions, database_path)

    catalog = load_catalog_for_run(
        catalog_path,
        source_name=catalog_source_name,
        database_path=database_path,
        settings=active_settings,
    )
    estimate = load_estimate(
        estimate_path,
        settings=active_settings,
        config=config,
        selected_sheet_title=selected_sheet_title,
    )
    row_numbers = [row_number for row_number, _ in estimate.positioned_rows]
    estimate_rows = [estimate_row for _, estimate_row in estimate.positioned_rows]

    coefficient, coefficient_method, resolved_region = _resolve_coefficient(
        estimate_path,
        estimate.sheet_title,
        regional_coefficient,
        config,
    )
    effective_region = target_region if target_region is not None else resolved_region

    result = run_matching(
        catalog.rows,
        estimate_rows,
        name_exclusion_rules=exclusion_rules,
        gesn_exceptions=resolved_exceptions,
        demontazh_filter_enabled=demontazh_filter_enabled,
        price_spread_limit=spread_limit,
        regional_coefficient=coefficient,
    )

    _persist_risk_log(
        result,
        row_numbers,
        coefficient,
        database_path=database_path,
    )

    destination = _resolve_output_path(estimate_path, output_path)
    writer_columns = _writer_columns(
        estimate,
        active_settings,
        estimate_path,
        estimate.sheet_title,
    )
    write_report = write_run_result(
        estimate_path,
        destination,
        result,
        row_numbers,
        columns=writer_columns,
        regional_coefficient=coefficient,
        sheet_title=estimate.sheet_title,
        task_color_entries=task_colors,
        task_highlight_reasons=task_highlight_reasons,
        target_region=effective_region,
    )

    return RunAndWriteResult(
        result=result,
        write_report=write_report,
        regional_coefficient=coefficient,
        coefficient_method=coefficient_method,
        output_path=destination,
        sheet_title=estimate.sheet_title,
        read_method=estimate.method,
        macro_workbook=macro_workbook,
        name_exclusion_rule_count=len(exclusion_rules),
        catalog_source=catalog.source_label,
        catalog_row_count=catalog.row_count,
    )


def _resolve_macro_settings(
    name_exclusion_rules: list[NameExclusionRule] | None,
    task_color_entries: list[TaskColorEntry] | None,
) -> tuple[list[NameExclusionRule], list[TaskColorEntry], Path | None]:
    if name_exclusion_rules is not None and task_color_entries is not None:
        return name_exclusion_rules, task_color_entries, None

    macro = load_default_macro_settings()
    rules = macro.name_exclusion_rules if name_exclusion_rules is None else name_exclusion_rules
    colors = macro.task_color_entries if task_color_entries is None else task_color_entries
    macro_path = macro.workbook_path if (name_exclusion_rules is None or task_color_entries is None) else None
    return rules, colors, macro_path


def _writer_columns(
    estimate: EstimateData,
    settings: Settings,
    estimate_path: str | Path,
    sheet_title: str,
) -> WriterColumns:
    workbook = load_workbook(estimate_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_title]
        section = estimate.layout.column(FIELD_SECTION)
        if section is None:
            section = settings.col_section
            analog_start = settings.col_analog_start
        else:
            analog_start = section + 1

        code_column = estimate.code_column
        code_kr = resolve_kr_column(
            worksheet,
            estimate.header_row,
            code_column,
            settings.col_kr,
            settings.col_search,
        )
        base_price = (
            settings.col_f
            if estimate.method == METHOD_TEMPLATE
            else estimate.base_price_column
        )
    finally:
        workbook.close()

    return WriterColumns(
        base_price=base_price,
        code=code_column,
        code_kr=code_kr,
        section=section,
        analog_start=analog_start,
        header_row=estimate.header_row,
    )


def _resolve_coefficient(
    estimate_path: str | Path,
    sheet_title: str,
    explicit: float | None,
    config: LayoutConfig,
) -> tuple[float, str, str | None]:
    workbook = load_workbook(estimate_path, data_only=True)
    try:
        worksheet = workbook[sheet_title]
        resolution = resolve_regional_coefficient(worksheet, config)
    finally:
        workbook.close()

    if explicit is not None:
        return explicit, "explicit", resolution.region

    return resolution.value, resolution.method, resolution.region


def _resolve_output_path(
    estimate_path: str | Path,
    output_path: str | Path | None,
) -> Path:
    if output_path is not None:
        return Path(output_path)

    source = Path(estimate_path)
    return source.with_name(f"{source.stem}{WA_SUFFIX}{source.suffix}")


def _resolve_gesn_exceptions(
    gesn_exceptions: dict[str, GesnException] | None,
    database_path: str | Path | None,
) -> dict[str, GesnException]:
    if gesn_exceptions is not None:
        return gesn_exceptions
    if not database_is_available(database_path):
        return {}

    connection = connect(database_path)
    try:
        init_database(connection)
        return load_gesn_exceptions(connection)
    finally:
        connection.close()


def _persist_risk_log(
    result: MatchingRunResult,
    row_numbers: list[int],
    coefficient: float,
    *,
    database_path: str | Path | None,
) -> None:
    if not database_is_available(database_path):
        return

    snapshots = [
        _flagged_snapshot(row_number, row, coefficient)
        for row_number, row in zip(row_numbers, result.rows)
        if row.risk_result.is_flagged and row.exception_key
    ]
    if not snapshots:
        return

    connection = connect(database_path)
    try:
        init_database(connection)
        persist_flagged_risks(connection, snapshots)
    finally:
        connection.close()


def _flagged_snapshot(
    row_number: int,
    row: EstimateRowResult,
    coefficient: float,
) -> FlaggedRiskSnapshot:
    min_price, max_price = _risk_min_max(row, coefficient)
    return FlaggedRiskSnapshot(
        exception_key=row.exception_key,
        reason=row.risk_result.reason,
        code=row.norm_code,
        unit=row.norm_unit,
        min_price=min_price,
        max_price=max_price,
        ratio=row.risk_result.ratio or None,
        recommended_price=row.recommended_price,
        estimate_row=row_number,
    )


def _risk_min_max(
    row: EstimateRowResult,
    coefficient: float,
) -> tuple[float | None, float | None]:
    risk = row.risk_result
    entries = risk.flagged_entries

    if risk.min_entry is not None:
        min_price = risk.min_entry.price
    elif entries:
        min_price = min(entry.price for entry in entries)
    else:
        min_price = None

    if risk.max_entry is not None:
        max_price = risk.max_entry.price
    elif entries:
        max_price = max(entry.price for entry in entries)
    else:
        max_price = None

    scaled_min = None if min_price is None else min_price * coefficient
    scaled_max = None if max_price is None else max_price * coefficient
    return scaled_min, scaled_max

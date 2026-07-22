from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from core.storage import (
    connect,
    count_tkp_items,
    import_tkp_catalog_workbook,
    init_database,
    list_tkp_catalog_page,
    list_tkp_items,
    list_tkp_sources,
)
from core.tkp_ingest import FILE_CATALOG_HEADERS, FILE_CATALOG_SHEET, WOR_CATALOG_HEADERS, WOR_CATALOG_SHEET

RUN_ID = "20260624_172128"
FILE_PATH = "C:\\KL20\\sample.xlsx"
FILE_NAME = "sample.xlsx"
ITEM_NAME = "\u0420\u0430\u0437\u0431\u043e\u0440\u043a\u0430 \u0442\u0435\u043f\u043b\u043e\u0432\u043e\u0439 \u0438\u0437\u043e\u043b\u044f\u0446\u0438\u0438"
WINNER = '\u041e\u041e\u041e "\u041f\u0440\u0438\u043c\u0435\u0440"'


def _file_row(modified: datetime, **overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID, "FilePath": FILE_PATH, "FileName": FILE_NAME,
        "ModifiedDate": modified, "SheetName": "\u041a\u041b 2.0",
        "ParseStatus": "OK", "ParseMessage": "Parsed.",
        "UsedRows": 10, "UsedCols": 10, "WorStartRow": 1, "WorEndRow": 2,
        "WorEndMethod": "advance_limit", "WorPositionCount": 1,
        "WorSectionCount": 1, "ParticipantCount": 1, "WinnerGroupIndex": 1,
        "WinnerStartCol": 1, "WinnerStartColLetter": "A",
        "WinnerMethod": "block10_recommended", "WinnerName": WINNER,
        "WinnerINN": "1234567890", "WinnerUIN": "uin-1",
        "WinnerTaskNoBK": 111, "WinnerHeader1": "", "WinnerHeader2": "",
        "WinnerHeader3": "", "WinnerHeader4": "", "WinnerBlockName": WINNER,
        "WinnerBlockUIN": "", "WinnerBlockTotalVat": 1000.0,
        "WinnerBlockReason": "", "WinnerBlockSource": "block10",
        "TaskNo": "111", "RequestDate": datetime(2026, 5, 1), "Version": 1,
        "Customer": "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a", "GeneralContractor": "",
        "ProcedureName": "\u041e\u0431\u044a\u0435\u043a\u0442", "WinnerTotalNoVat": 900.0,
        "WinnerTotalVat": 1080.0, "RnmcTotalNoVat": 950.0,
    }
    values.update(overrides)
    return tuple(values[header] for header in FILE_CATALOG_HEADERS)


def _wor_row(price: float, **overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID, "FilePath": FILE_PATH, "FileName": FILE_NAME,
        "SheetName": "\u041a\u041b 2.0", "SourceRow": 5, "SectionCode": "1.",
        "SectionName": "\u0420\u0430\u0437\u0434\u0435\u043b 1", "SubsectionName": "",
        "ItemCode": "1.1", "ItemName": ITEM_NAME, "Unit": "\u043c2",
        "Qty": 10.0, "QtySourceText": "10", "RnmcUnitPriceNoVat": 100.0,
        "RnmcLineTotalNoVat": 1000.0, "WinnerUnitPriceNoVat": price,
        "WinnerLineTotalNoVat": price * 10, "WinnerName": WINNER,
        "WinnerINN": "1234567890", "WinnerUIN": "uin-1", "WinnerGroupIndex": 1,
        "WinnerStartCol": 1, "WinnerStartColLetter": "A",
        "WinnerUnitHeader": "", "WinnerTotalHeader": "", "TaskNo": "111",
        "RequestDate": datetime(2026, 5, 1), "Version": 1,
        "Customer": "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a", "GeneralContractor": "",
        "ProcedureName": "\u041e\u0431\u044a\u0435\u043a\u0442",
        "WinnerMethod": "block10_recommended", "WinnerBlockName": WINNER,
        "WinnerBlockUIN": "", "WinnerBlockTotalVat": 1000.0,
        "WinnerBlockReason": "", "WinnerBlockSource": "block10",
    }
    values.update(overrides)
    return tuple(values[header] for header in WOR_CATALOG_HEADERS)


def _make_workbook(path: Path, file_rows: list[tuple], wor_rows: list[tuple]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    files_sheet = workbook.create_sheet(FILE_CATALOG_SHEET)
    files_sheet.append(FILE_CATALOG_HEADERS)
    for row in file_rows:
        files_sheet.append(row)
    wor_sheet = workbook.create_sheet(WOR_CATALOG_SHEET)
    wor_sheet.append(WOR_CATALOG_HEADERS)
    for row in wor_rows:
        wor_sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def test_import_stores_source_and_items(tmp_path: Path) -> None:
    workbook_path = _make_workbook(
        tmp_path / "catalog.xlsm",
        [_file_row(datetime(2026, 6, 17, 10, 0, 0))],
        [_wor_row(500.0)],
    )
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_tkp_catalog_workbook(connection, workbook_path)

        assert result.files_seen == 1
        assert result.files_imported == 1
        assert result.files_updated == 0
        assert result.files_skipped == 0
        assert result.items_imported == 1
        assert count_tkp_items(connection) == 1

        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].file_name == FILE_NAME
        assert sources[0].item_count == 1
        assert sources[0].winner_name == WINNER

        items = list_tkp_items(connection)
        assert len(items) == 1
        assert items[0].item_name == ITEM_NAME
        assert items[0].qty_source_text == "10"
        assert items[0].rnmc_unit_price_no_vat == 100.0
        assert items[0].rnmc_line_total_no_vat == 1000.0
        assert items[0].winner_unit_price_no_vat == 500.0
        assert items[0].winner_line_total_no_vat == 5000.0
        assert items[0].winner_group_index == 1
        assert items[0].winner_start_col == 1
        assert items[0].winner_start_col_letter == "A"
        assert items[0].winner_unit_header == ""
        assert items[0].winner_total_header == ""
        assert items[0].version == "1"
        assert items[0].winner_method == "block10_recommended"
        assert items[0].winner_block_name == WINNER
        assert items[0].winner_block_uin == ""
        assert items[0].winner_block_total_vat == 1000.0
        assert items[0].winner_block_reason == ""
        assert items[0].source_file_name == FILE_NAME
    finally:
        connection.close()


def test_catalog_page_filters_sorts_and_paginates_items(tmp_path: Path) -> None:
    rows = [
        _wor_row(
            float(index),
            SourceRow=index + 2,
            ItemName=f"Работа {index:02d}",
            TaskNo="TASK-A" if index == 30 else "TASK-B",
            WinnerName="Нужный победитель" if index == 30 else "Другой победитель",
        )
        for index in range(1, 31)
    ]
    workbook_path = _make_workbook(
        tmp_path / "catalog.xlsm",
        [_file_row(datetime(2026, 6, 17, 10, 0, 0))],
        rows,
    )
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_tkp_catalog_workbook(connection, workbook_path)

        first_page = list_tkp_catalog_page(
            connection,
            page=1,
            page_size=25,
            sort="winner_unit_price_no_vat",
            direction="asc",
        )
        assert first_page.total_rows == 30
        assert first_page.total_pages == 2
        assert len(first_page.rows) == 25
        assert first_page.rows[0].winner_unit_price_no_vat == 1.0

        filtered = list_tkp_catalog_page(
            connection,
            filters={"q": "Нужный", "task_no": "TASK-A"},
        )
        assert filtered.total_rows == 1
        assert filtered.rows[0].item_name == "Работа 30"
    finally:
        connection.close()


def test_reimporting_the_same_workbook_skips_unchanged_files(tmp_path: Path) -> None:
    workbook_path = _make_workbook(
        tmp_path / "catalog.xlsm",
        [_file_row(datetime(2026, 6, 17, 10, 0, 0))],
        [_wor_row(500.0)],
    )
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_tkp_catalog_workbook(connection, workbook_path)
        second = import_tkp_catalog_workbook(connection, workbook_path)

        assert second.files_imported == 0
        assert second.files_skipped == 1
        assert count_tkp_items(connection) == 1
    finally:
        connection.close()


def test_reimport_backfills_rows_from_previous_tkp_schema(tmp_path: Path) -> None:
    workbook_path = _make_workbook(
        tmp_path / "catalog.xlsm",
        [_file_row(datetime(2026, 6, 17, 10, 0, 0))],
        [_wor_row(500.0)],
    )
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_tkp_catalog_workbook(connection, workbook_path)
        connection.execute("UPDATE tkp_sources SET details_version = 0")
        connection.execute("UPDATE tkp_items SET qty_source_text = ''")
        connection.commit()

        result = import_tkp_catalog_workbook(connection, workbook_path)

        assert result.files_updated == 1
        assert result.files_skipped == 0
        assert list_tkp_items(connection)[0].qty_source_text == "10"
    finally:
        connection.close()


def test_updated_file_replaces_old_items(tmp_path: Path) -> None:
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)

        first_path = _make_workbook(
            tmp_path / "v1.xlsm",
            [_file_row(datetime(2026, 6, 17, 10, 0, 0))],
            [_wor_row(500.0)],
        )
        import_tkp_catalog_workbook(connection, first_path)

        # Same file, refreshed content (new ModifiedDate + a different price).
        second_path = _make_workbook(
            tmp_path / "v2.xlsm",
            [_file_row(datetime(2026, 7, 1, 9, 0, 0))],
            [_wor_row(600.0)],
        )
        result = import_tkp_catalog_workbook(connection, second_path)

        assert result.files_updated == 1
        assert count_tkp_items(connection) == 1

        items = list_tkp_items(connection)
        assert items[0].winner_unit_price_no_vat == 600.0

        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].modified_date == datetime(2026, 7, 1, 9, 0, 0).isoformat()
    finally:
        connection.close()


def test_skipped_source_file_has_no_items(tmp_path: Path) -> None:
    workbook_path = _make_workbook(
        tmp_path / "catalog.xlsm",
        [_file_row(
            datetime(2026, 6, 17, 10, 0, 0),
            ParseStatus="Skipped", ParseMessage="KL20 sheet not found.",
            WinnerName="", WinnerINN="",
        )],
        [],
    )
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_tkp_catalog_workbook(connection, workbook_path)

        assert result.items_imported == 0
        sources = list_tkp_sources(connection)
        assert sources[0].parse_status == "Skipped"
        assert sources[0].item_count == 0
    finally:
        connection.close()


def test_schema_upgrade_adds_complete_tkp_detail_columns(tmp_path: Path) -> None:
    connection = connect(tmp_path / "estimate_ai.db")
    try:
        connection.executescript(
            """
            CREATE TABLE schema_migrations (version INTEGER PRIMARY KEY);
            INSERT INTO schema_migrations(version) VALUES (9);
            CREATE TABLE tkp_sources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT NOT NULL,
                modified_date TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE tkp_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_id INTEGER NOT NULL,
                item_name TEXT NOT NULL
            );
            """
        )

        init_database(connection)

        item_columns = {
            str(row["name"])
            for row in connection.execute("PRAGMA table_info(tkp_items)")
        }
        source_columns = {
            str(row["name"])
            for row in connection.execute("PRAGMA table_info(tkp_sources)")
        }
        assert {
            "qty_source_text", "rnmc_line_total_no_vat", "winner_group_index",
            "winner_start_col", "winner_start_col_letter", "winner_unit_header",
            "winner_total_header", "version", "winner_method", "winner_block_name",
            "winner_block_uin", "winner_block_total_vat", "winner_block_reason",
        } <= item_columns
        assert "details_version" in source_columns
    finally:
        connection.close()

"""Tests for authenticated users and durable post-processing."""

from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.services.postprocessing import build_postprocessed_workbook, resolve_preview_row_values
from app.web.app import create_app
from app.web.workspace import _render_workspace
from core.storage.accounts import (
    ROLE_ADMIN,
    ROLE_SPECIALIST,
    authenticate_user,
    create_session,
    create_user,
    user_from_session,
)
from core.storage.connection import connect, init_database
from core.storage.workspace import (
    REQUEST_APPROVED,
    REQUEST_BLUE_TASK,
    create_change_request,
    get_processing_job,
    list_cell_overrides,
    list_change_requests,
    list_processing_rows,
    replace_processing_rows,
    review_change_request,
    save_cell_override,
    upsert_processing_job,
)


def _seed_users(connection):
    admin_id = create_user(
        connection,
        full_name="Admin User",
        login="admin.user",
        password="admin-password",
        role=ROLE_ADMIN,
    )
    specialist_id = create_user(
        connection,
        full_name="Specialist User",
        login="specialist.user",
        password="specialist-password",
        role=ROLE_SPECIALIST,
    )
    return admin_id, specialist_id


def _seed_job(connection, tmp_path: Path, owner_user_id: int) -> str:
    job_id = "job-001"
    output_path = tmp_path / "base.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Estimate"
    worksheet.cell(row=1, column=1, value="Price")
    worksheet.cell(row=2, column=1, value=10.0)
    workbook.save(output_path)
    workbook.close()

    upsert_processing_job(
        connection,
        job_id=job_id,
        owner_user_id=owner_user_id,
        estimate_filename="estimate.xlsx",
        source_path=str(tmp_path / "source.xlsx"),
        output_path=str(output_path),
        sheet_title="Estimate",
        header_row=1,
        coefficient=1.0,
        region="",
        use_tkp_analogs=False,
        status="ready",
        total_rows=1,
        matched_rows=1,
        flagged_rows=0,
        tkp_matched_rows=0,
        column_schema=[
            {
                "index": 1,
                "letter": "A",
                "label": "Price",
                "sublabel": "",
                "kind": "source",
                "task_number": "",
                "editable": True,
            }
        ],
    )
    replace_processing_rows(
        connection,
        job_id,
        [
            (
                1,
                2,
                [10.0],
                {
                    "has_analogs": True,
                    "risk": False,
                    "has_tkp": False,
                    "status": "matched",
                    "reason": "matched",
                },
            )
        ],
    )
    return job_id


def test_v14_database_is_upgraded_with_auth_workspace_tables() -> None:
    connection = connect(":memory:")
    try:
        connection.execute(
            "CREATE TABLE schema_migrations (version INTEGER PRIMARY KEY)"
        )
        connection.execute(
            "INSERT INTO schema_migrations(version) VALUES (14)"
        )
        connection.commit()

        init_database(connection)

        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        version = connection.execute(
            "SELECT MAX(version) FROM schema_migrations"
        ).fetchone()[0]
    finally:
        connection.close()

    assert int(version) == 15
    assert "app_users" in tables
    assert "processing_jobs" in tables
    assert "specialist_change_requests" in tables


def test_schema_contains_account_and_workspace_tables() -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
    finally:
        connection.close()

    assert "app_users" in tables
    assert "user_sessions" in tables
    assert "processing_jobs" in tables
    assert "processing_rows" in tables
    assert "processing_cell_overrides" in tables
    assert "processing_edit_events" in tables
    assert "specialist_change_requests" in tables
    assert "audit_events" in tables


def test_preview_resolves_simple_source_formula_chain(tmp_path: Path) -> None:
    source_path = tmp_path / "source-formulas.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Estimate"
    worksheet["A2"] = 1
    worksheet["A3"] = "=A2+1"
    worksheet["A4"] = "=A3 + 1"
    workbook.save(source_path)
    workbook.close()

    job = SimpleNamespace(
        source_path=str(source_path),
        sheet_title="Estimate",
        column_schema=[{"index": 1, "kind": "source"}],
    )
    rows = [
        SimpleNamespace(id=1, excel_row_number=2, values=[1]),
        SimpleNamespace(id=2, excel_row_number=3, values=["=A2+1"]),
        SimpleNamespace(id=3, excel_row_number=4, values=["=A3 + 1"]),
    ]

    preview = resolve_preview_row_values(job, rows)

    assert preview[1][0] == 1
    assert preview[2][0] == 2
    assert preview[3][0] == 3
    assert rows[1].values[0] == "=A2+1"
    assert rows[2].values[0] == "=A3 + 1"


def test_workspace_wraps_text_when_columns_are_narrowed() -> None:
    job = SimpleNamespace(
        id="job-wrap",
        estimate_filename="estimate.xlsx",
        owner_name="Specialist User",
        sheet_title="Estimate",
        region="",
        total_rows=1,
        matched_rows=1,
        flagged_rows=0,
        tkp_matched_rows=0,
    )
    user = SimpleNamespace(
        id=2,
        full_name="Specialist User",
        role=ROLE_SPECIALIST,
    )

    page = _render_workspace(job, user)

    assert "overflow-wrap:anywhere" in page
    assert "Math.max(48" in page
    assert "function currentRowHeight(columns)" in page
    assert "makeRow(state.rows[i], columns, rowHeight)" in page
    assert "grid-template-rows:46px 14px 26px" in page
    assert "head-stack" in page
    assert "stack.append(head, sub, filter)" in page
    assert 'id="gridHScroll"' in page
    assert "syncHorizontalScrollbar()" in page
    assert 'id="withAnalogOnly"' in page
    assert "row.metadata.has_analogs || row.metadata.has_tkp" in page


def test_user_password_and_session_round_trip() -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        admin_id, _specialist_id = _seed_users(connection)

        assert authenticate_user(connection, "ADMIN.USER", "wrong") is None
        authenticated = authenticate_user(
            connection,
            "ADMIN.USER",
            "admin-password",
        )
        assert authenticated is not None
        assert authenticated.id == admin_id
        assert authenticated.role == ROLE_ADMIN

        token = create_session(connection, admin_id, lifetime_days=7)
        from_session = user_from_session(connection, token)
        assert from_session is not None
        assert from_session.id == admin_id
    finally:
        connection.close()


def test_cell_autosave_is_idempotent_and_keeps_history(tmp_path: Path) -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        _admin_id, specialist_id = _seed_users(connection)
        job_id = _seed_job(connection, tmp_path, specialist_id)
        row = list_processing_rows(connection, job_id)[0]

        first = save_cell_override(
            connection,
            job_id=job_id,
            row_id=row.id,
            column_index=1,
            new_value=20.0,
            editor_user_id=specialist_id,
            client_change_id="client-change-1",
        )
        repeated = save_cell_override(
            connection,
            job_id=job_id,
            row_id=row.id,
            column_index=1,
            new_value=999.0,
            editor_user_id=specialist_id,
            client_change_id="client-change-1",
        )
        second = save_cell_override(
            connection,
            job_id=job_id,
            row_id=row.id,
            column_index=1,
            new_value=30.0,
            editor_user_id=specialist_id,
            client_change_id="client-change-2",
        )

        events = connection.execute(
            "SELECT COUNT(*) FROM processing_edit_events WHERE job_id = ?",
            (job_id,),
        ).fetchone()[0]
        overrides = list_cell_overrides(connection, job_id)
        job = get_processing_job(connection, job_id)
    finally:
        connection.close()

    assert first.id == repeated.id
    assert repeated.current_value == 20.0
    assert second.revision == 2
    assert second.current_value == 30.0
    assert events == 2
    assert len(overrides) == 1
    assert job is not None
    assert job.status == "draft"


def test_reviewed_excel_uses_overrides_without_modifying_base(tmp_path: Path) -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        _admin_id, specialist_id = _seed_users(connection)
        job_id = _seed_job(connection, tmp_path, specialist_id)
        row = list_processing_rows(connection, job_id)[0]
        save_cell_override(
            connection,
            job_id=job_id,
            row_id=row.id,
            column_index=1,
            new_value=25.5,
            editor_user_id=specialist_id,
            client_change_id="export-change",
        )

        final_path = build_postprocessed_workbook(connection, job_id)
        job = get_processing_job(connection, job_id)
        assert job is not None
        base_path = Path(job.output_path)
    finally:
        connection.close()

    base = load_workbook(base_path, data_only=False)
    reviewed = load_workbook(final_path, data_only=False)
    try:
        assert base["Estimate"]["A2"].value == 10.0
        assert reviewed["Estimate"]["A2"].value == 25.5
    finally:
        base.close()
        reviewed.close()


def test_blue_task_request_tracks_specialist_and_review(tmp_path: Path) -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        admin_id, specialist_id = _seed_users(connection)
        job_id = _seed_job(connection, tmp_path, specialist_id)
        row = list_processing_rows(connection, job_id)[0]

        request_id = create_change_request(
            connection,
            job_id=job_id,
            row_id=row.id,
            column_index=1,
            request_type=REQUEST_BLUE_TASK,
            task_number="TASK-41",
            comment="Needs manual review",
            submitted_by=specialist_id,
        )
        pending = list_change_requests(connection, status="pending")
        assert len(pending) == 1
        assert pending[0].id == request_id
        assert pending[0].submitted_by_name == "Specialist User"
        assert pending[0].estimate_filename == "estimate.xlsx"

        assert review_change_request(
            connection,
            request_id,
            status=REQUEST_APPROVED,
            reviewed_by=admin_id,
            review_comment="Approved",
        )
        approved = list_change_requests(connection, status=REQUEST_APPROVED)
    finally:
        connection.close()

    assert len(approved) == 1
    assert approved[0].reviewed_by_name == "Admin User"


def test_first_admin_setup_then_auth_guard(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "auth.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        assert client.get("/").status_code == 200

        setup = client.post(
            "/setup",
            data={
                "full_name": "Admin User",
                "login": "admin.user",
                "password": "admin-password",
            },
            follow_redirects=False,
        )
        assert setup.status_code == 303
        assert setup.headers["location"] == "/admin/users"
        assert "estimate_ai_session" in setup.headers.get("set-cookie", "")

        client.cookies.clear()
        guarded = client.get("/", follow_redirects=False)
        assert guarded.status_code == 303
        assert guarded.headers["location"].startswith("/login?next=")

        login = client.post(
            "/login",
            data={
                "login": "admin.user",
                "password": "admin-password",
                "next": "/admin/users",
            },
            follow_redirects=False,
        )
        assert login.status_code == 303
        assert login.headers["location"] == "/admin/users"

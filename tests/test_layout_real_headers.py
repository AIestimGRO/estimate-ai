"""Layout resolution against real estimate header wording (eV-grup format).

Encodes the actual headers observed in real client files (Step 4c bug: the
base-price column 'Cena edinicy rabot...' and the code column 'Perechen
GESN/FER/TER/KR' were not detected, and 'Kod razdela' was wrongly grabbed as
the code). Strings are ASCII-escaped per AGENTS.md rule 3.
"""

from openpyxl import Workbook

from app.services.read_estimate import METHOD_DETECTED, load_estimate
from core.layout import (
    load_layout_config,
    resolve_layout,
    resolve_regional_coefficient,
)

WORK = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
UNIT = "\u0415\u0434.\u0438\u0437\u043c."
QTY = "\u041a\u043e\u043b-\u0432\u043e"
BASE = "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421"
AVG = "\u0426\u0435\u043d\u0430 \u0441\u0440\u0435\u0434\u043d\u044f\u044f, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421"
CODEH = "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0422\u0415\u0420/\u041a\u0420"
KRH = "/\u041a\u0420"
SECTIONH = "\u041a\u043e\u0434 \u0440\u0430\u0437\u0434\u0435\u043b\u0430"
REGION_L = "\u0420\u0435\u0433\u0438\u043e\u043d:"
COEF_L = "\u041a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442:"
REGION_V = "\u041c\u043e\u0441\u043a\u043e\u0432\u0441\u043a\u0430\u044f \u043e\u0431\u043b."
CODEV = "\u0413\u042d\u0421\u041d\u043c10-10-001-02"
CODEV2 = "\u0413\u042d\u0421\u041d\u043c10-06-060-12"
UNITV = "\u0448\u0442"
NUM = "1"
ESTIMATE_TITLE = "\u041e\u0421_\u0421\u041c-\u041a"


def _write_header(worksheet, header_row, code_col, section_col):
    worksheet.cell(row=header_row, column=1, value="N")
    worksheet.cell(row=header_row, column=3, value=WORK)
    worksheet.cell(row=header_row, column=4, value=UNIT)
    worksheet.cell(row=header_row, column=5, value=QTY)
    worksheet.cell(row=header_row, column=6, value=BASE)
    worksheet.cell(row=header_row, column=7, value=AVG)
    worksheet.cell(row=header_row, column=code_col, value=CODEH)
    worksheet.cell(row=header_row, column=section_col, value=SECTIONH)


def test_resolves_real_headers_file_layout_a():
    workbook = Workbook()
    worksheet = workbook.active
    _write_header(worksheet, header_row=23, code_col=14, section_col=16)

    layout = resolve_layout(worksheet, load_layout_config())

    assert layout.ok
    assert layout.header_row == 23
    assert layout.column("work_name") == 3
    assert layout.column("unit") == 4
    assert layout.column("base_price") == 6
    assert layout.column("code") == 14
    assert layout.column("section") == 16


def test_resolves_real_headers_file_layout_b_offset():
    workbook = Workbook()
    worksheet = workbook.active
    _write_header(worksheet, header_row=26, code_col=15, section_col=17)
    worksheet.cell(row=26, column=16, value=KRH)

    layout = resolve_layout(worksheet, load_layout_config())

    assert layout.ok
    assert layout.header_row == 26
    assert layout.column("base_price") == 6
    assert layout.column("code") == 15
    assert layout.column("section") == 17


def test_code_not_stolen_by_kod_razdela():
    # 'Kod razdela' must go to section, not code.
    workbook = Workbook()
    worksheet = workbook.active
    _write_header(worksheet, header_row=1, code_col=14, section_col=16)

    layout = resolve_layout(worksheet, load_layout_config())

    assert layout.column("code") == 14
    assert layout.column("section") == 16
    assert layout.column("code") != layout.column("section")


def test_stacked_regional_coefficient_with_colon_labels():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=19, column=2, value=REGION_L)
    worksheet.cell(row=19, column=3, value=REGION_V)
    worksheet.cell(row=20, column=2, value=COEF_L)
    worksheet.cell(row=20, column=3, value=1.7)

    resolution = resolve_regional_coefficient(worksheet, load_layout_config())

    assert resolution.value == 1.7
    assert resolution.method == "labeled_region"


def test_load_estimate_detected_on_offset_real_layout(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ESTIMATE_TITLE
    worksheet.cell(row=19, column=2, value=REGION_L)
    worksheet.cell(row=19, column=3, value=REGION_V)
    worksheet.cell(row=20, column=2, value=COEF_L)
    worksheet.cell(row=20, column=3, value=1)
    _write_header(worksheet, header_row=26, code_col=15, section_col=17)
    worksheet.cell(row=26, column=16, value=KRH)
    # numbering row (must be skipped as non-working)
    for col in range(1, 18):
        worksheet.cell(row=27, column=col, value=str(col))
    # section title row (blank key cells -> tolerated)
    worksheet.cell(row=28, column=1, value="1")
    worksheet.cell(row=28, column=3, value="Section title")
    # two real data rows
    worksheet.cell(row=30, column=3, value="work one")
    worksheet.cell(row=30, column=4, value=UNITV)
    worksheet.cell(row=30, column=6, value=3361.57)
    worksheet.cell(row=30, column=15, value=CODEV)
    worksheet.cell(row=31, column=3, value="work two")
    worksheet.cell(row=31, column=4, value=UNITV)
    worksheet.cell(row=31, column=6, value=5917.33)
    worksheet.cell(row=31, column=15, value=CODEV2)

    path = tmp_path / "offset.xlsx"
    workbook.save(path)
    workbook.close()

    data = load_estimate(path)

    assert data.method == METHOD_DETECTED
    assert data.sheet_title == ESTIMATE_TITLE
    assert data.header_row == 26
    assert data.base_price_column == 6
    assert data.code_column == 15
    assert [row_number for row_number, _ in data.positioned_rows] == [30, 31]

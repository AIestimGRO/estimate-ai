"""Tests for automatic RNMC workbook metadata detection."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook

from app.services.rnmc_excel import analyze_rnmc_zip_row_preview, import_rnmc_zip_catalog_rows
from core.storage import connect, init_database
from core.storage.catalog import get_imported_file, list_imported_files


def test_import_detects_lsr_quarter_and_planned_dates(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Moscow/metadata.xlsx": _workbook_with_metadata_and_rows()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        record = next(item for item in list_imported_files(connection) if item.filename == "metadata.xlsx")
        stored = get_imported_file(connection, record.id)
    finally:
        connection.close()

    assert result.success_count == 1
    entry = result.entries[0]
    assert entry.lsr_quarter == "2026 Q1"
    assert entry.planned_start == "2026-01-10"
    assert entry.planned_finish == "2026-02-20"
    assert stored is not None
    assert stored.lsr_quarter == "2026 Q1"
    assert stored.planned_start == "2026-01-10"
    assert stored.planned_finish == "2026-02-20"


def test_row_preview_shows_metadata_without_database_writes(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/preview.xlsx": _workbook_with_metadata_and_rows()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
        records = list_imported_files(connection)
    finally:
        connection.close()

    assert records == []
    assert result.preview_ok_count == 1
    entry = result.entries[0]
    assert entry.lsr_quarter == "2026 Q1"
    assert entry.planned_start == "2026-01-10"
    assert entry.planned_finish == "2026-02-20"


def test_metadata_is_detected_even_when_table_is_missing(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Tula/no-table.xlsx": _workbook_with_metadata_only()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        preview = analyze_rnmc_zip_row_preview(connection, str(zip_path))
        import_result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        record = next(item for item in list_imported_files(connection) if item.filename == "no-table.xlsx")
    finally:
        connection.close()

    assert preview.no_table_count == 1
    assert preview.entries[0].lsr_quarter == "2026 Q1"
    assert preview.entries[0].planned_start == "2026-01-10"
    assert preview.entries[0].planned_finish == "2026-02-20"
    assert import_result.no_data_count == 1
    assert record.lsr_quarter == "2026 Q1"
    assert record.planned_start == "2026-01-10"
    assert record.planned_finish == "2026-02-20"


def _write_zip(path: Path, files: dict[str, bytes]) -> None:
    with ZipFile(path, "w") as archive:
        for name, data in files.items():
            archive.writestr(name, data)


def _workbook_with_metadata_and_rows() -> bytes:
    workbook = _base_metadata_workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Ед.изм.",
        "Кол-во",
        "ГЭСН/ФЕР/Перечень",
        "Цена",
    ])
    sheet.append([1, "Work A", "м", 10, "GESN01-01-001-01", 100])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    return _to_bytes(workbook)


def _workbook_with_metadata_only() -> bytes:
    return _to_bytes(_base_metadata_workbook())


def _base_metadata_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-META"
    sheet["A2"] = "Год Квартал ЛСР"
    sheet["B2"] = "1 квартал 2026"
    sheet["A3"] = "Планирумый срок начала работ"
    sheet["B3"] = date(2026, 1, 10)
    sheet["A4"] = "Планируемый срок окончания работ: 20.02.2026"
    return workbook


def _to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_metadata_is_detected_from_legacy_note_text(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Moscow/legacy-note.xlsx": _workbook_with_legacy_note_metadata()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.lsr_quarter == "2025 Q4"
    assert entry.planned_start == "2026-06-01"
    assert entry.planned_finish == "2026-09-01"


def test_metadata_accepts_excel_serial_dates_near_labels(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/serial-dates.xlsx": _workbook_with_serial_date_metadata()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.lsr_quarter == "2025 Q2"
    assert entry.planned_start == "2026-07-01"
    assert entry.planned_finish == "2026-09-01"


def _workbook_with_legacy_note_metadata() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-LEGACY"
    sheet["A2"] = (
        "Расчет выполнен на основании предоставленных смет (в ценах IV кв.2025 г.) "
        "с применением индексов фактической инфляции и дефлятора, рассчитанных "
        "на период выполнения работ – с июня 2026 г. по сентябрь 2026 г."
    )
    sheet.append([])
    sheet.append(["№ п/п", "Наименование работ", "Ед.изм.", "Кол-во"])
    sheet.append([1, "Work A", "м", 10])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    return _to_bytes(workbook)


def _workbook_with_serial_date_metadata() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-SERIAL"
    sheet["A2"] = "1.1. Предоставленных смет в ценах:"
    sheet["B2"] = "2 кв. 25г."
    sheet["A3"] = "начало работ:"
    sheet["B3"] = 46204
    sheet["A4"] = "окончание работ:"
    sheet["B4"] = 46266
    sheet.append([])
    sheet.append(["№ п/п", "Наименование работ", "Ед.изм.", "Кол-во"])
    sheet.append([1, "Work A", "м", 10])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    return _to_bytes(workbook)


def test_consolidation_region_and_coefficient_are_detected_and_stored(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"FolderRegion/consolidation.xlsx": _workbook_with_consolidation_metadata()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        preview = analyze_rnmc_zip_row_preview(connection, str(zip_path))
        import_result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        record = next(
            item for item in list_imported_files(connection) if item.filename == "consolidation.xlsx"
        )
        from core.storage.catalog import RNMC_ZIP_SOURCE_NAME, list_catalog_rows

        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    preview_entry = preview.entries[0]
    assert preview_entry.region_folder == "Якутия"
    assert preview_entry.regional_coefficient == 1.4
    assert import_result.success_count == 1
    import_entry = import_result.entries[0]
    assert import_entry.region_folder == "Якутия"
    assert import_entry.regional_coefficient == 1.4
    assert record.region_folder == "Якутия"
    assert record.regional_coefficient == 1.4
    assert len(rows) == 1
    assert rows[0].region == "Якутия"
    assert rows[0].regional_coefficient == 1.4


def test_manual_region_override_wins_over_workbook_region(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"FolderRegion/manual-region.xlsx": _workbook_with_consolidation_metadata()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        preview = analyze_rnmc_zip_row_preview(
            connection,
            str(zip_path),
            region_override="ManualRegion",
        )
    finally:
        connection.close()

    entry = preview.entries[0]
    assert entry.region_folder == "ManualRegion"
    assert entry.regional_coefficient == 1.4


def _workbook_with_consolidation_metadata() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-CONSOLIDATION"
    sheet["A2"] = "Данные для консолидации:"
    sheet["A3"] = "Регион расположения объекта:"
    sheet["B3"] = "Якутия"
    sheet["A4"] = "Региональный коэффициент:"
    sheet["B4"] = "1,4"
    sheet["A5"] = "Год/квартал ЛСР"
    sheet["B5"] = "IV кв.2025 г."
    sheet["A6"] = "начало работ:"
    sheet["B6"] = 46204
    sheet["A7"] = "окончание работ:"
    sheet["B7"] = 46266
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Единица измерения",
        "Кол-во",
        "ГЭСН/ФЕР/Перечень",
        "Цена единицы работ, руб. без НДС",
    ])
    sheet.append([1, "Work A", "м", 10, "GESN01-01-001-01", 100])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    return _to_bytes(workbook)


def test_metadata_does_not_treat_lsr_section_rows_as_lsr_quarter(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Buryatia/local-lsr.xlsx": _workbook_with_local_lsr_section_only()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.lsr_quarter == ""
    assert entry.planned_start == ""
    assert entry.planned_finish == ""


def test_metadata_rejects_implausible_regional_coefficient_values(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"SPb/bad-coef.xlsx": _workbook_with_bad_coefficient_metadata()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    assert result.entries[0].regional_coefficient is None


def _workbook_with_local_lsr_section_only() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-LOCAL-LSR"
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Ед.изм.",
        "Кол-во",
        "Перечень ГЭСН/ФЕР/ТЕР/КР",
        "Цена единицы работ (с учетом вспомогательных материалов), руб. без НДС",
        "Итого стоимость, руб. без НДС",
    ])
    sheet.append([1, 2, 3, 4, 5, 6, 7])
    sheet.append([1, "ЛСР 02-01-01 Оперативно-технологическая радиосвязь", "", "", "", "", ""])
    sheet.append([2, "Монтаж материалов", "шт", 1, "ГЭСНм08-03-572-06 /КР", 100, 100])
    sheet.append([None] * 7)
    sheet.append([None] * 7)
    sheet.append([None] * 7)
    return _to_bytes(workbook)


def _workbook_with_bad_coefficient_metadata() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-BAD-COEF"
    sheet["A2"] = "Региональный коэффициент:"
    sheet["B2"] = 2026
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Ед.изм.",
        "Кол-во",
        "Перечень ГЭСН",
        "Цена единицы работ, руб. без НДС",
    ])
    sheet.append([1, "Work A", "шт", 1, "ГЭСН01-01-001-01", 100])
    sheet.append([None] * 6)
    sheet.append([None] * 6)
    sheet.append([None] * 6)
    return _to_bytes(workbook)

"""Tests for importing RNMC zip workbook rows into catalog_items."""

from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.services.rnmc_excel import import_rnmc_zip_catalog_rows
from app.web.app import create_app
from core.storage import connect, init_database
from core.storage.catalog import (
    RNMC_ZIP_SOURCE_NAME,
    STATUS_DUPLICATE_NAME,
    STATUS_LEGACY_IMPORTED,
    STATUS_SUCCESS,
    import_legacy_file_log,
    list_catalog_rows,
    list_imported_files,
    record_imported_file,
)


def test_import_rnmc_zip_catalog_rows_updates_catalog_and_import_log(tmp_path: Path) -> None:
    file_log = tmp_path / "File_Log.xlsx"
    _write_file_log(file_log, [["Moscow", "old.xlsx", 3, "", "", ""]])
    zip_path = tmp_path / "rnmc.zip"
    workbook_bytes = _workbook_bytes(task_number="TASK-42")
    _write_zip(
        zip_path,
        {
            "Moscow/old.xlsx": workbook_bytes,
            "Kazan/new.xlsx": workbook_bytes,
            "Samara/new.xlsx": workbook_bytes,
            "Tula/no-data.xlsx": _workbook_without_valid_rows(task_number="TASK-99"),
            "readme.txt": b"ignored",
        },
    )

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_legacy_file_log(connection, file_log)
        result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        records = list_imported_files(connection)
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert result.total_excel_files == 4
    assert result.success_count == 1
    assert result.skipped_count == 1
    assert result.duplicate_name_count == 1
    assert result.no_data_count == 1
    assert result.rows_imported_total == 2
    assert result.rows_rejected_total == 3
    assert result.ignored_files == 1

    by_region_file = {(record.region_folder, record.filename): record for record in records}
    assert by_region_file[("Moscow", "old.xlsx")].status == STATUS_LEGACY_IMPORTED
    assert by_region_file[("Kazan", "new.xlsx")].status == STATUS_SUCCESS
    assert by_region_file[("Kazan", "new.xlsx")].rows_ok == 2
    assert by_region_file[("Kazan", "new.xlsx")].rows_rejected == 2
    assert by_region_file[("Samara", "new.xlsx")].status == STATUS_DUPLICATE_NAME
    assert by_region_file[("Tula", "no-data.xlsx")].status == "no_data"

    assert len(rows) == 2
    assert rows[0].task_id == "TASK-42"
    assert rows[0].region == "Kazan"
    assert rows[0].work_name == "Work A"
    assert rows[0].code == "GESN01-01-001-01"
    assert rows[0].unit == "м"
    assert rows[0].price == 100.5
    assert rows[1].price == 1234.56


def test_import_rnmc_zip_catalog_rows_replaces_existing_rows_for_pending_file(
    tmp_path: Path,
) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/pending.xlsx": _workbook_bytes(task_number="TASK-1")})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        record_imported_file(
            connection,
            region_folder="Kazan",
            filename="pending.xlsx",
            status="pending",
        )
        import_rnmc_zip_catalog_rows(connection, str(zip_path))
        import_rnmc_zip_catalog_rows(connection, str(zip_path))
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert len(rows) == 2


def test_admin_can_import_rnmc_zip_rows_into_catalog(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Folder/admin-new.xlsx": _workbook_bytes(task_number="ADMIN-7")})

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        with zip_path.open("rb") as handle:
            response = client.post(
                "/admin/imports/rnmc-import",
                data={"region_override": "Manual Region"},
                files={"rnmc_zip": ("rnmc.zip", handle, "application/zip")},
            )

    assert response.status_code == 200
    assert "ZIP импортирован в каталог" in response.text
    assert "admin-new.xlsx" in response.text
    assert "Manual Region" in response.text
    assert "success" in response.text
    assert "Строк добавлено" in response.text

    connection = connect(db_path)
    try:
        init_database(connection)
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()
    assert len(rows) == 2
    assert rows[0].region == "Manual Region"


def _write_zip(path: Path, files: dict[str, bytes]) -> None:
    with ZipFile(path, "w") as archive:
        for name, data in files.items():
            archive.writestr(name, data)


def _write_file_log(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FileLog"
    sheet.append(
        [
            "Folder",
            "File",
            "Status",
            "Год Квартал ЛСР",
            "Планирумый срок начала работ",
            "Планируемый срок окончания работ",
        ]
    )
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _workbook_bytes(*, task_number: str) -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = f"№ задачи 1Ф: {task_number}"
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Ед.изм.",
        "Кол-во",
        "ГЭСН/ФЕР/Перечень",
        "Цена",
    ])
    sheet.append([1, "Work A", "м", 10, "GESN01-01-001-01", 100.5])
    sheet.append([2, "Work B", "шт", 5, "GESN02-01-001-01", "1 234,56"])
    sheet.append([3, "Bad price", "м", 1, "GESN03-01-001-01", ""])
    sheet.append([4, "Work without unit and qty", "", "", "GESN04-01-001-01", 10])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _workbook_without_valid_rows(*, task_number: str) -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = f"№ задачи 1Ф: {task_number}"
    sheet.append([])
    sheet.append(["№ п/п", "Наименование работ", "Ед.изм.", "Кол-во", "Цена"])
    sheet.append([1, "No code", "м", 1, 100])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    sheet.append([None, None, None, None, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

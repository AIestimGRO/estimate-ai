import pytest
from openpyxl import Workbook

from app.services.read_estimate import (
    METHOD_DETECTED,
    METHOD_TEMPLATE,
    KeyDataNotFoundError,
    MultipleSheetsError,
    load_estimate,
)

METER = "\u043c"
CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
INSTALLATION = "\u043c\u043e\u043d\u0442\u0430\u0436"

HEADER_CODE = "\u0428\u0438\u0444\u0440"  # "Shifr"
HEADER_NAME = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
HEADER_UNIT = "\u0415\u0434. \u0438\u0437\u043c."
HEADER_PRICE = "\u0426\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434."

ESTIMATE_TITLE = "\u041e\u0421 estimate"
SHEET_A = "\u0421\u043c\u0435\u0442\u0430 A"
SHEET_B = "\u0421\u043c\u0435\u0442\u0430 B"


def _template_estimate(path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ESTIMATE_TITLE
    worksheet.cell(row=7, column=14).value = CODE
    worksheet.cell(row=9, column=3).value = INSTALLATION
    worksheet.cell(row=9, column=4).value = METER
    worksheet.cell(row=9, column=6).value = 50.0
    worksheet.cell(row=9, column=14).value = CODE
    workbook.save(path)
    workbook.close()
    return path


def _detected_sheet(worksheet):
    worksheet.cell(row=1, column=1).value = HEADER_CODE
    worksheet.cell(row=1, column=2).value = HEADER_NAME
    worksheet.cell(row=1, column=3).value = HEADER_UNIT
    worksheet.cell(row=1, column=4).value = HEADER_PRICE
    worksheet.cell(row=2, column=1).value = CODE
    worksheet.cell(row=2, column=2).value = INSTALLATION
    worksheet.cell(row=2, column=3).value = METER
    worksheet.cell(row=2, column=4).value = 50.0


def _detected_estimate(path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_A
    _detected_sheet(worksheet)
    workbook.save(path)
    workbook.close()
    return path


def test_template_read(tmp_path):
    estimate = _template_estimate(tmp_path / "estimate.xlsx")

    data = load_estimate(estimate)

    assert data.method == METHOD_TEMPLATE
    assert data.sheet_title == ESTIMATE_TITLE
    assert len(data.positioned_rows) == 1
    row_number, estimate_row = data.positioned_rows[0]
    assert row_number == 9
    assert estimate_row.base_price == 50.0


def test_detected_read(tmp_path):
    estimate = _detected_estimate(tmp_path / "estimate.xlsx")

    data = load_estimate(estimate)

    assert data.method == METHOD_DETECTED
    assert data.sheet_title == SHEET_A
    assert data.header_row == 1
    assert data.code_column == 1
    assert data.base_price_column == 4
    assert len(data.positioned_rows) == 1
    assert data.positioned_rows[0][0] == 2


def test_detected_read_tolerates_blank_body_rows(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_A
    _detected_sheet(worksheet)
    # blank row 3, then a second data row 4
    worksheet.cell(row=4, column=1).value = CODE
    worksheet.cell(row=4, column=3).value = METER
    worksheet.cell(row=4, column=4).value = 70.0
    path = tmp_path / "estimate.xlsx"
    workbook.save(path)
    workbook.close()

    data = load_estimate(path)

    assert [row_number for row_number, _ in data.positioned_rows] == [2, 4]


def test_key_data_not_found(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Junk"
    worksheet.cell(row=1, column=1).value = "abc"
    worksheet.cell(row=2, column=1).value = "def"
    path = tmp_path / "estimate.xlsx"
    workbook.save(path)
    workbook.close()

    with pytest.raises(KeyDataNotFoundError) as info:
        load_estimate(path)

    assert info.value.sheet_title == "Junk"
    assert "header_row" in info.value.report


def test_multiple_resolvable_sheets_require_choice(tmp_path):
    workbook = Workbook()
    first = workbook.active
    first.title = SHEET_A
    _detected_sheet(first)
    second = workbook.create_sheet(title=SHEET_B)
    _detected_sheet(second)
    path = tmp_path / "estimate.xlsx"
    workbook.save(path)
    workbook.close()

    with pytest.raises(MultipleSheetsError) as info:
        load_estimate(path)

    assert set(info.value.candidates) == {SHEET_A, SHEET_B}


def test_load_estimate_reads_cached_values(tmp_path, monkeypatch):
    # Guards the formula fix: reading must use data_only=True so formula-driven
    # price cells resolve to the numbers Excel cached (OPEN_ITEMS #2).
    import app.services.read_estimate as read_estimate_module

    captured = {}
    real_load = read_estimate_module.load_workbook

    def spy(path, **kwargs):
        captured["data_only"] = kwargs.get("data_only")
        return real_load(path, **kwargs)

    monkeypatch.setattr(read_estimate_module, "load_workbook", spy)

    estimate = _template_estimate(tmp_path / "estimate.xlsx")
    load_estimate(estimate)

    assert captured["data_only"] is True


def test_selected_sheet_title_forces_choice(tmp_path):
    workbook = Workbook()
    first = workbook.active
    first.title = SHEET_A
    _detected_sheet(first)
    second = workbook.create_sheet(title=SHEET_B)
    _detected_sheet(second)
    path = tmp_path / "estimate.xlsx"
    workbook.save(path)
    workbook.close()

    data = load_estimate(path, selected_sheet_title=SHEET_B)

    assert data.sheet_title == SHEET_B
    assert data.method == METHOD_DETECTED

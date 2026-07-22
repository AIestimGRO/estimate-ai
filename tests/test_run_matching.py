from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from app.services.run_matching import (
    KR_END,
    run_matching,
    run_matching_from_files,
)
from core.catalog import CatalogRow
from core.exclusions import NameExclusionRule
from core.matching import (
    REASON_EXCLUDED_BY_NAME,
    REASON_INVALID_INPUT,
    REASON_MATCHED,
    REASON_NO_MATCH,
    EstimateRow,
)
from core.normalize import NormCode, NormUnit
from core.risk import (
    REASON_NONE,
    REASON_RATIO_EXCEEDED,
    GesnException,
    build_dem_key,
    build_gesn_exception_key,
)
from core.tkp_matching import build_tkp_catalog_index


METER = "\u043c"
CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
OTHER_CODE = "\u0413\u042d\u0421\u041d05-01-001-01"
INSTALLATION = "\u043c\u043e\u043d\u0442\u0430\u0436"
DEMOLITION = "\u0434\u0435\u043c\u043e\u043d\u0442\u0430\u0436"
EXCLUDED = "\u0438\u0441\u043a\u043b\u044e\u0447\u0438\u0442\u044c"


def catalog_row(
    *,
    task_id: object = "task-1",
    price: object = 100.0,
    code: object = CODE,
    unit: object = METER,
    work_name: object = INSTALLATION,
    region: object = "region-1",
) -> CatalogRow:
    return CatalogRow(
        task_id=task_id,
        price=price,
        code=code,
        unit=unit,
        work_name=work_name,
        region=region,
    )


def estimate_row(
    *,
    code: object = CODE,
    unit: object = METER,
    work_name: object = INSTALLATION,
    base_price: object = 50.0,
) -> EstimateRow:
    return EstimateRow(
        code=code,
        unit=unit,
        work_name=work_name,
        base_price=base_price,
    )


def exception_key_for(
    *,
    unit: object = METER,
    code: object = CODE,
    source_has_demolition: bool = False,
    demontazh_filter_enabled: bool = True,
) -> str:
    dem_key = build_dem_key(source_has_demolition, demontazh_filter_enabled)
    return build_gesn_exception_key(NormUnit(unit), NormCode(code), dem_key)


def test_matched_row_produces_analogs_section_recommended_and_kr() -> None:
    catalog_rows = [
        catalog_row(task_id="task-1", price=100),
        catalog_row(task_id="task-2", price=120),
    ]

    result = run_matching(catalog_rows, [estimate_row(base_price=50)])
    row = result.rows[0]

    assert row.status == REASON_MATCHED
    assert row.has_analogs
    assert [analog.entry.price for analog in row.analogs] == [100, 120]
    assert row.section_code == "01"
    assert row.kr_code is not None and row.kr_code.endswith(KR_END)
    # MAX(50, AVERAGE(50, 100, 120)) = MAX(50, 90) = 90
    assert row.recommended_price == 90
    assert row.exception_key == exception_key_for()


def test_regional_coefficient_applies_to_analogs_but_not_base() -> None:
    result = run_matching(
        [catalog_row(price=100)],
        [estimate_row(base_price=50)],
        regional_coefficient=2.0,
    )

    # Base stays 50, analog becomes 100 * 2 = 200 -> AVERAGE(50, 200) = 125.
    assert result.rows[0].recommended_price == 125


def test_tkp_best_match_is_separate_from_rnmc_and_included_in_average() -> None:
    tkp_index = build_tkp_catalog_index(
        [
            SimpleNamespace(
                id=1,
                item_name=INSTALLATION,
                unit=METER,
                winner_unit_price_no_vat=200.0,
                winner_name="winner",
                source_file_name="source.xlsx",
                task_no="tkp-42",
            )
        ]
    )

    result = run_matching(
        [catalog_row(price=100)],
        [estimate_row(base_price=50)],
        regional_coefficient=2.0,
        tkp_catalog_index=tkp_index,
        use_tkp_analogs=True,
    )
    row = result.rows[0]

    assert [analog.entry.price for analog in row.analogs] == [100]
    assert row.tkp_match is not None
    assert row.tkp_match.entry.task_no == "tkp-42"
    assert result.matched_row_count == 1
    assert result.tkp_matched_row_count == 1
    # RNMC is scaled to 200, TKP remains 200: AVERAGE(50, 200, 200) = 150.
    assert row.recommended_price == 150


def test_tkp_catalog_is_ignored_when_toggle_is_off() -> None:
    tkp_index = build_tkp_catalog_index(
        [
            SimpleNamespace(
                id=1,
                item_name=INSTALLATION,
                unit=METER,
                winner_unit_price_no_vat=200.0,
                winner_name="winner",
                source_file_name="source.xlsx",
                task_no="tkp-42",
            )
        ]
    )

    result = run_matching(
        [catalog_row(price=100)],
        [estimate_row(base_price=50)],
        tkp_catalog_index=tkp_index,
        use_tkp_analogs=False,
    )

    assert result.rows[0].tkp_match is None
    assert result.rows[0].recommended_price == 75
    assert result.tkp_matched_row_count == 0


def test_no_match_row_keeps_base_price_and_gets_plain_code_in_kr() -> None:
    result = run_matching(
        [catalog_row(code=CODE)],
        [estimate_row(code=OTHER_CODE, base_price=50)],
    )
    row = result.rows[0]

    assert row.status == REASON_NO_MATCH
    assert not row.has_analogs
    # 2026-07 rule: no analog found -> /КР gets the plain ГЭСН code, no suffix.
    assert row.kr_code == OTHER_CODE
    assert row.recommended_price == 50
    assert row.section_code == "02"


def test_name_excluded_estimate_row_is_skipped_but_still_priced() -> None:
    rule = NameExclusionRule(
        enabled=True,
        scope="SMETA",
        match_mode="CONTAINS",
        pattern=EXCLUDED,
    )

    result = run_matching(
        [catalog_row()],
        [estimate_row(work_name=f"{EXCLUDED} {INSTALLATION}", base_price=50)],
        name_exclusion_rules=[rule],
    )
    row = result.rows[0]

    assert row.status == REASON_EXCLUDED_BY_NAME
    assert not row.has_analogs
    assert row.kr_code == CODE
    assert row.recommended_price == 50
    assert row.section_code == "01"


def test_invalid_input_row_has_no_recommended_price() -> None:
    result = run_matching([catalog_row()], [estimate_row(base_price=0)])
    row = result.rows[0]

    assert row.status == REASON_INVALID_INPUT
    assert row.recommended_price is None
    assert row.kr_code == CODE


def test_ratio_spread_below_default_limit_is_not_flagged() -> None:
    result = run_matching(
        [
            catalog_row(task_id="task-1", price=100),
            catalog_row(task_id="task-2", price=250),
        ],
        [estimate_row()],
    )
    row = result.rows[0]

    assert not row.risk_result.is_flagged
    assert result.flagged_row_count == 0


def test_ratio_spread_is_flagged() -> None:
    result = run_matching(
        [
            catalog_row(task_id="task-1", price=100),
            catalog_row(task_id="task-2", price=300),
        ],
        [estimate_row()],
    )
    row = result.rows[0]

    assert row.risk_result.is_flagged
    assert row.risk_result.reason == REASON_RATIO_EXCEEDED
    assert result.flagged_row_count == 1


def test_approved_exception_overrides_ratio_check() -> None:
    exceptions = {
        exception_key_for(): GesnException(
            exception_key=exception_key_for(),
            approved_min=90,
            approved_max=110,
            last_range_update_date=10,
        )
    }

    result = run_matching(
        [
            catalog_row(task_id="task-1", price=100),
            catalog_row(task_id="task-2", price=300),
        ],
        [estimate_row()],
        gesn_exceptions=exceptions,
    )
    row = result.rows[0]

    # Analogs have no added-date (serial 0), so the approved-range check treats
    # them as not new enough to re-flag, short-circuiting the ratio problem.
    assert not row.risk_result.is_flagged
    assert row.risk_result.reason == REASON_NONE
    assert result.flagged_row_count == 0


def test_demolition_filter_selects_matching_analogs() -> None:
    catalog_rows = [
        catalog_row(price=100, work_name=INSTALLATION),
        catalog_row(price=200, work_name=DEMOLITION),
    ]

    result = run_matching(catalog_rows, [estimate_row(work_name=DEMOLITION)])
    row = result.rows[0]

    assert row.status == REASON_MATCHED
    assert [analog.entry.price for analog in row.analogs] == [200]
    assert row.is_demolition


def test_run_aggregates_counts() -> None:
    catalog_rows = [
        catalog_row(task_id="task-1", price=100),
        catalog_row(task_id="task-2", price=300),
    ]
    estimate_rows = [
        estimate_row(),
        estimate_row(code=OTHER_CODE),
    ]

    result = run_matching(catalog_rows, estimate_rows)

    assert result.catalog_key_count == 1
    assert result.matched_row_count == 1
    assert result.flagged_row_count == 1


def _fixture_path(tmp_path: Path, name: str) -> Path:
    fixtures_dir = tmp_path / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)
    return fixtures_dir / name


def _save(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    workbook.close()
    return path


def _make_catalog_file(tmp_path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2).value = "task-1"
    worksheet.cell(row=4, column=3).value = INSTALLATION
    worksheet.cell(row=4, column=4).value = METER
    worksheet.cell(row=4, column=7).value = 100
    worksheet.cell(row=4, column=14).value = CODE
    worksheet.cell(row=5, column=2).value = "task-2"
    worksheet.cell(row=5, column=3).value = INSTALLATION
    worksheet.cell(row=5, column=4).value = METER
    worksheet.cell(row=5, column=7).value = 120
    worksheet.cell(row=5, column=14).value = CODE
    return _save(workbook, _fixture_path(tmp_path, "catalog.xlsx"))


def _make_estimate_file(tmp_path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041e\u0421 estimate"
    worksheet.cell(row=7, column=14).value = CODE
    worksheet.cell(row=9, column=3).value = INSTALLATION
    worksheet.cell(row=9, column=4).value = METER
    worksheet.cell(row=9, column=6).value = 50
    worksheet.cell(row=9, column=14).value = CODE
    return _save(workbook, _fixture_path(tmp_path, "estimate.xlsx"))


def test_run_matching_from_files_end_to_end(tmp_path: Path) -> None:
    catalog_path = _make_catalog_file(tmp_path)
    estimate_path = _make_estimate_file(tmp_path)

    result = run_matching_from_files(catalog_path, estimate_path)

    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.status == REASON_MATCHED
    assert [analog.entry.price for analog in row.analogs] == [100, 120]
    assert row.recommended_price == 90
    assert row.kr_code is not None and row.kr_code.endswith(KR_END)


def test_second_run_and_write_updates_single_open_risk_row(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    from app.services.write_result import run_and_write
    from core.storage import connect, init_database, list_price_risks

    catalog_path = tmp_path / "catalog.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2).value = "task-1"
    worksheet.cell(row=4, column=3).value = INSTALLATION
    worksheet.cell(row=4, column=4).value = METER
    worksheet.cell(row=4, column=7).value = 100
    worksheet.cell(row=4, column=14).value = CODE
    worksheet.cell(row=5, column=2).value = "task-2"
    worksheet.cell(row=5, column=3).value = INSTALLATION
    worksheet.cell(row=5, column=4).value = METER
    worksheet.cell(row=5, column=7).value = 350
    worksheet.cell(row=5, column=14).value = CODE
    workbook.save(catalog_path)
    workbook.close()

    estimate_path = _make_estimate_file(tmp_path)
    db_path = tmp_path / "risks.db"
    output = tmp_path / "out.xlsx"

    run_and_write(catalog_path, estimate_path, output, database_path=db_path)
    run_and_write(catalog_path, estimate_path, output, database_path=db_path)

    connection = connect(db_path)
    try:
        init_database(connection)
        open_rows = list_price_risks(connection, status="open")
        total = connection.execute("SELECT COUNT(*) FROM price_risk_log").fetchone()[0]
    finally:
        connection.close()

    assert total == 1
    assert len(open_rows) == 1
    assert open_rows[0].max_price == pytest.approx(350)

    workbook = load_workbook(output, data_only=False)
    try:
        assert "Price_Check_Log" not in workbook.sheetnames
    finally:
        workbook.close()

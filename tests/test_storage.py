"""Tests for SQLite storage layer."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from core.catalog import BuildCatalog
from core.exclusions import is_name_excluded
from core.macro_workbook import NAME_EXCLUSIONS_SHEET, load_all_rules_from_workbook
from core.storage import (
    connect,
    filename_is_processed,
    import_catalog_from_excel,
    import_rules_from_workbook,
    init_database,
    list_catalog_rows,
    list_catalog_sources,
    list_name_exclusion_rules,
    list_task_color_entries,
)
CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
METER = "\u043c"


def _write_catalog_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2, value="TASK-100")
    worksheet.cell(row=4, column=3, value="\u0440\u0430\u0431\u043e\u0442\u0430")
    worksheet.cell(row=4, column=4, value=METER)
    worksheet.cell(row=4, column=7, value=125.5)
    worksheet.cell(row=4, column=14, value=CODE)
    workbook.save(path)
    workbook.close()


def _write_rules_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = NAME_EXCLUSIONS_SHEET
    worksheet.cell(row=2, column=1, value=1)
    worksheet.cell(row=2, column=2, value="SMETA")
    worksheet.cell(row=2, column=3, value="ALL_WORDS")
    worksheet.cell(row=2, column=4, value="\u043c\u043c|\u0438\u0437\u043c\u0435\u043d")
    worksheet.cell(row=3, column=1, value=0)
    worksheet.cell(row=3, column=2, value="BOTH")
    worksheet.cell(row=3, column=3, value="ALL_WORDS")
    worksheet.cell(row=3, column=4, value="\u0441\u043b\u043e\u0439")
    worksheet.cell(row=2, column=8, value=1)
    worksheet.cell(row=2, column=9, value="999")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def test_init_database_creates_schema() -> None:
    connection = connect(":memory:")
    try:
        init_database(connection)
        tables = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
        assert "catalog_sources" in tables
        assert "catalog_items" in tables
        assert "name_exclusion_rules" in tables
        assert "price_risk_log" in tables
        assert "gesn_exceptions" in tables
    finally:
        connection.close()


def test_import_catalog_from_excel_round_trip(tmp_path: Path) -> None:
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)

    connection = connect(":memory:")
    try:
        init_database(connection)
        result = import_catalog_from_excel(connection, catalog_path, source_name="main")
        rows = list_catalog_rows(connection, source_name="main")
    finally:
        connection.close()

    assert result.rows_imported == 1
    assert len(rows) == 1
    assert rows[0].code == CODE
    assert float(rows[0].price) == pytest.approx(125.5)


def test_import_catalog_replace_clears_previous_source(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _write_catalog_workbook(first)

    second.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2, value="TASK-200")
    worksheet.cell(row=4, column=3, value="\u0440\u0430\u0431\u043e\u0442\u0430")
    worksheet.cell(row=4, column=4, value=METER)
    worksheet.cell(row=4, column=7, value=200)
    worksheet.cell(row=4, column=14, value=CODE)
    workbook.save(second)
    workbook.close()

    connection = connect(":memory:")
    try:
        init_database(connection)
        import_catalog_from_excel(connection, first, source_name="main")
        import_catalog_from_excel(connection, second, source_name="main", replace=True)
        rows = list_catalog_rows(connection, source_name="main")
        sources = list_catalog_sources(connection)
    finally:
        connection.close()

    assert len(sources) == 1
    assert len(rows) == 1
    assert rows[0].task_id == "TASK-200"


def test_imported_catalog_builds_matching_catalog(tmp_path: Path) -> None:
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)

    connection = connect(":memory:")
    try:
        init_database(connection)
        import_catalog_from_excel(connection, catalog_path)
        rows = list_catalog_rows(connection)
        catalog = BuildCatalog(rows, [])
    finally:
        connection.close()

    assert len(catalog) == 1


def test_import_rules_from_workbook(tmp_path: Path) -> None:
    xlsm = tmp_path / "macro.xlsm"
    _write_rules_workbook(xlsm)

    connection = connect(":memory:")
    try:
        init_database(connection)
        rule_count, color_count = import_rules_from_workbook(connection, xlsm)
        rules = list_name_exclusion_rules(connection)
        colors = list_task_color_entries(connection)
    finally:
        connection.close()

    assert rule_count == 2
    assert color_count == 1
    enabled = [rule for rule in rules if rule.enabled]
    assert len(enabled) == 1
    assert is_name_excluded(enabled, "SMETA", "\u0442\u0435\u043a\u0441\u0442 \u043c\u043c \u0438\u0437\u043c\u0435\u043d")


def test_load_all_rules_from_workbook_includes_disabled(tmp_path: Path) -> None:
    xlsm = tmp_path / "macro.xlsm"
    _write_rules_workbook(xlsm)

    rules, _colors = load_all_rules_from_workbook(xlsm)

    assert len(rules) == 2
    assert any(not rule.enabled for rule in rules)


def test_cli_import_catalog(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "test.db"
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    from app.cli.__main__ import main

    assert main(["init-db"]) == 0
    assert main(["import-catalog", str(catalog_path)]) == 0
    assert main(["status"]) == 0
    assert db_path.is_file()


def test_upsert_open_risk_does_not_duplicate_on_second_run(tmp_path: Path) -> None:
    from core.risk import REASON_RATIO_EXCEEDED
    from core.storage.risk_log import (
        STATUS_OPEN,
        list_price_risks,
        upsert_open_risk,
    )

    db_path = tmp_path / "risks.db"
    connection = connect(db_path)
    try:
        init_database(connection)
        key = f"{METER}||{CODE}||NO_DEM"
        upsert_open_risk(
            connection,
            exception_key=key,
            reason=REASON_RATIO_EXCEEDED,
            code=CODE,
            unit=METER,
            min_price=100,
            max_price=300,
            ratio=3.0,
            recommended_price=150,
            estimate_row=9,
        )
        upsert_open_risk(
            connection,
            exception_key=key,
            reason=REASON_RATIO_EXCEEDED,
            code=CODE,
            unit=METER,
            min_price=110,
            max_price=320,
            ratio=2.9,
            recommended_price=160,
            estimate_row=10,
        )
        open_rows = list_price_risks(connection, status=STATUS_OPEN)
        total = connection.execute("SELECT COUNT(*) FROM price_risk_log").fetchone()[0]
    finally:
        connection.close()

    assert total == 1
    assert len(open_rows) == 1
    assert open_rows[0].min_price == pytest.approx(110)
    assert open_rows[0].max_price == pytest.approx(320)
    assert open_rows[0].ratio == pytest.approx(2.9)
    assert open_rows[0].estimate_row == 10


def test_approve_risk_writes_gesn_exception_and_clears_open_flag(tmp_path: Path) -> None:
    from core.catalog import CatalogRow
    from core.matching import EstimateRow
    from core.normalize import NormCode, NormUnit
    from core.risk import (
        REASON_NONE,
        REASON_RATIO_EXCEEDED,
        build_dem_key,
        build_gesn_exception_key,
    )
    from core.storage.risk_log import (
        STATUS_APPROVED,
        STATUS_OPEN,
        approve_risk,
        load_gesn_exceptions,
        list_price_risks,
        upsert_open_risk,
    )
    from app.services.run_matching import run_matching

    db_path = tmp_path / "risks.db"
    dem_key = build_dem_key(source_has_demolition=False, demontazh_filter_enabled=True)
    key = build_gesn_exception_key(NormUnit(METER), NormCode(CODE), dem_key)
    connection = connect(db_path)
    try:
        init_database(connection)
        upsert_open_risk(
            connection,
            exception_key=key,
            reason=REASON_RATIO_EXCEEDED,
            code=CODE,
            unit=METER,
            min_price=100,
            max_price=300,
            ratio=3.0,
            recommended_price=150,
            estimate_row=9,
        )
        approve_risk(connection, key, proposed_min=90, proposed_max=310, proposed_date_serial=20)
        exceptions = load_gesn_exceptions(connection)
        open_rows = list_price_risks(connection, status=STATUS_OPEN)
        approved_rows = list_price_risks(connection, status=STATUS_APPROVED)
    finally:
        connection.close()

    assert key in exceptions
    assert exceptions[key].approved_min == pytest.approx(90)
    assert exceptions[key].approved_max == pytest.approx(310)
    assert len(open_rows) == 0
    assert len(approved_rows) == 1

    catalog_rows = [
        CatalogRow(
            task_id="task-1",
            price=100,
            code=CODE,
            unit=METER,
            work_name="\u043c\u043e\u043d\u0442\u0430\u0436",
            region="region-1",
        ),
        CatalogRow(
            task_id="task-2",
            price=300,
            code=CODE,
            unit=METER,
            work_name="\u043c\u043e\u043d\u0442\u0430\u0436",
            region="region-1",
        ),
    ]
    result = run_matching(
        catalog_rows,
        [
            EstimateRow(
                code=CODE,
                unit=METER,
                work_name="\u043c\u043e\u043d\u0442\u0430\u0436",
                base_price=50.0,
            )
        ],
        gesn_exceptions=exceptions,
    )
    row = result.rows[0]

    assert not row.risk_result.is_flagged
    assert row.risk_result.reason == REASON_NONE
    assert result.flagged_row_count == 0


def _write_zlvl_catalog_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Каталог"
    worksheet.append([])
    worksheet.append([None, None, None, None, None, None, "ZLVL"])
    worksheet.append(
        [
            "№_пп",
            "Номер задачи",
            "Наименование работ",
            "Ед.изм.",
            "Кол-во",
            "Цена единицы работ (с учетом вспомогательных материалов), руб. без НДС",
            "Цена единицы работ (с учетом вспомогательных материалов), руб. без НДС ZLVL",
            "Итого стоимость, руб. с НДС",
            "Итого стоимость, руб. без НДС",
            "ТЗ на ед., чел-час",
            "ТЗ всего, чел-час",
            "ТЗм на ед., чел-час",
            "ТЗм всего, чел-час",
            "Перечень ГЭСН/ФЕР/ТЕР/КР",
            "source_file",
            "Регион",
            "Год Квартал ЛСР",
            "Планирумый срок начала работ",
            "Планируемый срок окончания работ",
            "Региональный коэффициент",
            "Дата добавления в каталог",
        ]
    )
    worksheet.append(
        [
            1,
            "TASK-ZLVL",
            "Legacy work",
            "шт",
            "2,5",
            120,
            100,
            300,
            250,
            1.1,
            2.2,
            0.3,
            0.6,
            CODE,
            "source-rnmc.xlsx",
            "Москва",
            "2026 Q1",
            "2026-07-01",
            "2026-09-01",
            1.2,
            "2026-07-10",
        ]
    )
    workbook.save(path)
    workbook.close()


def test_import_zlvl_catalog_maps_all_required_columns(tmp_path: Path) -> None:
    catalog_path = tmp_path / "РНМЦ_КА_ЖО_ZLVL_V3.xlsx"
    _write_zlvl_catalog_workbook(catalog_path)

    connection = connect(":memory:")
    try:
        init_database(connection)
        result = import_catalog_from_excel(connection, catalog_path, source_name="main")
        rows = list_catalog_rows(connection, source_name="main")
        db_row = connection.execute(
            """
            SELECT * FROM catalog_items
            INNER JOIN catalog_sources ON catalog_sources.id = catalog_items.source_id
            WHERE catalog_sources.name = 'main'
            """
        ).fetchone()
        source_import = connection.execute(
            """
            SELECT * FROM imported_files
            WHERE filename = 'source-rnmc.xlsx'
            """
        ).fetchone()
        consolidated_import = connection.execute(
            """
            SELECT * FROM imported_files
            WHERE filename = ?
            """,
            (catalog_path.name,),
        ).fetchone()
        source_is_processed = filename_is_processed(connection, "source-rnmc.xlsx")
    finally:
        connection.close()

    assert result.rows_imported == 1
    assert len(rows) == 1
    row = rows[0]
    assert row.task_id == "TASK-ZLVL"
    assert row.work_name == "Legacy work"
    assert row.unit == "шт"
    assert row.quantity == pytest.approx(2.5)
    assert row.price_original == pytest.approx(120)
    assert row.price_zlvl == pytest.approx(100)
    assert row.price == pytest.approx(100)
    assert row.total_price == pytest.approx(250)
    assert row.labor_unit == pytest.approx(1.1)
    assert row.labor_total == pytest.approx(2.2)
    assert row.machine_labor_unit == pytest.approx(0.3)
    assert row.machine_labor_total == pytest.approx(0.6)
    assert row.code == CODE
    assert row.region == "Москва"
    assert row.source_filename == "source-rnmc.xlsx"
    assert row.lsr_quarter == "2026 Q1"
    assert row.planned_start == "2026-07-01"
    assert row.planned_finish == "2026-09-01"
    assert row.regional_coefficient == pytest.approx(1.2)
    assert db_row["source_filename"] == "source-rnmc.xlsx"
    assert db_row["added_date"] == "2026-07-10"
    assert source_import is not None
    assert source_import["status"] == "legacy_imported"
    assert source_import["region_folder"] == "Москва"
    assert source_import["filename_key"] == "source-rnmc.xlsx"
    assert source_import["rows_ok"] == 1
    assert source_import["lsr_quarter"] == "2026 Q1"
    assert source_import["planned_start"] == "2026-07-01"
    assert source_import["planned_finish"] == "2026-09-01"
    assert source_import["regional_coefficient"] == pytest.approx(1.2)
    assert consolidated_import is not None
    assert consolidated_import["status"] == "success"
    assert source_is_processed is True


def test_final_filename_keys_for_preview_recovers_from_catalog_source_filename(tmp_path):
    from core.storage.catalog import final_filename_keys_for_preview, replace_catalog_rows_for_file
    from core.storage.connection import connect, init_database
    from core.catalog import CatalogRow

    connection = connect(tmp_path / "catalog.db")
    init_database(connection)
    replace_catalog_rows_for_file(
        connection,
        [
            __import__('core.storage.catalog', fromlist=['CatalogRowStorageItem']).CatalogRowStorageItem(
                catalog_row=CatalogRow(task_id="1", code="GESN01", unit="m", work_name="Work", price=10.0),
                source_region_folder="Region",
                source_filename="Recovered RNMC.xlsx",
                source_row_number=2,
            )
        ],
        region_folder="Region",
        filename="Recovered RNMC.xlsx",
    )
    connection.execute("DELETE FROM imported_files")
    connection.commit()

    assert "recovered rnmc.xlsx" in final_filename_keys_for_preview(connection)
    connection.close()


def test_clear_catalog_for_rebuild_preserves_imported_file_history(tmp_path):
    from core.storage.catalog import clear_catalog_for_rebuild, record_imported_file
    from core.storage.connection import connect, init_database

    connection = connect(tmp_path / "catalog.db")
    init_database(connection)
    record_imported_file(
        connection,
        region_folder="Region",
        filename="History.xlsx",
        status="legacy_imported",
    )
    clear_catalog_for_rebuild(connection)

    count = connection.execute("SELECT COUNT(*) AS c FROM imported_files").fetchone()["c"]
    assert count == 1
    connection.close()

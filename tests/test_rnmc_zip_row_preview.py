"""Tests for RNMC zip workbook row previews."""

from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.services.rnmc_excel import (
    STATUS_DUPLICATE_NAME,
    STATUS_PREVIEW_OK,
    STATUS_SKIPPED_PROCESSED,
    analyze_rnmc_zip_row_preview,
)
from app.web.app import create_app
from core.storage import connect, init_database
from core.storage.catalog import STATUS_PENDING, import_legacy_file_log, record_imported_file


def test_rnmc_zip_row_preview_counts_rows_and_keeps_pending_previewable(tmp_path: Path) -> None:
    file_log = tmp_path / "File_Log.xlsx"
    _write_file_log(file_log, [["Moscow", "old.xlsx", 3, "", "", ""]])
    valid = _workbook_bytes(task_number="TASK-42")
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(
        zip_path,
        {
            "Moscow/old.xlsx": valid,
            "Kazan/new.xlsx": valid,
            "Samara/new.xlsx": valid,
            "Tula/pending.xlsx": valid,
            "notes/readme.txt": b"ignored",
        },
    )

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_legacy_file_log(connection, file_log)
        record_imported_file(
            connection,
            region_folder="Tula",
            filename="pending.xlsx",
            status=STATUS_PENDING,
        )
        connection.commit()
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    by_path = {entry.archive_path: entry for entry in result.entries}
    assert result.total_excel_files == 4
    assert result.preview_ok_count == 2
    assert result.rows_ok_total == 4
    assert result.rows_rejected_total == 0
    assert result.skipped_processed_count == 1
    assert result.duplicate_name_count == 1
    assert result.ignored_files == 1

    assert by_path["Moscow/old.xlsx"].status == STATUS_SKIPPED_PROCESSED
    assert by_path["Kazan/new.xlsx"].status == STATUS_PREVIEW_OK
    assert by_path["Kazan/new.xlsx"].region_folder == "Kazan"
    assert by_path["Kazan/new.xlsx"].sheet_name == "Data"
    assert by_path["Kazan/new.xlsx"].header_row == 3
    assert by_path["Kazan/new.xlsx"].task_number == "TASK-42"
    assert by_path["Kazan/new.xlsx"].rows_ok == 2
    assert by_path["Kazan/new.xlsx"].rows_rejected == 0
    assert by_path["Kazan/new.xlsx"].sample_rows[0].work_name == "Work A"
    assert by_path["Samara/new.xlsx"].status == STATUS_DUPLICATE_NAME
    assert by_path["Tula/pending.xlsx"].status == STATUS_PREVIEW_OK



def test_row_preview_skips_processed_workbook_without_opening_it(tmp_path: Path) -> None:
    file_log = tmp_path / "File_Log.xlsx"
    _write_file_log(file_log, [["Moscow", "old.xlsx", 3, "", "", ""]])
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Moscow/old.xlsx": b"not a real workbook"})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_legacy_file_log(connection, file_log)
        connection.commit()
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    assert result.skipped_processed_count == 1
    assert result.parse_error_count == 0
    assert result.entries[0].status == STATUS_SKIPPED_PROCESSED
    assert "not opened" in result.entries[0].reason


def test_row_preview_limits_table_scan_to_30_rows_by_default(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/large.xlsx": _large_workbook_bytes(rows=45)})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert result.preview_ok_count == 1
    assert result.rows_ok_total == 30
    assert result.limited_count == 1
    assert entry.rows_ok == 30
    assert entry.rows_rejected == 0
    assert entry.is_limited is True
    assert len(entry.sample_rows) == 30
    assert entry.sample_rows[0].row_number == 4
    assert entry.sample_rows[-1].row_number == 33
    assert entry.sample_rows[-1].work_name == "Work 30"
    assert "preview stopped at 30 table rows" in entry.reason


def test_row_preview_includes_detected_headers_and_value_samples(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/values.xlsx": _value_preview_workbook_bytes()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.header_preview.code == "Перечень ГЭСН"
    assert entry.header_preview.unit == "Единица измерения"
    assert entry.header_preview.unit_price == "Цена единицы работ, руб. с НДС"
    assert entry.header_preview.total_price == "Итого стоимость, руб. с НДС"
    assert len(entry.sample_rows) == 2
    assert entry.sample_rows[0].code == "ГЭСН01-01-001-01"
    assert entry.sample_rows[0].unit_price == "100"
    assert entry.sample_rows[0].total_price == "200"
    assert entry.sample_rows[0].labor_unit == "1.5"
    assert entry.sample_rows[1].issue == "missing_unit_and_quantity"




def test_row_preview_skips_numbering_and_section_rows_and_uses_gesn_column(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/sections.xlsx": _sectioned_workbook_bytes()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.status == STATUS_PREVIEW_OK
    assert entry.header_preview.code == "Перечень ГЭСН/ФЕР/ТЕР/КР"
    assert entry.rows_ok == 2
    assert entry.rows_rejected == 0
    assert [row.row_number for row in entry.sample_rows] == [34, 35]
    assert entry.sample_rows[0].code == "ГЭСНм08-03-572-06 /КР"
    assert entry.sample_rows[0].work_name == "Valid work A"
    assert "11" not in [row.code for row in entry.sample_rows]


def test_row_preview_does_not_use_section_code_as_catalog_code(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/no-code.xlsx": _section_code_only_workbook_bytes()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = analyze_rnmc_zip_row_preview(connection, str(zip_path))
    finally:
        connection.close()

    entry = result.entries[0]
    assert entry.header_preview.code == ""
    assert entry.rows_ok == 1
    assert entry.sample_rows[0].code == ""

def test_row_preview_limit_does_not_change_catalog_import(tmp_path: Path) -> None:
    from app.services.rnmc_excel import import_rnmc_zip_catalog_rows
    from core.storage.catalog import RNMC_ZIP_SOURCE_NAME, count_catalog_rows

    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/large.xlsx": _large_workbook_bytes(rows=45, with_import_columns=True)})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        preview = analyze_rnmc_zip_row_preview(connection, str(zip_path))
        imported = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        total_rows = count_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert preview.rows_ok_total == 30
    assert imported.rows_imported_total == 45
    assert total_rows == 45

def test_admin_can_preview_rnmc_zip_rows(tmp_path: Path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Folder/admin-new.xlsx": _workbook_bytes(task_number="ADMIN-7")})

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        with zip_path.open("rb") as handle:
            response = client.post(
                "/admin/imports/rnmc-row-preview",
                data={"region_override": "Manual Region"},
                files={"rnmc_zip": ("rnmc.zip", handle, "application/zip")},
            )

    assert response.status_code == 200
    assert "Предпросмотр строк РНМЦ" in response.text
    assert "admin-new.xlsx" in response.text
    assert "Manual Region" in response.text
    assert "ADMIN-7" in response.text
    assert "preview_ok" in response.text
    assert "OK-строк в preview" in response.text
    assert "Файлы и статусы" in response.text
    assert "Метаданные" in response.text
    assert "Заголовки" in response.text
    assert "Строки предпросмотра" in response.text
    assert 'data-rnmc-search="rnmc-files"' in response.text
    assert 'data-rnmc-hide-processed="rnmc-files"' in response.text
    assert 'data-rnmc-only-problems="rnmc-rows"' in response.text
    assert "data-rnmc-zoom" in response.text
    assert "Масштаб таблиц" in response.text


def _value_preview_workbook_bytes() -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-PREVIEW"
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Единица измерения",
        "Кол-во",
        "Перечень ГЭСН",
        "Цена единицы работ, руб. с НДС",
        "Итого стоимость, руб. с НДС",
        "ТЗ на ед., чел-час",
        "ТЗ всего, чел-час",
        "ТЗм на ед., чел-час",
        "ТЗм всего, чел-час",
    ])
    sheet.append([1, "Work A", "м", 2, "ГЭСН01-01-001-01", 120, 240, 1.5, 3, 0.5, 1])
    sheet.append([2, "Rejected sample", "", "", "ГЭСН01-01-001-02", 300, 600, 2, 4, 1, 2])
    sheet.append([None] * 11)
    sheet.append([None] * 11)
    sheet.append([None] * 11)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()



def _sectioned_workbook_bytes() -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-SECTIONS"
    for _ in range(2, 29):
        sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Единица измерения",
        "Кол-во",
        "Цена единицы работ, руб. без НДС",
        "Итого стоимость, руб. без НДС",
        "ТЗ на ед., чел-час",
        "ТЗм на ед., чел-час",
        "ТЗ всего, чел-час",
        "ТЗм всего, чел-час",
        "Перечень ГЭСН/ФЕР/ТЕР/КР",
        "Код раздела",
    ])
    sheet.append(list(range(1, 13)))
    sheet.append([None, "1102-01UYF-2029 object title", None, None, None, None, None, None, None, None, None, None])
    sheet.append([None, "Раздел 1. Контрольно-пропускной пункт №2", None, None, None, None, None, None, None, None, None, 14])
    sheet.append([None, "Оборудование", None, None, None, None, None, None, None, None, None, 15])
    sheet.append([1, "Valid work A", "шт", 1, 8371.13, 8371.13, 3.09, 0.54, 3.09, 0.54, "ГЭСНм08-03-572-06 /КР", 15])
    sheet.append([2, "Valid work B", "шт", 1, 1614.79, 1614.79, 1.03, 0.01, 1.03, 0.01, "ГЭСНм11-04-008-01 /КР", 15])
    sheet.append([None] * 12)
    sheet.append([None] * 12)
    sheet.append([None] * 12)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _section_code_only_workbook_bytes() -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-NO-CODE"
    sheet.append([])
    sheet.append([
        "№ п/п",
        "Наименование работ",
        "Единица измерения",
        "Кол-во",
        "Цена единицы работ, руб. без НДС",
        "Код раздела",
    ])
    sheet.append([1, "Work with section code only", "шт", 1, 100, 14])
    sheet.append([None] * 6)
    sheet.append([None] * 6)
    sheet.append([None] * 6)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_zip(path: Path, files: dict[str, bytes]) -> None:
    with ZipFile(path, "w") as archive:
        for name, data in files.items():
            archive.writestr(name, data)


def _write_file_log(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FileLog"
    sheet.append(
        [
            "Folder",
            "File",
            "Status",
            "Год Квартал ЛСР",
            "Планирумый срок начала работ",
            "Планируемый срок окончания работ",
        ]
    )
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _workbook_bytes(*, task_number: str) -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = f"№ задачи 1Ф: {task_number}"
    sheet.append([])
    sheet.append(["№ п/п", "Наименование работ", "Ед.изм.", "Кол-во"])
    sheet.append([1, "Work A", "м", 10])
    sheet.append([2, "Work B", "шт", 5])
    sheet.append([3, "Work without unit and qty", "", ""])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _large_workbook_bytes(*, rows: int, with_import_columns: bool = False) -> bytes:
    from io import BytesIO

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-LARGE"
    sheet.append([])
    if with_import_columns:
        sheet.append([
            "№ п/п",
            "Наименование работ",
            "Ед.изм.",
            "Кол-во",
            "Перечень ГЭСН",
            "Цена единицы работ, руб. без НДС",
        ])
        for index in range(1, rows + 1):
            sheet.append([index, f"Work {index}", "м", index, "ГЭСН01-01-001-01", 100 + index])
    else:
        sheet.append(["№ п/п", "Наименование работ", "Ед.изм.", "Кол-во"])
        for index in range(1, rows + 1):
            sheet.append([index, f"Work {index}", "м", index])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    sheet.append([None, None, None, None])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

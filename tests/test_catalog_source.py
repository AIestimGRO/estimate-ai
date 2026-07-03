"""Tests for catalog resolution (database vs Excel upload)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.catalog_source import (
    CatalogNotAvailableError,
    database_has_catalog,
    load_catalog_for_run,
)
from core.storage import connect, import_catalog_from_excel, init_database

CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
METER = "\u043c"


def _write_catalog_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2, value="TASK-100")
    worksheet.cell(row=4, column=3, value="\u0440\u0430\u0431\u043e\u0442\u0430")
    worksheet.cell(row=4, column=4, value=METER)
    worksheet.cell(row=4, column=7, value=125.5)
    worksheet.cell(row=4, column=14, value=CODE)
    workbook.save(path)
    workbook.close()


def test_load_catalog_from_excel_when_path_given(tmp_path: Path) -> None:
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)

    result = load_catalog_for_run(catalog_path, database_path=tmp_path / "unused.db")

    assert result.row_count == 1
    assert result.source_label == "file:catalog.xlsx"
    assert result.rows[0].code == CODE


def test_load_catalog_from_database_when_no_path(tmp_path: Path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)

    connection = connect(db_path)
    try:
        init_database(connection)
        import_catalog_from_excel(connection, catalog_path, source_name="main")
    finally:
        connection.close()

    result = load_catalog_for_run(None, database_path=db_path)

    assert result.row_count == 1
    assert result.source_label == "database:main"
    assert result.rows[0].code == CODE


def test_excel_upload_overrides_database(tmp_path: Path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    db_catalog = tmp_path / "db_catalog.xlsx"
    upload_catalog = tmp_path / "upload.xlsx"
    _write_catalog_workbook(db_catalog)
    _write_catalog_workbook(upload_catalog)

    connection = connect(db_path)
    try:
        init_database(connection)
        import_catalog_from_excel(connection, db_catalog, source_name="main")
    finally:
        connection.close()

    result = load_catalog_for_run(upload_catalog, database_path=db_path)

    assert result.source_label == "file:upload.xlsx"


def test_empty_database_raises(tmp_path: Path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    connection = connect(db_path)
    try:
        init_database(connection)
    finally:
        connection.close()

    with pytest.raises(CatalogNotAvailableError):
        load_catalog_for_run(None, database_path=db_path)


def test_database_has_catalog(tmp_path: Path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    catalog_path = tmp_path / "catalog.xlsx"
    _write_catalog_workbook(catalog_path)

    assert database_has_catalog(database_path=db_path) is False

    connection = connect(db_path)
    try:
        init_database(connection)
        import_catalog_from_excel(connection, catalog_path, source_name="main")
    finally:
        connection.close()

    assert database_has_catalog(database_path=db_path) is True

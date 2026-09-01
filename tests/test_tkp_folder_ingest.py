"""Direct folder-upload parsing for original KL workbooks."""

from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.web.app import create_app
from core.storage import connect, count_tkp_items, init_database, list_tkp_sources
from core.storage.tkp import import_tkp_parse_result, list_tkp_items
from core.tkp_folder_ingest import parse_tkp_source_workbook


KL_SHEET = "\u041a\u041b 4"
WINNER = '\u041e\u041e\u041e "\u041f\u043e\u0431\u0435\u0434\u0430"'
OTHER = '\u041e\u041e\u041e "\u0414\u0440\u0443\u0433\u043e\u0439"'
ITEM = "\u041c\u043e\u043d\u0442\u0430\u0436 \u043f\u0430\u043d\u0435\u043b\u0435\u0439"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _source_bytes(
    *,
    winner_price: float | None = 90.0,
    reserve_price: float | None = 100.0,
    task_no: object = 12345,
    winner_total: float | None = None,
    reserve_total: float | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = KL_SHEET
    sheet["B1"] = "\u041c\u043e\u043d\u0442\u0430\u0436 \u043e\u0431\u044a\u0435\u043a\u0442\u0430"
    metadata = (
        (6, "1.1.", "\u0414\u0430\u0442\u0430", datetime(2026, 7, 1)),
        (7, "1.2.", "\u0412\u0435\u0440\u0441\u0438\u044f", 1),
        (8, "1.3.", "\u041d\u043e\u043c\u0435\u0440 \u0437\u0430\u0434\u0430\u0447\u0438", task_no),
        (10, "1.5.", "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a", "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a"),
        (12, "1.7.", "\u0413\u0435\u043d\u043f\u043e\u0434\u0440\u044f\u0434\u0447\u0438\u043a", "\u0413\u041f"),
    )
    for row, code, label, value in metadata:
        sheet.cell(row, 1, code)
        sheet.cell(row, 2, label)
        sheet.cell(row, 11, value)
        sheet.cell(row, 15, value)

    sheet["A17"] = "2.2."
    sheet["B17"] = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
    sheet["K17"] = OTHER
    sheet["O17"] = WINNER
    sheet["A18"] = "2.3."
    sheet["B18"] = "\u0418\u041d\u041d"
    sheet["K18"] = "1111111111"
    sheet["O18"] = "2222222222"

    sheet["A37"] = 4
    sheet["B37"] = "\u0411\u043b\u043e\u043a \u0412\u041e\u0420 \u0438 \u0426\u0435\u043d\u0430"
    sheet["K37"] = "\u0446\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434., \u0431\u0435\u0437 \u041d\u0414\u0421"
    sheet["L37"] = "\u0421\u0442-\u0442\u044c \u0432\u0441\u0435\u0433\u043e, \u0431\u0435\u0437 \u041d\u0414\u0421"
    sheet["O37"] = "\u0446\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434., \u0431\u0435\u0437 \u041d\u0414\u0421"
    sheet["P37"] = "\u0421\u0442-\u0442\u044c \u0432\u0441\u0435\u0433\u043e, \u0431\u0435\u0437 \u041d\u0414\u0421"
    sheet["A38"] = "4.1."
    sheet["B38"] = "\u0420\u0430\u0437\u0434\u0435\u043b"
    sheet["A39"] = "4.1.1."
    sheet["B39"] = ITEM
    sheet["C39"] = "\u043c2"
    sheet["D39"] = 10
    sheet["I39"] = 95
    sheet["J39"] = 950
    effective_winner_total = (
        winner_price * 10 if winner_total is None and winner_price is not None else winner_total
    )
    effective_reserve_total = (
        reserve_price * 10 if reserve_total is None and reserve_price is not None else reserve_total
    )
    sheet["K39"] = reserve_price
    sheet["L39"] = effective_reserve_total
    sheet["O39"] = winner_price
    sheet["P39"] = effective_winner_total
    sheet["CM39"] = "\u0421\u043b\u0443\u0436\u0435\u0431\u043d\u043e\u0435 \u043f\u0440\u0438\u043c\u0435\u0447\u0430\u043d\u0438\u0435"

    sheet["A40"] = "4.2."
    sheet["B40"] = "\u043f\u0440\u0435\u0434\u0435\u043b\u044c\u043d\u044b\u0439 \u0440\u0430\u0437\u043c\u0435\u0440 \u0430\u0432\u0430\u043d\u0441\u0430, %"
    sheet["A41"] = "4.3."
    sheet["B41"] = "\u0418\u0442\u043e\u0433\u043e\u0432\u0430\u044f \u0441\u0443\u043c\u043c\u0430 \u043f\u0440\u0435\u0434\u043b\u043e\u0436\u0435\u043d\u0438\u044f, \u0440\u0443\u0431 \u0431\u0435\u0437 \u041d\u0414\u0421"
    sheet["L41"] = effective_reserve_total
    sheet["P41"] = effective_winner_total
    sheet["A42"] = "4.3.1."
    sheet["B42"] = "\u0418\u0442\u043e\u0433\u043e\u0432\u0430\u044f \u0441\u0443\u043c\u043c\u0430 \u043f\u0440\u0435\u0434\u043b\u043e\u0436\u0435\u043d\u0438\u044f, \u0440\u0443\u0431 \u0441 \u041d\u0414\u0421"
    sheet["L42"] = (
        effective_reserve_total * 1.22 if effective_reserve_total is not None else None
    )
    sheet["P42"] = (
        effective_winner_total * 1.22 if effective_winner_total is not None else None
    )

    sheet["A83"] = "10.1."
    sheet["B83"] = "\u0420\u0435\u043a\u043e\u043c\u0435\u043d\u0434\u0443\u0435\u043c\u044b\u0439 \u043f\u043e\u0431\u0435\u0434\u0438\u0442\u0435\u043b\u044c \u041a\u041f"
    sheet["E83"] = "=O17"
    sheet["G83"] = "=P42"
    sheet["H83"] = "\u041b\u0443\u0447\u0448\u0430\u044f \u0446\u0435\u043d\u0430"
    sheet["A84"] = "10.2."
    sheet["B84"] = "\u0420\u0435\u0437\u0435\u0440\u0432\u043d\u044b\u0439 \u043f\u043e\u0431\u0435\u0434\u0438\u0442\u0435\u043b\u044c \u041a\u041f"
    sheet["E84"] = "=K17"
    sheet["G84"] = "=L42"
    sheet["H84"] = "\u0420\u0435\u0437\u0435\u0440\u0432"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _write_source(path: Path, **kwargs) -> None:
    path.write_bytes(_source_bytes(**kwargs))


def _rnmc_like_bytes() -> bytes:
    """Build a non-KL workbook that used to satisfy the loose structural score."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "OS"
    sheet["A17"] = "2.2."
    sheet["B17"] = "Наименование"
    sheet["K17"] = "Поставщик"
    sheet["A37"] = 4
    sheet["B37"] = "Блок ВОР и Цена"
    sheet["K37"] = "цена за ед., без НДС"
    sheet["L37"] = "Ст-ть всего, без НДС"
    sheet["A39"] = "4.1.1."
    sheet["B39"] = "Посторонняя строка РНМЦ"
    sheet["C39"] = "шт"
    sheet["D39"] = 2
    sheet["K39"] = 50
    sheet["L39"] = 100
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_original_kl_parser_finds_structural_sheet_and_winner(tmp_path) -> None:
    source = tmp_path / "sample.xlsx"
    _write_source(source)

    result = parse_tkp_source_workbook(source, display_path="folder/sample.xlsx")

    assert len(result.files) == 1
    parsed_file = result.files[0]
    assert parsed_file.sheet_name == KL_SHEET
    assert parsed_file.parse_status == "OK"
    assert parsed_file.winner_name == WINNER
    assert len(result.items) == 1
    item = result.items[0]
    assert item.item_name == ITEM
    assert item.winner_group_index == 2
    assert item.winner_start_col == 15
    assert item.winner_unit_price_no_vat == 90.0
    assert item.rnmc_unit_price_no_vat == 95.0
    assert item.reserve_name == OTHER
    assert item.reserve_method == "block10_reserve"
    assert item.reserve_start_col == 11
    assert item.reserve_unit_price_no_vat == 100.0
    assert len(result.diagnostics) == 1
    diagnostics = result.diagnostics[0]
    assert diagnostics.wor_start_row == 37
    assert diagnostics.rows_found == 1
    assert diagnostics.rows_with_usable_price == 1
    blocks = {block.code: block for block in diagnostics.blocks}
    assert blocks["1.3"].found is True
    assert blocks["1.3"].value == "12345"
    assert blocks["10.1"].found is True
    assert blocks["10.2"].found is True




def test_original_kl_parser_reads_lot_schema_unit_prices(tmp_path) -> None:
    source = tmp_path / "lot-schema.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = KL_SHEET
    sheet["B1"] = "Монтаж объекта"
    for row, code, label, value in (
        (6, "1.1.", "Дата", datetime(2026, 7, 1)),
        (7, "1.2.", "Версия", 1),
        (8, "1.3.", "Номер задачи", 12345),
    ):
        sheet.cell(row, 1, code)
        sheet.cell(row, 2, label)
        sheet.cell(row, 13, value)
        sheet.cell(row, 25, value)
    sheet["A17"] = "2.2."
    sheet["B17"] = "Наименование"
    sheet["M17"] = WINNER
    sheet["Y17"] = OTHER
    sheet["A18"] = "2.3."
    sheet["B18"] = "ИНН"
    sheet["M18"] = "2222222222"
    sheet["Y18"] = "1111111111"
    sheet["A37"] = 4
    sheet["B37"] = "Блок ВОР и Цена"
    sheet["I37"] = "цена за ед., без НДС"
    sheet["J37"] = "ед.изм."
    sheet["K37"] = "кол-во"
    sheet["L37"] = "Ст-ть всего, без НДС"
    sheet["M37"] = "цена за ед., без НДС"
    sheet["N37"] = "Ст-ть всего, без НДС"
    sheet["Y37"] = "цена за ед., без НДС"
    sheet["Z37"] = "Ст-ть всего, без НДС"
    sheet["A38"] = "4.1.1"
    sheet["B38"] = ITEM
    sheet["C38"] = "шт"
    sheet["D38"] = 10
    sheet["I38"] = 95
    sheet["J38"] = "шт"
    sheet["K38"] = 10
    sheet["L38"] = 950
    sheet["M38"] = 90
    sheet["N38"] = 900
    sheet["Y38"] = 100
    sheet["Z38"] = 1000
    sheet["A39"] = "4.3."
    sheet["B39"] = "Итоговая сумма предложения, руб без НДС"
    sheet["N39"] = 900
    sheet["Z39"] = 1000
    sheet["A83"] = "10.1."
    sheet["B83"] = "Рекомендуемый победитель КП"
    sheet["E83"] = "=M17"
    sheet["A84"] = "10.2."
    sheet["B84"] = "Резервный победитель КП"
    sheet["E84"] = "=Y17"
    workbook.save(source)
    workbook.close()

    result = parse_tkp_source_workbook(source)

    assert result.files[0].parse_status == "OK"
    assert result.files[0].reserve_name == OTHER
    assert len(result.items) == 1
    item = result.items[0]
    assert item.wor_schema == "rnmc_lot_i_l"
    assert item.unit == "шт"
    assert item.qty == 10.0
    assert item.rnmc_unit_price_no_vat == 95.0
    assert item.rnmc_line_total_no_vat == 950.0
    assert item.winner_start_col == 13
    assert item.winner_unit_price_no_vat == 90.0
    assert item.reserve_start_col == 25
    assert item.reserve_unit_price_no_vat == 100.0



def test_rnmc_like_workbook_is_not_classified_as_kl(tmp_path) -> None:
    source = tmp_path / "rnmc-like.xlsx"
    source.write_bytes(_rnmc_like_bytes())

    parsed = parse_tkp_source_workbook(source)

    assert len(parsed.files) == 1
    assert parsed.files[0].parse_status == "Skipped"
    assert parsed.files[0].sheet_name == ""
    assert parsed.items == []

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_tkp_parse_result(connection, parsed)
        assert result.files_skipped == 1
        assert count_tkp_items(connection) == 0
        assert list_tkp_sources(connection) == []
    finally:
        connection.close()


def test_import_accepts_either_unit_price_and_rejects_rows_with_neither(tmp_path) -> None:
    reserve_only_path = tmp_path / "reserve-only.xlsx"
    winner_only_path = tmp_path / "winner-only.xlsx"
    no_price_path = tmp_path / "no-price.xlsx"
    _write_source(
        reserve_only_path,
        winner_price=None,
        reserve_price=100.0,
        winner_total=900.0,
    )
    _write_source(
        winner_only_path,
        winner_price=90.0,
        reserve_price=None,
        reserve_total=1000.0,
    )
    _write_source(
        no_price_path,
        winner_price=None,
        reserve_price=None,
        winner_total=900.0,
        reserve_total=1000.0,
    )

    reserve_only = parse_tkp_source_workbook(reserve_only_path)
    winner_only = parse_tkp_source_workbook(winner_only_path)
    no_price = parse_tkp_source_workbook(no_price_path)

    assert reserve_only.items[0].winner_unit_price_no_vat is None
    assert reserve_only.items[0].winner_line_total_no_vat == 900.0
    assert reserve_only.items[0].reserve_unit_price_no_vat == 100.0
    assert winner_only.items[0].winner_unit_price_no_vat == 90.0
    assert winner_only.items[0].reserve_unit_price_no_vat is None
    assert winner_only.items[0].reserve_line_total_no_vat == 1000.0
    assert no_price.items[0].winner_unit_price_no_vat is None
    assert no_price.items[0].winner_line_total_no_vat == 900.0
    assert no_price.items[0].reserve_unit_price_no_vat is None
    assert no_price.items[0].reserve_line_total_no_vat == 1000.0
    assert no_price.diagnostics[0].rows_rejected_missing_prices == 1

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_tkp_parse_result(connection, reserve_only)
        import_tkp_parse_result(connection, winner_only)
        import_tkp_parse_result(connection, no_price)

        rows = list_tkp_items(connection, limit=100)
        assert len(rows) == 2
        by_file = {row.source_file_name: row for row in rows}
        assert by_file["reserve-only.xlsx"].winner_unit_price_no_vat is None
        assert by_file["reserve-only.xlsx"].reserve_unit_price_no_vat == 100.0
        assert by_file["winner-only.xlsx"].winner_unit_price_no_vat == 90.0
        assert by_file["winner-only.xlsx"].reserve_unit_price_no_vat is None
        assert "no-price.xlsx" not in by_file
    finally:
        connection.close()


def test_storage_rejects_tkp_item_without_task_number(tmp_path) -> None:
    source = tmp_path / "missing-task.xlsx"
    _write_source(source, task_no=None)
    parsed = parse_tkp_source_workbook(source)

    assert parsed.files[0].parse_status == "Needs review"
    assert parsed.files[0].task_no == ""
    assert len(parsed.items) == 1

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_tkp_parse_result(connection, parsed)
        assert count_tkp_items(connection) == 0
    finally:
        connection.close()


def test_content_revision_skips_unchanged_and_updates_changed_file(tmp_path) -> None:
    source = tmp_path / "same-name.xlsx"
    db_path = tmp_path / "estimate_ai.db"
    _write_source(source, winner_price=90.0)
    first = parse_tkp_source_workbook(source)

    connection = connect(db_path)
    try:
        init_database(connection)
        imported = import_tkp_parse_result(connection, first)
        unchanged = import_tkp_parse_result(connection, first)
        assert imported.files_imported == 1
        assert unchanged.files_skipped == 1

        _write_source(source, winner_price=80.0)
        changed = parse_tkp_source_workbook(source)
        updated = import_tkp_parse_result(connection, changed)
        assert updated.files_updated == 1
        assert count_tkp_items(connection) == 1
        assert list_tkp_items(connection)[0].winner_unit_price_no_vat == 80.0
    finally:
        connection.close()


def test_admin_uploads_original_kl_folder_through_staged_preview(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        page = client.get("/admin/tkp")
        assert 'action="/admin/tkp/import-folder"' in page.text
        assert "webkitdirectory" in page.text
        assert "открыть предпросмотр" in page.text.lower()

        response = client.post(
            "/admin/tkp/import-folder",
            files=[
                (
                    "tkp_files",
                    ("nested/sample.xlsx", _source_bytes(), XLSX_MIME),
                )
            ],
            follow_redirects=False,
        )
        assert response.status_code == 303
        stage_url = response.headers["location"]
        assert stage_url.startswith("/admin/tkp/stage/")

        connection = connect(db_path)
        try:
            init_database(connection)
            assert count_tkp_items(connection) == 0
        finally:
            connection.close()

        preview = client.get(stage_url)
        assert preview.status_code == 200
        assert "Макропредпросмотр ТКП" in preview.text
        assert "10.1" in preview.text
        assert "10.2" in preview.text
        assert "Открыть исходный файл" in preview.text
        assert 'name="task_no__0"' in preview.text
        assert 'value="12345"' in preview.text

        rows = client.get(stage_url + "/rows")
        assert rows.status_code == 200
        assert "Построчный предпросмотр ТКП" in rows.text
        assert "READY" in rows.text
        assert "\u041f\u043e\u0431\u0435\u0434\u0430" in rows.text
        assert "\u0414\u0440\u0443\u0433\u043e\u0439" in rows.text

        source_file = client.get(stage_url + "/file/0")
        assert source_file.status_code == 200
        assert source_file.content[:2] == b"PK"

        commit = client.post(
            stage_url + "/commit",
            data={"task_no__0": "12345"},
            follow_redirects=False,
        )
        assert commit.status_code == 303
        assert commit.headers["location"].startswith("/admin/tkp?message=")

    connection = connect(db_path)
    try:
        assert count_tkp_items(connection) == 1
        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].file_name == "sample.xlsx"
        assert sources[0].winner_name == WINNER
        assert sources[0].task_no == "12345"
    finally:
        connection.close()


def test_admin_mixed_folder_skips_non_kl_files(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post(
            "/admin/tkp/import-folder",
            files=[
                ("tkp_files", ("mixed/valid-kl.xlsx", _source_bytes(), XLSX_MIME)),
                ("tkp_files", ("mixed/rnmc.xlsx", _rnmc_like_bytes(), XLSX_MIME)),
            ],
            follow_redirects=False,
        )
        stage_url = response.headers["location"]

        preview = client.get(stage_url)
        assert preview.status_code == 200
        assert "КЛ — готовы" in preview.text
        assert "Не КЛ — автоматически пропущены" in preview.text
        assert "NOT_KL" in preview.text
        assert "rnmc.xlsx" in preview.text

        commit = client.post(
            stage_url + "/commit",
            data={"task_no__0": "12345"},
            follow_redirects=False,
        )
        assert commit.status_code == 303

    connection = connect(db_path)
    try:
        rows = list_tkp_items(connection, limit=10)
        assert len(rows) == 1
        assert rows[0].source_file_name == "valid-kl.xlsx"
        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].file_name == "valid-kl.xlsx"
    finally:
        connection.close()


def test_admin_can_manually_exclude_kl_file_before_commit(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post(
            "/admin/tkp/import-folder",
            files=[
                ("tkp_files", ("mixed/keep.xlsx", _source_bytes(), XLSX_MIME)),
                (
                    "tkp_files",
                    ("mixed/exclude.xlsx", _source_bytes(task_no=None), XLSX_MIME),
                ),
            ],
            follow_redirects=False,
        )
        stage_url = response.headers["location"]

        preview = client.get(stage_url)
        assert preview.status_code == 200
        assert 'name="exclude__0"' in preview.text
        assert 'name="exclude__1"' in preview.text
        assert "NEEDS_TASK_NUMBER" in preview.text

        commit = client.post(
            stage_url + "/commit",
            data={
                "task_no__0": "12345",
                "task_no__1": "",
                "exclude__1": "1",
            },
            follow_redirects=False,
        )
        assert commit.status_code == 303

    connection = connect(db_path)
    try:
        rows = list_tkp_items(connection, limit=10)
        assert len(rows) == 1
        assert rows[0].source_file_name == "keep.xlsx"
        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].file_name == "keep.xlsx"
    finally:
        connection.close()


def test_admin_missing_task_can_be_filled_before_tkp_commit(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post(
            "/admin/tkp/import-folder",
            files=[
                (
                    "tkp_files",
                    ("nested/missing-task.xlsx", _source_bytes(task_no=None), XLSX_MIME),
                )
            ],
            follow_redirects=False,
        )
        stage_url = response.headers["location"]

        preview = client.get(stage_url)
        assert preview.status_code == 200
        assert "NEEDS_TASK_NUMBER" in preview.text
        assert 'name="task_no__0"' in preview.text
        assert 'value=""' in preview.text

        blocked = client.post(
            stage_url + "/commit",
            data={"task_no__0": ""},
        )
        assert blocked.status_code == 400
        assert "Номер задачи обязателен" in blocked.text

        connection = connect(db_path)
        try:
            init_database(connection)
            assert count_tkp_items(connection) == 0
        finally:
            connection.close()

        committed = client.post(
            stage_url + "/commit",
            data={"task_no__0": "777001"},
            follow_redirects=False,
        )
        assert committed.status_code == 303

    connection = connect(db_path)
    try:
        rows = list_tkp_items(connection, limit=10)
        assert len(rows) == 1
        assert rows[0].task_no == "777001"
        sources = list_tkp_sources(connection)
        assert sources[0].task_no == "777001"
        assert sources[0].parse_status == "OK"
    finally:
        connection.close()


def test_admin_folder_upload_ignores_non_excel_files(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post(
            "/admin/tkp/import-folder",
            files=[("tkp_files", ("notes.txt", b"not excel", "text/plain"))],
            follow_redirects=False,
        )

    assert response.status_code == 303
    assert "error=" in response.headers["location"]

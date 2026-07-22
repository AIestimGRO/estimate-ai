"""Tests for the admin TKP catalog page."""

from datetime import datetime
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.web.app import create_app
from core.storage import connect, count_tkp_items, init_database, list_tkp_sources
from core.tkp_ingest import FILE_CATALOG_HEADERS, FILE_CATALOG_SHEET, WOR_CATALOG_HEADERS, WOR_CATALOG_SHEET

RUN_ID = "20260624_172128"
FILE_NAME = "sample.xlsx"
FILE_PATH = "C:\\KL20\\sample.xlsx"
ITEM_NAME = "\u0420\u0430\u0437\u0431\u043e\u0440\u043a\u0430 \u0442\u0435\u043f\u043b\u043e\u0432\u043e\u0439 \u0438\u0437\u043e\u043b\u044f\u0446\u0438\u0438"
WINNER = '\u041e\u041e\u041e "\u041f\u0440\u0438\u043c\u0435\u0440"'
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"


def _file_row(**overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID, "FilePath": FILE_PATH, "FileName": FILE_NAME,
        "ModifiedDate": datetime(2026, 6, 17, 10, 0, 0), "SheetName": "\u041a\u041b 2.0",
        "ParseStatus": "OK", "ParseMessage": "Parsed.",
        "UsedRows": 10, "UsedCols": 10, "WorStartRow": 1, "WorEndRow": 2,
        "WorEndMethod": "advance_limit", "WorPositionCount": 1,
        "WorSectionCount": 1, "ParticipantCount": 1, "WinnerGroupIndex": 1,
        "WinnerStartCol": 1, "WinnerStartColLetter": "A",
        "WinnerMethod": "block10_recommended", "WinnerName": WINNER,
        "WinnerINN": "1234567890", "WinnerUIN": "uin-1",
        "WinnerTaskNoBK": 111, "WinnerHeader1": "", "WinnerHeader2": "",
        "WinnerHeader3": "", "WinnerHeader4": "", "WinnerBlockName": WINNER,
        "WinnerBlockUIN": "", "WinnerBlockTotalVat": 1000.0,
        "WinnerBlockReason": "", "WinnerBlockSource": "block10",
        "TaskNo": "111", "RequestDate": datetime(2026, 5, 1), "Version": 1,
        "Customer": "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a", "GeneralContractor": "",
        "ProcedureName": "\u041e\u0431\u044a\u0435\u043a\u0442", "WinnerTotalNoVat": 900.0,
        "WinnerTotalVat": 1080.0, "RnmcTotalNoVat": 950.0,
    }
    values.update(overrides)
    return tuple(values[header] for header in FILE_CATALOG_HEADERS)


def _wor_row(**overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID, "FilePath": FILE_PATH, "FileName": FILE_NAME,
        "SheetName": "\u041a\u041b 2.0", "SourceRow": 5, "SectionCode": "1.",
        "SectionName": "\u0420\u0430\u0437\u0434\u0435\u043b 1", "SubsectionName": "",
        "ItemCode": "1.1", "ItemName": ITEM_NAME, "Unit": "\u043c2",
        "Qty": 10.0, "QtySourceText": "10", "RnmcUnitPriceNoVat": 100.0,
        "RnmcLineTotalNoVat": 1000.0, "WinnerUnitPriceNoVat": 500.0,
        "WinnerLineTotalNoVat": 5000.0, "WinnerName": WINNER,
        "WinnerINN": "1234567890", "WinnerUIN": "uin-1", "WinnerGroupIndex": 1,
        "WinnerStartCol": 1, "WinnerStartColLetter": "A",
        "WinnerUnitHeader": "", "WinnerTotalHeader": "", "TaskNo": "111",
        "RequestDate": datetime(2026, 5, 1), "Version": 1,
        "Customer": "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a", "GeneralContractor": "",
        "ProcedureName": "\u041e\u0431\u044a\u0435\u043a\u0442",
        "WinnerMethod": "block10_recommended", "WinnerBlockName": WINNER,
        "WinnerBlockUIN": "", "WinnerBlockTotalVat": 1000.0,
        "WinnerBlockReason": "", "WinnerBlockSource": "block10",
    }
    values.update(overrides)
    return tuple(values[header] for header in WOR_CATALOG_HEADERS)


def _catalog_bytes(file_rows: list[tuple], wor_rows: list[tuple]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    files_sheet = workbook.create_sheet(FILE_CATALOG_SHEET)
    files_sheet.append(FILE_CATALOG_HEADERS)
    for row in file_rows:
        files_sheet.append(row)
    wor_sheet = workbook.create_sheet(WOR_CATALOG_SHEET)
    wor_sheet.append(WOR_CATALOG_HEADERS)
    for row in wor_rows:
        wor_sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_admin_tkp_shows_empty_state(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.get("/admin/tkp")

    assert response.status_code == 200
    assert "\u041a\u0430\u0442\u0430\u043b\u043e\u0433 \u0422\u041a\u041f" in response.text
    assert "\u0424\u0430\u0439\u043b\u044b \u0422\u041a\u041f \u043f\u043e\u043a\u0430 \u043d\u0435 \u0438\u043c\u043f\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u0430\u043d\u044b" in response.text
    assert 'action="/admin/tkp/import"' in response.text
    assert 'class="admin-nav-link active" href="/admin/tkp"' in response.text


def test_admin_tkp_import_uploads_and_stores_catalog(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    content = _catalog_bytes([_file_row()], [_wor_row()])

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post(
            "/admin/tkp/import",
            files={"tkp_catalog": ("catalog.xlsm", content, XLSM_MIME)},
            follow_redirects=False,
        )
        assert response.status_code == 303
        assert response.headers["location"].startswith("/admin/tkp?message=")

        page = client.get("/admin/tkp")

    assert page.status_code == 200
    assert FILE_NAME in page.text
    assert ITEM_NAME in page.text
    assert "\u041f\u0440\u0438\u043c\u0435\u0440" in page.text  # WINNER, minus the HTML-escaped quotes

    connection = connect(db_path)
    try:
        assert count_tkp_items(connection) == 1
        sources = list_tkp_sources(connection)
        assert len(sources) == 1
        assert sources[0].file_name == FILE_NAME
    finally:
        connection.close()


def test_admin_tkp_reimport_reports_skipped_files(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    content = _catalog_bytes([_file_row()], [_wor_row()])

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        client.post(
            "/admin/tkp/import",
            files={"tkp_catalog": ("catalog.xlsm", content, XLSM_MIME)},
        )
        response = client.post(
            "/admin/tkp/import",
            files={"tkp_catalog": ("catalog.xlsm", content, XLSM_MIME)},
            follow_redirects=False,
        )

    assert response.status_code == 303
    assert "message=" in response.headers["location"]

    connection = connect(db_path)
    try:
        assert count_tkp_items(connection) == 1
    finally:
        connection.close()


def test_admin_tkp_import_without_file_field_is_rejected(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.post("/admin/tkp/import")

    assert response.status_code == 422

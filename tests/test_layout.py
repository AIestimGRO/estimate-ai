from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.layout import (
    COEF_METHOD_CELL,
    COEF_METHOD_DEFAULT,
    COEF_METHOD_EXPLICIT,
    COEF_METHOD_LABEL,
    COEF_METHOD_REGION,
    CATALOG_FIELD_ADDED_DATE,
    CATALOG_FIELD_CODE,
    FIELD_BASE_PRICE,
    FIELD_CODE,
    FIELD_UNIT,
    FIELD_WORK_NAME,
    METHOD_DEFAULT,
    METHOD_DETECTED,
    METHOD_EXPLICIT,
    METHOD_MISSING,
    data_row_numbers,
    format_layout_report,
    load_catalog_layout_config,
    load_layout_config,
    rank_sheets,
    resolve_average_placement,
    resolve_catalog_layout,
    resolve_layout,
    resolve_regional_coefficient,
    select_sheets,
)

# Russian header wording is kept as unicode escapes to keep this file ASCII
# only (AGENTS.md rule 3); the real synonym dictionary lives in
# data/config/layout.json.
WORK_NAME_FULL = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
WORK_NAME_SHORT = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
UNIT_SHORT = "\u0415\u0434. \u0438\u0437\u043c."
UNIT_FULL = "\u0415\u0434\u0438\u043d\u0438\u0446\u0430 \u0438\u0437\u043c\u0435\u0440\u0435\u043d\u0438\u044f"
CODE_OBOSN = "\u041e\u0431\u043e\u0441\u043d\u043e\u0432\u0430\u043d\u0438\u0435"
CODE_SHIFR = "\u0428\u0438\u0444\u0440"
PRICE_PER_UNIT = "\u0426\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434."
PRICE_BASE = "\u0411\u0430\u0437\u043e\u0432\u0430\u044f \u0446\u0435\u043d\u0430"
TITLE = "\u0421\u043c\u0435\u0442\u0430 \u2116 1"
REGION_LABEL = "\u0420\u0435\u0433\u0438\u043e\u043d"
COEFFICIENT_LABEL = "\u041a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442"
REGION_NAME = "\u041c\u043e\u0441\u043a\u0432\u0430"

DEFAULT_COLUMNS = {
    FIELD_WORK_NAME: 3,
    FIELD_UNIT: 4,
    FIELD_CODE: 14,
    FIELD_BASE_PRICE: 6,
}


FULL_HEADERS = {1: WORK_NAME_FULL, 2: UNIT_SHORT, 3: CODE_OBOSN, 4: PRICE_PER_UNIT}


def _sheet(header_row: int, headers: dict[int, str]) -> Worksheet:
    workbook = Workbook()
    worksheet = workbook.active
    for column, text in headers.items():
        worksheet.cell(row=header_row, column=column).value = text
    return worksheet


def _workbook(sheets: list[tuple[str, int, dict[int, str]]]) -> Workbook:
    workbook = Workbook()
    for position, (title, header_row, headers) in enumerate(sheets):
        worksheet = workbook.active if position == 0 else workbook.create_sheet()
        worksheet.title = title
        for column, text in headers.items():
            worksheet.cell(row=header_row, column=column).value = text
    return workbook


def _config():
    return load_layout_config()


def test_load_layout_config_reads_real_file() -> None:
    config = _config()

    assert set(config.fields) >= {
        FIELD_BASE_PRICE,
        FIELD_UNIT,
        FIELD_WORK_NAME,
        FIELD_CODE,
    }
    assert config.field_priority[0] == FIELD_BASE_PRICE
    assert config.min_matched_fields >= 1


def test_load_catalog_layout_config_reads_real_file() -> None:
    config = load_catalog_layout_config()

    assert config is not None
    assert CATALOG_FIELD_ADDED_DATE in config.fields
    assert CATALOG_FIELD_CODE in config.fields
    assert config.data_start_offset == 1


def test_detects_standard_headers() -> None:
    worksheet = _sheet(
        3,
        {2: WORK_NAME_FULL, 4: UNIT_SHORT, 6: CODE_OBOSN, 8: PRICE_PER_UNIT},
    )

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert result.ok
    assert result.header_row == 3
    assert result.column(FIELD_WORK_NAME) == 2
    assert result.column(FIELD_UNIT) == 4
    assert result.column(FIELD_CODE) == 6
    assert result.column(FIELD_BASE_PRICE) == 8
    assert result.columns[FIELD_CODE].method == METHOD_DETECTED


def test_detects_synonyms_in_shuffled_columns() -> None:
    worksheet = _sheet(
        1,
        {5: PRICE_BASE, 9: UNIT_FULL, 12: CODE_SHIFR, 20: WORK_NAME_SHORT},
    )

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert result.ok
    assert result.column(FIELD_BASE_PRICE) == 5
    assert result.column(FIELD_UNIT) == 9
    assert result.column(FIELD_CODE) == 12
    assert result.column(FIELD_WORK_NAME) == 20


def test_header_row_located_below_title_rows() -> None:
    worksheet = _sheet(
        4,
        {2: WORK_NAME_FULL, 3: UNIT_SHORT, 4: CODE_OBOSN, 5: PRICE_PER_UNIT},
    )
    worksheet.cell(row=1, column=1).value = TITLE
    worksheet.cell(row=2, column=1).value = TITLE

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert result.header_row == 4
    assert result.ok


def test_missing_base_price_reports_key_data_not_found() -> None:
    worksheet = _sheet(
        2,
        {1: WORK_NAME_FULL, 2: UNIT_SHORT, 3: CODE_OBOSN},
    )

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert not result.ok
    assert result.missing_required == [FIELD_BASE_PRICE]
    assert result.columns[FIELD_BASE_PRICE].method == METHOD_MISSING


def test_explicit_override_wins_over_detection() -> None:
    worksheet = _sheet(
        1,
        {2: WORK_NAME_FULL, 4: UNIT_SHORT, 6: CODE_OBOSN, 8: PRICE_PER_UNIT},
    )

    result = resolve_layout(
        worksheet,
        _config(),
        default_columns=DEFAULT_COLUMNS,
        explicit_columns={FIELD_BASE_PRICE: 99},
    )

    assert result.ok
    assert result.column(FIELD_BASE_PRICE) == 99
    assert result.columns[FIELD_BASE_PRICE].method == METHOD_EXPLICIT


def test_optional_field_falls_back_to_default() -> None:
    worksheet = _sheet(
        1,
        {4: UNIT_SHORT, 6: CODE_OBOSN, 8: PRICE_PER_UNIT},
    )

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert result.ok
    work_name = result.columns[FIELD_WORK_NAME]
    assert work_name.method == METHOD_DEFAULT
    assert work_name.column == DEFAULT_COLUMNS[FIELD_WORK_NAME]


def test_no_headers_returns_missing_required() -> None:
    worksheet = _sheet(1, {})

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert not result.ok
    assert result.header_row == 0
    assert set(result.missing_required) == {
        FIELD_CODE,
        FIELD_UNIT,
        FIELD_BASE_PRICE,
    }


def test_code_column_not_confused_with_price() -> None:
    worksheet = _sheet(
        1,
        {1: WORK_NAME_FULL, 2: UNIT_SHORT, 3: PRICE_PER_UNIT, 4: CODE_OBOSN},
    )

    result = resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)

    assert result.column(FIELD_BASE_PRICE) == 3
    assert result.column(FIELD_CODE) == 4


def test_report_contains_header_row_and_fields() -> None:
    worksheet = _sheet(
        2,
        {1: WORK_NAME_FULL, 2: UNIT_SHORT, 3: CODE_OBOSN, 4: PRICE_PER_UNIT},
    )

    report = format_layout_report(
        resolve_layout(worksheet, _config(), default_columns=DEFAULT_COLUMNS)
    )

    assert "header_row: 2" in report
    assert FIELD_BASE_PRICE in report


def test_regional_coefficient_from_stacked_labels() -> None:
    worksheet = _sheet(2, {1: REGION_LABEL, 2: REGION_NAME})
    worksheet.cell(row=3, column=1).value = COEFFICIENT_LABEL
    worksheet.cell(row=3, column=2).value = 1.15

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.15
    assert result.method == COEF_METHOD_REGION
    assert result.region == REGION_NAME
    assert result.label_cell == (3, 1)


def test_regional_coefficient_from_standalone_label() -> None:
    worksheet = _sheet(5, {3: COEFFICIENT_LABEL})
    worksheet.cell(row=5, column=4).value = "1,2"

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.2
    assert result.method == COEF_METHOD_LABEL


def test_regional_coefficient_defaults_to_one_when_absent() -> None:
    worksheet = _sheet(1, {1: TITLE})

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.0
    assert result.method == COEF_METHOD_DEFAULT


def test_regional_coefficient_explicit_value_wins() -> None:
    worksheet = _sheet(2, {1: REGION_LABEL, 2: REGION_NAME})
    worksheet.cell(row=3, column=1).value = COEFFICIENT_LABEL
    worksheet.cell(row=3, column=2).value = 1.15

    result = resolve_regional_coefficient(worksheet, _config(), explicit_value="1,3")

    assert result.value == 1.3
    assert result.method == COEF_METHOD_EXPLICIT


def test_regional_coefficient_ignores_nonpositive_value() -> None:
    worksheet = _sheet(2, {1: REGION_LABEL, 2: REGION_NAME})
    worksheet.cell(row=3, column=1).value = COEFFICIENT_LABEL
    worksheet.cell(row=3, column=2).value = 0

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.0
    assert result.method == COEF_METHOD_DEFAULT


def test_regional_coefficient_reads_configured_value_cell() -> None:
    worksheet = _sheet(1, {1: COEFFICIENT_LABEL})
    worksheet["D20"].value = 1.4

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.4
    assert result.method == COEF_METHOD_CELL


def test_configured_coefficient_cell_keeps_adjacent_region() -> None:
    worksheet = _sheet(19, {3: REGION_LABEL, 4: REGION_NAME})
    worksheet.cell(row=20, column=3).value = COEFFICIENT_LABEL
    worksheet["D20"].value = 1.4

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.4
    assert result.method == COEF_METHOD_CELL
    assert result.region == REGION_NAME


def test_region_is_kept_when_coefficient_defaults() -> None:
    worksheet = _sheet(19, {3: REGION_LABEL, 4: REGION_NAME})

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.0
    assert result.method == COEF_METHOD_DEFAULT
    assert result.region == REGION_NAME


def test_regional_coefficient_skips_placeholder_in_value_cell() -> None:
    worksheet = _sheet(1, {1: COEFFICIENT_LABEL})
    worksheet["D20"].value = "(\u0437\u0430\u043f\u043e\u043b\u043d\u0438\u0442\u044c)"

    result = resolve_regional_coefficient(worksheet, _config())

    assert result.value == 1.0
    assert result.method == COEF_METHOD_DEFAULT


def test_average_placement_uses_adjacent_free_column() -> None:
    placement = resolve_average_placement(8, occupied_columns={6, 8})

    assert placement.column == 9
    assert placement.needs_insert is False


def test_average_placement_inserts_when_neighbour_occupied() -> None:
    placement = resolve_average_placement(8, occupied_columns={8, 9})

    assert placement.column == 9
    assert placement.needs_insert is True


def test_rank_sheets_puts_resolvable_first() -> None:
    workbook = _workbook(
        [
            ("empty", 1, {1: TITLE}),
            ("data", 1, FULL_HEADERS),
        ]
    )

    ranked = rank_sheets(workbook, _config(), default_columns=DEFAULT_COLUMNS)

    assert ranked[0].title == "data"
    assert ranked[0].ok
    assert ranked[0].score == 4
    assert not ranked[1].ok


def test_select_single_sheet_is_automatic() -> None:
    workbook = _workbook([("only", 1, FULL_HEADERS)])

    selection = select_sheets(workbook, _config(), default_columns=DEFAULT_COLUMNS)

    assert [c.title for c in selection.chosen] == ["only"]
    assert selection.needs_user_choice is False


def test_select_picks_the_single_resolvable_sheet() -> None:
    workbook = _workbook(
        [
            ("notes", 1, {1: TITLE}),
            ("data", 1, FULL_HEADERS),
        ]
    )

    selection = select_sheets(workbook, _config(), default_columns=DEFAULT_COLUMNS)

    assert [c.title for c in selection.chosen] == ["data"]
    assert selection.needs_user_choice is False


def test_select_flags_user_choice_when_several_resolvable() -> None:
    workbook = _workbook(
        [
            ("first", 1, FULL_HEADERS),
            ("second", 1, FULL_HEADERS),
        ]
    )

    selection = select_sheets(workbook, _config(), default_columns=DEFAULT_COLUMNS)

    assert selection.chosen == []
    assert selection.needs_user_choice is True
    assert {c.title for c in selection.candidates} == {"first", "second"}


def test_select_honours_explicit_titles() -> None:
    workbook = _workbook(
        [
            ("first", 1, FULL_HEADERS),
            ("second", 1, FULL_HEADERS),
        ]
    )

    selection = select_sheets(
        workbook,
        _config(),
        selected_titles={"second"},
        default_columns=DEFAULT_COLUMNS,
    )

    assert [c.title for c in selection.chosen] == ["second"]
    assert selection.needs_user_choice is False


def test_select_returns_empty_when_none_resolvable() -> None:
    workbook = _workbook(
        [
            ("a", 1, {1: TITLE}),
            ("b", 1, {1: TITLE}),
        ]
    )

    selection = select_sheets(workbook, _config(), default_columns=DEFAULT_COLUMNS)

    assert selection.chosen == []
    assert selection.needs_user_choice is False


def test_data_rows_tolerate_isolated_blank_rows() -> None:
    worksheet = _sheet(1, {})
    for row in (2, 3, 5, 6):
        worksheet.cell(row=row, column=1).value = "x"

    assert data_row_numbers(worksheet, 2, [1]) == [2, 3, 5, 6]


def test_data_rows_stop_after_blank_run() -> None:
    worksheet = _sheet(1, {})
    worksheet.cell(row=2, column=1).value = "x"
    worksheet.cell(row=3, column=1).value = "x"
    worksheet.cell(row=9, column=1).value = "x"

    rows = data_row_numbers(worksheet, 2, [1], max_blank_run=3, max_row=9)

    assert rows == [2, 3]


def test_data_rows_treat_any_key_column_value_as_data() -> None:
    worksheet = _sheet(1, {})
    worksheet.cell(row=2, column=1).value = "x"
    worksheet.cell(row=3, column=3).value = "y"

    assert data_row_numbers(worksheet, 2, [1, 3], max_row=3) == [2, 3]

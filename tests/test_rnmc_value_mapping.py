"""Tests for RNMC value column mapping."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from app.services import rnmc_excel
from app.services.rnmc_excel import import_rnmc_zip_catalog_rows
from core.storage import connect, init_database
from core.storage.catalog import (
    RNMC_ZIP_SOURCE_NAME,
    list_catalog_rows,
    list_import_row_logs,
    list_imported_files,
)


def test_import_maps_unit_price_total_and_labor_columns(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/values.xlsx": _value_workbook_bytes(price_header=_unit_price_with_vat())})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert result.success_count == 1
    assert len(rows) == 1
    row = rows[0]
    assert row.quantity == pytest.approx(10.0)
    assert row.price == pytest.approx(1000.0)
    assert row.total_price == pytest.approx(2000.0)
    assert row.labor_unit == pytest.approx(2.5)
    assert row.labor_total == pytest.approx(25.0)
    assert row.machine_labor_unit == pytest.approx(1.5)
    assert row.machine_labor_total == pytest.approx(15.0)


def test_import_accepts_unit_price_with_auxiliary_materials_without_vat(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Tula/aux.xlsx": _value_workbook_bytes(price_header=_unit_price_aux_without_vat())})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_rnmc_zip_catalog_rows(connection, str(zip_path))
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert len(rows) == 1
    assert rows[0].price == pytest.approx(1200.0)


def test_import_rejects_average_price_headers(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Moscow/average.xlsx": _average_only_workbook_bytes()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        record = next(item for item in list_imported_files(connection) if item.filename == "average.xlsx")
        logs = list_import_row_logs(connection, record.id)
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert result.no_data_count == 1
    assert rows == []
    assert [log.reason for log in logs] == ["missing_or_invalid_price"]


def test_import_reads_workbooks_with_data_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Kazan/data-only.xlsx": _value_workbook_bytes(price_header=_unit_price_aux_without_vat())})
    calls: list[dict[str, object]] = []
    original_load_workbook = rnmc_excel.load_workbook

    def wrapped_load_workbook(*args, **kwargs):
        calls.append(dict(kwargs))
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(rnmc_excel, "load_workbook", wrapped_load_workbook)

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        import_rnmc_zip_catalog_rows(connection, str(zip_path))
    finally:
        connection.close()

    assert calls
    assert all(call.get("data_only") is True for call in calls)


def test_import_maps_ztr_labor_headers_to_labor_columns(tmp_path: Path) -> None:
    zip_path = tmp_path / "rnmc.zip"
    _write_zip(zip_path, {"Moscow/ztr.xlsx": _ztr_workbook_bytes()})

    connection = connect(tmp_path / "estimate_ai.db")
    try:
        init_database(connection)
        result = import_rnmc_zip_catalog_rows(connection, str(zip_path))
        rows = list_catalog_rows(connection, source_name=RNMC_ZIP_SOURCE_NAME)
    finally:
        connection.close()

    assert result.success_count == 1
    assert len(rows) == 1
    assert rows[0].quantity == pytest.approx(2.5)
    assert rows[0].labor_unit == pytest.approx(3.5)
    assert rows[0].labor_total == pytest.approx(8.75)


def _write_zip(path: Path, files: dict[str, bytes]) -> None:
    with ZipFile(path, "w") as archive:
        for name, data in files.items():
            archive.writestr(name, data)


def _value_workbook_bytes(*, price_header: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438 1\u0424: TASK-VALUE"
    sheet.append([])
    sheet.append(
        [
            "\u2116 \u043f/\u043f",
            "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442",
            "\u0415\u0434\u0438\u043d\u0438\u0446\u0430 \u0438\u0437\u043c\u0435\u0440\u0435\u043d\u0438\u044f",
            "\u041a\u043e\u043b-\u0432\u043e",
            "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0413\u042d\u0421\u041d/\u041a\u0420",
            price_header,
            "\u0418\u0442\u043e\u0433\u043e \u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c, \u0440\u0443\u0431. \u0441 \u041d\u0414\u0421",
            "\u0422\u0417\u0440 \u043d\u0430 \u0435\u0434., \u0447\u0435\u043b-\u0447\u0430\u0441",
            "\u0422\u0417\u0440 \u0432\u0441\u0435\u0433\u043e, \u0447\u0435\u043b-\u0447\u0430\u0441",
            "\u0422\u0417\u043c \u043d\u0430 \u0435\u0434., \u0447\u0435\u043b-\u0447\u0430\u0441",
            "\u0422\u0417\u043c \u0432\u0441\u0435\u0433\u043e, \u0447\u0435\u043b-\u0447\u0430\u0441",
        ]
    )
    sheet.append(
        [
            1,
            "Work A",
            "\u043c",
            10,
            "GESN01-01-001-01",
            "1,200.00" if "\u0441 \u041d\u0414\u0421" in price_header else "1 200,00",
            2400,
            "2,5",
            25,
            1.5,
            15,
        ]
    )
    sheet.append([None] * 11)
    sheet.append([None] * 11)
    sheet.append([None] * 11)
    return _to_bytes(workbook)


def _ztr_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "№ задачи 1Ф: TASK-ZTR"
    sheet.append([])
    sheet.append(
        [
            "№ п/п",
            "Наименование работ",
            "Ед.изм.",
            "Кол-во",
            "Перечень ГЭСН/ФЕР/ТЕР/КР",
            "Цена единицы работ, руб. без НДС",
            "Итого стоимость, руб. без НДС",
            "ЗТР на ед., чел-час",
            "ЗТР всего, чел-час",
        ]
    )
    sheet.append([1, "Work ZTR", "шт", "2,5", "ГЭСН01-01-001-01", 100, 250, "3,5", "8,75"])
    sheet.append([None] * 9)
    sheet.append([None] * 9)
    sheet.append([None] * 9)
    return _to_bytes(workbook)


def _average_only_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet["A1"] = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438 1\u0424: TASK-AVG"
    sheet.append([])
    sheet.append(
        [
            "\u2116 \u043f/\u043f",
            "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442",
            "\u0415\u0434.\u0438\u0437\u043c",
            "\u041a\u043e\u043b-\u0432\u043e",
            "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d",
            "\u0426\u0435\u043d\u0430 \u0441\u0440\u0435\u0434\u043d\u044f\u044f, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421",
            "\u0418\u0442\u043e\u0433\u043e \u0441\u0442\u043e\u0438\u043c\u043e\u0441\u0442\u044c \u0441\u0440\u0435\u0434\u043d\u044f\u044f, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421",
        ]
    )
    sheet.append([1, "Average only", "\u043c", 1, "GESN01-01-001-01", 100, 100])
    sheet.append([None] * 7)
    sheet.append([None] * 7)
    sheet.append([None] * 7)
    return _to_bytes(workbook)


def _unit_price_aux_without_vat() -> str:
    return (
        "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442 "
        "(\u0441 \u0443\u0447\u0435\u0442\u043e\u043c \u0432\u0441\u043f\u043e\u043c\u043e\u0433\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0445 \u043c\u0430\u0442\u0435\u0440\u0438\u0430\u043b\u043e\u0432), "
        "\u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421 / \u043f\u043e \u0441\u043c\u0435\u0442\u0430\u043c"
    )


def _unit_price_with_vat() -> str:
    return (
        "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442 "
        "(\u0441 \u0443\u0447\u0435\u0442\u043e\u043c \u0432\u0441\u043f\u043e\u043c\u043e\u0433\u0430\u0442\u0435\u043b\u044c\u043d\u044b\u0445 \u043c\u0430\u0442\u0435\u0440\u0438\u0430\u043b\u043e\u0432), "
        "\u0440\u0443\u0431. \u0441 \u041d\u0414\u0421"
    )


def _to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

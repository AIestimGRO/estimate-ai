from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from core.tkp_ingest import (
    FILE_CATALOG_HEADERS,
    FILE_CATALOG_SHEET,
    WOR_CATALOG_HEADERS,
    WOR_CATALOG_SHEET,
    TkpCatalogFormatError,
    parse_tkp_catalog_workbook,
)

RUN_ID = "20260624_172128"
FILE_PATH_1 = "C:\\KL20\\!!!\u041a\u041b 2.0 + \u043a\u0440\u0430\u0442\u043a\u0438\u0439 \u0441\u0432\u043e\u0434.xlsx"
FILE_NAME_1 = "!!!\u041a\u041b 2.0 + \u043a\u0440\u0430\u0442\u043a\u0438\u0439 \u0441\u0432\u043e\u0434.xlsx"
FILE_PATH_2 = "C:\\KL20\\10-\u041a\u041b 2.xlsx"
FILE_NAME_2 = "10-\u041a\u041b 2.xlsx"
ITEM_NAME_1 = "\u0420\u0430\u0437\u0431\u043e\u0440\u043a\u0430 \u0442\u0435\u043f\u043b\u043e\u0432\u043e\u0439 \u0438\u0437\u043e\u043b\u044f\u0446\u0438\u0438: \u0438\u0437 \u0432\u0430\u0442\u044b \u043c\u0438\u043d\u0435\u0440\u0430\u043b\u044c\u043d\u043e\u0439"
WINNER_1 = '\u041e\u041e\u041e "\u041f\u0440\u0438\u043c\u0435\u0440"'


def _file_catalog_row(**overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID,
        "FilePath": FILE_PATH_1,
        "FileName": FILE_NAME_1,
        "ModifiedDate": datetime(2026, 6, 17, 10, 9, 54),
        "SheetName": "\u041a\u041b 2.0. ",
        "ParseStatus": "OK",
        "ParseMessage": "Parsed.",
        "UsedRows": 253, "UsedCols": 91, "WorStartRow": 37, "WorEndRow": 124,
        "WorEndMethod": "advance_limit", "WorPositionCount": 59,
        "WorSectionCount": 27, "ParticipantCount": 3, "WinnerGroupIndex": 1,
        "WinnerStartCol": 11, "WinnerStartColLetter": "K",
        "WinnerMethod": "block10_recommended", "WinnerName": WINNER_1,
        "WinnerINN": 2100006144, "WinnerUIN": "6224821-2026.05.25-1",
        "WinnerTaskNoBK": 5558522, "WinnerHeader1": "\u0446\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434., \u0431\u0435\u0437 \u041d\u0414\u0421",
        "WinnerHeader2": "", "WinnerHeader3": "", "WinnerHeader4": "",
        "WinnerBlockName": WINNER_1, "WinnerBlockUIN": "",
        "WinnerBlockTotalVat": 14186701.95,
        "WinnerBlockReason": "\u0426\u0435\u043d\u0430, \u043a\u0430\u0447\u0435\u0441\u0442\u0432\u043e",
        "WinnerBlockSource": "block10", "TaskNo": 6224821,
        "RequestDate": datetime(2026, 5, 25), "Version": 1,
        "Customer": '\u041e\u041e\u041e "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a"',
        "GeneralContractor": "", "ProcedureName": "\u0422\u0435\u0441\u0442\u043e\u0432\u044b\u0439 \u043e\u0431\u044a\u0435\u043a\u0442",
        "WinnerTotalNoVat": 9900708.03, "WinnerTotalVat": 12078863.80,
        "RnmcTotalNoVat": 20113772.29,
    }
    values.update(overrides)
    return tuple(values[header] for header in FILE_CATALOG_HEADERS)


def _wor_row(**overrides: object) -> tuple:
    values = {
        "RunId": RUN_ID, "FilePath": FILE_PATH_1, "FileName": FILE_NAME_1,
        "SheetName": "\u041a\u041b 2.0. ", "SourceRow": 39,
        "SectionCode": "4.1.", "SectionName": "\u0420\u0430\u0437\u0434\u0435\u043b 1. \u0414\u0435\u043c\u043e\u043d\u0442\u0430\u0436",
        "SubsectionName": "", "ItemCode": "4.1.1", "ItemName": ITEM_NAME_1,
        "Unit": "100 \u043c2", "Qty": 0.39, "QtySourceText": "0,39",
        "RnmcUnitPriceNoVat": 270.44, "RnmcLineTotalNoVat": 10547.24,
        "WinnerUnitPriceNoVat": 26547.77, "WinnerLineTotalNoVat": 10353.63,
        "WinnerName": WINNER_1, "WinnerINN": 2100006144,
        "WinnerUIN": "6224821-2026.05.25-1", "WinnerGroupIndex": 1,
        "WinnerStartCol": 11, "WinnerStartColLetter": "K",
        "WinnerUnitHeader": "\u0446\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434., \u0431\u0435\u0437 \u041d\u0414\u0421",
        "WinnerTotalHeader": "\u0421\u0442-\u0442\u044c \u0432\u0441\u0435\u0433\u043e, \u0431\u0435\u0437 \u041d\u0414\u0421",
        "TaskNo": 6224821, "RequestDate": datetime(2026, 5, 25), "Version": 1,
        "Customer": '\u041e\u041e\u041e "\u0417\u0430\u043a\u0430\u0437\u0447\u0438\u043a"',
        "GeneralContractor": "", "ProcedureName": "\u0422\u0435\u0441\u0442\u043e\u0432\u044b\u0439 \u043e\u0431\u044a\u0435\u043a\u0442",
        "WinnerMethod": "block10_recommended", "WinnerBlockName": WINNER_1,
        "WinnerBlockUIN": "", "WinnerBlockTotalVat": 14186701.95,
        "WinnerBlockReason": "\u0426\u0435\u043d\u0430, \u043a\u0430\u0447\u0435\u0441\u0442\u0432\u043e",
        "WinnerBlockSource": "block10",
    }
    values.update(overrides)
    return tuple(values[header] for header in WOR_CATALOG_HEADERS)


def _make_workbook(path: Path, file_rows: list[tuple], wor_rows: list[tuple]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    files_sheet = workbook.create_sheet(FILE_CATALOG_SHEET)
    files_sheet.append(FILE_CATALOG_HEADERS)
    for row in file_rows:
        files_sheet.append(row)

    wor_sheet = workbook.create_sheet(WOR_CATALOG_SHEET)
    wor_sheet.append(WOR_CATALOG_HEADERS)
    for row in wor_rows:
        wor_sheet.append(row)

    workbook.save(path)
    workbook.close()
    return path


def test_parses_file_catalog_and_wor_catalog(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path / "kl20.xlsm",
        [_file_catalog_row(), _file_catalog_row(
            FilePath=FILE_PATH_2, FileName=FILE_NAME_2, ParseStatus="Skipped",
            ParseMessage="KL20 sheet not found.", WinnerName="", WinnerINN="",
        )],
        [_wor_row()],
    )

    result = parse_tkp_catalog_workbook(path)

    assert result.run_ids == (RUN_ID,)
    assert len(result.files) == 2
    assert len(result.items) == 1

    ok_file = result.files[0]
    assert ok_file.is_parsed
    assert ok_file.winner_name == WINNER_1
    assert ok_file.winner_total_no_vat == 9900708.03

    skipped_file = result.files[1]
    assert not skipped_file.is_parsed
    assert skipped_file.parse_status == "Skipped"

    item = result.items[0]
    assert item.item_name == ITEM_NAME_1
    assert item.unit == "100 \u043c2"
    assert item.qty == 0.39
    assert item.qty_source_text == "0,39"
    assert item.winner_unit_price_no_vat == 26547.77
    assert item.rnmc_unit_price_no_vat == 270.44
    assert item.rnmc_line_total_no_vat == 10547.24
    assert item.winner_line_total_no_vat == 10353.63
    assert item.winner_group_index == 1
    assert item.winner_start_col == 11
    assert item.winner_start_col_letter == "K"
    assert item.winner_unit_header == "\u0446\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434., \u0431\u0435\u0437 \u041d\u0414\u0421"
    assert item.winner_total_header == "\u0421\u0442-\u0442\u044c \u0432\u0441\u0435\u0433\u043e, \u0431\u0435\u0437 \u041d\u0414\u0421"
    assert item.task_no == "6224821"
    assert item.version == "1"
    assert item.winner_method == "block10_recommended"
    assert item.winner_block_name == WINNER_1
    assert item.winner_block_uin == ""
    assert item.winner_block_total_vat == 14186701.95
    assert item.winner_block_reason == "\u0426\u0435\u043d\u0430, \u043a\u0430\u0447\u0435\u0441\u0442\u0432\u043e"


def test_parses_wor_only_catalog_and_derives_source_file(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WOR_CATALOG_SHEET
    worksheet.append(WOR_CATALOG_HEADERS)
    worksheet.append(_wor_row())
    path = tmp_path / "wor-only.xlsx"
    workbook.save(path)
    workbook.close()

    result = parse_tkp_catalog_workbook(path)

    assert result.run_ids == (RUN_ID,)
    assert len(result.files) == 1
    assert result.files[0].file_name == FILE_NAME_1
    assert result.files[0].parse_status == "OK"
    assert len(result.items) == 1


def test_rows_without_item_name_are_skipped(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path / "kl20.xlsm",
        [_file_catalog_row()],
        [_wor_row(), _wor_row(ItemName="", SourceRow=40)],
    )

    result = parse_tkp_catalog_workbook(path)

    assert len(result.items) == 1


def test_missing_wor_sheet_raises_format_error(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    files_sheet = workbook.create_sheet(FILE_CATALOG_SHEET)
    files_sheet.append(FILE_CATALOG_HEADERS)
    path = tmp_path / "broken.xlsm"
    workbook.save(path)
    workbook.close()

    try:
        parse_tkp_catalog_workbook(path)
        raise AssertionError("expected TkpCatalogFormatError")
    except TkpCatalogFormatError:
        pass


def test_missing_expected_column_raises_format_error(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    files_sheet = workbook.create_sheet(FILE_CATALOG_SHEET)
    files_sheet.append([h for h in FILE_CATALOG_HEADERS if h != "ParseStatus"])
    wor_sheet = workbook.create_sheet(WOR_CATALOG_SHEET)
    wor_sheet.append(WOR_CATALOG_HEADERS)
    path = tmp_path / "broken.xlsm"
    workbook.save(path)
    workbook.close()

    try:
        parse_tkp_catalog_workbook(path)
        raise AssertionError("expected TkpCatalogFormatError")
    except TkpCatalogFormatError:
        pass

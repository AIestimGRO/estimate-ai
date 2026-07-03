"""Integration: run_and_write applies macro Name_Exclusions by default."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from app.services.write_result import run_and_write
from core.macro_workbook import NAME_EXCLUSIONS_SHEET

MM = "\u043c\u043c"
CHANGED = "\u0438\u0437\u043c\u0435\u043d"
CODE = "\u0413\u042d\u0421\u041d46-03-001-17"
UNIT = "\u043e\u0442\u0432\u0435\u0440\u0441\u0442\u0438\u0439"
WORK = (
    f"\u041d\u0430 \u043a\u0430\u0436\u0434\u044b\u0435 10 {MM} {CHANGED}\u0435\u043d\u0438\u044f "
    f"\u0433\u043b\u0443\u0431\u0438\u043d\u044b \u0441\u0432\u0435\u0440\u043b\u0435\u043d\u0438\u044f"
)


def _catalog(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
    worksheet.cell(row=4, column=2, value="task-1")
    worksheet.cell(row=4, column=4, value=UNIT)
    worksheet.cell(row=4, column=7, value=25.0)
    worksheet.cell(row=4, column=14, value=f"{CODE}/ \u041a\u0420")
    workbook.save(path)
    workbook.close()


def _estimate(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u041e\u0421"
    worksheet.cell(row=7, column=14, value=CODE)
    worksheet.cell(row=9, column=3, value=WORK)
    worksheet.cell(row=9, column=4, value=UNIT)
    worksheet.cell(row=9, column=6, value=426.0)
    worksheet.cell(row=9, column=14, value=CODE)
    workbook.save(path)
    workbook.close()


def _macro(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = NAME_EXCLUSIONS_SHEET
    worksheet.cell(row=2, column=1, value=1)
    worksheet.cell(row=2, column=2, value="BOTH")
    worksheet.cell(row=2, column=3, value="ALL_WORDS")
    worksheet.cell(row=2, column=4, value=f"{MM}|{CHANGED}")
    workbook.save(path)
    workbook.close()


def test_run_and_write_skips_name_excluded_row_when_macro_on_disk(
    tmp_path: Path,
    monkeypatch,
) -> None:
    repo = tmp_path / "repo"
    config_dir = repo / "data" / "config"
    config_dir.mkdir(parents=True)
    macro = config_dir / "autopodbor.xlsm"
    _macro(macro)
    config_dir.joinpath("macro.json").write_text(
        json.dumps({"workbook": "data/config/autopodbor.xlsm"}),
        encoding="utf-8",
    )
    monkeypatch.setattr("core.macro_workbook.repo_root", lambda: repo)

    catalog = tmp_path / "catalog.xlsx"
    estimate = tmp_path / "estimate.xlsx"
    output = tmp_path / "estimate WA.xlsx"
    _catalog(catalog)
    _estimate(estimate)

    outcome = run_and_write(catalog, estimate, output)

    assert outcome.name_exclusion_rule_count == 1
    assert outcome.result.matched_row_count == 0
    assert outcome.macro_workbook == macro.resolve()

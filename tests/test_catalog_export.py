"""Tests for catalog export to Excel (2026-07: core writer, CLI, admin route)."""

import subprocess
import sys

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.web.app import create_app
from core.storage import connect, init_database
from core.storage.catalog import CATALOG_EXPORT_HEADERS, write_catalog_export_xlsx


def _seed_catalog_row(
    db_path,
    *,
    source_name: str = "rnmc_zip_upload",
    task_id: str = "6539705 / 66",
    region: str = "Якутия",
    code: str = "ГЭСН11-01-001-01",
    price_original: float = 353.4555208,
    price_zlvl: float = 252.4682291,
    total_price: float = 3880.9416,
    labor_unit: float = 0.082,
    labor_total: float = 0.90036,
    machine_labor_unit: float = 0.0024,
    machine_labor_total: float = 0.026352,
    regional_coefficient: float = 1.4,
    lsr_quarter: str = "2 кв. 2026",
    planned_start: str = "2026-07-01",
    planned_finish: str = "2026-12-01",
    source_filename: str = "OS_SM-K_6539705.xlsx",
    added_date: str = "2026-06-01",
    quantity: float = 10.98,
    unit: str = "м2",
    work_name: str = "Устройство гидроизоляции",
) -> int:
    connection = connect(db_path)
    try:
        init_database(connection)
        existing = connection.execute(
            "SELECT id FROM catalog_sources WHERE name = ?", (source_name,)
        ).fetchone()
        if existing is not None:
            source_id = int(existing["id"])
        else:
            cursor = connection.execute(
                "INSERT INTO catalog_sources(name, kind) VALUES (?, ?)",
                (source_name, "rnmc_zip"),
            )
            source_id = int(cursor.lastrowid)
        cursor = connection.execute(
            """
            INSERT INTO catalog_items (
                source_id, task_id, region, code, unit, quantity, work_name,
                price, price_original, price_zlvl, total_price,
                labor_unit, labor_total, machine_labor_unit, machine_labor_total,
                regional_coefficient, lsr_quarter, planned_start, planned_finish,
                added_date, source_region_folder, source_filename, source_row_number
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_id,
                task_id,
                region,
                code,
                unit,
                quantity,
                work_name,
                price_zlvl,
                price_original,
                price_zlvl,
                total_price,
                labor_unit,
                labor_total,
                machine_labor_unit,
                machine_labor_total,
                regional_coefficient,
                lsr_quarter,
                planned_start,
                planned_finish,
                added_date,
                region,
                source_filename,
                474,
            ),
        )
        row_id = int(cursor.lastrowid)
        connection.commit()
        return row_id
    finally:
        connection.close()


def test_write_catalog_export_xlsx_headers_and_values(tmp_path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    _seed_catalog_row(db_path)
    output_path = tmp_path / "export.xlsx"

    connection = connect(db_path)
    try:
        row_count = write_catalog_export_xlsx(connection, output_path)
    finally:
        connection.close()

    assert row_count == 1
    assert output_path.is_file()

    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook.active
        header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert tuple(header_row) == CATALOG_EXPORT_HEADERS

        data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        (
            pp,
            task_id,
            work_name,
            unit,
            quantity,
            price_without_vat,
            price_zlvl,
            total_with_vat,
            total_without_vat,
            labor_unit,
            labor_total,
            machine_labor_unit,
            machine_labor_total,
            code,
            source_file,
            region,
            lsr_quarter,
            planned_start,
            planned_finish,
            regional_coefficient,
            added_date,
        ) = data_row

        assert pp == 1
        assert task_id == "6539705 / 66"
        assert work_name == "Устройство гидроизоляции"
        assert unit == "м2"
        assert quantity == 10.98
        assert price_without_vat == 353.4555208
        assert price_zlvl == 252.4682291
        assert total_without_vat == 3880.9416
        assert abs(total_with_vat - 3880.9416 * 1.2) < 1e-6
        assert labor_unit == 0.082
        assert labor_total == 0.90036
        assert machine_labor_unit == 0.0024
        assert machine_labor_total == 0.026352
        assert code == "ГЭСН11-01-001-01"
        assert source_file == "OS_SM-K_6539705.xlsx"
        assert region == "Якутия"
        assert lsr_quarter == "2 кв. 2026"
        assert planned_start == "2026-07-01"
        assert planned_finish == "2026-12-01"
        assert regional_coefficient == 1.4
        assert added_date == "2026-06-01"
    finally:
        workbook.close()


def test_write_catalog_export_xlsx_respects_filters(tmp_path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    _seed_catalog_row(db_path, region="Якутия", code="A")
    _seed_catalog_row(db_path, region="Москва", code="B")
    output_path = tmp_path / "export.xlsx"

    connection = connect(db_path)
    try:
        row_count = write_catalog_export_xlsx(
            connection, output_path, filters={"region": "Якут"}
        )
    finally:
        connection.close()

    assert row_count == 1
    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook.active
        data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert data_row[15] == "Якутия"
    finally:
        workbook.close()


def test_write_catalog_export_xlsx_handles_missing_total_price(tmp_path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    connection = connect(db_path)
    try:
        init_database(connection)
        cursor = connection.execute(
            "INSERT INTO catalog_sources(name, kind) VALUES (?, ?)",
            ("main", "excel_bulk"),
        )
        source_id = int(cursor.lastrowid)
        connection.execute(
            """
            INSERT INTO catalog_items (source_id, task_id, region, code, unit, work_name, price)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (source_id, "T-1", "", "CODE", "шт", "work", 100.0),
        )
        connection.commit()
        output_path = tmp_path / "export.xlsx"
        row_count = write_catalog_export_xlsx(connection, output_path)
    finally:
        connection.close()

    assert row_count == 1
    workbook = load_workbook(output_path, data_only=False)
    try:
        sheet = workbook.active
        data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert data_row[7] is None
        assert data_row[8] is None
    finally:
        workbook.close()


def test_admin_catalog_export_route_downloads_xlsx(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    _seed_catalog_row(db_path)

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        page = client.get("/admin/catalog")
        assert page.status_code == 200
        assert "Экспорт в Excel" in page.text
        assert "/admin/catalog/export" in page.text

        response = client.get("/admin/catalog/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "1rows.xlsx" in response.headers["content-disposition"]
    assert response.content[:2] == b"PK"


def test_admin_catalog_export_route_respects_filters(tmp_path, monkeypatch) -> None:
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    _seed_catalog_row(db_path, region="Якутия")
    _seed_catalog_row(db_path, region="Москва")

    with TestClient(create_app(base_dir=tmp_path / "work")) as client:
        response = client.get("/admin/catalog/export", params={"region": "Москва"})

    assert response.status_code == 200
    assert "1rows.xlsx" in response.headers["content-disposition"]


def test_cli_export_catalog_writes_file(tmp_path) -> None:
    db_path = tmp_path / "estimate_ai.db"
    _seed_catalog_row(db_path)
    output_path = tmp_path / "cli_export.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "app.cli",
            "--db",
            str(db_path),
            "export-catalog",
            str(output_path),
        ],
        capture_output=True,
        text=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert output_path.is_file()
    assert "1 rows" in result.stdout

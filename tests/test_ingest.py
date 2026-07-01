from pathlib import Path

from openpyxl import Workbook

from core.ingest import FileIdentity, force_reimport_file, ingest_folder


NAME_HEADER = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
UNIT_HEADER = "\u0415\u0434.\u0438\u0437\u043c."
QTY_HEADER = "\u041a\u043e\u043b-\u0432\u043e"
CODE_HEADER = "\u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c"
TASK_LABEL = "\u2116 \u0437\u0430\u0434\u0430\u0447\u0438 1\u0424"
METER = "\u043c"


def write_rnmc_file(
    path: Path,
    *,
    task_number: str = "TASK-1",
    price: object = 100,
    code: object = "gesn01-01-001-01",
    unit: object = METER,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "RNMC"
    worksheet.cell(row=1, column=1).value = f"{TASK_LABEL}: {task_number}"
    worksheet.cell(row=5, column=1).value = "\u2116 \u043f/\u043f"
    worksheet.cell(row=5, column=2).value = NAME_HEADER
    worksheet.cell(row=5, column=3).value = UNIT_HEADER
    worksheet.cell(row=5, column=4).value = QTY_HEADER
    worksheet.cell(row=5, column=5).value = "\u0426\u0435\u043d\u0430"
    worksheet.cell(row=5, column=6).value = CODE_HEADER
    worksheet.cell(row=5, column=7).value = "CatalogAddedDate"
    worksheet.cell(row=6, column=1).value = 1
    worksheet.cell(row=6, column=2).value = "\u0440\u0430\u0431\u043e\u0442\u0430"
    worksheet.cell(row=6, column=3).value = unit
    worksheet.cell(row=6, column=4).value = 10
    worksheet.cell(row=6, column=5).value = price
    worksheet.cell(row=6, column=6).value = code
    workbook.save(path)
    workbook.close()
    return path


def write_no_header_file(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "NoHeader"
    worksheet.cell(row=1, column=1).value = "not an RNMC table"
    workbook.save(path)
    workbook.close()
    return path


def test_normal_run_skips_already_imported_region_filename(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx")
    imported = {FileIdentity("RegionA", "rnmc.xlsx")}

    results = ingest_folder(tmp_path, imported)

    assert len(results) == 1
    assert results[0].skipped
    assert results[0].rows == []


def test_normal_run_processes_new_region_filename(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx")
    imported: set[FileIdentity] = set()

    results = ingest_folder(tmp_path, imported)

    assert len(results) == 1
    assert not results[0].skipped
    assert not results[0].failed
    assert results[0].file_identity == FileIdentity("RegionA", "rnmc.xlsx")
    assert results[0].region == "RegionA"
    assert results[0].task_number == "TASK-1"
    assert len(results[0].rows) == 1
    assert results[0].rows[0].region == "RegionA"
    assert FileIdentity("RegionA", "rnmc.xlsx") in imported


def test_same_filename_in_different_regions_are_distinct(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx", task_number="TASK-A")
    write_rnmc_file(tmp_path / "RegionB" / "rnmc.xlsx", task_number="TASK-B")

    results = ingest_folder(tmp_path, set())

    assert {result.file_identity for result in results} == {
        FileIdentity("RegionA", "rnmc.xlsx"),
        FileIdentity("RegionB", "rnmc.xlsx"),
    }
    assert {result.task_number for result in results} == {"TASK-A", "TASK-B"}


def test_force_reimport_reprocesses_already_imported_file(tmp_path: Path) -> None:
    path = write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx")
    identity = FileIdentity("RegionA", "rnmc.xlsx")

    results = ingest_folder(tmp_path, {identity}, force_reimport_keys={identity})
    single_result = force_reimport_file(path)

    assert len(results) == 1
    assert not results[0].skipped
    assert results[0].force_reimport
    assert len(results[0].rows) == 1
    assert single_result.force_reimport
    assert len(single_result.rows) == 1


def test_missing_price_is_validation_issue_not_imported(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx", price=None)

    result = ingest_folder(tmp_path, set())[0]

    assert result.rows == []
    assert not result.failed
    assert [(issue.row_number, issue.reason) for issue in result.validation_issues] == [
        (6, "missing_or_invalid_price")
    ]


def test_missing_code_is_validation_issue_not_imported(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "rnmc.xlsx", code="")

    result = ingest_folder(tmp_path, set())[0]

    assert result.rows == []
    assert not result.failed
    assert [(issue.row_number, issue.reason) for issue in result.validation_issues] == [
        (6, "missing_or_invalid_code")
    ]


def test_file_without_matching_header_is_failed_file(tmp_path: Path) -> None:
    imported: set[FileIdentity] = set()
    write_no_header_file(tmp_path / "RegionA" / "bad.xlsx")

    first = ingest_folder(tmp_path, imported)[0]
    second = ingest_folder(tmp_path, imported)[0]

    assert first.failed
    assert "no matching header row found" in first.failure_reason
    assert first.validation_issues == []
    assert first.rows == []
    assert second.skipped


def test_region_is_immediate_parent_folder_name(tmp_path: Path) -> None:
    write_rnmc_file(tmp_path / "RegionA" / "Nested" / "rnmc.xlsx")

    result = ingest_folder(tmp_path, set())[0]

    assert result.file_identity == FileIdentity("Nested", "rnmc.xlsx")
    assert result.region == "Nested"
    assert result.rows[0].region == "Nested"

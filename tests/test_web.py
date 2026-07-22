import re
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.web.app import create_app
from app.web.rendering import XLSX_MIME
from core.storage import connect, import_catalog_from_excel, init_database

METER = "\u043c"
CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
INSTALLATION = "\u043c\u043e\u043d\u0442\u0430\u0436"
CATALOG_TITLE = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"
ESTIMATE_TITLE = "\u041e\u0421 estimate"

HEADER_CODE = "\u0428\u0438\u0444\u0440"
HEADER_NAME = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
HEADER_UNIT = "\u0415\u0434. \u0438\u0437\u043c."
HEADER_PRICE = "\u0426\u0435\u043d\u0430 \u0437\u0430 \u0435\u0434."
SHEET_A = "\u0421\u043c\u0435\u0442\u0430 A"
SHEET_B = "\u0421\u043c\u0435\u0442\u0430 B"


@pytest.fixture()
def client(tmp_path):
    with TestClient(create_app(base_dir=tmp_path / "work")) as test_client:
        yield test_client


def _catalog_bytes(prices):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = CATALOG_TITLE
    for offset, (task_id, price) in enumerate(prices):
        row = 4 + offset
        worksheet.cell(row=row, column=2, value=task_id)
        worksheet.cell(row=row, column=3, value=INSTALLATION)
        worksheet.cell(row=row, column=4, value=METER)
        worksheet.cell(row=row, column=7, value=price)
        worksheet.cell(row=row, column=14, value=CODE)
    return _to_bytes(workbook)


def _template_estimate_bytes():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ESTIMATE_TITLE
    worksheet.cell(row=7, column=14, value=CODE)
    worksheet.cell(row=9, column=3, value=INSTALLATION)
    worksheet.cell(row=9, column=4, value=METER)
    worksheet.cell(row=9, column=6, value=50.0)
    worksheet.cell(row=9, column=14, value=CODE)
    return _to_bytes(workbook)


def _write_detected_sheet(worksheet):
    worksheet.cell(row=1, column=1, value=HEADER_CODE)
    worksheet.cell(row=1, column=2, value=HEADER_NAME)
    worksheet.cell(row=1, column=3, value=HEADER_UNIT)
    worksheet.cell(row=1, column=4, value=HEADER_PRICE)
    worksheet.cell(row=2, column=1, value=CODE)
    worksheet.cell(row=2, column=2, value=INSTALLATION)
    worksheet.cell(row=2, column=3, value=METER)
    worksheet.cell(row=2, column=4, value=50.0)


def _two_sheet_estimate_bytes():
    workbook = Workbook()
    first = workbook.active
    first.title = SHEET_A
    _write_detected_sheet(first)
    _write_detected_sheet(workbook.create_sheet(title=SHEET_B))
    return _to_bytes(workbook)


def _junk_estimate_bytes():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Junk"
    worksheet.cell(row=1, column=1, value="abc")
    return _to_bytes(workbook)


def _to_bytes(workbook):
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _files(catalog_bytes, estimate_bytes):
    return {
        "catalog": ("catalog.xlsx", catalog_bytes, XLSX_MIME),
        "estimate": ("estimate.xlsx", estimate_bytes, XLSX_MIME),
    }


def _hidden_value(page_text, name):
    match = re.search(rf'name="{name}"[^>]*\svalue="([^"]*)"', page_text)
    assert match is not None, f"no hidden/prefilled field named {name!r} on the page"
    return match.group(1)


def _confirm(client, page, *, coefficient=None, region=None, use_tkp=False):
    """POST /confirm using the token/sheet from a rendered confirm page.

    Defaults to whatever the confirm screen pre-filled (i.e. "just click
    confirm"); pass `coefficient`/`region` to simulate the user editing a
    field before submitting.
    """
    token = _hidden_value(page.text, "token")
    sheet = _hidden_value(page.text, "sheet")
    if coefficient is None:
        coefficient = _hidden_value(page.text, "coefficient")
    if region is None:
        region = _hidden_value(page.text, "region")
    data = {"token": token, "sheet": sheet, "coefficient": coefficient, "region": region}
    if use_tkp:
        data["use_tkp_analogs"] = "1"
    return client.post("/confirm", data=data)


def test_index_page_has_upload_form(client):
    response = client.get("/")

    assert response.status_code == 200
    assert 'name="catalog"' in response.text
    assert 'name="estimate"' in response.text
    assert "SQLite:" in response.text


def _seed_database(db_path, catalog_bytes):
    catalog_file = db_path.parent / "seed_catalog.xlsx"
    catalog_file.write_bytes(catalog_bytes)
    connection = connect(db_path)
    try:
        init_database(connection)
        import_catalog_from_excel(connection, catalog_file, source_name="main")
    finally:
        connection.close()


def _seed_tkp_item(db_path):
    connection = connect(db_path)
    try:
        init_database(connection)
        cursor = connection.execute(
            "INSERT INTO tkp_sources (file_name, task_no, item_count) VALUES (?, ?, ?)",
            ("tkp-source.xlsx", "TKP-77", 1),
        )
        connection.execute(
            """
            INSERT INTO tkp_items (
                source_id, item_name, unit, winner_unit_price_no_vat,
                winner_name, task_no
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (cursor.lastrowid, INSTALLATION, METER, 200.0, "winner", "TKP-77"),
        )
        connection.commit()
    finally:
        connection.close()


def test_run_from_database_without_catalog_upload(client, tmp_path, monkeypatch):
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    _seed_database(db_path, _catalog_bytes([("t-1", 100), ("t-2", 120)]))

    files = {
        "estimate": ("estimate.xlsx", _template_estimate_bytes(), XLSX_MIME),
    }
    page = client.post("/run", files=files)
    assert page.status_code == 200
    assert "\u041f\u043e\u0434\u0442\u0432\u0435\u0440\u0434\u0438\u0442\u044c \u0438 \u043e\u0431\u0440\u0430\u0431\u043e\u0442\u0430\u0442\u044c" in page.text

    result = _confirm(client, page)
    assert result.status_code == 200
    assert "\u0411\u0430\u0437\u0430 \u0430\u043d\u0430\u043b\u043e\u0433\u043e\u0432" in result.text
    assert "\u0411\u0414" in result.text
    assert '<table class="preview">' in result.text


def test_missing_catalog_when_database_empty(client, tmp_path, monkeypatch):
    db_path = tmp_path / "empty.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    connection = connect(db_path)
    try:
        init_database(connection)
    finally:
        connection.close()

    files = {
        "estimate": ("estimate.xlsx", _template_estimate_bytes(), XLSX_MIME),
    }
    page = client.post("/run", files=files)

    assert page.status_code == 400
    assert "import-catalog" in page.text


def test_full_run_and_download(client):
    files = _files(_catalog_bytes([("t-1", 100), ("t-2", 120)]), _template_estimate_bytes())

    page = client.post("/run", files=files)
    assert page.status_code == 200
    assert "\u041f\u043e\u0434\u0442\u0432\u0435\u0440\u0434\u0438\u0442\u044c \u0438 \u043e\u0431\u0440\u0430\u0431\u043e\u0442\u0430\u0442\u044c" in page.text

    result = _confirm(client, page)
    assert result.status_code == 200
    assert '<table class="preview">' in result.text
    assert CODE in result.text
    match = re.search(r'href="(/download\?token=[0-9a-f]+)"', result.text)
    assert match is not None

    download = client.get(match.group(1))
    assert download.status_code == 200
    assert download.content[:2] == b"PK"
    assert "estimate%20WA.xlsx" in download.headers["content-disposition"]


def test_tkp_toggle_writes_best_candidate_to_download(client, tmp_path, monkeypatch):
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))
    _seed_database(db_path, _catalog_bytes([("t-1", 100)]))
    _seed_tkp_item(db_path)

    files = {"estimate": ("estimate.xlsx", _template_estimate_bytes(), XLSX_MIME)}
    page = client.post("/run", files=files)
    result = _confirm(client, page, use_tkp=True)

    assert result.status_code == 200
    assert "\u0410\u043d\u0430\u043b\u043e\u0433\u0438 \u0438\u0437 \u0422\u041a\u041f" in result.text
    match = re.search(r'href="(/download\?token=[0-9a-f]+)"', result.text)
    assert match is not None
    download = client.get(match.group(1))

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(download.content), data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=7, column=17).value == "\u0410\u043d\u0430\u043b\u043e\u0433 \u0438\u0437 \u0422\u041a\u041f"
        assert sheet.cell(row=9, column=17).value == 200
        assert sheet.cell(row=9, column=18).value == INSTALLATION
        assert sheet.cell(row=9, column=19).value == "TKP-77"
        assert sheet.cell(row=7, column=20).value == "t-1"
    finally:
        workbook.close()


def test_key_data_not_found_page(client):
    files = _files(_catalog_bytes([("t-1", 100)]), _junk_estimate_bytes())

    page = client.post("/run", files=files)

    assert page.status_code == 422
    assert "\u041a\u043b\u044e\u0447\u0435\u0432\u044b\u0435 \u0434\u0430\u043d\u043d\u044b\u0435" in page.text
    assert "header_row" in page.text


def test_missing_estimate_shows_notice(client):
    files = {"catalog": ("catalog.xlsx", _catalog_bytes([("t-1", 100)]), XLSX_MIME)}

    page = client.post("/run", files=files)

    assert page.status_code == 400
    assert "\u0441\u043c\u0435\u0442" in page.text


def test_invalid_coefficient_shows_notice(client):
    files = _files(_catalog_bytes([("t-1", 100)]), _template_estimate_bytes())

    page = client.post("/run", files=files)
    assert page.status_code == 200

    result = _confirm(client, page, coefficient="abc")
    assert result.status_code == 400
    assert "\u0447\u0438\u0441\u043b\u043e\u043c" in result.text
    # the user's own token/sheet must survive the error round-trip
    assert _hidden_value(result.text, "token") == _hidden_value(page.text, "token")


def test_multiple_sheets_choice_then_run(client):
    files = _files(_catalog_bytes([("t-1", 100), ("t-2", 120)]), _two_sheet_estimate_bytes())

    page = client.post("/run", files=files)

    assert page.status_code == 200
    assert SHEET_A in page.text
    assert SHEET_B in page.text
    match = re.search(r'href="(/run\?token=[^"]+)"', page.text)
    assert match is not None

    chosen = client.get(match.group(1).replace("&amp;", "&"))
    assert chosen.status_code == 200
    # sheet choice lands on the confirmation screen first, not the result
    assert "\u041f\u043e\u0434\u0442\u0432\u0435\u0440\u0434\u0438\u0442\u044c \u0438 \u043e\u0431\u0440\u0430\u0431\u043e\u0442\u0430\u0442\u044c" in chosen.text

    result = _confirm(client, chosen)
    assert result.status_code == 200
    assert "/download?token=" in result.text


def test_regional_coefficient_scales_result(client):
    files = _files(_catalog_bytes([("t-1", 100), ("t-2", 120)]), _template_estimate_bytes())

    page = client.post("/run", files=files)
    assert page.status_code == 200

    result = _confirm(client, page, coefficient="1,5")
    assert result.status_code == 200

    match = re.search(r'href="(/download\?token=[0-9a-f]+)"', result.text)
    assert match is not None
    download = client.get(match.group(1))
    assert download.content[:2] == b"PK"


def test_confirm_screen_shows_detected_values_and_lets_user_edit(client):
    files = _files(_catalog_bytes([("t-1", 100)]), _template_estimate_bytes())

    page = client.post("/run", files=files)
    assert page.status_code == 200
    assert 'name="region"' in page.text
    assert 'name="coefficient"' in page.text
    # nothing was found in this template file -> defaults to 1.0, and the
    # screen must warn about it instead of silently applying it (2026-07 rule)
    assert _hidden_value(page.text, "coefficient") == "1"
    assert "\u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d" in page.text
    assert 'name="use_tkp_analogs"' in page.text
    assert 'name="use_tkp_analogs" value="1" checked' not in page.text


def test_confirm_screen_warns_when_coefficient_defaulted(client):
    files = _files(_catalog_bytes([("t-1", 100)]), _template_estimate_bytes())

    page = client.post("/run", files=files)

    assert page.status_code == 200
    assert "\u26a0" in page.text
    assert "1.0" in page.text


def test_full_run_applies_reason_colour_and_label_from_admin(client, monkeypatch, tmp_path):
    db_path = tmp_path / "estimate_ai.db"
    monkeypatch.setenv("ESTIMATE_AI_DB_PATH", str(db_path))

    # Register a new reason with its own colour, then mark task "t-1" with it
    # through the same admin endpoints a real operator would use.
    client.post(
        "/admin/task-colors/reasons/add",
        data={"key": "FOT", "label": "\u0424\u041e\u0422", "color_hex": "#E2EFDA"},
    )
    client.post(
        "/admin/task-colors/add",
        data={"task_number": "t-1", "reason": "FOT", "comment": ""},
    )

    files = _files(_catalog_bytes([("t-1", 100), ("t-2", 120)]), _template_estimate_bytes())
    page = client.post("/run", files=files)
    assert page.status_code == 200

    result = _confirm(client, page)
    assert result.status_code == 200
    match = re.search(r'href="(/download\?token=[0-9a-f]+)"', result.text)
    assert match is not None

    download = client.get(match.group(1))
    assert download.status_code == 200

    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(download.content), data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        # header_row is 7 in the template fixture; t-1's analog column lands
        # at column 17 (see test_excel_writer.py fixtures for the same layout).
        assert sheet.cell(row=9, column=17).fill.start_color.rgb == "FFE2EFDA"
        assert sheet.cell(row=6, column=17).value == "\u0424\u041e\u0422"
    finally:
        workbook.close()

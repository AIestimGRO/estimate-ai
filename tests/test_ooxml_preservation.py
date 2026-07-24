from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from core.ooxml_preservation import preserve_workbook_package_features

SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def test_preserves_drawings_printer_settings_and_unmodified_sheets(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_feature_workbook(source)

    workbook = load_workbook(source)
    workbook["Estimate"]["A1"] = "changed"
    workbook.save(output)
    workbook.close()

    preserve_workbook_package_features(
        source,
        output,
        modified_sheet_title="Estimate",
    )

    with ZipFile(source) as source_zip, ZipFile(output) as output_zip:
        output_names = set(output_zip.namelist())
        assert "xl/drawings/drawing1.xml" in output_names
        assert "xl/media/image1.emf" in output_names
        assert "xl/printerSettings/printerSettings1.bin" in output_names
        assert output_zip.read("xl/worksheets/sheet2.xml") == source_zip.read(
            "xl/worksheets/sheet2.xml"
        )

        sheet1 = ET.fromstring(output_zip.read("xl/worksheets/sheet1.xml"))
        page_setup = sheet1.find(f"{{{SHEET_NS}}}pageSetup")
        assert page_setup is not None
        assert page_setup.attrib[f"{{{DOC_REL_NS}}}id"] == "rId1"

        relationships = output_zip.read("xl/worksheets/_rels/sheet1.xml.rels")
        assert b"printerSettings1.bin" in relationships
        assert output_zip.read("xl/media/image1.emf") == b"EMF-PLACEHOLDER"

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Estimate"]["A1"].value == "changed"
        assert workbook["Artwork"]["A1"].value == "untouched"
    finally:
        workbook.close()


def test_preserves_external_link_cache_without_duplicate_workbook_relationships(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    _make_feature_workbook(source)
    _add_external_link_and_shared_strings(source)

    workbook = load_workbook(source, data_only=False)
    workbook["Estimate"]["A1"] = "changed"
    workbook.save(output)
    workbook.close()

    preserve_workbook_package_features(
        source,
        output,
        modified_sheet_title="Estimate",
    )

    with ZipFile(source) as source_zip, ZipFile(output) as output_zip:
        assert output_zip.read("xl/externalLinks/externalLink1.xml") == source_zip.read(
            "xl/externalLinks/externalLink1.xml"
        )
        assert output_zip.read(
            "xl/externalLinks/_rels/externalLink1.xml.rels"
        ) == source_zip.read("xl/externalLinks/_rels/externalLink1.xml.rels")

        relationships = ET.fromstring(
            output_zip.read("xl/_rels/workbook.xml.rels")
        )
        worksheet_targets = [
            _normalized_workbook_target(relationship.attrib["Target"])
            for relationship in relationships
            if relationship.attrib.get("Type", "").endswith("/worksheet")
        ]
        assert len(worksheet_targets) == len(set(worksheet_targets))

    workbook = load_workbook(output, read_only=True, data_only=False)
    try:
        assert workbook["Estimate"]["A1"].value == "changed"
        assert workbook["Artwork"]["A1"].value == "untouched"
    finally:
        workbook.close()


def _make_feature_workbook(path: Path) -> None:
    workbook = Workbook()
    estimate = workbook.active
    estimate.title = "Estimate"
    estimate["A1"] = "original"
    estimate.page_setup.orientation = "landscape"
    artwork = workbook.create_sheet("Artwork")
    artwork["A1"] = "untouched"
    workbook.save(path)
    workbook.close()

    with ZipFile(path) as source_zip:
        infos = source_zip.infolist()
        entries = {info.filename: source_zip.read(info.filename) for info in infos}

    entries["xl/worksheets/sheet1.xml"] = _add_relationship_element(
        entries["xl/worksheets/sheet1.xml"],
        "pageSetup",
        "rId1",
    )
    entries["xl/worksheets/sheet2.xml"] = _add_relationship_element(
        entries["xl/worksheets/sheet2.xml"],
        "drawing",
        "rId1",
    )
    entries["xl/worksheets/_rels/sheet1.xml.rels"] = _relationships_xml(
        "printerSettings",
        "../printerSettings/printerSettings1.bin",
    )
    entries["xl/worksheets/_rels/sheet2.xml.rels"] = _relationships_xml(
        "drawing",
        "../drawings/drawing1.xml",
    )
    entries["xl/drawings/drawing1.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"/>'
    )
    entries["xl/media/image1.emf"] = b"EMF-PLACEHOLDER"
    entries["xl/printerSettings/printerSettings1.bin"] = b"PRINTER-PLACEHOLDER"
    entries["[Content_Types].xml"] = _add_content_types(
        entries["[Content_Types].xml"]
    )

    with ZipFile(path, "w", compression=ZIP_DEFLATED) as target_zip:
        original_names = {info.filename for info in infos}
        for info in infos:
            target_zip.writestr(info, entries[info.filename])
        for name, data in entries.items():
            if name not in original_names:
                target_zip.writestr(name, data)


def _add_external_link_and_shared_strings(path: Path) -> None:
    with ZipFile(path) as source_zip:
        infos = source_zip.infolist()
        entries = {info.filename: source_zip.read(info.filename) for info in infos}

    workbook_root = ET.fromstring(entries["xl/workbook.xml"])
    external_references = ET.SubElement(
        workbook_root,
        f"{{{SHEET_NS}}}externalReferences",
    )
    ET.SubElement(
        external_references,
        f"{{{SHEET_NS}}}externalReference",
        {f"{{{DOC_REL_NS}}}id": "rId5"},
    )
    entries["xl/workbook.xml"] = ET.tostring(
        workbook_root,
        encoding="utf-8",
        xml_declaration=True,
    )

    relationships = ET.fromstring(entries["xl/_rels/workbook.xml.rels"])
    ET.SubElement(
        relationships,
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship",
        Id="rId5",
        Type=f"{DOC_REL_NS}/externalLink",
        Target="externalLinks/externalLink1.xml",
    )
    ET.SubElement(
        relationships,
        "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship",
        Id="rId6",
        Type=f"{DOC_REL_NS}/sharedStrings",
        Target="sharedStrings.xml",
    )
    entries["xl/_rels/workbook.xml.rels"] = ET.tostring(
        relationships,
        encoding="utf-8",
        xml_declaration=True,
    )

    entries["xl/externalLinks/externalLink1.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        b'mc:Ignorable="x14 xxl21" '
        b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
        b'xmlns:xxl21="http://schemas.microsoft.com/office/spreadsheetml/2021/extlinks2021">'
        b'<externalBook xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        b'r:id="rId1"><sheetNames><sheetName val="Data"/></sheetNames>'
        b'<sheetDataSet><sheetData sheetId="0"><row r="1">'
        b'<cell r="A1" t="n"><v>42</v></cell>'
        b'</row></sheetData></sheetDataSet></externalBook></externalLink>'
    )
    entries["xl/externalLinks/_rels/externalLink1.xml.rels"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        b'<Relationship Id="rId1" '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" '
        b'Target="file:///C:/legacy/source.xlsx" TargetMode="External"/>'
        b"</Relationships>"
    )
    entries["xl/sharedStrings.xml"] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'count="1" uniqueCount="1"><si><t>unused</t></si></sst>'
    )
    entries["[Content_Types].xml"] = _add_external_link_content_types(
        entries["[Content_Types].xml"]
    )

    with ZipFile(path, "w", compression=ZIP_DEFLATED) as target_zip:
        original_names = {info.filename for info in infos}
        for info in infos:
            target_zip.writestr(info, entries[info.filename])
        for name, data in entries.items():
            if name not in original_names:
                target_zip.writestr(name, data)


def _add_external_link_content_types(content_types_xml: bytes) -> bytes:
    root = ET.fromstring(content_types_xml)
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES_NS}}}Override",
        PartName="/xl/externalLinks/externalLink1.xml",
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml",
    )
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES_NS}}}Override",
        PartName="/xl/sharedStrings.xml",
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
    )
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _normalized_workbook_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return f"xl/{target}"


def _add_relationship_element(sheet_xml: bytes, local_name: str, relation_id: str) -> bytes:
    root = ET.fromstring(sheet_xml)
    element = root.find(f"{{{SHEET_NS}}}{local_name}")
    if element is None:
        element = ET.SubElement(root, f"{{{SHEET_NS}}}{local_name}")
    element.attrib[f"{{{DOC_REL_NS}}}id"] = relation_id
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _relationships_xml(kind: str, target: str) -> bytes:
    relationship_type = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
        f"{kind}"
    )
    return (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + f'<Relationship Id="rId1" Type="{relationship_type}" Target="{target}"/>'.encode()
        + b"</Relationships>"
    )


def _add_content_types(content_types_xml: bytes) -> bytes:
    root = ET.fromstring(content_types_xml)
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES_NS}}}Default",
        Extension="bin",
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings",
    )
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES_NS}}}Default",
        Extension="emf",
        ContentType="image/x-emf",
    )
    ET.SubElement(
        root,
        f"{{{CONTENT_TYPES_NS}}}Override",
        PartName="/xl/drawings/drawing1.xml",
        ContentType="application/vnd.openxmlformats-officedocument.drawing+xml",
    )
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

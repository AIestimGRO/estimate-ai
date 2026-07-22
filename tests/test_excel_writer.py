from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.services.run_matching import KR_END
from app.services.write_result import run_and_write
from core.excel_writer import PROBLEM_FILL, TASK_FILL
from core.exclusions import TaskColorEntry, TaskHighlightReason
from core.risk import REASON_RATIO_EXCEEDED
from core.storage import connect, init_database

METER = "\u043c"
CODE = "\u0413\u042d\u0421\u041d01-01-001-01"
INSTALLATION = "\u043c\u043e\u043d\u0442\u0430\u0436"
REGION_LABEL = "\u0420\u0435\u0433\u0438\u043e\u043d"
COEFFICIENT_LABEL = "\u041a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442"
REGION_NAME = "\u041c\u043e\u0441\u043a\u0432\u0430"
ESTIMATE_TITLE = "\u041e\u0421 estimate"
CATALOG_TITLE = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433"

RED_RGB = "FFFFC7CE"
GREY_RGB = "FFD9D9D9"
BLUE_RGB = "FFDDEBF7"


def _seed_tkp_database(path: Path) -> None:
    connection = connect(path)
    try:
        init_database(connection)
        cursor = connection.execute(
            "INSERT INTO tkp_sources (file_name, task_no, item_count) VALUES (?, ?, ?)",
            ("tkp-source.xlsx", "TKP-77", 1),
        )
        connection.execute(
            """
            INSERT INTO tkp_items (
                source_id, item_name, unit, winner_unit_price_no_vat,
                winner_name, task_no
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (cursor.lastrowid, INSTALLATION, METER, 200.0, "winner", "TKP-77"),
        )
        connection.commit()
    finally:
        connection.close()


def _make_catalog_file(
    path: Path,
    prices: list[tuple[str, float]],
    *,
    region: str = "",
    regions: list[str] | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = CATALOG_TITLE
    for offset, (task_id, price) in enumerate(prices):
        row = 4 + offset
        worksheet.cell(row=row, column=2).value = task_id
        worksheet.cell(row=row, column=3).value = INSTALLATION
        worksheet.cell(row=row, column=4).value = METER
        worksheet.cell(row=row, column=7).value = price
        worksheet.cell(row=row, column=14).value = CODE
        row_region = regions[offset] if regions is not None else region
        if row_region:
            worksheet.cell(row=row, column=16).value = row_region
    workbook.save(path)
    workbook.close()
    return path


def _make_estimate_file(
    path: Path,
    *,
    base_price: float = 50.0,
    neighbour_value: object = None,
    with_coefficient: float | None = None,
    region_name: str = REGION_NAME,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ESTIMATE_TITLE
    worksheet.cell(row=7, column=14).value = CODE
    worksheet.cell(row=9, column=3).value = INSTALLATION
    worksheet.cell(row=9, column=4).value = METER
    worksheet.cell(row=9, column=6).value = base_price
    worksheet.cell(row=9, column=14).value = CODE
    if neighbour_value is not None:
        worksheet.cell(row=9, column=7).value = neighbour_value
    if with_coefficient is not None:
        worksheet.cell(row=1, column=1).value = REGION_LABEL
        worksheet.cell(row=1, column=2).value = region_name
        worksheet.cell(row=2, column=1).value = COEFFICIENT_LABEL
        worksheet.cell(row=2, column=2).value = with_coefficient
    workbook.save(path)
    workbook.close()
    return path


def test_writes_analogs_formula_kr_and_section(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"

    outcome = run_and_write(catalog, estimate, output)

    assert outcome.write_report.written_rows == 1
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=17).value == 100
        assert sheet.cell(row=9, column=18).value == 120
        assert sheet.cell(row=7, column=17).value == "task-1"
        assert sheet.cell(row=7, column=18).value == "task-2"
        assert sheet.cell(row=9, column=7).value == (
            "=MAX(F9, IFERROR(AVERAGE(F9, Q9:R9), F9))"
        )
        assert sheet.cell(row=9, column=14).value == CODE
        assert str(sheet.cell(row=9, column=15).value).endswith(KR_END)
        assert sheet.cell(row=9, column=16).value == "01"
    finally:
        workbook.close()


def test_writes_one_tkp_candidate_block_and_includes_price_in_formula(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    database = tmp_path / "estimate_ai.db"
    _seed_tkp_database(database)
    output = tmp_path / "out.xlsx"

    outcome = run_and_write(
        catalog,
        estimate,
        output,
        database_path=database,
        use_tkp_analogs=True,
    )

    assert outcome.result.matched_row_count == 1
    assert outcome.result.tkp_matched_row_count == 1
    assert outcome.write_report.tkp_start_column == 17
    assert outcome.write_report.analog_start_column == 20
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=7, column=17).value == "\u0410\u043d\u0430\u043b\u043e\u0433 \u0438\u0437 \u0422\u041a\u041f"
        assert sheet.cell(row=7, column=18).value == "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0438\u0437 \u0422\u041a\u041f"
        assert sheet.cell(row=7, column=19).value == "\u041d\u043e\u043c\u0435\u0440 \u0437\u0430\u0434\u0430\u0447\u0438 \u0422\u041a\u041f"
        assert sheet.cell(row=7, column=20).value == "task-1"
        assert sheet.cell(row=9, column=17).value == 200
        assert sheet.cell(row=9, column=18).value == INSTALLATION
        assert sheet.cell(row=9, column=19).value == "TKP-77"
        assert sheet.cell(row=9, column=20).value == 100
        assert sheet.cell(row=9, column=7).value == (
            "=MAX(F9, IFERROR(AVERAGE(F9, Q9:T9), F9))"
        )
    finally:
        workbook.close()


def test_source_file_is_not_modified(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")

    run_and_write(catalog, estimate, tmp_path / "out.xlsx")

    workbook = load_workbook(estimate, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=16).value is None
        assert sheet.cell(row=9, column=7).value is None
    finally:
        workbook.close()


def test_default_output_name_gets_wa_suffix(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")

    outcome = run_and_write(catalog, estimate)

    assert outcome.output_path.name == "estimate WA.xlsx"
    assert outcome.output_path.exists()


def test_ratio_risk_colours_red_without_price_check_log_sheet(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 300)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"
    db_path = tmp_path / "risks.db"

    outcome = run_and_write(catalog, estimate, output, database_path=db_path)

    assert outcome.result.flagged_row_count == 1
    workbook = load_workbook(output, data_only=False)
    try:
        assert "Price_Check_Log" not in workbook.sheetnames

        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=17).fill.start_color.rgb == RED_RGB
        assert sheet.cell(row=9, column=18).fill.start_color.rgb == RED_RGB
    finally:
        workbook.close()

    from core.storage import connect, init_database, list_price_risks

    connection = connect(db_path)
    try:
        init_database(connection)
        risks = list_price_risks(connection, status="open")
    finally:
        connection.close()

    assert len(risks) == 1
    assert risks[0].reason == REASON_RATIO_EXCEEDED
    assert risks[0].min_price == 100
    assert risks[0].max_price == 300


def test_second_price_within_task_is_greyed(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-1", 150)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"

    run_and_write(catalog, estimate, output)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=18).fill.start_color.rgb == GREY_RGB
    finally:
        workbook.close()


def test_regional_coefficient_by_label_scales_analogs(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx", with_coefficient=1.5)
    output = tmp_path / "out.xlsx"

    outcome = run_and_write(catalog, estimate, output)

    assert outcome.regional_coefficient == 1.5
    assert outcome.coefficient_method == "labeled_region"
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=17).value == 150
        assert sheet.cell(row=9, column=18).value == 180
    finally:
        workbook.close()


def test_average_column_inserted_when_neighbour_occupied(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx", neighbour_value=999)
    output = tmp_path / "out.xlsx"

    outcome = run_and_write(catalog, estimate, output)

    assert outcome.write_report.inserted_average_column is True
    assert outcome.write_report.average_column == 7
    assert outcome.write_report.analog_start_column == 18
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=18).value == 100
        assert sheet.cell(row=9, column=19).value == 120
        assert sheet.cell(row=9, column=7).value == (
            "=MAX(F9, IFERROR(AVERAGE(F9, R9:S9), F9))"
        )
        assert sheet.cell(row=9, column=15).value == CODE
        assert str(sheet.cell(row=9, column=16).value).endswith(KR_END)
        assert sheet.cell(row=9, column=17).value == "01"
    finally:
        workbook.close()


def test_existing_formulas_are_repointed_when_average_column_is_inserted(
    tmp_path: Path,
) -> None:
    """Regression test for the eV-grup web-app bug (2026-07).

    ``Worksheet.insert_cols()`` moves cell values but does not rewrite
    formulas living in *other* cells, so a pre-existing "ИТОГО" total (or a
    leftover average-price formula from a prior run) silently ends up
    summing/averaging the wrong column once the insert has shifted the real
    data one column to the right. Real files showed this as totals summing
    the neighbouring column and average-price formulas missing the first or
    last analog column.
    """
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx", neighbour_value=999)

    # A pre-existing "ИТОГО" total that (before the fix) already lives in a
    # column to the right of the insertion point (average column = 7), and
    # therefore must be re-pointed one column to the right after the insert.
    workbook = load_workbook(estimate, data_only=False)
    sheet = workbook[ESTIMATE_TITLE]
    sheet.cell(row=3, column=10).value = "=SUM($J$9:J9)"  # column J = 10
    workbook.save(estimate)
    workbook.close()

    output = tmp_path / "out.xlsx"
    run_and_write(catalog, estimate, output)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        # insert_cols() itself already moved the formula cell from J3 (10) to
        # K3 (11); the fix must also update its *text* so the SUM still
        # points at column K (its own, now-shifted column), not the stale J.
        assert sheet.cell(row=3, column=10).value is None
        assert sheet.cell(row=3, column=11).value == "=SUM($K$9:K9)"
    finally:
        workbook.close()


def test_analog_headers_include_task_number_and_region(tmp_path: Path) -> None:
    catalog = _make_catalog_file(
        tmp_path / "catalog.xlsx",
        [("task-1", 100), ("task-2", 120)],
        region=REGION_NAME,
    )
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"

    run_and_write(catalog, estimate, output)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=7, column=17).value == "task-1"
        assert sheet.cell(row=8, column=17).value == REGION_NAME
        assert sheet.cell(row=7, column=18).value == "task-2"
        assert sheet.cell(row=8, column=18).value == REGION_NAME
    finally:
        workbook.close()


def test_analog_columns_are_grouped_own_region_first_then_az(tmp_path: Path) -> None:
    """2026-07 rule: analog columns for the file's own region come first,
    remaining regions follow grouped together in А-Я order."""
    catalog = _make_catalog_file(
        tmp_path / "catalog.xlsx",
        [("task-kostroma", 100), ("task-yakutia", 120), ("task-tula", 140)],
        regions=["\u041a\u043e\u0441\u0442\u0440\u043e\u043c\u0430", "\u042f\u043a\u0443\u0442\u0438\u044f", "\u0422\u0443\u043b\u0430"],
    )
    # "71. Тульская область" should match the catalog's short "Тула" region
    # via the curated alias table (adjective softening: тул->туль-).
    estimate = _make_estimate_file(
        tmp_path / "estimate.xlsx",
        with_coefficient=1.0,
        region_name="71. \u0422\u0443\u043b\u044c\u0441\u043a\u0430\u044f \u043e\u0431\u043b\u0430\u0441\u0442\u044c",
    )
    output = tmp_path / "out.xlsx"

    run_and_write(catalog, estimate, output)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        # Own region (Тула) first, then the rest alphabetically: Кострома
        # before Якутия.
        assert sheet.cell(row=7, column=17).value == "task-tula"
        assert sheet.cell(row=7, column=18).value == "task-kostroma"
        assert sheet.cell(row=7, column=19).value == "task-yakutia"
    finally:
        workbook.close()


def test_kr_gets_plain_code_without_suffix_when_no_analog_found(tmp_path: Path) -> None:
    """2026-07 rule: rows with no analog still get /КР filled, but with the
    plain ГЭСН code -- no "/КР" suffix is appended."""
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"

    run_and_write(catalog, estimate, output)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=15).value == CODE
        assert not str(sheet.cell(row=9, column=15).value).endswith(KR_END)
    finally:
        workbook.close()


def test_reprocessing_reuses_existing_average_formula_column(tmp_path: Path) -> None:
    """2026-07 rule: re-running on an already-processed file must overwrite
    the existing "Цена средняя" formula in place, not insert a duplicate
    column next to it."""
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")

    # Simulate a prior run: the average-formula column already sits at
    # base_price + 1 (column 7), referencing a since-stale analog range.
    workbook = load_workbook(estimate, data_only=False)
    sheet = workbook[ESTIMATE_TITLE]
    sheet.cell(row=9, column=7).value = "=MAX(F9, IFERROR(AVERAGE(F9, Q9:Q9), F9))"
    workbook.save(estimate)
    workbook.close()

    output = tmp_path / "out.xlsx"
    outcome = run_and_write(catalog, estimate, output)

    assert outcome.write_report.inserted_average_column is False
    assert outcome.write_report.average_column == 7
    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        # Formula rewritten in place, no leftover duplicate column: KR/section
        # and the analog block land at their normal columns, same as a
        # first-time run -- nothing got shifted right by a phantom insert.
        assert sheet.cell(row=9, column=7).value == (
            "=MAX(F9, IFERROR(AVERAGE(F9, Q9:R9), F9))"
        )
        assert sheet.cell(row=9, column=17).value == 100
        assert sheet.cell(row=9, column=18).value == 120
    finally:
        workbook.close()


def test_task_color_list_tints_analog_column_blue(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"
    colors = [TaskColorEntry(enabled=True, task_number="task-1")]

    run_and_write(catalog, estimate, output, task_color_entries=colors)

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=17).fill.start_color.rgb == BLUE_RGB
        assert sheet.cell(row=9, column=18).fill.start_color.rgb != BLUE_RGB
    finally:
        workbook.close()


def test_task_reason_colours_analog_column_and_writes_label_above_header(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100), ("task-2", 120)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"
    colors = [TaskColorEntry(enabled=True, task_number="task-1", reason="FOT")]
    reasons = [TaskHighlightReason(key="FOT", label="\u0424\u041e\u0422", color_hex="E2EFDA")]

    run_and_write(
        catalog,
        estimate,
        output,
        task_color_entries=colors,
        task_highlight_reasons=reasons,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        # header_row is 7 in this fixture -> analog column fill uses the FOT colour
        assert sheet.cell(row=9, column=17).fill.start_color.rgb == "FFE2EFDA"
        assert sheet.cell(row=9, column=18).fill.start_color.rgb != "FFE2EFDA"
        # the label sits directly above the task_id header cell (row 7 - 1 = 6)
        assert sheet.cell(row=6, column=17).value == "\u0424\u041e\u0422"
        assert sheet.cell(row=6, column=17).fill.start_color.rgb == "FFE2EFDA"
        assert sheet.cell(row=6, column=18).value is None
    finally:
        workbook.close()


def test_marked_task_with_unregistered_reason_falls_back_to_legacy_colour(tmp_path: Path) -> None:
    catalog = _make_catalog_file(tmp_path / "catalog.xlsx", [("task-1", 100)])
    estimate = _make_estimate_file(tmp_path / "estimate.xlsx")
    output = tmp_path / "out.xlsx"
    # "reason" here is legacy free text that does not match any registered key.
    colors = [TaskColorEntry(enabled=True, task_number="task-1", reason="+1")]
    reasons = [TaskHighlightReason(key="FOT", label="\u0424\u041e\u0422", color_hex="E2EFDA")]

    run_and_write(
        catalog,
        estimate,
        output,
        task_color_entries=colors,
        task_highlight_reasons=reasons,
    )

    workbook = load_workbook(output, data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=9, column=17).fill.start_color.rgb == BLUE_RGB
        # no matching reason -> no label written above the header
        assert sheet.cell(row=6, column=17).value is None
    finally:
        workbook.close()


def test_kr_and_section_use_dedicated_columns_with_real_headers(tmp_path: Path) -> None:
    code_header = "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0422\u0415\u0420/\u041a\u0420"
    kr_header = "/\u041a\u0420"
    section_header = "\u041a\u043e\u0434 \u0440\u0430\u0437\u0434\u0435\u043b\u0430"
    unit_header = "\u0415\u0434.\u0438\u0437\u043c."
    base_header = "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421"

    catalog = _make_catalog_file(
        tmp_path / "catalog.xlsx",
        [("5818383", 100), ("4768644", 120)],
        region="\u042f\u043a\u0443\u0442\u0438\u044f",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ESTIMATE_TITLE
    header_row = 26
    sheet.cell(row=header_row, column=4, value=unit_header)
    sheet.cell(row=header_row, column=6, value=base_header)
    sheet.cell(row=header_row, column=14, value=code_header)
    sheet.cell(row=header_row, column=15, value=kr_header)
    sheet.cell(row=header_row, column=16, value=section_header)
    data_row = 28
    sheet.cell(row=data_row, column=3, value=INSTALLATION)
    sheet.cell(row=data_row, column=4, value=METER)
    sheet.cell(row=data_row, column=6, value=50.0)
    sheet.cell(row=data_row, column=14, value=CODE)
    estimate = tmp_path / "estimate.xlsx"
    workbook.save(estimate)
    workbook.close()

    run_and_write(catalog, estimate, tmp_path / "out.xlsx")

    workbook = load_workbook(tmp_path / "out.xlsx", data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=data_row, column=14).value == CODE
        assert str(sheet.cell(row=data_row, column=15).value).endswith(KR_END)
        assert sheet.cell(row=data_row, column=16).value == "01"
        assert sheet.cell(row=header_row, column=17).value == "5818383"
        assert sheet.cell(row=header_row + 1, column=17).value == "\u042f\u043a\u0443\u0442\u0438\u044f"
        assert sheet.cell(row=header_row, column=18).value == "4768644"
    finally:
        workbook.close()


def test_kr_reuses_free_column_without_inserting(tmp_path: Path) -> None:
    """eV-grup layout with blank col 15: /KR goes into 15, section stays at 16."""
    code_header = "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0422\u0415\u0420/\u041a\u0420"
    section_header = "\u041a\u043e\u0434 \u0440\u0430\u0437\u0434\u0435\u043b\u0430"
    unit_header = "\u0415\u0434.\u0438\u0437\u043c."
    base_header = "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442, \u0440\u0443\u0431. \u0431\u0435\u0437 \u041d\u0414\u0421"

    catalog = _make_catalog_file(
        tmp_path / "catalog.xlsx",
        [("5818383", 100), ("4768644", 120)],
        region="\u042f\u043a\u0443\u0442\u0438\u044f",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = ESTIMATE_TITLE
    header_row = 26
    sheet.cell(row=header_row, column=4, value=unit_header)
    sheet.cell(row=header_row, column=6, value=base_header)
    sheet.cell(row=header_row, column=14, value=code_header)
    sheet.cell(row=header_row, column=16, value=section_header)
    data_row = 28
    sheet.cell(row=data_row, column=3, value=INSTALLATION)
    sheet.cell(row=data_row, column=4, value=METER)
    sheet.cell(row=data_row, column=6, value=50.0)
    sheet.cell(row=data_row, column=14, value=CODE)
    estimate = tmp_path / "estimate.xlsx"
    workbook.save(estimate)
    workbook.close()

    outcome = run_and_write(catalog, estimate, tmp_path / "out.xlsx")

    assert outcome.write_report.analog_start_column == 17
    workbook = load_workbook(tmp_path / "out.xlsx", data_only=False)
    try:
        sheet = workbook[ESTIMATE_TITLE]
        assert sheet.cell(row=header_row, column=15).value == "/\u041a\u0420"
        assert sheet.cell(row=header_row, column=16).value == section_header
        assert sheet.cell(row=header_row, column=17).value == "5818383"
        assert str(sheet.cell(row=data_row, column=15).value).endswith(KR_END)
        assert sheet.cell(row=data_row, column=16).value == "01"
        assert sheet.cell(row=data_row, column=17).value == 100
    finally:
        workbook.close()

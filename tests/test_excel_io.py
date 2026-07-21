from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from core.catalog import _catalog_date_serial
from core.excel_io import read_catalog_rows, read_estimate_rows, read_estimate_rows_by_columns


CATALOG_SHEET = "\u041a\u0430\u0442\u0430\u043b\u043e\u0433 2026"
ESTIMATE_SHEET = "\u041e\u0421 estimate"
GESN_HEADER = "\u041a\u043e\u0434 \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c"
METER = "\u043c"
WORK_NAME = "\u0440\u0430\u0431\u043e\u0442\u0430"
REGION = "\u0440\u0435\u0433\u0438\u043e\u043d"
TASK_HEADER = "\u041d\u043e\u043c\u0435\u0440 \u0437\u0430\u0434\u0430\u0447\u0438"
WORK_HEADER = "\u041d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435 \u0440\u0430\u0431\u043e\u0442"
UNIT_HEADER = "\u0415\u0434.\u0438\u0437\u043c."
PRICE_HEADER = "\u0426\u0435\u043d\u0430 \u0435\u0434\u0438\u043d\u0438\u0446\u044b \u0440\u0430\u0431\u043e\u0442"
CODE_HEADER = "\u041f\u0435\u0440\u0435\u0447\u0435\u043d\u044c \u0413\u042d\u0421\u041d/\u0424\u0415\u0420/\u0422\u0415\u0420/\u041a\u0420"
REGION_HEADER = "\u0420\u0435\u0433\u0438\u043e\u043d"
ADDED_DATE_HEADER = "\u0414\u0430\u0442\u0430 \u0434\u043e\u0431\u0430\u0432\u043b\u0435\u043d\u0438\u044f \u0432 \u043a\u0430\u0442\u0430\u043b\u043e\u0433"
QUARTER_HEADER = "\u0413\u043e\u0434 \u041a\u0432\u0430\u0440\u0442\u0430\u043b \u041b\u0421\u0420"
REGIONAL_COEF_HEADER = "\u0420\u0435\u0433\u0438\u043e\u043d\u0430\u043b\u044c\u043d\u044b\u0439 \u043a\u043e\u044d\u0444\u0444\u0438\u0446\u0438\u0435\u043d\u0442"


def fixture_path(tmp_path: Path, name: str) -> Path:
    fixtures_dir = tmp_path / "fixtures"
    fixtures_dir.mkdir()
    return fixtures_dir / name


def save_workbook(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    workbook.close()
    return path


def make_catalog_workbook(sheet_name: str = CATALOG_SHEET) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    worksheet.cell(row=3, column=2).value = TASK_HEADER
    worksheet.cell(row=3, column=3).value = WORK_HEADER
    worksheet.cell(row=3, column=4).value = UNIT_HEADER
    worksheet.cell(row=3, column=7).value = PRICE_HEADER
    worksheet.cell(row=3, column=14).value = CODE_HEADER
    worksheet.cell(row=3, column=16).value = REGION_HEADER
    worksheet.cell(row=3, column=17).value = ADDED_DATE_HEADER

    worksheet.cell(row=4, column=2).value = "task-1"
    worksheet.cell(row=4, column=3).value = WORK_NAME
    worksheet.cell(row=4, column=4).value = METER
    worksheet.cell(row=4, column=7).value = 123.45
    worksheet.cell(row=4, column=14).value = "gesn01-01-001-01"
    worksheet.cell(row=4, column=16).value = REGION

    return workbook


def test_read_catalog_rows_finds_catalog_sheet_and_reads_from_row_4(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Other"
    worksheet = workbook.create_sheet(CATALOG_SHEET)
    worksheet.cell(row=3, column=2).value = "task"
    worksheet.cell(row=4, column=2).value = "task-1"
    worksheet.cell(row=4, column=3).value = WORK_NAME
    worksheet.cell(row=4, column=4).value = METER
    worksheet.cell(row=4, column=7).value = 123.45
    worksheet.cell(row=4, column=14).value = "gesn01-01-001-01"
    worksheet.cell(row=4, column=16).value = REGION

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "catalog.xlsx")))

    assert len(rows) == 1
    assert rows[0].task_id == "task-1"
    assert rows[0].work_name == WORK_NAME
    assert rows[0].unit == METER
    assert rows[0].price == 123.45
    assert rows[0].code == "gesn01-01-001-01"
    assert rows[0].region == REGION


def test_read_catalog_rows_falls_back_to_first_sheet(tmp_path: Path) -> None:
    workbook = make_catalog_workbook(sheet_name="First")

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "fallback.xlsx")))

    assert len(rows) == 1
    assert rows[0].task_id == "task-1"


def test_read_estimate_rows_detects_header_row_dynamically(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = ESTIMATE_SHEET
    worksheet.cell(row=7, column=14).value = GESN_HEADER
    worksheet.cell(row=9, column=3).value = WORK_NAME
    worksheet.cell(row=9, column=4).value = METER
    worksheet.cell(row=9, column=6).value = 100
    worksheet.cell(row=9, column=14).value = "gesn01-01-001-01"
    worksheet.cell(row=10, column=3).value = "ignored invalid row"
    worksheet.cell(row=10, column=4).value = METER
    worksheet.cell(row=10, column=6).value = 0
    worksheet.cell(row=10, column=14).value = "gesn01-01-001-02"

    rows = read_estimate_rows(save_workbook(workbook, fixture_path(tmp_path, "estimate.xlsx")))

    assert len(rows) == 1
    assert rows[0].code == "gesn01-01-001-01"
    assert rows[0].unit == METER
    assert rows[0].work_name == WORK_NAME
    assert rows[0].base_price == 100


def test_catalog_excel_date_reads_back_as_datetime_and_converts_to_serial(
    tmp_path: Path,
) -> None:
    workbook = make_catalog_workbook()
    worksheet = workbook[CATALOG_SHEET]
    worksheet.cell(row=4, column=17).value = datetime(2026, 6, 30)
    worksheet.cell(row=4, column=17).number_format = "dd.mm.yyyy"

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "date.xlsx")))

    assert type(rows[0].added_date) is datetime
    assert _catalog_date_serial(rows[0].added_date) > 0


def test_blank_catalog_added_date_resolves_to_serial_zero_downstream(
    tmp_path: Path,
) -> None:
    workbook = make_catalog_workbook()

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "blank-date.xlsx")))

    assert rows[0].added_date is None
    assert _catalog_date_serial(rows[0].added_date) == 0


def test_read_catalog_detects_added_date_column_offset_from_template(
    tmp_path: Path,
) -> None:
    workbook = make_catalog_workbook()
    worksheet = workbook[CATALOG_SHEET]
    worksheet.cell(row=3, column=17).value = QUARTER_HEADER
    worksheet.cell(row=3, column=20).value = REGIONAL_COEF_HEADER
    worksheet.cell(row=3, column=21).value = ADDED_DATE_HEADER
    worksheet.cell(row=4, column=21).value = datetime(2025, 6, 22)

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "offset-date.xlsx")))

    assert type(rows[0].added_date) is datetime
    assert rows[0].added_date == datetime(2025, 6, 22)


def test_read_catalog_prefers_template_price_column_when_headers_duplicate(
    tmp_path: Path,
) -> None:
    workbook = make_catalog_workbook()
    worksheet = workbook[CATALOG_SHEET]
    worksheet.cell(row=3, column=6).value = PRICE_HEADER
    worksheet.cell(row=4, column=6).value = 999.0
    worksheet.cell(row=4, column=7).value = 123.45

    rows = read_catalog_rows(save_workbook(workbook, fixture_path(tmp_path, "dup-price.xlsx")))

    assert rows[0].price == 123.45


def test_read_estimate_rows_by_columns_survives_a_run_of_section_headers(
    tmp_path: Path,
) -> None:
    """Regression test for a real production bug (2026-07): a real estimate
    had 5 consecutive rows holding only a section/object title (no code,
    unit or price -- e.g. object title, "Раздел N", "Оборудование",
    "Материал не требующий монтажа", "ЗИП") between two objects. That run
    exactly matched the old max_blank_run default of 5, so the scan stopped
    there and every row of the *next* object (700+ real line items in the
    reported file) was silently dropped -- no error, no warning, nothing in
    the output.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "\u0421\u043c\u0435\u0442\u0430"

    header_row = 1
    worksheet.cell(row=header_row, column=1).value = CODE_HEADER
    worksheet.cell(row=header_row, column=2).value = UNIT_HEADER
    worksheet.cell(row=header_row, column=3).value = WORK_HEADER
    worksheet.cell(row=header_row, column=4).value = PRICE_HEADER

    def write_item(row: int, code: str, price: float) -> None:
        worksheet.cell(row=row, column=1).value = code
        worksheet.cell(row=row, column=2).value = METER
        worksheet.cell(row=row, column=3).value = f"\u0420\u0430\u0431\u043e\u0442\u0430 {row}"
        worksheet.cell(row=row, column=4).value = price

    # First object: a couple of real line items.
    write_item(2, "\u0413\u042d\u0421\u041d01-01-001-01", 100.0)
    write_item(3, "\u0413\u042d\u0421\u041d01-01-001-02", 200.0)

    # Exactly 5 consecutive "section header only" rows -- the real-world
    # pattern: object title, раздел, then a couple of category labels.
    section_titles = [
        "\u041e\u0431\u044a\u0435\u043a\u0442 2",
        "\u0420\u0430\u0437\u0434\u0435\u043b 1",
        "\u041e\u0431\u043e\u0440\u0443\u0434\u043e\u0432\u0430\u043d\u0438\u0435",
        "\u041c\u0430\u0442\u0435\u0440\u0438\u0430\u043b \u043d\u0435 \u0442\u0440\u0435\u0431\u0443\u044e\u0449\u0438\u0439 \u043c\u043e\u043d\u0442\u0430\u0436\u0430",
        "\u0417\u0418\u041f",
    ]
    for offset, title in enumerate(section_titles):
        worksheet.cell(row=4 + offset, column=3).value = title

    # Second object: real line items that must NOT be dropped.
    write_item(9, "\u0413\u042d\u0421\u041d02-01-001-01", 300.0)
    write_item(10, "\u0413\u042d\u0421\u041d02-01-001-02", 400.0)

    rows = read_estimate_rows_by_columns(
        worksheet,
        header_row=header_row,
        code_column=1,
        unit_column=2,
        work_name_column=3,
        base_price_column=4,
    )

    codes = [row.code for _, row in rows]
    assert codes == [
        "\u0413\u042d\u0421\u041d01-01-001-01",
        "\u0413\u042d\u0421\u041d01-01-001-02",
        "\u0413\u042d\u0421\u041d02-01-001-01",
        "\u0413\u042d\u0421\u041d02-01-001-02",
    ], "rows after the section-header block were dropped -- the truncation bug is back"

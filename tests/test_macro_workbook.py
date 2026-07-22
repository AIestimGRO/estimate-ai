"""Tests for macro workbook settings loading."""

from __future__ import annotations

import json
import os
from pathlib import Path

import pytest
from openpyxl import Workbook

from core.exclusions import is_name_excluded
from core.macro_workbook import (
    NAME_EXCLUSIONS_SHEET,
    load_default_macro_settings,
    load_name_exclusions_workbook,
    resolve_macro_workbook_path,
)

MM = "\u043c\u043c"
CHANGED = "\u0438\u0437\u043c\u0435\u043d"
LAYER = "\u0441\u043b\u043e\u0439"
WORK_MM = f"\u041d\u0430 \u043a\u0430\u0436\u0434\u044b\u0435 10 {MM} {CHANGED}\u0435\u043d\u0438\u044f \u0433\u043b\u0443\u0431\u0438\u043d\u044b"
WORK_LAYER = f"\u041e\u043a\u043b\u0435\u0435\u0447\u043d\u0430\u044f \u0432 \u043e\u0434\u0438\u043d {LAYER}"


def _write_fixture_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = NAME_EXCLUSIONS_SHEET
    worksheet.cell(row=1, column=1, value="Enabled")
    worksheet.cell(row=1, column=4, value="Pattern")
    worksheet.cell(row=2, column=1, value=1)
    worksheet.cell(row=2, column=2, value="BOTH")
    worksheet.cell(row=2, column=3, value="ALL_WORDS")
    worksheet.cell(row=2, column=4, value=f"{MM}|{CHANGED}")
    worksheet.cell(row=3, column=1, value=1)
    worksheet.cell(row=3, column=2, value="BOTH")
    worksheet.cell(row=3, column=3, value="ALL_WORDS")
    worksheet.cell(row=3, column=4, value=LAYER)
    worksheet.cell(row=2, column=8, value=1)
    worksheet.cell(row=2, column=9, value="12345")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def test_load_name_exclusions_workbook_reads_rules_and_task_colors(tmp_path: Path) -> None:
    path = tmp_path / "macro.xlsm"
    _write_fixture_workbook(path)

    rules, colors = load_name_exclusions_workbook(path)

    assert len(rules) == 2
    assert len(colors) == 1
    assert is_name_excluded(rules, "SMETA", WORK_MM)
    assert is_name_excluded(rules, "SMETA", WORK_LAYER)
    assert colors[0].task_number == "12345"


def test_resolve_macro_workbook_path_uses_config_workbook(tmp_path: Path, monkeypatch) -> None:
    repo = tmp_path / "repo"
    config_dir = repo / "data" / "config"
    config_dir.mkdir(parents=True)
    workbook = config_dir / "autopodbor.xlsm"
    _write_fixture_workbook(workbook)
    config_dir.joinpath("macro.json").write_text(
        json.dumps({"workbook": "data/config/autopodbor.xlsm"}),
        encoding="utf-8",
    )
    monkeypatch.setattr("core.macro_workbook.repo_root", lambda: repo)

    resolved = resolve_macro_workbook_path(config_dir / "macro.json")

    assert resolved == workbook.resolve()


def test_resolve_macro_workbook_path_falls_back_to_newest_xlsm(tmp_path: Path, monkeypatch) -> None:
    repo = tmp_path / "repo"
    real_dir = repo / "data" / "real"
    real_dir.mkdir(parents=True)
    older = real_dir / "old.xlsm"
    newer = real_dir / "new.xlsm"
    _write_fixture_workbook(older)
    _write_fixture_workbook(newer)
    os.utime(older, ns=(1_700_000_000_000_000_000, 1_700_000_000_000_000_000))
    os.utime(newer, ns=(1_700_000_001_000_000_000, 1_700_000_001_000_000_000))
    config_dir = repo / "data" / "config"
    config_dir.mkdir()
    config_dir.joinpath("macro.json").write_text("{}", encoding="utf-8")
    monkeypatch.setattr("core.macro_workbook.repo_root", lambda: repo)

    resolved = resolve_macro_workbook_path(config_dir / "macro.json")

    assert resolved == newer.resolve()


def test_load_default_macro_settings_returns_empty_when_missing(
    tmp_path: Path,
    monkeypatch,
) -> None:
    repo = tmp_path / "repo"
    config_dir = repo / "data" / "config"
    config_dir.mkdir(parents=True)
    config_dir.joinpath("macro.json").write_text("{}", encoding="utf-8")
    monkeypatch.setattr("core.macro_workbook.repo_root", lambda: repo)

    settings = load_default_macro_settings()

    assert settings.workbook_path is None
    assert settings.name_exclusion_rules == []


@pytest.mark.skipif(
    not any(Path("data/real").glob("*.xlsm")),
    reason="local autopodbor workbook not present",
)
def test_real_autopodbor_workbook_excludes_fss_rows() -> None:
    settings = load_default_macro_settings()
    assert settings.workbook_path is not None
    assert len(settings.name_exclusion_rules) >= 4

    work101 = (
        "\u041d\u0430 \u043a\u0430\u0436\u0434\u044b\u0435 10 \u043c\u043c "
        "\u0438\u0437\u043c\u0435\u043d\u0435\u043d\u0438\u044f \u0433\u043b\u0443\u0431\u0438\u043d\u044b"
    )
    work211 = (
        "\u0423\u0441\u0442\u0440\u043e\u0439\u0441\u0442\u0432\u043e "
        "\u043f\u0430\u0440\u043e\u0438\u0437\u043e\u043b\u044f\u0446\u0438\u0438: "
        "\u043e\u043a\u043b\u0435\u0435\u0447\u043d\u043e\u0439 \u0432 \u043e\u0434\u0438\u043d \u0441\u043b\u043e\u0439"
    )
    assert is_name_excluded(settings.name_exclusion_rules, "SMETA", work101)
    assert is_name_excluded(settings.name_exclusion_rules, "SMETA", work211)

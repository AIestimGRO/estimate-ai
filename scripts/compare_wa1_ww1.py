"""Detailed comparison of WA1 vs WW1 outputs in data/real/."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

OUT = Path("data/real/_compare_wa1_ww1.txt"
)


def clean_code(value: object) -> str:
    text = re.sub(r"\s*/\s*\u041a\u0420\s*$", "", str(value or "").strip(), flags=re.I)
    return text.casefold()


def all_prices(ws, row: int, start: int) -> list[tuple[int, float]]:
    out: list[tuple[int, float]] = []
    for col in range(start, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        try:
            out.append((col, round(float(value), 2)))
        except (TypeError, ValueError):
            break
    return out


def data_rows(ws, code_col: int = 14, unit_col: int = 4, base_col: int = 6, start: int = 27):
    rows = []
    blank = 0
    for row in range(start, ws.max_row + 1):
        code = ws.cell(row=row, column=code_col).value
        unit = ws.cell(row=row, column=unit_col).value
        base = ws.cell(row=row, column=base_col).value
        if not code or not unit or not base:
            blank += 1
            if blank >= 5:
                break
            continue
        blank = 0
        if str(code).strip().isdigit():
            continue
        try:
            if float(base) <= 0:
                continue
        except (TypeError, ValueError):
            continue
        rows.append(row)
    return rows


def header_map(ws, header_row: int, start_col: int) -> dict[int, tuple[str, str]]:
    region_row = header_row + 1
    mapping: dict[int, tuple[str, str]] = {}
    for col in range(start_col, ws.max_column + 1):
        task = ws.cell(row=header_row, column=col).value
        region = ws.cell(row=region_row, column=col).value
        if task is None and region is None:
            if col > start_col + 5:
                break
            continue
        if task is not None and str(task).strip():
            mapping[col] = (str(task).strip(), "" if region is None else str(region).strip())
    return mapping


def main() -> None:
    real = Path("data/real")
    wa = next(p for p in real.glob("*.xlsx") if "WA1" in p.name.upper())
    ww = next(p for p in real.glob("*.xlsx") if "WW1" in p.name.upper())

    lines: list[str] = []
    lines.append(f"WA: {wa.name}")
    lines.append(f"WW: {ww.name}")
    lines.append("")

    wwb = load_workbook(ww, data_only=True)
    wab = load_workbook(wa, data_only=True)
    wwb_f = load_workbook(ww, data_only=False)
    wab_f = load_workbook(wa, data_only=False)
    wws, was = wwb.active, wab.active
    wws_f, was_f = wwb_f.active, wab_f.active

    header = 26
    analog_start = 17
    ww_headers = header_map(wws_f, header, analog_start)
    wa_headers = header_map(was_f, header, analog_start)

    lines.append(f"WW analog header cols: {len(ww_headers)} (from {min(ww_headers) if ww_headers else '-'})")
    lines.append(f"WA analog header cols: {len(wa_headers)} (from {min(wa_headers) if wa_headers else '-'})")
    lines.append("")

    ww_rows = data_rows(wws)
    wa_rows = data_rows(was)
    lines.append(f"data rows WW/WA: {len(ww_rows)} / {len(wa_rows)}")

    shared = sorted(set(ww_rows) & set(wa_rows))
    price_set_diff = 0
    col_placement_diff = 0
    avg_diff = 0
    kr_diff = 0
    section_diff = 0
    wa_no_prices = 0
    ww_no_prices = 0
    examples: list[str] = []

    for row in shared:
        code_ww = clean_code(wws.cell(row, 14).value)
        code_wa = clean_code(was.cell(row, 14).value)
        if code_ww != code_wa:
            kr_diff += 1

        sec_ww = wws.cell(row, 16).value
        sec_wa = was.cell(row, 16).value
        if str(sec_ww or "") != str(sec_wa or ""):
            section_diff += 1

        wp = all_prices(wws, row, analog_start)
        ap = all_prices(was, row, analog_start)
        if not wp:
            ww_no_prices += 1
        if not ap:
            wa_no_prices += 1
        if wp and not ap:
            if len(examples) < 8:
                examples.append(f"row {row} {code_ww}: WW has {len(wp)} prices, WA none")
            continue
        if not wp and ap:
            if len(examples) < 8:
                examples.append(f"row {row} {code_ww}: WA has {len(ap)} prices, WW none")
            continue

        wset = sorted(v for _, v in wp)
        aset = sorted(v for _, v in ap)
        if wset != aset:
            price_set_diff += 1
            if len(examples) < 12:
                examples.append(f"row {row} {code_ww}: price sets differ WW={wset[:6]} WA={aset[:6]}")

        # same task column placement?
        def by_task(prices, headers):
            mapped = {}
            for col, val in prices:
                task = headers.get(col, ("?", ""))[0]
                mapped.setdefault(task, []).append(val)
            return mapped

        if wp and ap and by_task(wp, ww_headers) != by_task(ap, wa_headers):
            col_placement_diff += 1

        wavg = wws.cell(row, 7).value
        aavg = was.cell(row, 7).value
        if isinstance(wavg, (int, float)) and isinstance(aavg, (int, float)):
            if abs(float(wavg) - float(aavg)) > 1:
                avg_diff += 1
        elif isinstance(wavg, (int, float)) and aavg is None:
            avg_diff += 1

    lines.append(f"shared rows: {len(shared)}")
    lines.append(f"code mismatch: {kr_diff}")
    lines.append(f"section mismatch: {section_diff}")
    lines.append(f"rows WW without prices: {ww_no_prices}")
    lines.append(f"rows WA without prices: {wa_no_prices}")
    lines.append(f"price set mismatch: {price_set_diff}")
    lines.append(f"task-column placement mismatch: {col_placement_diff}")
    lines.append(f"avg cached value mismatch: {avg_diff}")
    lines.append("")
    lines.append("sample differences:")
    lines.extend(examples)

    # formula comparison on matched rows with analogs
    lines.append("")
    lines.append("formula samples (rows with analogs):")
    shown = 0
    for row in shared:
        wp = all_prices(wws, row, analog_start)
        ap = all_prices(was, row, analog_start)
        if not wp or not ap:
            continue
        wf = wws_f.cell(row, 7).value
        af = was_f.cell(row, 7).value
        lines.append(f"  row {row}: WW avg={wws.cell(row,7).value} formula={str(wf)[:55]}")
        lines.append(f"           WA avg={was.cell(row,7).value} formula={str(af)[:55]}")
        shown += 1
        if shown >= 5:
            break

    # header task order first 10 cols
    lines.append("")
    lines.append("WW headers 17-26:")
    for col in range(17, 27):
        if col in ww_headers:
            lines.append(f"  c{col}: {ww_headers[col]}")
    lines.append("WA headers 17-26:")
    for col in range(17, 27):
        if col in wa_headers:
            lines.append(f"  c{col}: {wa_headers[col]}")

    wwb.close()
    wab.close()
    wwb_f.close()
    wab_f.close()
    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()

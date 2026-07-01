"""Dump the top-left structure of a workbook for mapping diagnostics.

Prints every worksheet title, then a preview grid of the first rows/columns
with their spreadsheet column letters, so column mapping can be verified
against core/excel_io.py Settings.

Usage:
    python scripts/dump_sheet.py <file.xlsx> [--rows 25] [--cols 20]
"""

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _preview(value: object, width: int = 40) -> str:
    text = "" if value is None else str(value).replace("\n", " ").replace("\r", " ")
    text = text.strip()
    if len(text) > width:
        text = text[: width - 1] + "\u2026"
    return text


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Dump workbook structure.")
    parser.add_argument("path", type=Path, help="path to the .xlsx file")
    parser.add_argument("--rows", type=int, default=25, help="preview row count")
    parser.add_argument("--cols", type=int, default=20, help="preview column count")
    args = parser.parse_args(argv)

    workbook = load_workbook(args.path, data_only=True, read_only=True)
    try:
        print(f"file   : {args.path}")
        print(f"sheets : {[ws.title for ws in workbook.worksheets]}")

        for worksheet in workbook.worksheets:
            print()
            print(f"=== sheet: {worksheet.title!r} ===")
            for row_number in range(1, args.rows + 1):
                cells: list[str] = []
                for col_number in range(1, args.cols + 1):
                    value = worksheet.cell(row=row_number, column=col_number).value
                    if value is None or str(value).strip() == "":
                        continue
                    letter = get_column_letter(col_number)
                    cells.append(f"{letter}={_preview(value)}")
                if cells:
                    print(f"row {row_number:>3}: " + " | ".join(cells))
    finally:
        workbook.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""Compare 6356153 WAW (web) vs WAM (macro)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

REAL = Path("data/real")
OUT = REAL / "_compare_6356153_waw_wam.txt"

CODE_COL = 14
UNIT_COL = 4
BASE_COL = 6
KR_COL = 15
SECTION_COL = 16
ANALOG_START = 17
DATA_START = 29
HEADER_ROW = 25


def clean_code(value: object) -> str:
    text = re.sub(r"\s*/\s*\u041a\u0420\s*$", "", str(value or "").strip(), flags=re.I)
    return text.casefold()


def gesn_rows(ws) -> list[int]:
    rows: list[int] = []
    blank = 0
    for row in range(DATA_START, ws.max_row + 1):
        unit = ws.cell(row=row, column=UNIT_COL).value
        base = ws.cell(row=row, column=BASE_COL).value
        code = ws.cell(row=row, column=CODE_COL).value
        if not unit or not base or not code:
            blank += 1
            if blank >= 8:
                break
            continue
        blank = 0
        code_text = str(code).strip()
        if not code_text.upper().startswith("\u0413\u042d\u0421\u041d"):
            continue
        try:
            if float(base) <= 0:
                continue
        except (TypeError, ValueError):
            continue
        rows.append(row)
    return rows


def analog_prices(ws, row: int) -> list[float]:
    out: list[float] = []
    for col in range(ANALOG_START, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        try:
            out.append(round(float(value), 2))
        except (TypeError, ValueError):
            break
    return out


def kr_value(ws, row: int) -> str:
    return clean_code(ws.cell(row=row, column=KR_COL).value)


def section_value(ws, row: int) -> str:
    value = ws.cell(row=row, column=SECTION_COL).value
    return "" if value is None else str(value).strip()


def recommended_value(ws, row: int) -> float | None:
    value = ws.cell(row=row, column=7).value
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def log_entries(wb) -> list[list[object]]:
    if "Price_Check_Log" not in wb.sheetnames:
        return []
    ws = wb["Price_Check_Log"]
    rows: list[list[object]] = []
    for row in range(2, ws.max_row + 1):
        entry = [ws.cell(row=row, column=col).value for col in range(1, 8)]
        if any(entry):
            rows.append(entry)
    return rows


def main() -> None:
    wam = next(p for p in REAL.iterdir() if "6356153" in p.name and "WAM" in p.name.upper())
    waw = next(p for p in REAL.iterdir() if "6356153" in p.name and "WAW" in p.name.upper())

    wb_m = load_workbook(wam, data_only=True)
    wb_w = load_workbook(waw, data_only=True)
    sheet = wb_w.sheetnames[0]
    ws_m = wb_m[sheet]
    ws_w = wb_w[sheet]

    rows_m = gesn_rows(ws_m)
    rows_w = gesn_rows(ws_w)
    shared = sorted(set(rows_m) & set(rows_w))

    with_prices_m = sum(1 for row in shared if analog_prices(ws_m, row))
    with_prices_w = sum(1 for row in shared if analog_prices(ws_w, row))

    presence_mismatch: list[str] = []
    price_mismatch: list[str] = []
    kr_mismatch: list[str] = []
    section_mismatch: list[str] = []
    rec_mismatch: list[str] = []

    for row in shared:
        code = ws_m.cell(row=row, column=CODE_COL).value
        pm = analog_prices(ws_m, row)
        pw = analog_prices(ws_w, row)
        if (len(pm) > 0) != (len(pw) > 0):
            presence_mismatch.append(
                f"row {row} {code}: macro {len(pm)} web {len(pw)}"
            )
        elif pm and pw and pm != pw:
            price_mismatch.append(f"row {row} {code}: macro {pm[:6]} web {pw[:6]}")
        km, kw = kr_value(ws_m, row), kr_value(ws_w, row)
        if km != kw:
            kr_mismatch.append(f"row {row} {code}: macro {km!r} web {kw!r}")
        sm, sw = section_value(ws_m, row), section_value(ws_w, row)
        if sm != sw:
            section_mismatch.append(f"row {row} {code}: macro {sm!r} web {sw!r}")
        rm, rw = recommended_value(ws_m, row), recommended_value(ws_w, row)
        if rm is not None and rw is not None and abs(rm - rw) > 0.51:
            rec_mismatch.append(f"row {row} {code}: macro {rm} web {rw}")

    log_m = log_entries(wb_m)
    log_w = log_entries(wb_w)

    lines = [
        f"Macro: {wam.name}",
        f"Web:   {waw.name}",
        "",
        f"GESN rows macro/web: {len(rows_m)}/{len(rows_w)}",
        f"shared rows: {len(shared)}",
        f"rows with analog prices macro/web: {with_prices_m}/{with_prices_w}",
        f"price presence mismatches: {len(presence_mismatch)}",
        f"price value mismatches: {len(price_mismatch)}",
        f"/KR mismatches: {len(kr_mismatch)}",
        f"section mismatches: {len(section_mismatch)}",
        f"recommended price mismatches (>0.5): {len(rec_mismatch)}",
        f"Price_Check_Log macro/web: {len(log_m)}/{len(log_w)}",
        "",
    ]

    for title, items in [
        ("Presence", presence_mismatch),
        ("Prices", price_mismatch),
        ("/KR", kr_mismatch),
        ("Section", section_mismatch),
        ("Recommended", rec_mismatch),
    ]:
        if items:
            lines.append(f"{title} examples:")
            lines.extend(f"  {item}" for item in items[:20])
            if len(items) > 20:
                lines.append(f"  ... +{len(items) - 20} more")
            lines.append("")

    if log_m or log_w:
        lines.append("Price_Check_Log macro:")
        for entry in log_m[:15]:
            lines.append(f"  {entry}")
        lines.append("Price_Check_Log web:")
        for entry in log_w[:15]:
            lines.append(f"  {entry}")

    text = "\n".join(lines)
    OUT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()

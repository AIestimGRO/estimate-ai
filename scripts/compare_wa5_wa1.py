"""Compare WA5 web output vs WA1 macro reference."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

REAL = Path("data/real")
OUT = REAL / "_compare_wa5_wa1.txt"


def clean_code(value: object) -> str:
    text = re.sub(r"\s*/\s*\u041a\u0420\s*$", "", str(value or "").strip(), flags=re.I)
    return text.casefold()


def data_rows(ws, start: int = 41, code_col: int = 14, unit_col: int = 4, base_col: int = 6) -> list[int]:
    rows: list[int] = []
    blank = 0
    for row in range(start, ws.max_row + 1):
        code = ws.cell(row=row, column=code_col).value
        unit = ws.cell(row=row, column=unit_col).value
        base = ws.cell(row=row, column=base_col).value
        if not code or not unit or not base:
            blank += 1
            if blank >= 5:
                break
            continue
        blank = 0
        if str(code).strip().isdigit():
            continue
        try:
            if float(base) <= 0:
                continue
        except (TypeError, ValueError):
            continue
        rows.append(row)
    return rows


def analog_prices(ws, row: int, start: int = 17) -> list[float]:
    out: list[float] = []
    for col in range(start, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        try:
            out.append(float(value))
        except (TypeError, ValueError):
            break
    return out


def is_red(ws, row: int, max_col: int = 20) -> bool:
    for col in range(1, max_col + 1):
        fill = ws.cell(row=row, column=col).fill
        if not fill or not fill.fgColor or not fill.fgColor.rgb:
            continue
        rgb = str(fill.fgColor.rgb)
        if rgb.endswith("FF0000") or rgb == "FFFF0000":
            return True
    return False


def main() -> None:
    wa1 = next(p for p in REAL.glob("*.xlsx") if "WA1" in p.name.upper())
    wa5 = next(p for p in REAL.glob("*.xlsx") if "WA5" in p.name.upper())

    wb1 = load_workbook(wa1, data_only=False)
    wb5 = load_workbook(wa5, data_only=False)
    sheet = wb5.sheetnames[0]
    ws1 = wb1[sheet]
    ws5 = wb5[sheet]

    rows1 = data_rows(ws1)
    rows5 = data_rows(ws5)
    shared = sorted(set(rows1) & set(rows5))

    with_prices = lambda ws, r: len(analog_prices(ws, r)) > 0
    count1 = sum(1 for r in shared if with_prices(ws1, r))
    count5 = sum(1 for r in shared if with_prices(ws5, r))

    red1 = {r for r in shared if is_red(ws1, r)}
    red5 = {r for r in shared if is_red(ws5, r)}

    mismatches: list[tuple[int, object, int, int]] = []
    for row in shared:
        p1 = analog_prices(ws1, row)
        p5 = analog_prices(ws5, row)
        if (len(p1) > 0) != (len(p5) > 0):
            mismatches.append((row, ws1.cell(row=row, column=14).value, len(p1), len(p5)))
        elif p1 and p5 and abs(p1[0] - p5[0]) > 0.5:
            mismatches.append((row, ws1.cell(row=row, column=14).value, len(p1), len(p5)))

    lines = [
        f"Macro: {wa1.name}",
        f"Web:   {wa5.name}",
        "",
        f"shared rows: {len(shared)}",
        f"rows with analogs macro/web: {count1}/{count5}",
        f"red rows macro/web: {len(red1)}/{len(red5)}",
        f"both red: {len(red1 & red5)}",
        f"macro only red: {sorted(red1 - red5)}",
        f"web only red: {sorted(red5 - red1)}",
        "",
    ]
    for row in (101, 211):
        lines.append(
            f"row {row}: code={ws5.cell(row=row, column=14).value!r} "
            f"macro_prices={len(analog_prices(ws1, row))} "
            f"web_prices={len(analog_prices(ws5, row))} "
            f"/KR={ws5.cell(row=row, column=15).value!r}"
        )
    lines.append("")
    lines.append(f"price presence mismatches: {len(mismatches)}")
    for item in mismatches[:20]:
        lines.append(f"  {item}")

    text = "\n".join(lines)
    OUT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()

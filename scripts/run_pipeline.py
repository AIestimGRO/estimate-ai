"""End-to-end pipeline run over the real files in data/real/ (validation).

Auto-detects which file is the catalog (sheet name contains 'Katal') and which
is the estimate, runs run_and_write, and writes a UTF-8 summary. Throwaway
diagnostic; the produced workbook goes to data/real/_out.xlsx (gitignored).
"""

from pathlib import Path

from openpyxl import load_workbook

from app.services.write_result import run_and_write

REAL_DIR = Path("data/real")


def main() -> None:
    files = sorted(p for p in REAL_DIR.glob("*.xlsx") if not p.name.startswith("_"))
    catalog = None
    estimate = None
    for path in files:
        workbook = load_workbook(path, read_only=True)
        names = workbook.sheetnames
        workbook.close()
        if any("\u041a\u0430\u0442\u0430\u043b" in name for name in names):
            catalog = path
        else:
            estimate = path

    lines: list[str] = []
    lines.append(f"catalog:  {catalog.name if catalog else None}")
    lines.append(f"estimate: {estimate.name if estimate else None}")

    if catalog is None or estimate is None:
        lines.append("ERROR: could not identify both catalog and estimate")
        (REAL_DIR / "_run.txt").write_text("\n".join(lines), encoding="utf-8")
        print("missing files")
        return

    output = REAL_DIR / "_out.xlsx"
    outcome = run_and_write(catalog, estimate, output)
    result = outcome.result

    lines.append("")
    lines.append(f"sheet:                {outcome.sheet_title}")
    lines.append(f"read method:          {outcome.read_method}")
    lines.append(f"coefficient:          {outcome.regional_coefficient} ({outcome.coefficient_method})")
    lines.append(f"catalog keys:         {result.catalog_key_count}")
    lines.append(f"estimate rows:        {len(result.rows)}")
    lines.append(f"rows with analogs:    {result.matched_row_count}")
    lines.append(f"rows flagged (risk):  {result.flagged_row_count}")
    lines.append(f"written rows:         {outcome.write_report.written_rows}")
    lines.append(f"avg column:           {outcome.write_report.average_column} "
                 f"(inserted={outcome.write_report.inserted_average_column})")
    lines.append(f"analog start column:  {outcome.write_report.analog_start_column}")
    lines.append(f"risk log rows:        {outcome.write_report.risk_log_rows}")
    lines.append(f"output:               {outcome.output_path}")

    lines.append("")
    lines.append("first rows with analogs:")
    shown = 0
    for row in result.rows:
        if not row.has_analogs:
            continue
        codes = [f"{a.entry.price:.2f}" for a in row.analogs[:6]]
        lines.append(
            f"  [{row.row_index}] {str(row.estimate_row.code)[:24]:24} "
            f"base={row.estimate_row.base_price} rec={row.recommended_price} "
            f"analogs({len(row.analogs)})={codes}"
        )
        shown += 1
        if shown >= 12:
            break

    (REAL_DIR / "_run.txt").write_text("\n".join(lines), encoding="utf-8")
    print("done; wrote data/real/_run.txt")


if __name__ == "__main__":
    main()

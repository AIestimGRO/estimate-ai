"""Compare 6356153 WAM (macro) vs WAW (web+Excel) vs WAWS (web+DB)."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

REAL = Path("data/real")
OUT = REAL / "_compare_6356153_wam_waw_waws.txt"

CODE_COL = 14
UNIT_COL = 4
BASE_COL = 6
REC_COL = 7
KR_COL = 15
SECTION_COL = 16
ANALOG_START = 17
DATA_START = 29


def clean_code(value: object) -> str:
    text = re.sub(r"\s*/\s*\u041a\u0420\s*$", "", str(value or "").strip(), flags=re.I)
    return text.casefold()


def find_file(tag: str) -> Path:
    tag_upper = tag.upper()
    for path in REAL.iterdir():
        if "6356153" not in path.name or path.suffix.lower() != ".xlsx":
            continue
        if path.name.startswith("_"):
            continue
        name_upper = path.name.upper()
        if tag_upper == "WAM" and name_upper.endswith("WAM.XLSX"):
            return path
        if tag_upper == "WAW" and name_upper.endswith("WAW.XLSX"):
            return path
        if tag_upper == "WAWS" and name_upper.endswith("WAWS.XLSX"):
            return path
    raise FileNotFoundError(f"6356153 {tag} file not found under {REAL}")


def gesn_rows(ws) -> list[int]:
    rows: list[int] = []
    blank = 0
    for row in range(DATA_START, ws.max_row + 1):
        unit = ws.cell(row=row, column=UNIT_COL).value
        base = ws.cell(row=row, column=BASE_COL).value
        code = ws.cell(row=row, column=CODE_COL).value
        if not unit or not base or not code:
            blank += 1
            if blank >= 8:
                break
            continue
        blank = 0
        code_text = str(code).strip()
        if not code_text.upper().startswith("\u0413\u042d\u0421\u041d"):
            continue
        try:
            if float(base) <= 0:
                continue
        except (TypeError, ValueError):
            continue
        rows.append(row)
    return rows


def analog_prices(ws, row: int) -> list[float]:
    out: list[float] = []
    for col in range(ANALOG_START, ws.max_column + 1):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        try:
            out.append(round(float(value), 2))
        except (TypeError, ValueError):
            break
    return out


def kr_value(ws, row: int) -> str:
    return clean_code(ws.cell(row=row, column=KR_COL).value)


def section_value(ws, row: int) -> str:
    value = ws.cell(row=row, column=SECTION_COL).value
    return "" if value is None else str(value).strip()


def recommended_value(ws, row: int) -> float | None:
    value = ws.cell(row=row, column=REC_COL).value
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return None


def log_entries(wb) -> list[list[object]]:
    if "Price_Check_Log" not in wb.sheetnames:
        return []
    ws = wb["Price_Check_Log"]
    rows: list[list[object]] = []
    for row in range(2, ws.max_row + 1):
        entry = [ws.cell(row=row, column=col).value for col in range(1, 8)]
        if any(entry):
            rows.append(entry)
    return rows


def compare_pair(
    label: str,
    ws_a,
    ws_b,
    shared: list[int],
) -> dict[str, object]:
    presence: list[str] = []
    prices: list[str] = []
    kr: list[str] = []
    section: list[str] = []
    rec: list[str] = []

    with_a = sum(1 for row in shared if analog_prices(ws_a, row))
    with_b = sum(1 for row in shared if analog_prices(ws_b, row))

    for row in shared:
        code = ws_a.cell(row=row, column=CODE_COL).value
        pa = analog_prices(ws_a, row)
        pb = analog_prices(ws_b, row)
        if (len(pa) > 0) != (len(pb) > 0):
            presence.append(f"row {row} {code}: A {len(pa)} B {len(pb)}")
        elif pa and pb and pa != pb:
            prices.append(f"row {row} {code}: A {pa[:6]} B {pb[:6]}")
        ka, kb = kr_value(ws_a, row), kr_value(ws_b, row)
        if ka != kb:
            kr.append(f"row {row} {code}: A {ka!r} B {kb!r}")
        sa, sb = section_value(ws_a, row), section_value(ws_b, row)
        if sa != sb:
            section.append(f"row {row} {code}: A {sa!r} B {sb!r}")
        ra, rb = recommended_value(ws_a, row), recommended_value(ws_b, row)
        if ra is not None and rb is not None and abs(ra - rb) > 0.51:
            rec.append(f"row {row} {code}: A {ra} B {rb}")
        elif (ra is None) != (rb is None) and (pa or pb):
            rec.append(f"row {row} {code}: A rec={ra} B rec={rb}")

    return {
        "label": label,
        "with_a": with_a,
        "with_b": with_b,
        "presence": presence,
        "prices": prices,
        "kr": kr,
        "section": section,
        "rec": rec,
    }


def format_pair(result: dict[str, object]) -> list[str]:
    lines = [
        f"=== {result['label']} ===",
        f"rows with analogs: {result['with_a']}/{result['with_b']}",
        f"presence: {len(result['presence'])}",
        f"price values: {len(result['prices'])}",
        f"/KR: {len(result['kr'])}",
        f"section: {len(result['section'])}",
        f"recommended: {len(result['rec'])}",
        "",
    ]
    for title, key in [
        ("Presence", "presence"),
        ("Prices", "prices"),
        ("/KR", "kr"),
        ("Section", "section"),
        ("Recommended", "rec"),
    ]:
        items = result[key]
        if items:
            lines.append(f"{title} examples:")
            lines.extend(f"  {item}" for item in items[:15])
            if len(items) > 15:
                lines.append(f"  ... +{len(items) - 15} more")
            lines.append("")
    return lines


def main() -> None:
    wam_path = find_file("WAM")
    waw_path = find_file("WAW")
    waws_path = find_file("WAWS")

    wb_m = load_workbook(wam_path, data_only=True)
    wb_w = load_workbook(waw_path, data_only=True)
    wb_s = load_workbook(waws_path, data_only=True)
    sheet = wb_w.sheetnames[0]
    ws_m = wb_m[sheet]
    ws_w = wb_w[sheet]
    ws_s = wb_s[sheet]

    rows_m = gesn_rows(ws_m)
    rows_w = gesn_rows(ws_w)
    rows_s = gesn_rows(ws_s)
    shared_all = sorted(set(rows_m) & set(rows_w) & set(rows_s))

    lines = [
        f"Macro:     {wam_path.name}",
        f"Web Excel: {waw_path.name}",
        f"Web DB:    {waws_path.name}",
        "",
        f"GESN rows macro/web-excel/web-db: {len(rows_m)}/{len(rows_w)}/{len(rows_s)}",
        f"shared rows (all three): {len(shared_all)}",
        "",
    ]

    for result in [
        compare_pair("WAM vs WAW (macro vs web+Excel)", ws_m, ws_w, shared_all),
        compare_pair("WAM vs WAWS (macro vs web+DB)", ws_m, ws_s, shared_all),
        compare_pair("WAW vs WAWS (web+Excel vs web+DB)", ws_w, ws_s, shared_all),
    ]:
        lines.extend(format_pair(result))

    log_m = log_entries(wb_m)
    log_w = log_entries(wb_w)
    log_s = log_entries(wb_s)
    lines.extend([
        f"Price_Check_Log: macro={len(log_m)} web-excel={len(log_w)} web-db={len(log_s)}",
        "",
    ])
    if log_w or log_s:
        lines.append("Price_Check_Log web-excel:")
        for entry in log_w[:10]:
            lines.append(f"  {entry}")
        lines.append("Price_Check_Log web-db:")
        for entry in log_s[:10]:
            lines.append(f"  {entry}")

    text = "\n".join(lines)
    OUT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()

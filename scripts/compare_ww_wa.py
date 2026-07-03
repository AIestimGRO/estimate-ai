"""Compare macro (WW) vs web (WA) estimate outputs in data/real/."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REAL_DIR = Path("data/real")
OUT = REAL_DIR / "_compare_ww_wa.txt"

# Template-ish columns from Settings / typical eV-grup layout
COL_WORK = 3
COL_UNIT = 4
COL_BASE = 6
COL_AVG = 7
COL_CODE = 14
COL_SECTION = 15
COL_ANALOG_START = 16
MAX_ANALOG_COLS = 20


@dataclass
class RowSnapshot:
    sheet_row: int
    code: str
    unit: str
    base: object
    avg: object
    analogs: list[object]
    section: object


def _find_files() -> tuple[Path, Path, Path | None]:
    files = sorted(REAL_DIR.glob("*.xlsx"))
    ww = wa = source = None
    for path in files:
        name = path.name.casefold()
        if name.startswith("_"):
            continue
        if " ww" in name or name.endswith("ww.xlsx"):
            ww = path
        elif " wa" in name or name.endswith("wa.xlsx"):
            wa = path
        elif "\u0441\u043c-\u043a" in name.casefold() and "ww" not in name and "wa" not in name:
            source = path
    if ww is None or wa is None:
        raise SystemExit("Could not find WW and WA files in data/real/")
    return ww, wa, source


def _estimate_sheet(workbook) -> Any:
    for worksheet in workbook.worksheets:
        title = worksheet.title.casefold()
        if "\u043e\u0441" in title or "\u043b\u0441\u0440" in title:
            return worksheet
    return workbook.worksheets[0]


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _num(value: object) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_data_row(worksheet, row: int) -> bool:
    code = _norm(worksheet.cell(row=row, column=COL_CODE).value)
    unit = _norm(worksheet.cell(row=row, column=COL_UNIT).value)
    base = _num(worksheet.cell(row=row, column=COL_BASE).value)
    if not code or not unit or base is None or base <= 0:
        return False
    if code.isdigit():
        return False
    return True


def _header_row(worksheet) -> int:
    for row in range(1, 40):
        text = _norm(worksheet.cell(row=row, column=COL_CODE).value).casefold()
        if "\u0433\u044d\u0441\u043d" in text or "\u043f\u0435\u0440\u0435\u0447\u0435\u043d\u044c" in text:
            return row
    return 0


def _read_rows(path: Path) -> tuple[str, int, list[RowSnapshot]]:
    workbook = load_workbook(path, data_only=True)
    try:
        worksheet = _estimate_sheet(workbook)
        header = _header_row(worksheet)
        start = header + 2 if header else 4
        rows: list[RowSnapshot] = []
        blank_run = 0
        for row in range(start, worksheet.max_row + 1):
            if not _is_data_row(worksheet, row):
                blank_run += 1
                if blank_run >= 5:
                    break
                continue
            blank_run = 0
            analogs = []
            for col in range(COL_ANALOG_START, COL_ANALOG_START + MAX_ANALOG_COLS):
                value = worksheet.cell(row=row, column=col).value
                if value is None or _norm(value) == "":
                    break
                analogs.append(value)
            rows.append(
                RowSnapshot(
                    sheet_row=row,
                    code=_norm(worksheet.cell(row=row, column=COL_CODE).value),
                    unit=_norm(worksheet.cell(row=row, column=COL_UNIT).value),
                    base=worksheet.cell(row=row, column=COL_BASE).value,
                    avg=worksheet.cell(row=row, column=COL_AVG).value,
                    analogs=analogs,
                    section=worksheet.cell(row=row, column=COL_SECTION).value,
                )
            )
        return worksheet.title, header, rows
    finally:
        workbook.close()


def _avg_number(avg: object, base: object) -> float | None:
    if isinstance(avg, (int, float)):
        return float(avg)
    if isinstance(avg, str) and avg.startswith("="):
        return _num(base)
    return _num(avg)


def main() -> None:
    ww_path, wa_path, source_path = _find_files()
    lines: list[str] = []
    lines.append(f"WW (macro):  {ww_path.name}")
    lines.append(f"WA (web):    {wa_path.name}")
    if source_path:
        lines.append(f"Source:      {source_path.name}")
    lines.append("")

    ww_title, ww_header, ww_rows = _read_rows(ww_path)
    wa_title, wa_header, wa_rows = _read_rows(wa_path)
    lines.append(f"WW sheet={ww_title!r} header_row={ww_header} data_rows={len(ww_rows)}")
    lines.append(f"WA sheet={wa_title!r} header_row={wa_header} data_rows={len(wa_rows)}")
    lines.append("")

    ww_by_code = {(r.code.casefold(), r.unit.casefold()): r for r in ww_rows}
    wa_by_code = {(r.code.casefold(), r.unit.casefold()): r for r in wa_rows}

    ww_keys = set(ww_by_code)
    wa_keys = set(wa_by_code)
    only_ww = sorted(ww_keys - wa_keys)
    only_wa = sorted(wa_keys - ww_keys)
    common = sorted(ww_keys & wa_keys)

    lines.append(f"rows only in WW: {len(only_ww)}")
    for key in only_ww[:15]:
        r = ww_by_code[key]
        lines.append(f"  WW row {r.sheet_row}: {r.code} | {r.unit}")
    if len(only_ww) > 15:
        lines.append(f"  ... +{len(only_ww) - 15} more")

    lines.append(f"rows only in WA: {len(only_wa)}")
    for key in only_wa[:15]:
        r = wa_by_code[key]
        lines.append(f"  WA row {r.sheet_row}: {r.code} | {r.unit}")
    if len(only_wa) > 15:
        lines.append(f"  ... +{len(only_wa) - 15} more")

    avg_diff = 0
    analog_count_diff = 0
    kr_diff = 0
    base_diff = 0
    examples: list[str] = []

    for key in common:
        ww = ww_by_code[key]
        wa = wa_by_code[key]
        ww_code_clean = ww.code.replace(" /КР", "").replace(" /KR", "").casefold()
        wa_code_clean = wa.code.replace(" /КР", "").replace(" /KR", "").casefold()
        if ww_code_clean != wa_code_clean or ("/кр" in ww.code.casefold()) != ("/кр" in wa.code.casefold()):
            kr_diff += 1

        ww_base = _num(ww.base)
        wa_base = _num(wa.base)
        if ww_base is not None and wa_base is not None and abs(ww_base - wa_base) > 0.01:
            base_diff += 1

        ww_avg = _avg_number(ww.avg, ww.base)
        wa_avg = _avg_number(wa.avg, wa.base)
        if ww_avg is not None and wa_avg is not None and abs(ww_avg - wa_avg) > 0.5:
            avg_diff += 1
            if len(examples) < 12:
                examples.append(
                    f"  [{ww.code}] avg WW={ww_avg:.2f} WA={wa_avg:.2f} "
                    f"analogs WW={len(ww.analogs)} WA={len(wa.analogs)}"
                )

        if len(ww.analogs) != len(wa.analogs):
            analog_count_diff += 1

    lines.append("")
    lines.append(f"common rows: {len(common)}")
    lines.append(f"base price differs: {base_diff}")
    lines.append(f"/KR marking differs: {kr_diff}")
    lines.append(f"average differs (>0.5): {avg_diff}")
    lines.append(f"analog count differs: {analog_count_diff}")
    if examples:
        lines.append("")
        lines.append("sample average differences:")
        lines.extend(examples)

    # Sheets present
    for label, path in (("WW", ww_path), ("WA", wa_path)):
        workbook = load_workbook(path, read_only=True)
        lines.append("")
        lines.append(f"{label} sheets: {workbook.sheetnames}")
        workbook.close()

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT} ({len(lines)} lines)")


if __name__ == "__main__":
    main()
